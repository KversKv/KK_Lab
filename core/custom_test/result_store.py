"""Custom Test result model and export helpers."""

from __future__ import annotations

import csv
import json
import os
import re
from collections import OrderedDict
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from log_config import get_logger

logger = get_logger(__name__)

_NUMERIC_DTYPES = {"int", "float", "number"}
_NON_PLOT_TOKENS = (
    "addr",
    "address",
    "reg",
    "gpio",
    "pin",
    "channel",
    "width",
    "len",
    "index",
    "total",
    "status",
    "instrument",
)


@dataclass
class ResultField:
    name: str
    unit: str = ""
    dtype: str = "text"
    precision: Optional[int] = None
    plot: bool = False
    export: bool = True


@dataclass
class ResultRow:
    values: Dict[str, Any]
    source_node_uid: Optional[str] = None
    timestamp: datetime = field(default_factory=datetime.now)
    status: str = "ok"

    def to_record(self) -> Dict[str, Any]:
        return dict(self.values)


@dataclass
class ResultViewState:
    visible_columns: Optional[List[str]] = None
    display_names: Dict[str, str] = field(default_factory=dict)
    order: List[str] = field(default_factory=list)
    formats: Dict[str, str] = field(default_factory=dict)
    hidden_columns: set[str] = field(default_factory=set)
    row_order: Optional[List[int]] = None

    def is_visible(self, field_name: str) -> bool:
        if field_name in self.hidden_columns:
            return False
        if self.visible_columns is None:
            return True
        return field_name in self.visible_columns

    def display_name(self, field_name: str) -> str:
        return self.display_names.get(field_name, field_name)


class ResultStore:
    """Canonical result rows plus display/export state."""

    def __init__(self) -> None:
        self.fields: "OrderedDict[str, ResultField]" = OrderedDict()
        self.rows: List[ResultRow] = []
        self.view_state = ResultViewState()
        self.created_at = datetime.now()
        self.updated_at: Optional[datetime] = None

    @property
    def records(self) -> List[Dict[str, Any]]:
        return [row.to_record() for row in self.rows]

    def clear(self) -> None:
        self.fields.clear()
        self.rows.clear()
        self.view_state = ResultViewState()
        self.created_at = datetime.now()
        self.updated_at = None

    def append(
        self,
        values: Dict[str, Any],
        *,
        source_node_uid: Optional[str] = None,
        timestamp: Optional[datetime] = None,
        status: str = "ok",
    ) -> ResultRow:
        record = dict(values)
        for name, value in record.items():
            self.register_field(name, value=value)
        row = ResultRow(
            values=record,
            source_node_uid=source_node_uid,
            timestamp=timestamp or datetime.now(),
            status=status,
        )
        self.rows.append(row)
        self.updated_at = row.timestamp
        self.view_state.row_order = None
        return row

    def register_field(
        self,
        name: str,
        *,
        value: Any = None,
        unit: str = "",
        dtype: Optional[str] = None,
        precision: Optional[int] = None,
        plot: Optional[bool] = None,
        export: Optional[bool] = None,
    ) -> ResultField:
        inferred_dtype = dtype or self._infer_dtype(value)
        inferred_precision = precision
        if inferred_precision is None and isinstance(value, float):
            inferred_precision = _decimal_places(value)
        if plot is None:
            plot = self._infer_plot(name, inferred_dtype)

        if name not in self.fields:
            field_obj = ResultField(
                name=name,
                unit=unit,
                dtype=inferred_dtype,
                precision=inferred_precision,
                plot=plot,
                export=True if export is None else export,
            )
            self.fields[name] = field_obj
            self.view_state.order.append(name)
            return field_obj

        field_obj = self.fields[name]
        field_obj.dtype = self._merge_dtype(field_obj.dtype, inferred_dtype)
        if unit:
            field_obj.unit = unit
        if inferred_precision is not None:
            if field_obj.precision is None:
                field_obj.precision = inferred_precision
            else:
                field_obj.precision = max(field_obj.precision, inferred_precision)
        if plot is not None:
            field_obj.plot = field_obj.plot or plot
        if export is not None:
            field_obj.export = export
        return field_obj

    def get_visible_fields(self) -> List[ResultField]:
        ordered_names = self._ordered_field_names(view=True)
        return [self.fields[name] for name in ordered_names]

    def set_display_name(self, field_name: str, display_name: str) -> None:
        if field_name in self.fields:
            self.view_state.display_names[field_name] = display_name.strip() or field_name

    def hide_field(self, field_name: str) -> None:
        if field_name in self.fields:
            self.view_state.hidden_columns.add(field_name)

    def set_field_format(self, field_name: str, fmt: str) -> None:
        if field_name in self.fields:
            self.view_state.formats[field_name] = fmt

    def set_field_order(self, ordered_names: Sequence[str]) -> None:
        seen = set()
        new_order: List[str] = []
        for name in ordered_names:
            if name in self.fields and name not in seen:
                new_order.append(name)
                seen.add(name)
        for name in self.fields:
            if name not in seen:
                new_order.append(name)
        self.view_state.order = new_order

    def sort_by(self, field_name: str, *, ascending: bool = True) -> None:
        if field_name not in self.fields:
            return
        indexed = list(enumerate(self.rows))

        def _sort_key(item: Tuple[int, ResultRow]) -> Tuple[int, Any]:
            value = item[1].values.get(field_name, "")
            numeric = _to_number(value)
            if numeric is not None:
                return (0, numeric)
            return (1, str(value))

        indexed.sort(key=_sort_key, reverse=not ascending)
        self.view_state.row_order = [idx for idx, _ in indexed]

    def view_table(self) -> Tuple[List[str], List[List[str]]]:
        fields = self.get_visible_fields()
        headers = [self.view_state.display_name(field_obj.name) for field_obj in fields]
        rows = []
        for row in self._iter_rows(view=True):
            rows.append([
                self.format_value(field_obj.name, row.values.get(field_obj.name, ""))
                for field_obj in fields
            ])
        return headers, rows

    def canonical_table(self) -> Tuple[List[str], List[List[Any]]]:
        names = self._ordered_field_names(view=False)
        headers = list(names)
        rows = []
        for row in self.rows:
            rows.append([row.values.get(name, "") for name in names])
        return headers, rows

    def plot_series(self) -> Dict[str, List[float]]:
        series: Dict[str, List[float]] = {}
        for field_obj in self.fields.values():
            if not field_obj.plot:
                continue
            values: List[float] = []
            for row in self.rows:
                numeric = _to_number(row.values.get(field_obj.name))
                if numeric is None:
                    continue
                values.append(float(numeric))
            if values:
                series[field_obj.name] = values
        return series

    def format_value(self, field_name: str, value: Any) -> str:
        if value is None:
            return ""
        fmt = self.view_state.formats.get(field_name, "auto")
        field_obj = self.fields.get(field_name)
        numeric = _to_number(value)

        if fmt == "hex":
            if numeric is not None:
                return f"0x{int(numeric):X}"
            return str(value)
        if fmt == "sci":
            if numeric is not None:
                return f"{numeric:.6e}"
            return str(value)
        if fmt not in ("", "auto"):
            try:
                dp = int(fmt)
            except (TypeError, ValueError):
                dp = None
            if dp is not None and numeric is not None:
                return f"{numeric:.{dp}f}"

        if isinstance(value, float):
            precision = field_obj.precision if field_obj and field_obj.precision is not None else 2
            precision = max(0, min(precision, 10))
            return f"{value:.{precision}f}"
        return str(value)

    def export_csv(
        self,
        file_path: str,
        *,
        view: bool = False,
        manifest: Optional[Dict[str, Any]] = None,
    ) -> str:
        headers, rows = self.view_table() if view else self.canonical_table()
        os.makedirs(os.path.dirname(os.path.abspath(file_path)), exist_ok=True)
        with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(rows)
        if manifest:
            manifest_path = f"{os.path.splitext(file_path)[0]}.manifest.json"
            with open(manifest_path, "w", encoding="utf-8") as f:
                json.dump(manifest, f, ensure_ascii=False, indent=2, default=str)
        return file_path

    def export_xlsx(
        self,
        file_path: str,
        *,
        view: bool = False,
        manifest: Optional[Dict[str, Any]] = None,
    ) -> str:
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            raise RuntimeError("缺少 openpyxl 依赖，无法导出 XLSX。") from exc

        headers, rows = self.view_table() if view else self.canonical_table()
        os.makedirs(os.path.dirname(os.path.abspath(file_path)), exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="2F4070")
        header_align = Alignment(horizontal="center", vertical="center")
        cell_align = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="B4C7E7")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        even_fill = PatternFill("solid", fgColor="F2F6FC")

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        for row_idx, row_values in enumerate(rows, start=2):
            for col_idx, value in enumerate(row_values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_align
                cell.border = border
                if isinstance(value, (int, float)):
                    cell.number_format = "0.######"
                if row_idx % 2 == 0:
                    cell.fill = even_fill

        for col_idx, header in enumerate(headers, start=1):
            max_len = len(str(header))
            for row_values in rows:
                cell_text = row_values[col_idx - 1] if col_idx - 1 < len(row_values) else ""
                max_len = max(max_len, len(str(cell_text)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 4, 10), 40)
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

        if manifest:
            manifest_ws = wb.create_sheet("Manifest")
            manifest_ws.append(["key", "value"])
            for key, value in manifest.items():
                manifest_ws.append([key, json.dumps(value, ensure_ascii=False, default=str)])

        wb.save(file_path)
        return file_path

    def export(
        self,
        file_path: str,
        *,
        fmt: Optional[str] = None,
        view: bool = False,
        manifest: Optional[Dict[str, Any]] = None,
    ) -> str:
        fmt_name = (fmt or os.path.splitext(file_path)[1].lstrip(".") or "csv").lower()
        if fmt_name in ("xlsx", "excel"):
            if not file_path.lower().endswith(".xlsx"):
                file_path = f"{file_path}.xlsx"
            return self.export_xlsx(file_path, view=view, manifest=manifest)
        if fmt_name != "csv":
            raise ValueError(f"不支持的导出格式: {fmt_name}")
        if not file_path.lower().endswith(".csv"):
            file_path = f"{file_path}.csv"
        return self.export_csv(file_path, view=view, manifest=manifest)

    def build_manifest(
        self,
        *,
        sequence_hash: str = "",
        instrument_snapshot: Optional[Any] = None,
        started_at: Optional[datetime] = None,
        finished_at: Optional[datetime] = None,
    ) -> Dict[str, Any]:
        return {
            "sequence_hash": sequence_hash,
            "start_time": started_at or self.created_at,
            "end_time": finished_at or self.updated_at or datetime.now(),
            "instrument_snapshot": instrument_snapshot,
            "row_count": len(self.rows),
            "fields": [field_obj.__dict__ for field_obj in self.fields.values()],
        }

    def _ordered_field_names(self, *, view: bool) -> List[str]:
        ordered = [name for name in self.view_state.order if name in self.fields]
        for name in self.fields:
            if name not in ordered:
                ordered.append(name)
        if not view:
            return [name for name in ordered if self.fields[name].export]
        return [name for name in ordered if self.fields[name].export and self.view_state.is_visible(name)]

    def _iter_rows(self, *, view: bool) -> Iterable[ResultRow]:
        if not view or self.view_state.row_order is None:
            return iter(self.rows)
        ordered_rows = [
            self.rows[index]
            for index in self.view_state.row_order
            if 0 <= index < len(self.rows)
        ]
        return iter(ordered_rows)

    @staticmethod
    def _infer_dtype(value: Any) -> str:
        if isinstance(value, bool):
            return "bool"
        if isinstance(value, int) and not isinstance(value, bool):
            return "int"
        if isinstance(value, float):
            return "float"
        if isinstance(value, str) and re.fullmatch(r"0[xX][0-9a-fA-F]+", value.strip()):
            return "hex"
        if value is None:
            return "text"
        return "text"

    @staticmethod
    def _merge_dtype(old: str, new: str) -> str:
        if old == new:
            return old
        if old in _NUMERIC_DTYPES and new in _NUMERIC_DTYPES:
            return "number"
        if old == "text":
            return new
        if new == "text":
            return old
        return "text"

    @staticmethod
    def _infer_plot(name: str, dtype: str) -> bool:
        if dtype not in _NUMERIC_DTYPES:
            return False
        lower = name.lower()
        return not any(token in lower for token in _NON_PLOT_TOKENS)


def build_default_result_path(
    project_root: str,
    *,
    chip_or_profile: str = "default",
    timestamp: Optional[datetime] = None,
    fmt: str = "csv",
) -> str:
    ts = (timestamp or datetime.now()).strftime("%Y%m%d_%H%M%S")
    profile = _safe_filename_part(chip_or_profile or "default")
    extension = "xlsx" if fmt.lower() in ("xlsx", "excel") else "csv"
    output_dir = os.path.join(project_root, "Results", "custom_test", ts)
    return os.path.join(output_dir, f"custom_test_{profile}_{ts}.{extension}")


def _decimal_places(value: float) -> int:
    text = f"{value:.12g}"
    if "." not in text:
        return 0
    return max(2, min(len(text.split(".")[1].rstrip("0")) or 0, 10))


def _safe_filename_part(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_.-]+", "_", value.strip())
    return cleaned.strip("._") or "default"


def _to_number(value: Any) -> Optional[float]:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if re.fullmatch(r"0[xX][0-9a-fA-F]+", text):
            return None
        try:
            return float(text)
        except ValueError:
            return None
    return None
