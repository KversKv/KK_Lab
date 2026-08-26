"""Module Test 结果导出 XLSX（数据表 + 波形/截图嵌入单元格）。

- ``export_result_xlsx(result, out_path)``：纯函数（无 Qt），Summary sheet
  （元信息 + 仪器 + 逐项一览）+ 每个测试项一个数据 sheet（原始 CSV 表格
  + 末列 Scope Shot 截图锚定；截图按 Iload/组号数值匹配数据行，未匹配的
  逐张锚在数据区下方，每行一张）。
- ``XlsxExportWorker``：QObject worker（仅 QtCore），UI 经 QThread +
  moveToThread 后台调用（图片读取 + 写 zip 随截图数增长，不阻塞主线程），
  线程编排模式与 ``ModuleConfigWorker`` 一致。

依赖：openpyxl + Pillow（openpyxl 嵌入图片必需），缺失时报
RuntimeError 交由 UI 层提示，不静默丢图。
"""
from __future__ import annotations

import csv
import os
import re
from typing import Any, Callable

from PySide6.QtCore import QObject, Signal

from log_config import get_logger

logger = get_logger(__name__)

# Excel sheet 名限制：≤31 字符，禁 [ ] : * ? / \
_SHEET_BAD_CHARS = re.compile(r"[\[\]:*?/\\]+")
# 嵌入单元格的行内缩略图宽（像素），下方独立大图稍宽
_THUMB_W = 260
_FULL_W = 420
_PX_TO_PT = 0.75  # Excel 行高单位为磅；96dpi 像素 → 磅


def _sheet_title(name: str, used: set[str]) -> str:
    """清洗 item 名为合法且唯一的 sheet 名（≤31 字符，非法字符直接删除）。"""
    cleaned = _SHEET_BAD_CHARS.sub("", str(name or "item")).strip() or "item"
    cleaned = cleaned[:31].rstrip()
    title, n = cleaned, 2
    while title in used:
        suffix = f" ({n})"
        title = cleaned[:31 - len(suffix)] + suffix
        n += 1
    used.add(title)
    return title


def _csv_rows(path: str | None) -> list[list[str]]:
    """读取原始数据 CSV（utf-8-sig，与 report._csv_to_rows 同语义）。"""
    if not path or not os.path.isfile(path):
        return []
    try:
        with open(path, "r", newline="", encoding="utf-8-sig") as f:
            return [row for row in csv.reader(f) if row]
    except OSError:
        logger.warning("XLSX 导出读取 CSV 失败: %s", path, exc_info=True)
        return []


def _cell_value(raw: str) -> Any:
    """CSV 字符串宽松转数值：纯 int/float 保留数值类型，其余原样。

    前导零（如 "007"）保留字符串避免丢信息。
    """
    s = str(raw).strip()
    if not s:
        return ""
    if s.isdigit():
        return int(s)
    if s[0] in "+-" and s[1:].isdigit():
        return int(s)
    if len(s) > 1 and s[0] == "0" and s[1].isdigit():
        return s  # 前导零整数（如 "007"）保留字符串，避免丢信息
    try:
        return float(s)
    except ValueError:
        return s


def _to_float(value: Any) -> float | None:
    try:
        if value is None or str(value).strip() == "":
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def _collect_shots(item) -> list[tuple[str, str]]:
    """收集截图 [(匹配键, png路径)]：measured["screenshots"] 优先，单波形图兜底。"""
    measured = item.measured if isinstance(item.measured, dict) else {}
    raw = measured.get("screenshots")
    shots: list[tuple[str, str]] = []
    if isinstance(raw, list):
        for s in raw:
            if isinstance(s, dict) and s.get("png"):
                shots.append((str(s.get("Iload (mA)", "")), str(s["png"])))
    if not shots and item.waveform_png:
        shots.append(("", str(item.waveform_png)))
    return [(k, p) for k, p in shots if os.path.isfile(p)]


def _scaled_img(img: Any, target_w: int) -> tuple[Any, int]:
    """openpyxl Image 按目标宽等比缩放，返回 (img, 显示高像素)。"""
    ratio = target_w / float(img.width or 1)
    img.width = target_w
    img.height = max(1, int(float(img.height or 1) * ratio))
    return img, int(img.height)


def _write_summary_sheet(ws, result) -> None:  # noqa: ANN001
    """Summary sheet：元信息 + 仪器 + 逐项一览（含数据 sheet 名导航列）。"""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    summary = result.build_summary()
    verdict = summary.get("overall", "N/A")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F4070")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="B4C7E7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    verdict_font = {
        "PASS": Font(bold=True, color="1A7F37"),
        "FAIL": Font(bold=True, color="CF222E"),
        "N/A": Font(color="6E7781"),
    }.get(verdict, Font())

    ws["A1"] = f"Module Test Report — {result.module_type.upper()}"
    ws["A1"].font = Font(bold=True, size=14)
    meta = [
        ("模块", result.module_type.upper()),
        ("芯片", result.chip_name or "—"),
        ("模块名称", result.module_name or "—"),
        ("操作员", result.operator or "—"),
        ("温度 (°C)", result.temperature or "常温"),
        ("开始时间", result.started_at or "—"),
        ("结束时间", result.finished_at or "—"),
        ("结论", verdict),
    ]
    for i, (k, v) in enumerate(meta, start=2):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        cell = ws.cell(row=i, column=2, value=v)
        if k == "结论":
            cell.font = verdict_font

    instruments = list(getattr(result, "instruments", []) or [])
    r = len(meta) + 3
    if instruments:
        ws.cell(row=r, column=1, value="仪器").font = Font(bold=True)
        r += 1
        for c, h in enumerate(("名称", "型号", "SN"), start=1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        for inst in instruments:
            r += 1
            for c, k in enumerate(("name", "model", "sn"), start=1):
                cell = ws.cell(row=r, column=c, value=str(inst.get(k, "") or ""))
                cell.border = border
        r += 2

    ws.cell(row=r, column=1, value="测试项一览").font = Font(bold=True)
    r += 1
    headers = ("#", "测试项", "判定", "备注", "完成时间", "数据 Sheet")
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    for idx, (it, sheet) in enumerate(
            zip(result.items, _item_sheet_names(result)), start=1):
        r += 1
        v = "PASS" if it.passed is True else (
            "FAIL" if it.passed is False else "N/A")
        values = (idx, it.name, v, it.notes or "", it.ts or "", sheet)
        for c, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            if c == 3:
                cell.font = {
                    "PASS": Font(bold=True, color="1A7F37"),
                    "FAIL": Font(bold=True, color="CF222E"),
                }.get(v, Font())
            elif c == 1:
                cell.alignment = center
    for col, w in zip("ABCDEF", (6, 34, 10, 46, 20, 30)):
        ws.column_dimensions[col].width = w


def _item_sheet_names(result) -> list[str]:  # noqa: ANN001
    """预先生成与 result.items 同序的 sheet 名（Summary 导航列复用）。"""
    used: set[str] = {"Summary"}
    return [_sheet_title(it.name, used) for it in result.items]


def _write_item_sheet(ws, item, log_fn: Callable[[str], None]) -> None:  # noqa: ANN001
    """单个测试项数据 sheet：CSV 表格 + 截图（行内匹配 / 下方堆叠）。"""
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F4070")
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    thin = Side(style="thin", color="B4C7E7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    rows = _csv_rows(item.raw_csv_path)
    start = 1  # 数据表起始行（无 CSV 时兜底 measured，也直接从第 1 行写）
    has_table = bool(rows) and len(rows) >= 2
    if not has_table and isinstance(item.measured, list) and item.measured \
            and isinstance(item.measured[0], dict):
        keys: list[str] = []
        for row in item.measured:
            for k in row:
                if k not in keys:
                    keys.append(k)
        rows = [keys] + [[str(row.get(k, "")) for k in keys]
                         for row in item.measured]
        has_table = len(rows) >= 2
    if has_table:
        for r, row in enumerate(rows, start=start):
            for c, raw in enumerate(row, start=1):
                value = _cell_value(raw) if r > start else str(raw)
                cell = ws.cell(row=r, column=c, value=value)
                cell.border = border
                if r == start:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center
                elif isinstance(value, (int, float)):
                    cell.alignment = right
        n_cols = max(len(r) for r in rows)
        for c in range(1, n_cols + 1):
            width = max((len(str(r[c - 1])) if c <= len(r) else 0)
                        for r in rows) + 4
            ws.column_dimensions[get_column_letter(c)].width = min(width, 40)
        ws.row_dimensions[start].height = 22
        ws.freeze_panes = ws.cell(row=start + 1, column=1).coordinate

    # ---- 截图：优先按首列数值匹配数据行（Iload / 组号），否则堆叠在下方 ----
    shots = _collect_shots(item)
    if not shots:
        return
    if has_table:
        first_col = {_to_float(r[0]) for r in rows[1:] if r}
        matched: dict[float, tuple[str, str]] = {}
        used_keys: set[float] = set()
        unmatched: list[tuple[str, str]] = []
        for key, png in shots:
            kv = _to_float(key)
            if kv is not None and kv in first_col and kv not in used_keys:
                matched[kv] = (key, png)
                used_keys.add(kv)
            else:
                unmatched.append((key, png))
        if matched:
            shot_col = n_cols + 1
            col_letter = get_column_letter(shot_col)
            cell = ws.cell(row=start, column=shot_col, value="Scope Shot")
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
            ws.column_dimensions[col_letter].width = 44
            for r, row in enumerate(rows[1:], start=start + 1):
                kv = _to_float(row[0]) if row else None
                if kv in matched:
                    key, png = matched[kv]
                    try:
                        img, h = _scaled_img(XLImage(png), _THUMB_W)
                        ws.add_image(img, f"{col_letter}{r}")
                        ws.row_dimensions[r].height = h * _PX_TO_PT + 4
                    except OSError:
                        log_fn(f"[EXPORT] 截图读取失败，已跳过：{png}")
            shots = unmatched

    r = (start + len(rows) - 1) if has_table else start + 1
    for key, png in shots:
        r += 2
        if key:
            ws.cell(row=r, column=1, value=f"Scope shot — Iload={key} mA").font = \
                Font(bold=True)
        else:
            ws.cell(row=r, column=1, value="Waveform").font = Font(bold=True)
        try:
            img, h = _scaled_img(XLImage(png), _FULL_W)
            ws.add_image(img, f"A{r + 1}")
            ws.row_dimensions[r + 1].height = h * _PX_TO_PT + 4
        except OSError:
            log_fn(f"[EXPORT] 截图读取失败，已跳过：{png}")


def export_result_xlsx(result, out_path: str,
                       log_fn: Callable[[str], None] | None = None) -> str:
    """把 ModuleTestResult 导出为单文件 XLSX（数据 + 截图入格），返回路径。

    Summary sheet + 每项一个数据 sheet；缺 openpyxl/Pillow 抛 RuntimeError
    （不静默降级——截图入格是本导出的核心价值）。
    """
    try:
        import openpyxl
        from openpyxl.drawing.image import Image as XLImage  # noqa: F401
    except ImportError as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法导出 XLSX。") from exc
    try:
        from PIL import Image as PILImage  # noqa: F401
    except ImportError as exc:
        raise RuntimeError(
            "缺少 Pillow 依赖，无法在 XLSX 中嵌入截图。"
            "请安装：pip install Pillow") from exc

    log = log_fn or (lambda _msg: None)
    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    _write_summary_sheet(ws, result)

    used: set[str] = {"Summary"}
    for item in result.items:
        sheet = _sheet_title(item.name, used)
        _write_item_sheet(wb.create_sheet(sheet), item, log)
    wb.save(out_path)
    logger.info("Module Test XLSX 已导出: %s", out_path)
    return out_path


# ---------------------------------------------------------------------- Worker
class XlsxExportWorker(QObject):
    """QThread Worker：后台导出 XLSX，日志/结果经 Signal 回 UI 线程。"""

    log = Signal(str)
    finished = Signal(bool, str)  # ok, path 或错误信息

    def __init__(self, *, result, out_path: str, parent: QObject | None = None):
        super().__init__(parent)
        self._result = result
        self._out_path = out_path

    def run(self) -> None:
        try:
            path = export_result_xlsx(self._result, self._out_path,
                                      log_fn=self.log.emit)
            self.finished.emit(True, path)
        except Exception as exc:  # noqa: BLE001 - 统一回 UI 层提示
            logger.error("XLSX 导出失败: %s", exc, exc_info=True)
            self.finished.emit(False, str(exc))
