"""
GPADC测试UI组件
修复左侧滚动区域宽度与显示不完整问题
"""
# run cmd:
# python -m ui d:\CodeProject\TRAE_Projects\KK_Lab\ui\gpadc_test_ui.py


from ui.widgets.dark_combobox import DarkComboBox
from ui.styles import SCROLL_AREA_STYLE
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QLineEdit, QGridLayout, QSpinBox, QDoubleSpinBox, QFrame, QRadioButton,
    QButtonGroup, QApplication, QSizePolicy, QStackedWidget, QScrollArea,
    QTextEdit, QSplitter
)
from PySide6.QtCore import Qt, Signal, QThread, QObject
from PySide6.QtGui import QFont
import time

import sys
from pathlib import Path

i2c_lib_path = Path(__file__).parent.parent.parent.parent / "lib" / "i2c"
sys.path.insert(0, str(i2c_lib_path))

from i2c_interface_x64 import I2CInterface
from Bes_I2CIO_Interface import I2CSpeedMode, I2CWidthFlag

from log_config import get_logger
from debug_config import DEBUG_MOCK
from instruments.mock.mock_instruments import MockI2C, MockN6705C, MockVT6002

logger = get_logger(__name__)


class _SearchN6705CWorker(QObject):
    finished = Signal(list)
    error = Signal(str)

    def run(self):
        rm = None
        try:
            import pyvisa
            rm = pyvisa.ResourceManager()
            resources = list(rm.list_resources()) or []
            n6705c_devices = []
            for dev in resources:
                try:
                    instr = rm.open_resource(dev, timeout=1000)
                    idn = instr.query('*IDN?').strip()
                    instr.close()
                    if "N6705C" in idn:
                        n6705c_devices.append(dev)
                except Exception:
                    pass
            self.finished.emit(n6705c_devices)
        except Exception as e:
            self.error.emit(str(e))
        finally:
            if rm is not None:
                try:
                    rm.close()
                except Exception:
                    pass


class _SearchSerialWorker(QObject):
    finished = Signal(list)
    error = Signal(str)

    def run(self):
        try:
            import serial.tools.list_ports
            ports = serial.tools.list_ports.comports()
            result = [f"{p.device} ({p.description})" for p in ports]
            self.finished.emit(result)
        except Exception as e:
            self.error.emit(str(e))


class _TestWorker(QObject):
    finished = Signal(object)
    error = Signal(str)
    log = Signal(str)

    def __init__(self, fn, kwargs):
        super().__init__()
        self._fn = fn
        self._kwargs = kwargs
        self._stop_requested = False

    def request_stop(self):
        self._stop_requested = True

    def is_stop_requested(self):
        return self._stop_requested

    def run(self):
        try:
            result = self._fn(stop_check=self.is_stop_requested, **self._kwargs)
            self.finished.emit(result)
        except Exception as e:
            self.error.emit(str(e))


class GPADCTestUI(QWidget):
    """GPADC测试UI组件"""

    connection_status_changed = Signal(bool)

    TEST_1000CNT = "1000CNT TEST"
    TEST_FORCE_VOLTAGE = "Force Voltage Test"
    TEST_HIGH_LOW_TEMP = "High-Low Temp Test"
    TEST_TEMP_CONSISTENCY = "Temp Consistency Test"

    def __init__(self, n6705c_top=None):
        super().__init__()

        self._n6705c_top = n6705c_top
        self.rm = None
        self.n6705c = None
        self.is_n6705c_connected = False
        self.available_n6705c_devices = []

        self.vt6002 = None
        self.is_vt6002_connected = False
        self.available_vt6002_ports = []

        self.dut_serial = None
        self.is_dut_connected = False
        self.available_dut_ports = []

        self.is_test_running = False
        self._start_btn_text = "▶ START TEST"
        self.test_thread = None
        self._test_worker = None
        self._search_thread = None
        self._search_worker = None
        self._export_data = None
        self._chart_image_bytes = None

        self._setup_style()
        self._create_layout()
        self._init_ui_elements()
        self._sync_from_top()

    def _setup_style(self):
        font = QFont("Segoe UI", 9)
        self.setFont(font)

        self.setStyleSheet("""
            QWidget {
                background-color: #050b1a;
                color: #d8e3ff;
                font-family: "Segoe UI";
                font-size: 12px;
            }

            QLabel {
                background: transparent;
                color: #d8e3ff;
            }

            QFrame#page {
                background-color: #050b1a;
            }

            QFrame#panel,
            QFrame#chart_panel,
            QFrame#metric_card,
            QFrame#config_inner_panel,
            QFrame#action_panel {
                background-color: #0b1630;
                border: 1px solid #18284d;
                border-radius: 12px;
            }

            QFrame#instrument_inner {
                background-color: #071126;
                border: 1px solid #152240;
                border-radius: 10px;
            }

            QFrame#left_scroll_content {
                background: transparent;
                border: none;
            }

            QPushButton#test_item_btn {
                text-align: left;
                min-height: 52px;
                border: 1px solid #1c2a4a;
                border-radius: 8px;
                padding: 8px 10px;
                background-color: #020816;
                color: #dbe7ff;
            }

            QPushButton#test_item_btn:hover {
                background-color: #09132b;
                border: 1px solid #2a4175;
            }

            QPushButton#test_item_btn_checked {
                text-align: left;
                min-height: 52px;
                border: 1px solid #6b63ff;
                border-radius: 8px;
                padding: 8px 10px;
                background-color: rgba(98, 77, 255, 0.16);
                color: #dbe7ff;
            }

            QLabel#title_label {
                font-size: 18px;
                font-weight: 700;
                color: #ffffff;
            }

            QLabel#subtitle_label {
                font-size: 12px;
                color: #7e96bf;
            }

            QLabel#section_title {
                font-size: 13px;
                font-weight: 700;
                color: #ffffff;
            }

            QLabel#muted_label {
                color: #7e96bf;
                font-size: 11px;
            }

            QLabel#metric_name {
                color: #7e96bf;
                font-size: 11px;
                font-weight: 600;
            }

            QLabel#metric_value_green {
                color: #00d39a;
                font-size: 16px;
                font-weight: 700;
            }

            QLabel#metric_value_blue {
                color: #59a8ff;
                font-size: 16px;
                font-weight: 700;
            }

            QLabel#metric_value_yellow {
                color: #f2c94c;
                font-size: 16px;
                font-weight: 700;
            }

            QLineEdit, QSpinBox, QDoubleSpinBox {
                min-height: 32px;
                border: 1px solid #24365e;
                border-radius: 6px;
                padding: 4px 8px;
                background-color: #020816;
                color: #dbe7ff;
                selection-background-color: #4c6fff;
            }

            QLineEdit:focus, QComboBox:focus, QSpinBox:focus, QDoubleSpinBox:focus {
                border: 1px solid #5b7cff;
            }

            QComboBox::drop-down {
                border: none;
                width: 22px;
            }

            QComboBox QAbstractItemView {
                background-color: #0b1630;
                color: #dbe7ff;
                border: 1px solid #24365e;
                selection-background-color: #253a70;
            }

            QComboBox QAbstractItemView::item {
                background-color: #0b1630;
                color: #dbe7ff;
                padding: 4px 8px;
            }

            QComboBox QAbstractItemView::item:hover {
                background-color: #1a3260;
            }

            QComboBox QAbstractItemView::item:selected {
                background-color: #253a70;
            }

            QComboBox QFrame {
                background-color: #0b1630;
                border: 1px solid #24365e;
            }

            QPushButton {
                min-height: 32px;
                border: 1px solid #25355c;
                border-radius: 8px;
                padding: 6px 12px;
                background-color: #162544;
                color: #dbe7ff;
            }

            QPushButton:hover {
                background-color: #1c315b;
            }

            QPushButton:pressed {
                background-color: #10203d;
            }

            QPushButton:disabled {
                background-color: #0f1930;
                color: #5a6b8e;
                border: 1px solid #1b2847;
            }

            QPushButton#connect_btn {
                background-color: rgba(0, 211, 154, 0.14);
                color: #00d39a;
                border: 1px solid rgba(0, 211, 154, 0.25);
                font-weight: 600;
            }

            QPushButton#danger_btn {
                background-color: rgba(255, 90, 122, 0.14);
                color: #ff6f8e;
                border: 1px solid rgba(255, 90, 122, 0.25);
                font-weight: 600;
            }

            QPushButton#start_test_btn {
                background-color: #5d45ff;
                color: white;
                border: 1px solid #6d59ff;
                font-weight: 700;
                min-height: 38px;
            }

            QPushButton#stop_test_btn {
                background-color: rgba(255, 90, 122, 0.12);
                color: #ff7593;
                border: 1px solid rgba(255, 90, 122, 0.28);
                min-height: 38px;
                min-width: 42px;
                max-width: 42px;
                font-weight: 700;
            }

            QPushButton#tool_btn {
                min-height: 28px;
                border-radius: 6px;
                background-color: #162544;
                padding: 4px 10px;
            }

            QRadioButton {
                background: transparent;
                color: #dbe7ff;
                spacing: 8px;
            }

            QRadioButton::indicator {
                width: 14px;
                height: 14px;
            }

            QRadioButton::indicator:unchecked {
                border: 1px solid #4a5e8e;
                border-radius: 7px;
                background: #071126;
            }

            QRadioButton::indicator:checked {
                border: 1px solid #4ca8ff;
                border-radius: 7px;
                background: #4ca8ff;
            }

            QTextEdit {
                background-color: #050d1e;
                border: 1px solid #0e1e40;
                border-radius: 6px;
                color: #8abaff;
                font-size: 11px;
                font-family: 'Consolas', monospace;
                padding: 6px;
            }
        """ + SCROLL_AREA_STYLE)

    def _create_metric_card(self, title, value="---", value_object_name="metric_value_green"):
        card = QFrame()
        card.setObjectName("metric_card")
        card.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        card.setMinimumHeight(72)

        layout = QVBoxLayout(card)
        layout.setContentsMargins(16, 8, 16, 8)
        layout.setSpacing(4)

        name_label = QLabel(title)
        name_label.setObjectName("metric_name")
        name_label.setAlignment(Qt.AlignCenter)

        value_label = QLabel(value)
        value_label.setObjectName(value_object_name)
        value_label.setAlignment(Qt.AlignCenter)

        layout.addWidget(name_label)
        layout.addWidget(value_label)
        return card, value_label

    def _create_instrument_card(self, title, subtitle, combo_attr_name, search_btn_attr_name,
                                connect_btn_attr_name, disconnect_btn_attr_name, status_attr_name):
        card = QFrame()
        card.setObjectName("instrument_inner")
        card.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        layout = QVBoxLayout(card)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(8)

        top_row = QHBoxLayout()
        top_row.setSpacing(8)

        title_box = QVBoxLayout()
        title_box.setSpacing(2)

        name_label = QLabel(title)
        name_label.setStyleSheet("font-size: 13px; font-weight: 700; color: #ffffff; border: none;")    
        name_label.setWordWrap(True)

        desc_label = QLabel(subtitle)
        desc_label.setObjectName("muted_label")
        desc_label.setStyleSheet("border: none;")
        desc_label.setWordWrap(True)

        title_box.addWidget(name_label)
        title_box.addWidget(desc_label)

        status_label = QLabel("Not Connected")
        status_label.setStyleSheet("color: #ff5a7a; font-weight: 600;")
        status_label.setWordWrap(True)
        setattr(self, status_attr_name, status_label)

        btn_connect = QPushButton("Connect")
        btn_connect.setObjectName("connect_btn")
        btn_connect.setFixedWidth(90)

        setattr(self, connect_btn_attr_name, btn_connect)
        setattr(self, disconnect_btn_attr_name, btn_connect)

        top_row.addLayout(title_box, 1)
        top_row.addWidget(btn_connect, 0, Qt.AlignTop)

        select_row = QHBoxLayout()
        select_row.setSpacing(6)

        combo = DarkComboBox(bg="#0b1630", border="#24365e")
        combo.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        search_btn = QPushButton("⌕")
        search_btn.setObjectName("tool_btn")
        search_btn.setFixedWidth(34)

        setattr(self, combo_attr_name, combo)
        setattr(self, search_btn_attr_name, search_btn)

        select_row.addWidget(combo, 1)
        select_row.addWidget(search_btn)

        layout.addLayout(top_row)
        layout.addWidget(status_label)
        layout.addLayout(select_row)

        return card

    def _create_test_item_button(self, title, desc):
        btn = QPushButton(f"●  {title}\n{desc}")
        btn.setObjectName("test_item_btn")
        btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        btn.setMinimumHeight(56)
        return btn

    def _create_layout(self):
        root_layout = QVBoxLayout(self)
        root_layout.setContentsMargins(8, 6, 8, 8)
        root_layout.setSpacing(8)

        self.page = QFrame()
        self.page.setObjectName("page")
        page_layout = QVBoxLayout(self.page)
        page_layout.setContentsMargins(0, 0, 0, 0)
        page_layout.setSpacing(10)

        header_layout = QVBoxLayout()
        header_layout.setSpacing(2)


        title_label = QLabel("🛢 GPADC Automated Test")
        title_label.setObjectName("title_label")
        title_label.setStyleSheet("border: none")

        subtitle_label = QLabel("Evaluate GPADC performance including Linearity, ENOB, and Temperature Drift.")
        subtitle_label.setObjectName("subtitle_label")
        subtitle_label.setStyleSheet("border: none")

        header_layout.addWidget(title_label)
        header_layout.addWidget(subtitle_label)
        page_layout.addLayout(header_layout)

        body_layout = QHBoxLayout()
        body_layout.setSpacing(12)

        # 左侧滚动区
        self.left_scroll = QScrollArea()
        self.left_scroll.setWidgetResizable(True)
        self.left_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.left_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.left_scroll.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Expanding)
        self.left_scroll.setMinimumWidth(320)
        self.left_scroll.setMaximumWidth(320)

        left_content = QFrame()
        left_content.setObjectName("left_scroll_content")
        left_content.setMinimumWidth(300)
        left_content.setMaximumWidth(300)
        left_content.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Preferred)

        left_col = QVBoxLayout(left_content)
        left_col.setContentsMargins(0, 0, 6, 0)
        left_col.setSpacing(12)

        # Instruments
        instruments_panel = QFrame()
        instruments_panel.setObjectName("panel")
        instruments_panel.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        instruments_layout = QVBoxLayout(instruments_panel)
        instruments_layout.setContentsMargins(12, 12, 12, 12)
        instruments_layout.setSpacing(10)

        instruments_title = QLabel("🔗 INSTRUMENTS")
        instruments_title.setObjectName("section_title")
        instruments_title.setStyleSheet("border: none")
        instruments_layout.addWidget(instruments_title)

        instruments_layout.addWidget(self._create_instrument_card(
            "N6705C Power Analyzer",
            "Voltage Source",
            "n6705c_combo",
            "n6705c_search_btn",
            "n6705c_connect_btn",
            "n6705c_disconnect_btn",
            "n6705c_status"
        ))
        instruments_layout.addWidget(self._create_instrument_card(
            "VT6002 Chamber",
            "Thermal Control",
            "vt6002_combo",
            "vt6002_search_btn",
            "vt6002_connect_btn",
            "vt6002_disconnect_btn",
            "vt6002_status"
        ))
        left_col.addWidget(instruments_panel)

        # Data Acquisition
        data_panel = QFrame()
        data_panel.setObjectName("panel")
        data_panel.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        data_layout = QVBoxLayout(data_panel)
        data_layout.setContentsMargins(12, 12, 12, 12)
        data_layout.setSpacing(10)

        data_title = QLabel("⇄ DATA ACQUISITION")
        data_title.setObjectName("section_title")
        data_layout.addWidget(data_title)

        radio_row = QHBoxLayout()
        radio_row.setSpacing(12)

        self.iic_radio = QRadioButton("I2C")
        self.uart_radio = QRadioButton("UART Log")
        self.iic_radio.setChecked(True)

        self.data_acquisition_group = QButtonGroup()
        self.data_acquisition_group.addButton(self.iic_radio)
        self.data_acquisition_group.addButton(self.uart_radio)

        radio_row.addWidget(self.iic_radio)
        radio_row.addWidget(self.uart_radio)
        radio_row.addStretch()
        data_layout.addLayout(radio_row)

        self.data_stack = QStackedWidget()
        self.data_stack.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        self.iic_group = QFrame()
        self.iic_group.setObjectName("config_inner_panel")
        iic_layout = QGridLayout(self.iic_group)
        iic_layout.setContentsMargins(10, 10, 10, 10)
        iic_layout.setHorizontalSpacing(6)
        iic_layout.setVerticalSpacing(6)

        iic_layout.addWidget(QLabel("Device Address (Hex)"), 0, 0)
        self.iic_device_address = QLineEdit("0x17")
        iic_layout.addWidget(self.iic_device_address, 1, 0)

        iic_layout.addWidget(QLabel("Raw Data Register (Hex)"), 2, 0)
        self.iic_data_address = QLineEdit("0x57")
        iic_layout.addWidget(self.iic_data_address, 3, 0)

        self.uart_group = QFrame()
        self.uart_group.setObjectName("config_inner_panel")
        uart_layout = QGridLayout(self.uart_group)
        uart_layout.setContentsMargins(10, 10, 10, 10)
        uart_layout.setHorizontalSpacing(6)
        uart_layout.setVerticalSpacing(6)

        uart_layout.addWidget(QLabel("DUT Serial Port"), 0, 0, 1, 2)
        self.dut_combo = DarkComboBox(bg="#0b1630", border="#24365e")
        self.dut_combo.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        self.dut_search_btn = QPushButton("⌕")
        self.dut_search_btn.setObjectName("tool_btn")
        self.dut_search_btn.setFixedWidth(34)

        uart_layout.addWidget(self.dut_combo, 1, 0)
        uart_layout.addWidget(self.dut_search_btn, 1, 1)

        uart_layout.addWidget(QLabel("Search Keyword"), 2, 0, 1, 2)
        self.uart_keyword = QLineEdit("GPADC RAW")
        uart_layout.addWidget(self.uart_keyword, 3, 0, 1, 2)

        self.data_stack.addWidget(self.iic_group)
        self.data_stack.addWidget(self.uart_group)
        data_layout.addWidget(self.data_stack)
        left_col.addWidget(data_panel)

        # Test Item
        test_item_panel = QFrame()
        test_item_panel.setObjectName("panel")
        test_item_panel.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        test_item_layout = QVBoxLayout(test_item_panel)
        test_item_layout.setContentsMargins(12, 12, 12, 12)
        test_item_layout.setSpacing(10)

        test_item_title = QLabel("✦ TEST ITEM")
        test_item_title.setObjectName("section_title")
        test_item_title.setStyleSheet("border: none")
        test_item_layout.addWidget(test_item_title)

        self.cnt1000_test_btn = self._create_test_item_button(
            "1000CNT TEST",
            "Test 1000 times average, min and max"
        )
        self.force_voltage_test_btn = self._create_test_item_button(
            "Force Voltage Test",
            "Sweep voltage at current temperature"
        )
        self.high_low_temp_test_btn = self._create_test_item_button(
            "High-Low Temp Test",
            "Sweep temperature across multiple temperatures"
        )
        self.temp_consistency_test_btn = self._create_test_item_button(
            "Temp Consistency Test",
            "Evaluate consistency across temperature cycles"
        )

        test_item_layout.addWidget(self.cnt1000_test_btn)
        test_item_layout.addWidget(self.force_voltage_test_btn)
        test_item_layout.addWidget(self.high_low_temp_test_btn)
        test_item_layout.addWidget(self.temp_consistency_test_btn)
        left_col.addWidget(test_item_panel)

        # Test Parameters
        params_panel = QFrame()
        params_panel.setObjectName("panel")
        params_panel.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        params_layout = QVBoxLayout(params_panel)
        params_layout.setContentsMargins(12, 12, 12, 12)
        params_layout.setSpacing(8)

        self.params_title = QLabel("☷ TEST PARAMETERS")
        self.params_title.setObjectName("section_title")
        self.params_title.setStyleSheet("border: none;")
        params_layout.addWidget(self.params_title)

        self.params_mode_label = QLabel("VOLTAGE SWEEP")
        self.params_mode_label.setStyleSheet("color: #7e96bf; font-size: 11px; font-weight: 700; border: none;")
        params_layout.addWidget(self.params_mode_label)

        self.voltage_params_frame = QFrame()
        self.voltage_params_frame.setStyleSheet("background: transparent; border: none;")
        voltage_layout = QGridLayout(self.voltage_params_frame)
        voltage_layout.setContentsMargins(0, 0, 0, 0)
        voltage_layout.setHorizontalSpacing(6)
        voltage_layout.setVerticalSpacing(6)

        voltage_layout.addWidget(QLabel("Start (V)"), 0, 0)
        voltage_layout.addWidget(QLabel("End (V)"), 0, 1)
        voltage_layout.addWidget(QLabel("Step (V)"), 0, 2)

        self.voltage_min = QDoubleSpinBox()
        self.voltage_min.setRange(0.0, 5.0)
        self.voltage_min.setValue(4.0)
        self.voltage_min.setSingleStep(0.01)
        self.voltage_min.setDecimals(3)

        self.voltage_max = QDoubleSpinBox()
        self.voltage_max.setRange(0.0, 5.0)
        self.voltage_max.setValue(4.2)
        self.voltage_max.setSingleStep(0.01)
        self.voltage_max.setDecimals(3)

        self.voltage_step = QDoubleSpinBox()
        self.voltage_step.setRange(0.001, 1.0)
        self.voltage_step.setValue(0.05)
        self.voltage_step.setSingleStep(0.010)
        self.voltage_step.setDecimals(3)

        voltage_layout.addWidget(self.voltage_min, 1, 0)
        voltage_layout.addWidget(self.voltage_max, 1, 1)
        voltage_layout.addWidget(self.voltage_step, 1, 2)

        self.temp_params_frame = QFrame()
        self.temp_params_frame.setStyleSheet("background: transparent; border: none;")
        temp_layout = QGridLayout(self.temp_params_frame)
        temp_layout.setContentsMargins(0, 0, 0, 0)
        temp_layout.setHorizontalSpacing(6)
        temp_layout.setVerticalSpacing(6)

        temp_layout.addWidget(QLabel("Start (°C)"), 0, 0)
        temp_layout.addWidget(QLabel("End (°C)"), 0, 1)
        temp_layout.addWidget(QLabel("Step (°C)"), 0, 2)

        self.temp_min = QDoubleSpinBox()
        self.temp_min.setRange(-40.0, 125.0)
        self.temp_min.setValue(-40.0)
        self.temp_min.setSingleStep(1.0)
        self.temp_min.setDecimals(1)

        self.temp_max = QDoubleSpinBox()
        self.temp_max.setRange(-40.0, 125.0)
        self.temp_max.setValue(125.0)
        self.temp_max.setSingleStep(1.0)
        self.temp_max.setDecimals(1)

        self.temp_step = QDoubleSpinBox()
        self.temp_step.setRange(0.1, 50.0)
        self.temp_step.setValue(5)
        self.temp_step.setSingleStep(0.1)
        self.temp_step.setDecimals(2)

        temp_layout.addWidget(self.temp_min, 1, 0)
        temp_layout.addWidget(self.temp_max, 1, 1)
        temp_layout.addWidget(self.temp_step, 1, 2)

        params_layout.addWidget(self.voltage_params_frame)
        params_layout.addWidget(self.temp_params_frame)



        self.voltage_channel_label = QLabel("Voltage Channel")
        self.voltage_channel_label.setObjectName("muted_label")
        params_layout.addWidget(self.voltage_channel_label)

        self.voltage_channel = QSpinBox()
        self.voltage_channel.setRange(1, 4)
        self.voltage_channel.setValue(4)
        params_layout.addWidget(self.voltage_channel)

        self.temp_hint_label = QLabel("Connect VT6002 to enable temperature testing.")
        self.temp_hint_label.setStyleSheet("color: #ff5a7a; font-size: 11px;")
        self.temp_hint_label.setWordWrap(True)
        params_layout.addWidget(self.temp_hint_label)

        left_col.addWidget(params_panel)

        # Action
        action_panel = QFrame()
        action_panel.setObjectName("action_panel")
        action_panel.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        action_layout = QHBoxLayout(action_panel)
        action_layout.setContentsMargins(12, 12, 12, 12)
        action_layout.setSpacing(8)

        self.start_test_btn = QPushButton("▶ START TEST")
        self.start_test_btn.setObjectName("start_test_btn")
        self.start_test_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        self.stop_test_btn = QPushButton("■")
        self.stop_test_btn.setObjectName("stop_test_btn")
        self.stop_test_btn.setEnabled(False)
        self.stop_test_btn.hide()

        action_layout.addWidget(self.start_test_btn, 1)
        left_col.addWidget(action_panel)
        left_col.addStretch()

        self.left_scroll.setWidget(left_content)

        # 右侧
        right_col = QVBoxLayout()
        right_col.setSpacing(12)

        metrics_layout = QHBoxLayout()
        metrics_layout.setSpacing(8)

        inl_card, self.inl_value = self._create_metric_card("INL", "---", "metric_value_green")
        dnl_card, self.dnl_value = self._create_metric_card("DNL", "---", "metric_value_green")
        enob_card, self.enob_value = self._create_metric_card("ENOB", "---", "metric_value_blue")
        offset_card, self.offset_error_value = self._create_metric_card("OFFSET ERR", "---", "metric_value_yellow")
        gain_card, self.gain_error_value = self._create_metric_card("GAIN ERR", "---", "metric_value_yellow")
        
        # 添加1000CNT TEST的指标卡片
        avg_card, self.avg_value = self._create_metric_card("AVG", "---", "metric_value_green")
        min_card, self.min_value = self._create_metric_card("MIN", "---", "metric_value_blue")
        max_card, self.max_value = self._create_metric_card("MAX", "---", "metric_value_yellow")

        self.linearity_value = QLabel("---")
        self.linearity_value.hide()

        metrics_layout.addWidget(inl_card)
        metrics_layout.addWidget(dnl_card)
        metrics_layout.addWidget(enob_card)
        metrics_layout.addWidget(offset_card)
        metrics_layout.addWidget(gain_card)
        metrics_layout.addWidget(avg_card)
        metrics_layout.addWidget(min_card)
        metrics_layout.addWidget(max_card)

        right_col.addLayout(metrics_layout)

        chart_panel = QFrame()
        chart_panel.setObjectName("chart_panel")
        chart_panel.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        chart_layout = QVBoxLayout(chart_panel)
        chart_layout.setContentsMargins(14, 14, 14, 14)
        chart_layout.setSpacing(10)

        chart_top = QHBoxLayout()
        chart_title = QLabel("ADC Transfer Curve")
        chart_title.setObjectName("section_title")
        chart_title.setStyleSheet("border: none;")

        self.export_result_btn = QPushButton("🗎 Export Result")
        self.export_result_btn.setObjectName("tool_btn")

        chart_top.addWidget(chart_title)
        chart_top.addStretch()
        chart_top.addWidget(self.export_result_btn)
        chart_layout.addLayout(chart_top)

        self.chart_placeholder = QFrame()
        self.chart_placeholder.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.chart_placeholder.setMinimumHeight(200)
        self.chart_placeholder.setStyleSheet("""
            QFrame {
                background-color: #0a1735;
                border: none;
                border-radius: 8px;
            }
        """)

        chart_placeholder_layout = QVBoxLayout(self.chart_placeholder)
        chart_placeholder_layout.setContentsMargins(22, 22, 22, 18)
        chart_placeholder_layout.setSpacing(12)

        legend_row = QHBoxLayout()
        legend_row.addStretch()

        actual_legend = QLabel("↔ Actual Code")
        actual_legend.setStyleSheet("color: #00d39a; font-size: 12px;")
        ideal_legend = QLabel("↔ Ideal Code")
        ideal_legend.setStyleSheet("color: #7e96bf; font-size: 12px;")

        legend_row.addWidget(actual_legend)
        legend_row.addWidget(ideal_legend)
        legend_row.addStretch()
        chart_placeholder_layout.addLayout(legend_row)

        plot_area = QFrame()
        plot_area.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        plot_area.setStyleSheet("""
            QFrame {
                background-color: transparent;
                border-left: 1px solid #6f7fa5;
                border-bottom: 1px solid #6f7fa5;
                border-top: 1px dashed rgba(126,150,191,0.18);
                border-right: none;
                border-radius: 0px;
            }
        """)
        chart_placeholder_layout.addWidget(plot_area, 1)

        x_label = QLabel("Input Voltage (V)")
        x_label.setAlignment(Qt.AlignCenter)
        x_label.setObjectName("muted_label")
        chart_placeholder_layout.addWidget(x_label)

        chart_layout.addWidget(self.chart_placeholder, 1)

        log_panel = QFrame()
        log_panel.setObjectName("panel")
        log_layout = QVBoxLayout(log_panel)
        log_layout.setContentsMargins(10, 10, 10, 10)
        log_layout.setSpacing(6)

        log_title_row = QHBoxLayout()
        log_title = QLabel("TEST LOG")
        log_title.setObjectName("section_title")
        self.clear_log_btn = QPushButton("Clear")
        self.clear_log_btn.setObjectName("tool_btn")
        self.clear_log_btn.setFixedWidth(60)
        log_title_row.addWidget(log_title)
        log_title_row.addStretch()
        log_title_row.addWidget(self.clear_log_btn)
        log_layout.addLayout(log_title_row)

        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setTabStopDistance(100)
        log_layout.addWidget(self.log_text, 1)

        right_splitter = QSplitter(Qt.Vertical)
        right_splitter.setStyleSheet("""
            QSplitter::handle {
                background-color: #18284d;
                height: 4px;
                border-radius: 2px;
            }
            QSplitter::handle:hover {
                background-color: #5b7cff;
            }
        """)
        right_splitter.addWidget(chart_panel)
        right_splitter.addWidget(log_panel)
        right_splitter.setStretchFactor(0, 3)
        right_splitter.setStretchFactor(1, 2)
        right_splitter.setCollapsible(0, False)
        right_splitter.setCollapsible(1, False)

        right_col.addWidget(right_splitter, 1)

        self.export_params_btn = QPushButton("Export Parameters")
        self.export_params_btn.hide()

        self.save_config_btn = QPushButton("Save Config")
        self.load_config_btn = QPushButton("Load Config")
        self.save_config_btn.hide()
        self.load_config_btn.hide()

        body_layout.addWidget(self.left_scroll, 0)
        body_layout.addLayout(right_col, 1)

        page_layout.addLayout(body_layout, 1)
        root_layout.addWidget(self.page, 1)

    def _init_ui_elements(self):
        self.current_test_item = self.TEST_FORCE_VOLTAGE

        self.iic_radio.toggled.connect(self._update_data_acquisition_ui)
        self.uart_radio.toggled.connect(self._update_data_acquisition_ui)

        self.cnt1000_test_btn.clicked.connect(
            lambda: self._set_test_item(self.TEST_1000CNT)
        )
        self.force_voltage_test_btn.clicked.connect(
            lambda: self._set_test_item(self.TEST_FORCE_VOLTAGE)
        )
        self.high_low_temp_test_btn.clicked.connect(
            lambda: self._set_test_item(self.TEST_HIGH_LOW_TEMP)
        )
        self.temp_consistency_test_btn.clicked.connect(
            lambda: self._set_test_item(self.TEST_TEMP_CONSISTENCY)
        )

        self.n6705c_search_btn.clicked.connect(self._search_n6705c)
        self.n6705c_connect_btn.clicked.connect(self._toggle_n6705c)

        self.vt6002_search_btn.clicked.connect(self._search_vt6002)
        self.vt6002_connect_btn.clicked.connect(self._toggle_vt6002)

        self.dut_search_btn.clicked.connect(self._search_dut_ports)

        self.start_test_btn.clicked.connect(self._on_start_or_stop)
        self.stop_test_btn.clicked.connect(self._stop_test)
        self.export_result_btn.clicked.connect(self.export_result)
        self.clear_log_btn.clicked.connect(self.log_text.clear)

        self._update_data_acquisition_ui()
        self._set_test_item(self.TEST_1000CNT)

        self._search_vt6002()
        self._search_dut_ports()
        self._search_n6705c()

    def _set_status_label(self, label, text, status_type="err"):
        if status_type == "ok":
            label.setStyleSheet("color: #00d39a; font-weight: 600; border: none;")
        elif status_type == "warn":
            label.setStyleSheet("color: #ffb84d; font-weight: 600; border: none;")
        elif status_type == "warn":
            label.setStyleSheet("color: #ffb84d; font-weight: 600; border: none;")
        else:
            label.setStyleSheet("color: #ff5a7a; font-weight: 600; border: none;")
        label.setText(text)

    def _update_data_acquisition_ui(self):
        if self.iic_radio.isChecked():
            self.data_stack.setCurrentWidget(self.iic_group)
        else:
            self.data_stack.setCurrentWidget(self.uart_group)

    def _set_test_item(self, test_item):
        self.current_test_item = test_item

        self.cnt1000_test_btn.setObjectName(
            "test_item_btn_checked" if test_item == self.TEST_1000CNT else "test_item_btn"
        )
        self.force_voltage_test_btn.setObjectName(
            "test_item_btn_checked" if test_item == self.TEST_FORCE_VOLTAGE else "test_item_btn"
        )
        self.high_low_temp_test_btn.setObjectName(
            "test_item_btn_checked" if test_item == self.TEST_HIGH_LOW_TEMP else "test_item_btn"
        )
        self.temp_consistency_test_btn.setObjectName(
            "test_item_btn_checked" if test_item == self.TEST_TEMP_CONSISTENCY else "test_item_btn"
        )

        for btn in (
            self.cnt1000_test_btn,
            self.force_voltage_test_btn,
            self.high_low_temp_test_btn,
            self.temp_consistency_test_btn
        ):
            btn.style().unpolish(btn)
            btn.style().polish(btn)
            btn.update()

        if test_item == self.TEST_1000CNT:
            self.params_mode_label.setText("1000 COUNT TEST")
            self.voltage_params_frame.hide()
            self.temp_params_frame.hide()
            self.temp_hint_label.hide()
            self.start_test_btn.setText("▶ START 1000CNT TEST")
            self._start_btn_text = "▶ START 1000CNT TEST"
        elif test_item == self.TEST_FORCE_VOLTAGE:
            self.params_mode_label.setText("VOLTAGE SWEEP")
            self.voltage_params_frame.show()
            self.temp_params_frame.hide()
            self.temp_hint_label.hide()
            self.start_test_btn.setText("▶ START VOLT TEST")
            self._start_btn_text = "▶ START VOLT TEST"
        elif test_item == self.TEST_HIGH_LOW_TEMP:
            self.params_mode_label.setText("TEMPERATURE SWEEP")
            self.voltage_params_frame.hide()
            self.temp_params_frame.show()
            self.temp_hint_label.show()
            self.start_test_btn.setText("▶ START TEMP TEST")
            self._start_btn_text = "▶ START TEMP TEST"
            # 在温度扫描测试中重置avg、min、max参数显示
            self.avg_value.setText("---")
            self.min_value.setText("---")
            self.max_value.setText("---")
        else:
            self.params_mode_label.setText("VOLTAGE + TEMPERATURE")
            self.voltage_params_frame.show()
            self.temp_params_frame.show()
            self.temp_hint_label.show()
            self.start_test_btn.setText("▶ START CONSISTENCY TEST")
            self._start_btn_text = "▶ START CONSISTENCY TEST"
            # 在一致性测试中也重置avg、min、max参数显示
            self.avg_value.setText("---")
            self.min_value.setText("---")
            self.max_value.setText("---")

    def _search_n6705c(self):
        if self._n6705c_top and self._n6705c_top.is_connected_a:
            return
        self._set_status_label(self.n6705c_status, "Searching...", "warn")
        self.n6705c_search_btn.setEnabled(False)
        self.n6705c_connect_btn.setEnabled(False)

        worker = _SearchN6705CWorker()
        thread = QThread()
        worker.moveToThread(thread)

        thread.started.connect(worker.run)
        worker.finished.connect(self._on_n6705c_search_done)
        worker.error.connect(self._on_n6705c_search_error)
        worker.finished.connect(thread.quit)
        worker.error.connect(thread.quit)
        thread.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)

        self._n6705c_search_thread = thread
        self._n6705c_search_worker = worker
        thread.start()

    def _on_n6705c_search_done(self, n6705c_devices):
        default_device = "TCPIP0::K-N6705C-06098.local::hislip0::INSTR"
        self.n6705c_combo.clear()
        if n6705c_devices:
            for dev in n6705c_devices:
                self.n6705c_combo.addItem(dev)
            if DEBUG_MOCK and default_device not in n6705c_devices:
                self.n6705c_combo.addItem(default_device)
            self._set_status_label(self.n6705c_status, "Available", "ok")
        else:
            if DEBUG_MOCK:
                self.n6705c_combo.addItem(default_device)
                self._set_status_label(self.n6705c_status, "Default Device Available", "ok")
            else:
                self._set_status_label(self.n6705c_status, "No Device Found", "err")
        self.n6705c_search_btn.setEnabled(True)
        self.n6705c_connect_btn.setEnabled(DEBUG_MOCK or self.n6705c_combo.count() > 0)

    def _on_n6705c_search_error(self, err):
        self._append_log(f"[WARN] Search N6705C error: {err}")
        self._set_status_label(self.n6705c_status, f"Error: {err}", "err")
        self.n6705c_search_btn.setEnabled(True)
        self.n6705c_connect_btn.setEnabled(False)

    def _toggle_n6705c(self):
        if self.is_n6705c_connected:
            self._disconnect_n6705c()
        else:
            self._connect_n6705c()

    def _toggle_vt6002(self):
        if self.is_vt6002_connected:
            self._disconnect_vt6002()
        else:
            self._connect_vt6002()

    def _set_btn_connected(self, btn):
        btn.setText("Disconnect")
        btn.setObjectName("danger_btn")
        btn.setEnabled(True)
        btn.style().unpolish(btn)
        btn.style().polish(btn)

    def _set_btn_disconnected(self, btn):
        btn.setText("Connect")
        btn.setObjectName("connect_btn")
        btn.setEnabled(True)
        btn.style().unpolish(btn)
        btn.style().polish(btn)

    def _connect_n6705c(self):
        self._set_status_label(self.n6705c_status, "Connecting...", "warn")
        self.n6705c_connect_btn.setEnabled(False)
        try:
            from instruments.power.keysight.n6705c import N6705C
            device_address = self.n6705c_combo.currentText()
            self.n6705c = N6705C(device_address)
            idn = self.n6705c.instr.query("*IDN?")
            if "N6705C" in idn:
                self.is_n6705c_connected = True
                self._set_status_label(self.n6705c_status, "Connected", "ok")
                self._set_btn_connected(self.n6705c_connect_btn)
                self.n6705c_search_btn.setEnabled(False)

                if self._n6705c_top:
                    self._n6705c_top.connect_a(device_address, self.n6705c)
            else:
                self._set_status_label(self.n6705c_status, "Device Mismatch", "err")
                self._set_btn_disconnected(self.n6705c_connect_btn)
        except Exception as e:
            self._append_log(f"[ERROR] Connect N6705C error: {e}")
            self._set_status_label(self.n6705c_status, f"Error: {e}", "err")
            self._set_btn_disconnected(self.n6705c_connect_btn)

    def _disconnect_n6705c(self):
        self._set_status_label(self.n6705c_status, "Disconnecting...", "warn")
        self.n6705c_connect_btn.setEnabled(False)
        try:
            if self._n6705c_top:
                self._n6705c_top.disconnect_a()
                self.n6705c = None
            else:
                if self.n6705c:
                    if hasattr(self.n6705c, 'instr') and self.n6705c.instr:
                        self.n6705c.instr.close()
                    if hasattr(self.n6705c, 'rm') and self.n6705c.rm:
                        self.n6705c.rm.close()
                    self.n6705c = None

            self.is_n6705c_connected = False
            self._set_status_label(self.n6705c_status, "Disconnected", "err")
            self._set_btn_disconnected(self.n6705c_connect_btn)
            self.n6705c_search_btn.setEnabled(True)
        except Exception as e:
            self._append_log(f"[ERROR] Disconnect N6705C error: {e}")
            self._set_status_label(self.n6705c_status, f"Error: {e}", "err")
            self._set_btn_connected(self.n6705c_connect_btn)

    def _sync_from_top(self):
        if not self._n6705c_top:
            return
        if self._n6705c_top.is_connected_a and self._n6705c_top.n6705c_a:
            self.n6705c = self._n6705c_top.n6705c_a
            self.is_n6705c_connected = True
            self._set_status_label(self.n6705c_status, "Connected", "ok")
            self._set_btn_connected(self.n6705c_connect_btn)
            self.n6705c_search_btn.setEnabled(False)
            if self._n6705c_top.visa_resource_a:
                self.n6705c_combo.clear()
                self.n6705c_combo.addItem(self._n6705c_top.visa_resource_a)
        elif not self.is_n6705c_connected:
            self._set_status_label(self.n6705c_status, "Disconnected", "err")
            self._set_btn_disconnected(self.n6705c_connect_btn)

    def _search_vt6002(self):
        self._set_status_label(self.vt6002_status, "Searching...", "warn")
        self.vt6002_search_btn.setEnabled(False)
        self.vt6002_connect_btn.setEnabled(False)

        worker = _SearchSerialWorker()
        thread = QThread()
        worker.moveToThread(thread)

        thread.started.connect(worker.run)
        worker.finished.connect(self._on_vt6002_search_done)
        worker.error.connect(self._on_vt6002_search_error)
        worker.finished.connect(thread.quit)
        worker.error.connect(thread.quit)
        thread.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)

        self._vt6002_search_thread = thread
        self._vt6002_search_worker = worker
        thread.start()

    def _on_vt6002_search_done(self, ports):
        self.available_vt6002_ports = ports
        self.vt6002_combo.clear()
        if ports:
            for port in ports:
                self.vt6002_combo.addItem(port)
            self._set_status_label(self.vt6002_status, "Available", "ok")
            self.vt6002_connect_btn.setEnabled(True)
        else:
            self.vt6002_combo.addItem("No serial ports found")
            self._set_status_label(self.vt6002_status, "Not Available", "err")
            self.vt6002_connect_btn.setEnabled(False)
        self.vt6002_search_btn.setEnabled(True)

    def _on_vt6002_search_error(self, err):
        self._append_log(f"[WARN] Search VT6002 error: {err}")
        self._set_status_label(self.vt6002_status, f"Error: {err}", "err")
        self.vt6002_search_btn.setEnabled(True)
        self.vt6002_connect_btn.setEnabled(False)

    def _connect_vt6002(self):
        self._set_status_label(self.vt6002_status, "Connecting...", "warn")
        self.vt6002_connect_btn.setEnabled(False)
        try:
            from instruments.chambers.vt6002_chamber import VT6002
            port_str = self.vt6002_combo.currentText()
            device_port = port_str.split()[0]

            self.vt6002 = VT6002(device_port)
            self.is_vt6002_connected = True

            self._set_status_label(self.vt6002_status, "Connected", "ok")
            self._set_btn_connected(self.vt6002_connect_btn)
            self.vt6002_search_btn.setEnabled(False)
        except Exception as e:
            self._append_log(f"[ERROR] Connect VT6002 error: {e}")
            self._set_status_label(self.vt6002_status, f"Error: {e}", "err")
            self._set_btn_disconnected(self.vt6002_connect_btn)

    def _disconnect_vt6002(self):
        self._set_status_label(self.vt6002_status, "Disconnecting...", "warn")
        self.vt6002_connect_btn.setEnabled(False)
        try:
            if self.vt6002:
                self.vt6002.close()
                self.vt6002 = None

            self.is_vt6002_connected = False
            self._set_status_label(self.vt6002_status, "Disconnected", "err")
            self._set_btn_disconnected(self.vt6002_connect_btn)
            self.vt6002_search_btn.setEnabled(True)
        except Exception as e:
            self._append_log(f"[ERROR] Disconnect VT6002 error: {e}")
            self._set_status_label(self.vt6002_status, f"Error: {e}", "err")
            self._set_btn_connected(self.vt6002_connect_btn)

    def _search_dut_ports(self):
        worker = _SearchSerialWorker()
        thread = QThread()
        worker.moveToThread(thread)

        thread.started.connect(worker.run)
        worker.finished.connect(self._on_dut_search_done)
        worker.error.connect(lambda err: self._append_log(f"[WARN] Search DUT ports error: {err}"))
        worker.finished.connect(thread.quit)
        worker.error.connect(thread.quit)
        thread.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)

        self._dut_search_thread = thread
        self._dut_search_worker = worker
        thread.start()

    def _on_dut_search_done(self, ports):
        self.available_dut_ports = ports
        self.dut_combo.clear()
        if ports:
            for port in ports:
                self.dut_combo.addItem(port)
        else:
            self.dut_combo.addItem("No serial ports found")

    def _on_start_or_stop(self):
        if self.is_test_running:
            self._stop_test()
        else:
            self._start_test()

    def _start_test(self):
        if self.is_test_running:
            return

        self.is_test_running = True
        self.start_test_btn.setEnabled(True)
        self.stop_test_btn.setEnabled(True)
        self._update_test_button_state(True)
        self._set_ui_enabled(False)
        self._append_log(f"[INFO] Starting GPADC test... mode={self.current_test_item}")

        test_item = self.current_test_item
        iic_device_addr = int(self.iic_device_address.text(), 16)
        iic_reg_addr = int(self.iic_data_address.text(), 16)

        if test_item == self.TEST_1000CNT:
            fn = self._run_1000cnt_test
            kwargs = dict(
                device_addr=iic_device_addr,
                reg_addr=iic_reg_addr,
            )
        elif test_item == self.TEST_FORCE_VOLTAGE:
            fn = self._run_force_voltage_test
            kwargs = dict(
                device_addr=iic_device_addr,
                reg_addr=iic_reg_addr,
                voltage_min=self.voltage_min.value(),
                voltage_max=self.voltage_max.value(),
                voltage_step=self.voltage_step.value(),
                voltage_channel=self.voltage_channel.value(),
            )
        elif test_item == self.TEST_HIGH_LOW_TEMP:
            fn = self._run_high_low_temp_test
            kwargs = dict(
                device_addr=iic_device_addr,
                reg_addr=iic_reg_addr,
                temp_min=self.temp_min.value(),
                temp_max=self.temp_max.value(),
                temp_step=self.temp_step.value(),
                voltage_channel=self.voltage_channel.value(),
            )
        elif test_item == self.TEST_TEMP_CONSISTENCY:
            fn = self._run_temp_consistency_test
            kwargs = dict(
                device_addr=iic_device_addr,
                reg_addr=iic_reg_addr,
                temp_min=self.temp_min.value(),
                temp_max=self.temp_max.value(),
                temp_step=self.temp_step.value(),
                voltage_min=self.voltage_min.value(),
                voltage_max=self.voltage_max.value(),
                voltage_step=self.voltage_step.value(),
                voltage_channel=self.voltage_channel.value(),
            )
        else:
            self._stop_test()
            return

        worker = _TestWorker(fn, kwargs)
        thread = QThread()
        worker.moveToThread(thread)

        thread.started.connect(worker.run)
        worker.finished.connect(self._on_test_done)
        worker.error.connect(self._on_test_error)
        worker.log.connect(self._append_log)
        worker.finished.connect(thread.quit)
        worker.error.connect(thread.quit)
        thread.finished.connect(worker.deleteLater)
        thread.finished.connect(self._on_test_thread_finished)

        self._test_worker = worker
        self.test_thread = thread
        thread.start()

    def _run_1000cnt_test(self, device_addr, reg_addr, stop_check=None):
        self._test_worker.log.emit(f"[INFO] Starting 1000CNT TEST with I2C address: 0x{device_addr:x} Register: 0x{reg_addr:x}")
        avg, max_val, min_val = self.gpadc_reg_read_by_cnts(
            device_addr=device_addr,
            reg_addr=reg_addr,
            iic_weight=10,
            get_reg_cnt=1000,
            stop_check=stop_check,
        )
        return ('1000cnt', {'avg': avg, 'max': max_val, 'min': min_val})

    def _run_force_voltage_test(self, device_addr, reg_addr, voltage_min, voltage_max,
                                voltage_step, voltage_channel, stop_check=None):
        result = self.gpadc_force_voltage_test(
            n6705c=self.n6705c,
            device_addr=device_addr,
            reg_addr=reg_addr,
            iic_weight=10,
            voltage_min=voltage_min,
            voltage_max=voltage_max,
            voltage_step=voltage_step,
            voltage_channel=voltage_channel,
            stop_check=stop_check,
        )
        return ('force_voltage', result)

    def _run_high_low_temp_test(self, device_addr, reg_addr, temp_min, temp_max,
                                temp_step, voltage_channel, stop_check=None):
        self._test_worker.log.emit("[INFO] RUN TEST_HIGH_LOW_TEMP TEST")
        result = self.gpadc_high_low_temp_test(
            device_addr=device_addr,
            reg_addr=reg_addr,
            iic_weight=10,
            temp_min=temp_min,
            temp_max=temp_max,
            temp_step=temp_step,
            voltage_channel=voltage_channel,
            stop_check=stop_check,
        )
        return ('high_low_temp', result)

    def _run_temp_consistency_test(self, device_addr, reg_addr, temp_min, temp_max,
                                   temp_step, voltage_min, voltage_max, voltage_step,
                                   voltage_channel, stop_check=None):
        self._test_worker.log.emit("[INFO] RUN TEST_TEMP_CONSISTENCY TEST")
        result = self.gpadc_temp_consistency_test(
            device_addr=device_addr,
            reg_addr=reg_addr,
            iic_weight=10,
            temp_min=temp_min,
            temp_max=temp_max,
            temp_step=temp_step,
            voltage_min=voltage_min,
            voltage_max=voltage_max,
            voltage_step=voltage_step,
            voltage_channel=voltage_channel,
            stop_check=stop_check,
        )
        return ('temp_consistency', result)

    def _on_test_done(self, payload):
        if payload is None:
            return
        kind, result = payload

        if kind == '1000cnt':
            self.update_test_result(result)
            self._append_log(f"[RESULT] 1000CNT TEST: AVG={result.get('avg', '---'):.3f}, MIN={result.get('min', '---'):.3f}, MAX={result.get('max', '---'):.3f}")

        elif kind == 'force_voltage':
            if result is not None:
                result_after_calibration = self._calibration_data(result)
                voltage_data, mean_cali, adc_min_cali, adc_max_cali = result_after_calibration
                params = self._calculate_gpadc_parameters(result)
                self._plot_voltage_adc_curve(voltage_data, mean_cali, adc_min_cali, adc_max_cali)
                self.update_test_result(params)
                self._export_data = {
                    'params': params,
                    'raw': result,
                    'calibration': {
                        'voltage': voltage_data,
                        'mean_cali': mean_cali,
                        'min_cali': adc_min_cali,
                        'max_cali': adc_max_cali,
                    }
                }

        elif kind == 'high_low_temp':
            if result is not None:
                temp_data, mean_cali, adc_min_cali, adc_max_cali = self._calibration_data(result)
                params = self._calculate_gpadc_parameters(result)
                self._plot_voltage_adc_curve(temp_data, mean_cali, adc_min_cali, adc_max_cali, is_temp_mode=True)
                self.update_test_result(params)
                self._export_data = {
                    'params': params,
                    'raw': result,
                    'calibration': {
                        'voltage': temp_data,
                        'mean_cali': mean_cali,
                        'min_cali': adc_min_cali,
                        'max_cali': adc_max_cali,
                    }
                }
                self.set_system_status("GPADC温度测试完成")

        elif kind == 'temp_consistency':
            if result is not None:
                self._plot_temp_consistency_curves(result)
                self._export_data = {'raw': result}
                self._append_log("[RESULT] Temp Consistency Test completed")
                self.set_system_status("GPADC温度一致性测试完成")

    def _on_test_error(self, err):
        self._append_log(f"[ERROR] Test error: {err}")

    def _on_test_thread_finished(self):
        self._test_worker = None
        if self.test_thread is not None:
            self.test_thread.deleteLater()
            self.test_thread = None
        self.is_test_running = False
        self.start_test_btn.setEnabled(True)
        self.stop_test_btn.setEnabled(False)
        self._update_test_button_state(False)
        self._set_ui_enabled(True)

    def _update_test_button_state(self, running):
        if running:
            self.start_test_btn.setText("■ STOP")
            self.start_test_btn.setObjectName("stop_test_btn")
        else:
            self.start_test_btn.setText(self._start_btn_text)
            self.start_test_btn.setObjectName("start_test_btn")
        self.start_test_btn.style().unpolish(self.start_test_btn)
        self.start_test_btn.style().polish(self.start_test_btn)
        self.start_test_btn.update()

    def _stop_test(self):
        if self._test_worker is not None:
            self._test_worker.request_stop()
        if self.test_thread is not None and self.test_thread.isRunning():
            self.test_thread.quit()
            self.test_thread.wait()
        self._append_log("[INFO] Stopping GPADC test...")

    def export_result(self):
        import openpyxl
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                     Border, Side)
        from openpyxl.utils import get_column_letter
        from pathlib import Path
        import datetime

        if self._export_data is None:
            self._append_log("[WARN] No test result to export.")
            return

        params   = self._export_data['params']
        raw      = self._export_data['raw']
        calib    = self._export_data['calibration']

        results_dir = Path(__file__).parent.parent.parent.parent / "Results"
        results_dir.mkdir(exist_ok=True)

        ts       = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = results_dir / f"GPADC_Result_{ts}.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "GPADC Result"

        hdr_fill   = PatternFill("solid", fgColor="0A1735")
        hdr_font   = Font(bold=True, color="00D39A", size=11)
        val_font   = Font(color="000000", size=10)
        sub_fill   = PatternFill("solid", fgColor="0D1F40")
        sub_font   = Font(bold=True, color="7E96BF", size=10)
        thin_side  = Side(style="thin", color="1B2847")
        thin_border = Border(left=thin_side, right=thin_side,
                             top=thin_side, bottom=thin_side)

        def _hdr(ws, row, col, text):
            c = ws.cell(row=row, column=col, value=text)
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border
            return c

        def _val(ws, row, col, value):
            c = ws.cell(row=row, column=col, value=value)
            c.font = val_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border
            if isinstance(value, float):
                c.number_format = "0.000000"
            return c

        def _sub(ws, row, col, text):
            c = ws.cell(row=row, column=col, value=text)
            c.fill = sub_fill
            c.font = sub_font
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = thin_border
            return c

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 18

        row = 1
        _hdr(ws, row, 1, "GPADC ADC Parameters")
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=4)
        row += 1

        param_items = [
            ("Linearity (R²)",   params.get("linearity",    0.0)),
            ("INL (LSB)",        params.get("inl",          0.0)),
            ("DNL (LSB)",        params.get("dnl",          0.0)),
            ("ENOB (bits)",      params.get("enob",         0.0)),
            ("Gain Error (%)",   params.get("gain_error",   0.0)),
            ("Offset Error (LSB)", params.get("offset_error", 0.0)),
        ]
        _sub(ws, row, 1, "Parameter")
        _sub(ws, row, 2, "Value")
        _sub(ws, row, 3, "Parameter")
        _sub(ws, row, 4, "Value")
        row += 1
        for i in range(0, len(param_items), 2):
            _val(ws, row, 1, param_items[i][0])
            _val(ws, row, 2, param_items[i][1])
            if i + 1 < len(param_items):
                _val(ws, row, 3, param_items[i + 1][0])
                _val(ws, row, 4, param_items[i + 1][1])
            row += 1

        row += 1

        chart_data_start_row = row
        _hdr(ws, row, 1, "Chart Data")
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=4)
        row += 1
        _sub(ws, row, 1, "Input Voltage (V)")
        _sub(ws, row, 2, "Mean Cali (V)")
        _sub(ws, row, 3, "Min Cali (V)")
        _sub(ws, row, 4, "Max Cali (V)")
        row += 1
        chart_data_body_start = row
        v_list   = calib['voltage']
        mc_list  = calib['mean_cali']
        mn_list  = calib['min_cali']
        mx_list  = calib['max_cali']
        for i in range(len(v_list)):
            _val(ws, row, 1, float(v_list[i]))
            _val(ws, row, 2, float(mc_list[i]))
            _val(ws, row, 3, float(mn_list[i]))
            _val(ws, row, 4, float(mx_list[i]))
            row += 1
        chart_data_end = row - 1

        chart_anchor_col = get_column_letter(6)
        if self._chart_image_bytes is not None:
            try:
                from openpyxl.drawing.image import Image as XLImage
                self._chart_image_bytes.seek(0)
                img = XLImage(self._chart_image_bytes)
                img.width  = 600
                img.height = 350
                ws.add_image(img, f"{chart_anchor_col}{chart_data_start_row}")
            except Exception as ex:
                self._append_log(f"[WARN] Embed chart image failed: {ex}")
                self._embed_native_chart(ws, chart_data_start_row, chart_data_body_start,
                                         chart_data_end, chart_anchor_col)
        else:
            self._embed_native_chart(ws, chart_data_start_row, chart_data_body_start,
                                     chart_data_end, chart_anchor_col)

        row += 2

        _hdr(ws, row, 1, "Raw Data")
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=4)
        row += 1
        _sub(ws, row, 1, "Input Voltage (V)")
        _sub(ws, row, 2, "ADC Mean (LSB)")
        _sub(ws, row, 3, "ADC Min (LSB)")
        _sub(ws, row, 4, "ADC Max (LSB)")
        row += 1
        rv_list  = raw['voltage']
        rm_list  = raw['mean']
        rn_list  = raw['min']
        rx_list  = raw['max']
        for i in range(len(rv_list)):
            _val(ws, row, 1, float(rv_list[i]))
            _val(ws, row, 2, float(rm_list[i]))
            _val(ws, row, 3, float(rn_list[i]))
            _val(ws, row, 4, float(rx_list[i]))
            row += 1

        row += 2

        _hdr(ws, row, 1, "Calibrated Data")
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=4)
        row += 1
        _sub(ws, row, 1, "Input Voltage (V)")
        _sub(ws, row, 2, "Mean Cali (V)")
        _sub(ws, row, 3, "Min Cali (V)")
        _sub(ws, row, 4, "Max Cali (V)")
        row += 1
        for i in range(len(v_list)):
            _val(ws, row, 1, float(v_list[i]))
            _val(ws, row, 2, float(mc_list[i]))
            _val(ws, row, 3, float(mn_list[i]))
            _val(ws, row, 4, float(mx_list[i]))
            row += 1

        wb.save(str(filename))
        self._append_log(f"[INFO] Result exported to: {filename}")

    def _embed_native_chart(self, ws, chart_data_start_row, chart_data_body_start,
                            chart_data_end, chart_anchor_col):
        from openpyxl.chart import LineChart, Reference
        chart = LineChart()
        chart.title        = "ADC Transfer Curve (Calibrated)"
        chart.style        = 10
        chart.y_axis.title = "Calibrated Voltage (V)"
        chart.x_axis.title = "Input Voltage (V)"
        chart.width        = 22
        chart.height       = 14
        ref_x    = Reference(ws, min_col=1, min_row=chart_data_body_start, max_row=chart_data_end)
        ref_mean = Reference(ws, min_col=2, min_row=chart_data_body_start - 1, max_row=chart_data_end)
        ref_min  = Reference(ws, min_col=3, min_row=chart_data_body_start - 1, max_row=chart_data_end)
        ref_max  = Reference(ws, min_col=4, min_row=chart_data_body_start - 1, max_row=chart_data_end)
        chart.add_data(ref_mean, titles_from_data=True)
        chart.add_data(ref_min,  titles_from_data=True)
        chart.add_data(ref_max,  titles_from_data=True)
        chart.set_categories(ref_x)
        chart.series[0].graphicalProperties.line.solidFill = "00D39A"
        chart.series[0].graphicalProperties.line.width     = 20000
        chart.series[1].graphicalProperties.line.solidFill = "F0A040"
        chart.series[2].graphicalProperties.line.solidFill = "F0A040"
        ws.add_chart(chart, f"{chart_anchor_col}{chart_data_start_row}")

    def _set_ui_enabled(self, enabled):
        widgets = [
            self.n6705c_combo, self.n6705c_search_btn, self.n6705c_connect_btn,
            self.vt6002_combo, self.vt6002_search_btn, self.vt6002_connect_btn,
            self.dut_combo, self.dut_search_btn, self.uart_keyword,
            self.iic_radio, self.uart_radio,
            self.iic_device_address, self.iic_data_address,
            self.cnt1000_test_btn, self.force_voltage_test_btn, self.high_low_temp_test_btn, self.temp_consistency_test_btn,
            self.voltage_channel,
            self.voltage_min, self.voltage_max, self.voltage_step,
            self.temp_min, self.temp_max, self.temp_step
        ]
        for widget in widgets:
            widget.setEnabled(enabled)

    def get_test_config(self):
        acquisition_mode = 'IIC' if self.iic_radio.isChecked() else 'UART'
        dut_port = ""
        if self.dut_combo.currentText() and self.dut_combo.currentText() != "No serial ports found":
            dut_port = self.dut_combo.currentText().split()[0]

        return {
            'n6705c_connected': self.is_n6705c_connected,
            'vt6002_connected': self.is_vt6002_connected,
            'test_item': self.current_test_item,
            'data_acquisition_mode': acquisition_mode,
            'iic_device_address': self.iic_device_address.text(),
            'iic_data_address': self.iic_data_address.text(),
            'dut_port': dut_port,
            'uart_keyword': self.uart_keyword.text(),
            'voltage_channel': self.voltage_channel.value(),
            'voltage_min': self.voltage_min.value(),
            'voltage_max': self.voltage_max.value(),
            'voltage_step': self.voltage_step.value(),
            'temp_min': self.temp_min.value(),
            'temp_max': self.temp_max.value(),
            'temp_step': self.temp_step.value()
        }

    def update_test_result(self, result):
        if self.current_test_item == self.TEST_1000CNT:
            self.avg_value.setText(f"{result['avg']:.3f}" if 'avg' in result else "---")
            self.min_value.setText(f"{result['min']:.3f}" if 'min' in result else "---")
            self.max_value.setText(f"{result['max']:.3f}" if 'max' in result else "---")
            self.linearity_value.setText("---")
            self.enob_value.setText("---")
            self.dnl_value.setText("---")
            self.inl_value.setText("---")
            self.gain_error_value.setText("---")
            self.offset_error_value.setText("---")
        else:
            self.avg_value.setText("---")
            self.min_value.setText("---")
            self.max_value.setText("---")
            self.linearity_value.setText(f"{result['linearity']:.3f}" if 'linearity' in result else "---")
            self.enob_value.setText(f"{result['enob']:.3f}" if 'enob' in result else "---")
            self.dnl_value.setText(f"{result['dnl']:.3f}" if 'dnl' in result else "---")
            self.inl_value.setText(f"{result['inl']:.3f}" if 'inl' in result else "---")
            self.gain_error_value.setText(f"{result['gain_error']:.3f}" if 'gain_error' in result else "---")
            self.offset_error_value.setText(f"{result['offset_error']:.3f}" if 'offset_error' in result else "---")

    def clear_results(self):
        self.linearity_value.setText("---")
        self.enob_value.setText("---")
        self.dnl_value.setText("---")
        self.inl_value.setText("---")
        self.gain_error_value.setText("---")
        self.offset_error_value.setText("---")
        # 清除1000CNT TEST的结果
        self.avg_value.setText("---")
        self.min_value.setText("---")
        self.max_value.setText("---")

    def set_system_status(self, status, is_error=False):
        pass

    def _append_log(self, text):
        self.log_text.append(text)

    def update_instrument_info(self, instrument_info):
        pass

    def get_test_mode(self):
        return "GPADC Test"

    def set_test_mode(self, mode):
        pass

    def get_test_id(self):
        return "GPADC_TEST_001"

    def set_test_id(self, test_id):
        pass

    
    def gpadc_reg_read_by_cnts(
        self,
        device_addr=0x17,
        reg_addr=0x56,
        iic_weight=10,
        get_reg_cnt=1000,
        return_raw=False,
        stop_check=None,
    ):
        if DEBUG_MOCK:
            if not hasattr(self, "_mock_i2c"):
                self._mock_i2c = MockI2C()
            deviceI2C = self._mock_i2c
        else:
            if not hasattr(self, "deviceI2C"):
                self.deviceI2C = I2CInterface()
                self.set_system_status("I2C接口初始化成功")
            deviceI2C = self.deviceI2C

        raw_data = []

        for _ in range(get_reg_cnt):
            if stop_check and stop_check():
                break
            temp = deviceI2C.read(device_addr, reg_addr, iic_weight)
            raw_data.append(temp)

        # 排序用于统计
        sorted_data = sorted(raw_data)

        reg_min = sorted_data[0]
        reg_max = sorted_data[-1]

        # Trimmed Mean（去掉5%极值）
        trim = max(1, int(len(sorted_data) * 0.05))
        trimmed = sorted_data[trim:-trim] if len(sorted_data) > 2 * trim else sorted_data

        avg = sum(trimmed) / len(trimmed)

        # print(f"{get_reg_cnt} counts: avg={avg:.3f}; max={reg_max}; min={reg_min}")

        if return_raw:
            return avg, reg_max, reg_min, raw_data
        else:
            return avg, reg_max, reg_min

    def gpadc_force_voltage_test(
        self,
        n6705c=None,
        device_addr=0x17,
        reg_addr=0x56,
        iic_weight=10,
        voltage_min=0.1,
        voltage_max=1.8,
        voltage_step=0.05,
        voltage_channel=1,
        stop_check=None,
    ):
        self._test_worker.log.emit(f"[INFO] Running FORCE VOLTAGE TEST with I2C address: 0x{device_addr:x}, Register: 0x{reg_addr:x}")

        if DEBUG_MOCK:
            if not hasattr(self, "_mock_i2c"):
                self._mock_i2c = MockI2C()
            vol_source = MockN6705C()
            vol_source._mock_i2c = self._mock_i2c
        else:
            vol_source = n6705c if n6705c is not None else self.n6705c
            if vol_source is None or not self.is_n6705c_connected:
                self._test_worker.log.emit("[ERROR] N6705C not connected")
                self.set_system_status("错误: N6705C未连接", is_error=True)
                return None

        settle_time = 0.0 if DEBUG_MOCK else 0.5
        step_time   = 0.0 if DEBUG_MOCK else 0.2

        voltage_data = []
        adc_mean = []
        adc_min = []
        adc_max = []

        vol_source.set_voltage(voltage_channel, voltage_min)
        time.sleep(settle_time)

        current_voltage = voltage_min

        while current_voltage <= voltage_max + voltage_step * 0.001:
            if stop_check and stop_check():
                self._test_worker.log.emit("[INFO] Force voltage test stopped by user.")
                break
            vol_source.set_voltage(voltage_channel, current_voltage)
            time.sleep(step_time)

            avg, max_val, min_val = self.gpadc_reg_read_by_cnts(
                device_addr,
                reg_addr,
                iic_weight,
                get_reg_cnt=1000,
                return_raw=False,
                stop_check=stop_check,
            )

            voltage_data.append(current_voltage)
            adc_mean.append(avg)
            adc_min.append(min_val)
            adc_max.append(max_val)

            current_voltage = round(current_voltage + voltage_step, 6)

        self._test_worker.log.emit("===== FORCE VOLTAGE TEST 结果 =====")

        result = {
            "voltage": voltage_data,
            "mean": adc_mean,
            "min": adc_min,
            "max": adc_max,
        }
        return result

    def _calibration_data(self, result):
        adc_raw_data = result["voltage"]
        adc_mean     = result["mean"]
        adc_min      = result["min"]
        adc_max      = result["max"]

        n        = len(adc_raw_data)
        idx_low  = n // 4
        idx_high = (3 * n) // 4

        v_low,  m_low  = adc_raw_data[idx_low],  adc_mean[idx_low]
        v_high, m_high = adc_raw_data[idx_high], adc_mean[idx_high]
        k = (m_high - m_low) / (v_high - v_low)
        b = m_low - k * v_low

        self._append_log(f"[INFO] Calibration: k={k:.6f} (LSB/V), b={b:.6f} (LSB)")

        mean_cali    = [(adc - b) / k for adc in adc_mean]
        adc_min_cali = [(adc - b) / k for adc in adc_min]
        adc_max_cali = [(adc - b) / k for adc in adc_max]

        self._append_log(
            'Voltage,RawMean,mean_cali,Δmean(mV),RawMin,RawMax,min_cali,max_cali,Δmin(mV),Δmax(mV)'
        )
        for i in range(n):
            v_in = adc_raw_data[i]
            delta_mean_mv = (mean_cali[i] - v_in) * 1000
            delta_min_mv = (adc_min_cali[i] - v_in) * 1000
            delta_max_mv = (adc_max_cali[i] - v_in) * 1000
            self._append_log(
                f'{v_in:.4f},{adc_mean[i]:.3f},{mean_cali[i]:.6f},{delta_mean_mv:+.3f}'
                f',{adc_min[i]},{adc_max[i]}'
                f',{adc_min_cali[i]:.6f},{adc_max_cali[i]:.6f}'
                f',{delta_min_mv:+.3f},{delta_max_mv:+.3f}'
            )

        self._append_log(f"[INFO] calibration points: v_low={v_low:.4f}, m_low={m_low:.3f}, v_high={v_high:.4f}, m_high={m_high:.3f}")
        self._append_log(f"[INFO] Calibration: k={k:.6f} (LSB/V), b={b:.6f} (LSB)")

        return adc_raw_data, mean_cali, adc_min_cali, adc_max_cali

    def gpadc_high_low_temp_test(
        self,
        device_addr=0x17,
        reg_addr=0x56,
        iic_weight=10,
        temp_min=0,
        temp_max=100,
        temp_step=1,
        voltage_channel=100,
        stop_check=None,
    ):
        try:
            if DEBUG_MOCK:
                if not hasattr(self, "_mock_i2c"):
                    self._mock_i2c = MockI2C()
                chamber = MockVT6002()
                deviceI2C = self._mock_i2c
            else:
                if not hasattr(self, 'vt6002') or not self.is_vt6002_connected:
                    self._test_worker.log.emit("[ERROR] VT6002 chamber not connected")
                    self.set_system_status("错误: VT6002温箱未连接", is_error=True)
                    return None
                if not hasattr(self, "deviceI2C"):
                    self.deviceI2C = I2CInterface()
                    self.set_system_status("I2C接口初始化成功")
                chamber = self.vt6002
                deviceI2C = self.deviceI2C

            test_data = deviceI2C.read(device_addr, reg_addr, iic_weight)
            self._test_worker.log.emit(f"[INFO] Test data: {test_data:x}")

            temp_data = []
            adc_mean = []
            adc_min = []
            adc_max = []
            adc_raw_all = []

            current_temp = temp_min

            while current_temp <= temp_max + 0.001:
                if stop_check and stop_check():
                    self._test_worker.log.emit("[INFO] High/Low temp test stopped by user.")
                    break

                chamber.set_temperature(current_temp)
                self.set_system_status(f"设置温箱温度到 {current_temp:.1f}°C")

                if DEBUG_MOCK:
                    self._test_worker.log.emit(f"[DEBUG] Temp set to {current_temp:.1f}°C (instant)")
                else:
                    history = []
                    stable_count = 0

                    while True:
                        if stop_check and stop_check():
                            break
                        actual_temp = chamber.get_current_temp()
                        history.append(actual_temp)

                        if len(history) > 10:
                            history.pop(0)

                        if len(history) >= 5:
                            if max(history) - min(history) < 0.2:
                                stable_count += 1
                            else:
                                stable_count = 0

                            if stable_count >= 3:
                                break

                        self._test_worker.log.emit(f"[INFO] Temp stabilizing: target={current_temp:.1f}, actual={actual_temp:.2f}")
                        time.sleep(30)

                    if stop_check and stop_check():
                        self._test_worker.log.emit("[INFO] High/Low temp test stopped by user.")
                        break

                    self.set_system_status(f"DUT温度均衡中: {current_temp:.1f}°C")
                    for _ in range(180):
                        if stop_check and stop_check():
                            break
                        time.sleep(1)

                    if stop_check and stop_check():
                        self._test_worker.log.emit("[INFO] High/Low temp test stopped by user.")
                        break

                if DEBUG_MOCK:
                    self._mock_i2c.set_mock_voltage(current_temp / 100.0)

                avg, max_val, min_val, raw_data = self.gpadc_reg_read_by_cnts(
                    device_addr,
                    reg_addr,
                    iic_weight,
                    get_reg_cnt=1000,
                    return_raw=True,
                    stop_check=stop_check,
                )

                temp_data.append(current_temp)
                adc_mean.append(avg)
                adc_min.append(min_val)
                adc_max.append(max_val)
                adc_raw_all.append(raw_data)

                self._test_worker.log.emit(f"[INFO] T={current_temp:.1f}°C, avg={avg:.3f}, min={min_val}, max={max_val}")

                current_temp += temp_step
                time.sleep(1)
            chamber.set_temperature(25.0)
            self._test_worker.log.emit("===== HIGH/LOW TEMP TEST 结果 =====")
            self._test_worker.log.emit("Temp, RawMean, RawMin, RawMax")
            for i in range(len(temp_data)):
                self._test_worker.log.emit(f"{temp_data[i]:.2f}, {adc_mean[i]:.3f}, {adc_min[i]}, {adc_max[i]}")

            return {
                "voltage": temp_data,
                "mean": adc_mean,
                "min": adc_min,
                "max": adc_max,
            }

        except Exception as e:
            self._test_worker.log.emit(f"[ERROR] {e}")
            logger.error("测试执行错误: %s", e, exc_info=True)
            self.set_system_status(f"错误: {e}", is_error=True)
            return None
    
    def gpadc_temp_consistency_test(
        self,
        device_addr=0x17,
        reg_addr=0x56,
        iic_weight=10,
        temp_min=0,
        temp_max=100,
        temp_step=1,
        voltage_min=0.1,
        voltage_max=1.8,
        voltage_step=0.05,
        voltage_channel=1,
        stop_check=None,
    ):
        self._test_worker.log.emit(f"[INFO] Running TEMP CONSISTENCY TEST with I2C address: 0x{device_addr:x}, Register: 0x{reg_addr:x}")

        if DEBUG_MOCK:
            if not hasattr(self, "_mock_i2c"):
                self._mock_i2c = MockI2C()
            chamber = MockVT6002()
            vol_source = MockN6705C()
            vol_source._mock_i2c = self._mock_i2c
        else:
            if not hasattr(self, 'vt6002') or not self.is_vt6002_connected:
                self._test_worker.log.emit("[ERROR] VT6002 chamber not connected")
                self.set_system_status("错误: VT6002温箱未连接", is_error=True)
                return None
            if self.n6705c is None or not self.is_n6705c_connected:
                self._test_worker.log.emit("[ERROR] N6705C not connected")
                self.set_system_status("错误: N6705C未连接", is_error=True)
                return None
            if not hasattr(self, "deviceI2C"):
                self.deviceI2C = I2CInterface()
                self.set_system_status("I2C接口初始化成功")
            chamber = self.vt6002
            vol_source = self.n6705c

        settle_time = 0.0 if DEBUG_MOCK else 0.5
        step_time   = 0.0 if DEBUG_MOCK else 0.2

        voltage_points = []
        v = voltage_min
        while v <= voltage_max + voltage_step * 0.001:
            voltage_points.append(round(v, 6))
            v = round(v + voltage_step, 6)

        temp_list   = []
        mean_matrix = []
        min_matrix  = []
        max_matrix  = []

        current_temp = temp_min
        while current_temp <= temp_max + 0.001:
            if stop_check and stop_check():
                self._test_worker.log.emit("[INFO] Temp consistency test stopped by user.")
                break

            chamber.set_temperature(current_temp)
            self.set_system_status(f"设置温箱温度到 {current_temp:.1f}°C")

            if DEBUG_MOCK:
                self._test_worker.log.emit(f"[DEBUG] Temp set to {current_temp:.1f}°C (instant)")
            else:
                history = []
                stable_count = 0
                while True:
                    if stop_check and stop_check():
                        break
                    actual_temp = chamber.get_current_temp()
                    history.append(actual_temp)
                    if len(history) > 10:
                        history.pop(0)
                    if len(history) >= 5:
                        if max(history) - min(history) < 0.2:
                            stable_count += 1
                        else:
                            stable_count = 0
                        if stable_count >= 3:
                            break
                    self._test_worker.log.emit(f"[INFO] Temp stabilizing: target={current_temp:.1f}, actual={actual_temp:.2f}")
                    time.sleep(30)

                if stop_check and stop_check():
                    self._test_worker.log.emit("[INFO] Temp consistency test stopped by user.")
                    break

                self.set_system_status(f"DUT温度均衡中: {current_temp:.1f}°C")
                for _ in range(180):
                    if stop_check and stop_check():
                        break
                    time.sleep(1)

                if stop_check and stop_check():
                    self._test_worker.log.emit("[INFO] Temp consistency test stopped by user.")
                    break

            mean_row = []
            min_row  = []
            max_row  = []

            vol_source.set_voltage(voltage_channel, voltage_min)
            time.sleep(settle_time)

            for vpt in voltage_points:
                if stop_check and stop_check():
                    break
                vol_source.set_voltage(voltage_channel, vpt)
                time.sleep(step_time)

                avg, max_val, min_val = self.gpadc_reg_read_by_cnts(
                    device_addr,
                    reg_addr,
                    iic_weight,
                    get_reg_cnt=1000,
                    return_raw=False,
                    stop_check=stop_check,
                )
                mean_row.append(avg)
                min_row.append(min_val)
                max_row.append(max_val)

            if stop_check and stop_check():
                break

            temp_list.append(current_temp)
            mean_matrix.append(mean_row)
            min_matrix.append(min_row)
            max_matrix.append(max_row)

            self._test_worker.log.emit(f"[INFO] T={current_temp:.1f}°C  voltage sweep done ({len(mean_row)} points)")
            current_temp = round(current_temp + temp_step, 6)

        vol_header = "  ".join(f"{v:.3f}" for v in voltage_points)

        self._test_worker.log.emit("===== TEMP CONSISTENCY TEST 结果 (Mean) =====")
        self._test_worker.log.emit("Temp\\Voltage  " + vol_header)
        for i, t in enumerate(temp_list):
            if i < len(mean_matrix):
                row = "  ".join(f"{mean_matrix[i][j]:.1f}" for j in range(len(mean_matrix[i])))
                self._test_worker.log.emit(f"T={t:.1f}  {row}")

        self._test_worker.log.emit("===== TEMP CONSISTENCY TEST 结果 (Min) =====")
        self._test_worker.log.emit("Temp\\Voltage  " + vol_header)
        for i, t in enumerate(temp_list):
            if i < len(min_matrix):
                row = "  ".join(f"{min_matrix[i][j]:.1f}" for j in range(len(min_matrix[i])))
                self._test_worker.log.emit(f"T={t:.1f}  {row}")

        self._test_worker.log.emit("===== TEMP CONSISTENCY TEST 结果 (Max) =====")
        self._test_worker.log.emit("Temp\\Voltage  " + vol_header)
        for i, t in enumerate(temp_list):
            if i < len(max_matrix):
                row = "  ".join(f"{max_matrix[i][j]:.1f}" for j in range(len(max_matrix[i])))
                self._test_worker.log.emit(f"T={t:.1f}  {row}")

        mean_cali_matrix = []
        min_cali_matrix = []
        max_cali_matrix = []
        for i, t in enumerate(temp_list):
            if i < len(mean_matrix):
                row_result = {
                    "voltage": voltage_points,
                    "mean": mean_matrix[i],
                    "min": min_matrix[i],
                    "max": max_matrix[i],
                }
                _, mean_cali, min_cali, max_cali = self._calibration_data(row_result)
                mean_cali_matrix.append(mean_cali)
                min_cali_matrix.append(min_cali)
                max_cali_matrix.append(max_cali)

        self._test_worker.log.emit("===== TEMP CONSISTENCY TEST 校准结果 (Mean Cali) =====")
        self._test_worker.log.emit("Temp\\Voltage  " + vol_header)
        for i, t in enumerate(temp_list):
            if i < len(mean_cali_matrix):
                row = "  ".join(f"{mean_cali_matrix[i][j]:.6f}" for j in range(len(mean_cali_matrix[i])))
                self._test_worker.log.emit(f"T={t:.1f}  {row}")

        self._test_worker.log.emit("===== TEMP CONSISTENCY TEST 校准结果 (Min Cali) =====")
        self._test_worker.log.emit("Temp\\Voltage  " + vol_header)
        for i, t in enumerate(temp_list):
            if i < len(min_cali_matrix):
                row = "  ".join(f"{min_cali_matrix[i][j]:.6f}" for j in range(len(min_cali_matrix[i])))
                self._test_worker.log.emit(f"T={t:.1f}  {row}")

        self._test_worker.log.emit("===== TEMP CONSISTENCY TEST 校准结果 (Max Cali) =====")
        self._test_worker.log.emit("Temp\\Voltage  " + vol_header)
        for i, t in enumerate(temp_list):
            if i < len(max_cali_matrix):
                row = "  ".join(f"{max_cali_matrix[i][j]:.6f}" for j in range(len(max_cali_matrix[i])))
                self._test_worker.log.emit(f"T={t:.1f}  {row}")

        return {
            "temp":    temp_list,
            "voltage": voltage_points,
            "mean":    mean_matrix,
            "min":     min_matrix,
            "max":     max_matrix,
            "mean_cali": mean_cali_matrix,
            "min_cali":  min_cali_matrix,
            "max_cali":  max_cali_matrix,
        }

    def _calculate_gpadc_parameters(self, force_voltage_result):
        """
        基于 gpadc_force_voltage_test 的返回值计算 GPADC 性能参数。

        force_voltage_result 格式:
            {
                "voltage": [v0, v1, ...],   # 输入电压 (V)
                "mean":    [m0, m1, ...],   # 每个电压点的 ADC 均值
                "min":     [n0, n1, ...],   # 每个电压点的 ADC 最小值
                "max":     [x0, x1, ...],   # 每个电压点的 ADC 最大值
            }

        返回 params dict，包含:
            linearity, dnl, inl, enob, gain_error, offset_error,
            avg, min, max
        """
        import numpy as np

        voltage = np.array(force_voltage_result["voltage"], dtype=float)
        adc_mean = np.array(force_voltage_result["mean"], dtype=float)
        adc_min = np.array(force_voltage_result["min"], dtype=float)
        adc_max = np.array(force_voltage_result["max"], dtype=float)

        n = len(voltage)
        params = {}

        params['avg'] = float(np.mean(adc_mean))
        params['max'] = float(np.max(adc_max))
        params['min'] = float(np.min(adc_min))

        if n < 2:
            params.update({'linearity': 0.0, 'dnl': 0.0, 'inl': 0.0,
                           'enob': 0.0, 'gain_error': 0.0, 'offset_error': 0.0})
            return params

        # ── 线性拟合 voltage → adc_mean ──────────────────────────────
        coeffs = np.polyfit(voltage, adc_mean, 1)
        slope = coeffs[0]       # LSB/V
        intercept = coeffs[1]
        adc_ideal = np.polyval(coeffs, voltage)

        # Linearity: R²（越接近1越好）
        ss_res = np.sum((adc_mean - adc_ideal) ** 2)
        ss_tot = np.sum((adc_mean - np.mean(adc_mean)) ** 2)
        params['linearity'] = float(1.0 - ss_res / ss_tot) if ss_tot > 0 else 1.0

        # ── INL (Integral Non-Linearity) ──────────────────────────────
        # 每个采样点与理想线性值的偏差，取峰峰值的一半（LSB）
        inl_values = adc_mean - adc_ideal
        params['inl'] = float(np.max(np.abs(inl_values)))

        # ── DNL (Differential Non-Linearity) ─────────────────────────
        # 相邻步进的实际ADC增量 vs 理想增量之差（LSB）
        ideal_step = slope * (voltage[1] - voltage[0])   # 理想单步增量（LSB）
        actual_steps = np.diff(adc_mean)
        if ideal_step != 0:
            dnl_values = (actual_steps - ideal_step) / abs(ideal_step)
            params['dnl'] = float(np.max(np.abs(dnl_values)))
        else:
            params['dnl'] = 0.0

        # ── ENOB (Effective Number of Bits) ──────────────────────────
        # 用每个采样点的噪声范围（max-min）估算 RMS 噪声，再计算 SINAD→ENOB
        noise_pp = adc_max - adc_min          # peak-to-peak 噪声
        noise_rms = np.mean(noise_pp) / (2 * np.sqrt(3))   # 均匀分布近似
        adc_full_range = params['max'] - params['min']
        if noise_rms > 0 and adc_full_range > 0:
            snr_db = 20 * np.log10(adc_full_range / (noise_rms * 2 * np.sqrt(3)))
            params['enob'] = float(max(0.0, (snr_db - 1.76) / 6.02))
        else:
            params['enob'] = 0.0

        # ── Gain Error ────────────────────────────────────────────────
        # 理想斜率 = (ADC满量程范围) / (电压满量程范围)
        volt_range = float(voltage[-1] - voltage[0])
        adc_range = float(adc_mean[-1] - adc_mean[0])
        if volt_range > 0:
            ideal_slope = adc_range / volt_range
            params['gain_error'] = float(((slope - ideal_slope) / ideal_slope) * 100) if ideal_slope != 0 else 0.0
        else:
            params['gain_error'] = 0.0

        # ── Offset Error ──────────────────────────────────────────────
        # 拟合线在最小电压处的预测值与实际测量值的偏差（LSB）
        params['offset_error'] = float(adc_ideal[0] - adc_mean[0])
        self._append_log(f"[RESULT] INL={params['inl']:.3f}, DNL={params['dnl']:.3f}, ENOB={params['enob']:.3f}, Gain Error: {params['gain_error']:.3f}%, Offset Error: {params['offset_error']:.3f} LSB")
        return params
    
    def _clear_chart_placeholder(self):
        existing = self.chart_placeholder.layout()
        if existing is not None:
            while existing.count():
                item = existing.takeAt(0)
                widget = item.widget()
                if widget is not None:
                    widget.deleteLater()
                else:
                    sub = item.layout()
                    if sub is not None:
                        while sub.count():
                            sub_item = sub.takeAt(0)
                            sub_widget = sub_item.widget()
                            if sub_widget is not None:
                                sub_widget.deleteLater()

    def _plot_voltage_adc_curve(self, voltage_data, mean_cali, adc_min_cali=None, adc_max_cali=None, is_temp_mode=False):
        try:
            import pyqtgraph as pg
            import numpy as np

            self._clear_chart_placeholder()
            layout = self.chart_placeholder.layout()
            if layout is None:
                layout = QVBoxLayout(self.chart_placeholder)
            layout.setContentsMargins(14, 14, 14, 10)
            layout.setSpacing(8)

            legend_row = QHBoxLayout()
            legend_row.addStretch()
            if is_temp_mode:
                cali_legend = QLabel("● Calibrated Temperature (°C)")
                x_axis_title = "Input Temperature (°C)"
                y_axis_title = "Calibrated Temperature (°C)"
            else:
                cali_legend = QLabel("● Calibrated Voltage (V)")
                x_axis_title = "Input Voltage (V)"
                y_axis_title = "Calibrated Voltage (V)"
            cali_legend.setStyleSheet("color: #00d39a; font-size: 12px;")
            ideal_legend = QLabel("● Ideal (y = x)")
            ideal_legend.setStyleSheet("color: #7e96bf; font-size: 12px;")
            band_legend = QLabel("▨ Max/Min Error Band")
            band_legend.setStyleSheet("color: #f0a040; font-size: 12px;")
            legend_row.addWidget(cali_legend)
            legend_row.addSpacing(16)
            legend_row.addWidget(ideal_legend)
            legend_row.addSpacing(16)
            legend_row.addWidget(band_legend)
            legend_row.addStretch()
            layout.addLayout(legend_row)

            pw = pg.PlotWidget()
            pw.setBackground("#0a1735")
            pw.showGrid(x=True, y=True, alpha=0.15)
            pw.setLabel("left",   y_axis_title, color="#a0b4d8")
            pw.setLabel("bottom", x_axis_title, color="#a0b4d8")

            for axis_name in ("left", "bottom"):
                axis = pw.getAxis(axis_name)
                axis.setTextPen(pg.mkPen("#a0b4d8"))
                axis.setPen(pg.mkPen("#3a4f7a"))

            x = np.array(voltage_data, dtype=float)
            y = np.array(mean_cali,    dtype=float)

            if adc_min_cali is not None and adc_max_cali is not None:
                y_min = np.array(adc_min_cali, dtype=float)
                y_max = np.array(adc_max_cali, dtype=float)

                x_band = np.concatenate([x, x[::-1]])
                y_band = np.concatenate([y_max, y_min[::-1]])
                fill = pg.PlotDataItem(x_band, y_band,
                                       pen=pg.mkPen(color="#f0a040", width=1))
                fill_under = pg.FillBetweenItem(
                    pg.PlotDataItem(x, y_max),
                    pg.PlotDataItem(x, y_min),
                    brush=pg.mkBrush(240, 160, 64, 50)
                )
                pw.addItem(fill_under)

                pw.plot(x, y_max, pen=pg.mkPen(color="#f0a040", width=1,
                        style=pg.QtCore.Qt.DashLine))
                pw.plot(x, y_min, pen=pg.mkPen(color="#f0a040", width=1,
                        style=pg.QtCore.Qt.DashLine))

            pw.plot(x, x, pen=pg.mkPen(color="#7e96bf", width=1,
                    style=pg.QtCore.Qt.DashLine))

            pw.plot(x, y, pen=pg.mkPen(color="#00d39a", width=2),
                    symbol="o", symbolSize=5,
                    symbolBrush="#00d39a", symbolPen=None)

            layout.addWidget(pw, 1)

            x_label = QLabel(x_axis_title)
            x_label.setAlignment(Qt.AlignCenter)
            x_label.setObjectName("muted_label")
            layout.addWidget(x_label)

            try:
                import io
                from pyqtgraph.exporters import ImageExporter
                from PySide6.QtCore import QBuffer, QIODevice
                from PySide6.QtGui import QImage
                exporter = ImageExporter(pw.plotItem)
                exporter.parameters()['width'] = 1200
                exporter.parameters()['height'] = 700
                result = exporter.export(toBytes=True)
                if isinstance(result, QImage):
                    qbuf = QBuffer()
                    qbuf.open(QIODevice.WriteOnly)
                    result.save(qbuf, "PNG")
                    raw = bytes(qbuf.data())
                    qbuf.close()
                else:
                    raw = bytes(result)
                buf = io.BytesIO(raw)
                self._chart_image_bytes = buf
            except Exception as ex:
                self._append_log(f"[WARN] Chart snapshot failed: {ex}")
                self._chart_image_bytes = None

        except Exception as e:
            self._append_log(f"[ERROR] Error plotting voltage-ADC curve: {e}")
            logger.error("Error plotting voltage-ADC curve: %s", e, exc_info=True)

    def _plot_temp_consistency_curves(self, result):
        try:
            import pyqtgraph as pg
            import numpy as np

            temp_list    = result["temp"]
            voltage_pts  = result["voltage"]
            mean_matrix  = result["mean"]
            min_matrix   = result["min"]
            max_matrix   = result["max"]

            self._clear_chart_placeholder()
            layout = self.chart_placeholder.layout()
            if layout is None:
                layout = QVBoxLayout(self.chart_placeholder)
            layout.setContentsMargins(14, 14, 14, 10)
            layout.setSpacing(8)

            palette = [
                "#00d39a", "#f0a040", "#7e96bf", "#e05c5c",
                "#a78bfa", "#34d399", "#fb923c", "#60a5fa",
            ]

            legend_row = QHBoxLayout()
            legend_row.addStretch()
            for i, t in enumerate(temp_list):
                color = palette[i % len(palette)]
                lbl = QLabel(f"● {t:.1f}°C")
                lbl.setStyleSheet(f"color: {color}; font-size: 11px;")
                legend_row.addWidget(lbl)
                legend_row.addSpacing(8)
            legend_row.addStretch()
            layout.addLayout(legend_row)

            pw = pg.PlotWidget()
            pw.setBackground("#0a1735")
            pw.showGrid(x=True, y=True, alpha=0.15)
            pw.setLabel("left",   "ADC Code",         color="#a0b4d8")
            pw.setLabel("bottom", "Input Voltage (V)", color="#a0b4d8")

            for axis_name in ("left", "bottom"):
                axis = pw.getAxis(axis_name)
                axis.setTextPen(pg.mkPen("#a0b4d8"))
                axis.setPen(pg.mkPen("#3a4f7a"))

            x = np.array(voltage_pts, dtype=float)

            for i, t in enumerate(temp_list):
                if i >= len(mean_matrix):
                    break
                color = palette[i % len(palette)]
                mean_row = np.array(mean_matrix[i], dtype=float)
                min_row  = np.array(min_matrix[i],  dtype=float)
                max_row  = np.array(max_matrix[i],  dtype=float)

                n = min(len(x), len(mean_row))

                fill = pg.FillBetweenItem(
                    pg.PlotDataItem(x[:n], max_row[:n]),
                    pg.PlotDataItem(x[:n], min_row[:n]),
                    brush=pg.mkBrush(
                        int(color[1:3], 16),
                        int(color[3:5], 16),
                        int(color[5:7], 16),
                        35,
                    )
                )
                pw.addItem(fill)

                pw.plot(x[:n], max_row[:n],
                        pen=pg.mkPen(color=color, width=1,
                                     style=pg.QtCore.Qt.DashLine))
                pw.plot(x[:n], min_row[:n],
                        pen=pg.mkPen(color=color, width=1,
                                     style=pg.QtCore.Qt.DashLine))
                pw.plot(x[:n], mean_row[:n],
                        pen=pg.mkPen(color=color, width=2),
                        symbol="o", symbolSize=4,
                        symbolBrush=color, symbolPen=None,
                        name=f"{t:.1f}°C")

            layout.addWidget(pw, 1)

            x_label = QLabel("Input Voltage (V)")
            x_label.setAlignment(Qt.AlignCenter)
            x_label.setObjectName("muted_label")
            layout.addWidget(x_label)

            try:
                import io
                from pyqtgraph.exporters import ImageExporter
                from PySide6.QtCore import QBuffer, QIODevice
                from PySide6.QtGui import QImage
                exporter = ImageExporter(pw.plotItem)
                exporter.parameters()['width'] = 1200
                exporter.parameters()['height'] = 700
                snap = exporter.export(toBytes=True)
                if isinstance(snap, QImage):
                    qbuf = QBuffer()
                    qbuf.open(QIODevice.WriteOnly)
                    snap.save(qbuf, "PNG")
                    raw = bytes(qbuf.data())
                    qbuf.close()
                else:
                    raw = bytes(snap)
                self._chart_image_bytes = io.BytesIO(raw)
            except Exception as ex:
                self._append_log(f"[WARN] Chart snapshot failed: {ex}")
                self._chart_image_bytes = None

        except Exception as e:
            self._append_log(f"[ERROR] Error plotting temp consistency curves: {e}")
            logger.error("Error plotting temp consistency curves: %s", e, exc_info=True)

    def _plot_temperature_adc_curve(self, temp_data, adc_data):
        """绘制温度-ADC曲线到UI"""
        try:
            self._clear_chart_placeholder()
            chart_placeholder_layout = self.chart_placeholder.layout()
            if chart_placeholder_layout is None:
                chart_placeholder_layout = QVBoxLayout(self.chart_placeholder)
            chart_placeholder_layout.setContentsMargins(22, 22, 22, 18)
            chart_placeholder_layout.setSpacing(12)
            
            # 添加图例
            legend_row = QHBoxLayout()
            legend_row.addStretch()
            
            temp_legend = QLabel("↔ Temperature (°C)")
            temp_legend.setStyleSheet("color: #00d39a; font-size: 12px;")
            adc_legend = QLabel("↔ ADC Value")
            adc_legend.setStyleSheet("color: #59a8ff; font-size: 12px;")
            
            legend_row.addWidget(temp_legend)
            legend_row.addWidget(adc_legend)
            legend_row.addStretch()
            chart_placeholder_layout.addLayout(legend_row)
            
            # 创建绘图区域
            plot_area = QFrame()
            plot_area.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            plot_area.setStyleSheet("""
                QFrame {
                    background-color: transparent;
                    border-left: 1px solid #6f7fa5;
                    border-bottom: 1px solid #6f7fa5;
                    border-top: 1px dashed rgba(126,150,191,0.18);
                    border-right: none;
                    border-radius: 0px;
                }
            """)
            
            # 简单的曲线绘制（使用QLabel模拟，实际项目中应该使用专业绘图库）
            import numpy as np
            
            # 创建一个简单的文本显示，显示曲线数据
            curve_data = """
Temperature (°C) | ADC Value
----------------|----------
"""
            for t, adc in zip(temp_data, adc_data):
                curve_data += f"{t:>15.1f} | {adc:>10.3f}\n"
            
            # 添加统计信息
            curve_data += "\n" + "="*40 + "\n"
            curve_data += f"Min Temp: {min(temp_data):.1f}°C\n"
            curve_data += f"Max Temp: {max(temp_data):.1f}°C\n"
            curve_data += f"Min ADC:  {min(adc_data):.3f}\n"
            curve_data += f"Max ADC:  {max(adc_data):.3f}\n"
            curve_data += f"Avg ADC:  {np.mean(adc_data):.3f}\n"
            
            data_label = QLabel(curve_data)
            data_label.setStyleSheet("color: #d8e3ff; font-family: 'Consolas', monospace; font-size: 10px;")
            data_label.setWordWrap(True)
            
            # 创建滚动区域以显示大量数据
            scroll_area = QScrollArea()
            scroll_area.setWidget(data_label)
            scroll_area.setWidgetResizable(True)
            scroll_area.setStyleSheet("background: transparent; border: none;" + SCROLL_AREA_STYLE)
            
            # 添加到布局
            chart_placeholder_layout.addWidget(scroll_area, 1)
            
            # 添加轴标签
            x_label = QLabel("Temperature (°C)")
            x_label.setAlignment(Qt.AlignCenter)
            x_label.setObjectName("muted_label")
            chart_placeholder_layout.addWidget(x_label)
            
            # 在实际项目中，应该使用matplotlib或PyQtGraph等专业绘图库
            # 这里只是一个简单的模拟实现
            
        except Exception as e:
            self._append_log(f"[ERROR] Error plotting temperature-ADC curve: {e}")
            logger.error("Error plotting temperature-ADC curve: %s", e, exc_info=True)
            
            # 如果绘图失败，显示错误信息
            error_label = QLabel(f"绘图失败: {str(e)}")
            error_label.setStyleSheet("color: #ff5a7a; font-size: 14px;")
            error_label.setAlignment(Qt.AlignCenter)

            self._clear_chart_placeholder()
            chart_placeholder_layout = self.chart_placeholder.layout()
            if chart_placeholder_layout is None:
                chart_placeholder_layout = QVBoxLayout(self.chart_placeholder)
            chart_placeholder_layout.addWidget(error_label, 1, Qt.AlignCenter)


if __name__ == "__main__":
    import sys
    import logging
    from PySide6.QtCore import qInstallMessageHandler, QtMsgType
    from log_config import setup_logging
    setup_logging()

    app = QApplication(sys.argv)
    app.setStyle("Fusion")

    def custom_message_handler(msg_type, context, message):
        if msg_type == QtMsgType.QtWarningMsg and "QPainter::end" in message:
            return
        logging.getLogger(__name__).debug("%s:%s - %s", context.file, context.line, message)
    qInstallMessageHandler(custom_message_handler)

    gpadc_test_ui = GPADCTestUI()
    gpadc_test_ui.setWindowTitle("GPADC Test System")
    gpadc_test_ui.resize(1920, 900)
    gpadc_test_ui.show()

    sys.exit(app.exec())
