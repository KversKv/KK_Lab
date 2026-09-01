"""
GPADC测试UI组件
修复左侧滚动区域宽度与显示不完整问题
"""
# run cmd:
# python -m ui d:\CodeProject\TRAE_Projects\KK_Lab\ui\gpadc_test_ui.py


from ui.widgets.dark_combobox import DarkComboBox
from ui.styles import SCROLL_AREA_STYLE, START_BTN_STYLE, update_start_btn_state
from ui.widgets.button import update_connect_button_state
from ui.modules.execution_logs_module_frame import ExecutionLogsFrame
from ui.modules.n6705c_module_frame import N6705CConnectionMixin
from ui.modules.chamber_module_frame import ChamberConnectionMixin
from ui.modules.serialCom_module.serialCom_module_frame import SerialComMixin, MODE_FULL
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QLineEdit, QGridLayout, QSpinBox, QDoubleSpinBox, QFrame, QRadioButton,
    QButtonGroup, QApplication, QSizePolicy, QStackedWidget, QScrollArea,
    QTextEdit, QProgressBar, QListWidget, QListWidgetItem, QAbstractItemView,
    QSplitter, QMenu, QInputDialog
)
from PySide6.QtCore import Qt, Signal, QThread, QTimer
from PySide6.QtGui import QFont, QColor, QBrush, QAction
import datetime
import math
import queue
import random
import time
import threading
from typing import Any

import sys
from pathlib import Path

from lib.i2c.i2c_interface_x64 import I2CInterface

from log_config import get_logger
from debug_config import DEBUG_MOCK
from instruments.mock.mock_instruments import MockChamber, MockI2C, MockN6705C
from instruments.chambers import TemperatureStabilizer
from ui.theme import Colors, FontSizes, Radius, Spacing, FONT_MONO
from ui.styles import get_page_base_qss
from core.pmu_test.gpadc import (
    TestWorker as _TestWorker,
    compute_reg_stats,
    compute_calibration,
    compute_detailed_stats,
    parse_uart_gpadc_raw,
)
from core.ai.page_contract import (
    CAP_APPLY_CONFIG,
    CAP_GET_CONFIG,
    CAP_GET_RESULT,
    CAP_START_TEST,
    CAP_STOP_TEST,
)
from core.ai.ui_action_registry import UIActionSpec

logger = get_logger(__name__)

# AI 回填可视化（AIAssist_PageScopedControlPlan.md §4.2 / Phase 3）：
# 被 AI 修改的控件临时高亮边框色 + 持续时长。
_AI_HIGHLIGHT_QSS = "border: 1px solid #15d1a3;"
_AI_HIGHLIGHT_MS = 1500


class GPADCTestUI(N6705CConnectionMixin, ChamberConnectionMixin, SerialComMixin, QWidget):
    """GPADC测试UI组件"""

    connection_status_changed = Signal(bool)
    # 测试结束 → AI 异步动作回灌续跑（与 Orchestrator 同契约，§4 / S3-2）。
    # MainWindow._ai_on_sequence_finished_resume 监听本信号，回灌 pending 任务。
    sequence_execution_finished = Signal(bool, str)
    # worker 线程 → 主线程的系统状态栏更新通道（见 set_system_status 重写）
    system_status_requested = Signal(str, bool)

    TEST_1000CNT = "1000CNT TEST"
    TEST_FORCE_VOLTAGE = "Force Voltage Test"
    TEST_HIGH_LOW_TEMP = "High-Low Temp Test"
    TEST_TEMP_CONSISTENCY = "Temp Consistency Test"

    # 最近测试记录上限（超出丢弃最旧）
    RECENT_TEST_LIMIT = 10
    # 支持曲线对比的测试类型
    _CURVE_KINDS = ('force_voltage', 'high_low_temp', 'temp_consistency')
    # 对比曲线调色板（与 temp_consistency 图共用色系）
    _COMPARE_PALETTE = [
        "#00d39a", "#f0a040", "#5b9cf5", "#e05c5c",
        "#a78bfa", "#34d399", "#fb923c", "#60a5fa",
    ]

    # 扫描响应保护：电压扫描前 N 点 Raw 均值极差（LSB）低于阈值或无净增量 → 判定 DUT 无响应
    _SWEEP_GUARD_POINTS = 3
    _SWEEP_GUARD_MIN_SPREAD_LSB = 2

    INSTRUMENT_MAP = {
        TEST_1000CNT: [],
        TEST_FORCE_VOLTAGE: ["n6705c"],
        TEST_HIGH_LOW_TEMP: ["n6705c", "chamber"],
        TEST_TEMP_CONSISTENCY: ["n6705c", "chamber"],
    }

    def __init__(self, n6705c_top=None, instrument_manager=None, ui_action_registry=None):
        super().__init__()

        self._instrument_manager = instrument_manager
        self._ui_action_registry = ui_action_registry
        # 状态栏跨线程更新通道：worker 线程的 set_system_status 经此队列化回主线程
        self._gui_thread = QThread.currentThread()
        self.system_status_requested.connect(self._apply_system_status)
        self.init_n6705c_connection(n6705c_top, instrument_manager=instrument_manager)
        self.init_chamber_connection(instrument_manager=instrument_manager)
        self.init_serial_connection(mode=MODE_FULL, prefix="DUT")

        self.dut_serial = None
        self.is_dut_connected = False
        self.available_dut_ports = []

        self._uart_rx_queue = queue.Queue()
        self._uart_keyword_snapshot = ""
        self._acq_mode_snapshot = 'IIC'
        self._calib_points_snapshot = None

        self.is_test_running = False
        self._start_btn_text = "▶ START TEST"
        self.test_thread = None
        self._test_worker = None
        self._search_thread = None
        self._search_worker = None
        self._export_data = None
        self._chart_image_bytes = None

        # 最近测试记录（本会话内，用于历史曲线对比与载入）
        self._recent_test_records: list[dict] = []
        self._recent_test_seq = 0
        # 当前从 Recent 载入的记录（切换 Curve View 时据此重绘）
        self._loaded_record = None
        # Recent 管理栏折叠前的宽度记忆
        self._recent_panel_sizes = [1200, 260]
        # 列表当前选中（高亮目标）的记录 id，对比图中该记录加粗、其余淡化
        self._highlight_record_id = None

        self._setup_style()
        self._create_layout()
        self._init_ui_elements()
        self._register_ai_ui_actions()
        self.sync_n6705c_from_top()

    def _register_ai_ui_actions(self):
        """§5b.5：登记本页无专用接口的按钮为 AI 可触发的具名 UI 动作。

        handler 复用按钮原槽；启停已有专用契约，不在此重复登记。
        """
        registry = self._ui_action_registry
        if registry is None:
            return

        def _wrap(label, fn):
            def _run() -> tuple[bool, str]:
                try:
                    fn()
                    return True, f"{label} 已执行。"
                except Exception as exc:  # noqa: BLE001
                    logger.error("%s 执行失败", label, exc_info=True)
                    return False, f"{label} 执行失败：{exc}"
            return _run

        registry.register_many([
            UIActionSpec(
                id="pmu_gpadc.export_result",
                label="导出结果",
                page_key="pmu_gpadc",
                handler=_wrap("导出结果", self.export_result),
                risk="low",
                confirm=False,
                enabled_when=lambda: self._export_data is not None,
                description="导出当前 GPADC 测试结果为 Excel。需已有测试结果数据。",
            ),
        ])

    def _setup_style(self):
        font = QFont("Segoe UI", 9)
        self.setFont(font)

        page_extra = f"""
            QFrame#page {{
                background-color: {Colors.bg_secondary};
            }}

            QFrame#panel,
            QFrame#chart_panel,
            QFrame#metric_card,
            QFrame#config_inner_panel {{
                background-color: {Colors.bg_card};
                border: 1px solid {Colors.border_primary};
                border-radius: {Radius.card}px;
            }}

            QFrame#left_scroll_content {{
                background: transparent;
                border: none;
            }}

            QLabel#title_label {{
                font-size: {FontSizes.title};
                font-weight: 700;
                color: #ffffff;
            }}

            QLabel#subtitle_label {{
                font-size: {FontSizes.subtitle};
                color: {Colors.text_muted};
            }}

            QLabel#section_title {{
                font-size: {FontSizes.body};
                font-weight: 700;
                color: #ffffff;
            }}

            QLabel#muted_label {{
                color: {Colors.text_muted};
                font-size: {FontSizes.caption};
            }}

            QLabel#metric_name {{
                color: {Colors.text_muted};
                font-size: {FontSizes.caption};
                font-weight: 600;
            }}

            QLabel#metric_value_green {{
                color: {Colors.success};
                font-size: 16px;
                font-weight: 700;
            }}

            QLabel#metric_value_blue {{
                color: {Colors.info};
                font-size: 16px;
                font-weight: 700;
            }}

            QLabel#metric_value_yellow {{
                color: {Colors.warning};
                font-size: 16px;
                font-weight: 700;
            }}

            QSpinBox::up-button, QSpinBox::down-button,
            QDoubleSpinBox::up-button, QDoubleSpinBox::down-button {{
                width: 0px; height: 0px; border: none;
            }}

            QPushButton#connect_btn {{
                background-color: rgba(0, 211, 154, 0.14);
                color: {Colors.success};
                border: 1px solid rgba(0, 211, 154, 0.25);
                font-weight: 600;
            }}

            QPushButton#danger_btn {{
                background-color: rgba(255, 90, 122, 0.14);
                color: #ff6f8e;
                border: 1px solid rgba(255, 90, 122, 0.25);
                font-weight: 600;
            }}
""" + START_BTN_STYLE + f"""
            QPushButton#tool_btn {{
                min-height: 28px;
                border-radius: {Radius.small}px;
                background-color: {Colors.bg_card};
                padding: 4px 10px;
            }}

            QPushButton#tool_btn:checked {{
                background-color: rgba(91, 156, 245, 0.18);
                border: 1px solid {Colors.info};
            }}

            /* 手柄与 log_splitter.qss 同规范：默认透明，悬停/拖拽才显色反馈 */
            QSplitter#recent_curve_splitter::handle {{
                background-color: transparent;
            }}

            QSplitter#recent_curve_splitter::handle:horizontal {{
                width: 2px;
            }}

            QSplitter#recent_curve_splitter::handle:horizontal:hover {{
                background-color: #18284d;
            }}

            QSplitter#recent_curve_splitter::handle:horizontal:pressed {{
                background-color: #5b7cff;
            }}

            QRadioButton {{
                background: transparent;
                color: {Colors.text_secondary};
                spacing: 8px;
            }}

            QRadioButton::indicator {{
                width: 14px;
                height: 14px;
            }}

            QRadioButton::indicator:unchecked {{
                border: 1px solid #4a5e8e;
                border-radius: 7px;
                background: #071126;
            }}

            QRadioButton::indicator:checked {{
                border: 1px solid #4ca8ff;
                border-radius: 7px;
                background: #4ca8ff;
            }}

            QTextEdit {{
                background-color: #050d1e;
                border: 1px solid #0e1e40;
                border-radius: {Radius.small}px;
                color: #8abaff;
                font-size: {FontSizes.caption};
                font-family: {FONT_MONO};
                padding: 6px;
            }}

            QProgressBar {{
                background-color: #152749;
                border: none;
                border-radius: 4px;
                text-align: center;
                color: #b7c8ea;
                min-height: 8px;
                max-height: 8px;
            }}

            QProgressBar::chunk {{
                background-color: {Colors.accent_primary};
                border-radius: 4px;
            }}

            QListWidget#recent_test_list {{
                background-color: #050d1e;
                border: 1px solid #0e1e40;
                border-radius: {Radius.small}px;
                color: #b7c8ea;
                font-size: {FontSizes.caption};
                padding: 4px;
                outline: none;
            }}

            QListWidget#recent_test_list::item {{
                background: transparent;
                border: none;
                padding: 3px 6px;
                border-radius: 4px;
            }}

            QListWidget#recent_test_list::item:hover {{
                background: rgba(91, 156, 245, 0.12);
            }}

            QListWidget#recent_test_list::item:selected {{
                background: rgba(91, 156, 245, 0.22);
            }}

            QLabel#recent_hint_label {{
                color: {Colors.text_muted};
                font-size: {FontSizes.caption};
            }}
        """ + SCROLL_AREA_STYLE
        self.setStyleSheet(get_page_base_qss() + page_extra)

    def _create_metric_card(self, title, value="---", value_object_name="metric_value_green"):
        card = QFrame()
        card.setObjectName("metric_card")
        card.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        card.setMinimumHeight(72)

        layout = QVBoxLayout(card)
        layout.setContentsMargins(16, 8, 16, 8)
        layout.setSpacing(4)

        name_label = QLabel(title)
        name_label.setObjectName("metric_name")
        name_label.setAlignment(Qt.AlignCenter)

        value_label = QLabel(value)
        value_label.setObjectName(value_object_name)
        value_label.setAlignment(Qt.AlignCenter)

        layout.addWidget(name_label)
        layout.addWidget(value_label)
        return card, value_label

    def _create_recent_tests_panel(self):
        """Curve 右侧最近测试管理栏：列表 + Curve View 选项 + 对比/载入/清空。"""
        panel = QFrame()
        panel.setObjectName("panel")
        panel.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
        panel.setMinimumWidth(220)

        layout = QVBoxLayout(panel)
        layout.setContentsMargins(12, 10, 12, 10)
        layout.setSpacing(6)

        title_row = QHBoxLayout()
        title_row.setSpacing(6)
        title = QLabel("Recent Tests")
        title.setObjectName("section_title")
        title.setStyleSheet("border: none;")
        title_row.addWidget(title)
        title_row.addStretch()

        self.compare_recent_btn = QPushButton("Compare")
        self.compare_recent_btn.setObjectName("tool_btn")
        self.compare_recent_btn.setToolTip("对比勾选的多条测试记录曲线")
        self.load_recent_btn = QPushButton("Load")
        self.load_recent_btn.setObjectName("tool_btn")
        self.load_recent_btn.setToolTip("载入勾选记录的曲线与指标")
        self.clear_recent_btn = QPushButton("Clear")
        self.clear_recent_btn.setObjectName("tool_btn")
        self.clear_recent_btn.setToolTip("清空最近测试记录")

        title_row.addWidget(self.compare_recent_btn)
        title_row.addWidget(self.load_recent_btn)
        title_row.addWidget(self.clear_recent_btn)
        layout.addLayout(title_row)

        self.recent_test_list = QListWidget()
        self.recent_test_list.setObjectName("recent_test_list")
        self.recent_test_list.setSelectionMode(QAbstractItemView.SingleSelection)
        self.recent_test_list.setToolTip(
            "单击选中高亮对应曲线；勾选多条后 Compare 对比；双击载入曲线与指标；右键重命名/删除"
        )
        self.recent_test_list.setContextMenuPolicy(Qt.CustomContextMenu)
        layout.addWidget(self.recent_test_list, 1)

        # Curve View：控制载入/对比图显示哪些曲线（Mean / Min-Max / Error 可组合）
        view_title = QLabel("Curve View")
        view_title.setObjectName("section_title")
        view_title.setStyleSheet("border: none;")
        layout.addWidget(view_title)

        self.curve_view_group = QButtonGroup()
        self.curve_view_group.setExclusive(False)
        self.show_mean_btn = QPushButton("Mean")
        self.show_band_btn = QPushButton("Min-Max")
        self.show_error_btn = QPushButton("Error")
        for btn, name, tip in (
            (self.show_mean_btn, "Mean", "显示均值校准曲线"),
            (self.show_band_btn, "Min-Max", "显示最大/最小值包络带"),
            (self.show_error_btn, "Error", "仅显示误差曲线（Actual - Ideal，右轴）"),
        ):
            btn.setObjectName("tool_btn")
            btn.setCheckable(True)
            btn.setToolTip(tip)
            self.curve_view_group.addButton(btn)

        self.show_mean_btn.setChecked(True)

        view_row = QHBoxLayout()
        view_row.setSpacing(4)
        view_row.addWidget(self.show_mean_btn)
        view_row.addWidget(self.show_band_btn)
        view_row.addWidget(self.show_error_btn)
        layout.addLayout(view_row)

        hint = QLabel("勾选记录进行对比；双击记录载入曲线与指标")
        hint.setObjectName("recent_hint_label")
        hint.setWordWrap(True)
        layout.addWidget(hint)

        return panel

    def _is_curve_view_enabled(self, name: str) -> bool:
        """Curve View 选项开关：mean / band / error。"""
        mapping = {
            'mean': self.show_mean_btn,
            'band': self.show_band_btn,
            'error': self.show_error_btn,
        }
        btn = mapping.get(name)
        return btn is not None and btn.isChecked()

    def _set_curve_view_all(self, enabled: bool = True):
        """一次性设置全部 Curve View 选项（blockSignals 防止自动重绘副作用）。"""
        for btn in (self.show_mean_btn, self.show_band_btn, self.show_error_btn):
            btn.blockSignals(True)
            btn.setChecked(enabled)
            btn.blockSignals(False)

    def _attach_curve_context_menu(self, pw):
        """给 PlotWidget 挂 Curve View 右键菜单（同时禁用 pyqtgraph 自带菜单避免冲突）。"""
        pw.setContextMenuPolicy(Qt.CustomContextMenu)
        pw.customContextMenuRequested.connect(self._show_curve_view_menu)
        try:
            pw.plotItem.vb.setMenuEnabled(False)
        except Exception:
            pass

    def _show_curve_view_menu(self, pos):
        """图表右键菜单：勾选启用单次数据显示项（与 Curve View 按钮同步）。"""
        sender = self.sender()
        menu = QMenu("Curve View", self)
        entries = (
            ("Voltage / Mean", self.show_mean_btn),
            ("Min-Max Band", self.show_band_btn),
            ("Error", self.show_error_btn),
        )
        for text, btn in entries:
            act = QAction(text, menu)
            act.setCheckable(True)
            act.setChecked(btn.isChecked())
            # 勾选变化直接同步按钮状态（按钮 toggled 会触发自动重绘）
            act.toggled.connect(btn.setChecked)
            menu.addAction(act)
        global_pos = sender.mapToGlobal(pos) if sender is not None else pos
        menu.exec(global_pos)

    def _on_curve_view_changed(self):
        """Curve View 选项切换后按当前显示内容重绘（优先对比图，其次载入记录）。"""
        checked = self._get_checked_records()
        curve_records = [r for r in checked if r['kind'] in self._CURVE_KINDS]
        if len(curve_records) >= 2:
            self._plot_comparison_curves(curve_records)
            return
        if self._loaded_record is not None and self._loaded_record['kind'] in self._CURVE_KINDS:
            self._load_recent_record(self._loaded_record)

    def _on_toggle_recent_panel(self, checked: bool):
        """折叠/展开 Curve 右侧 Recent 管理栏（记住展开宽度）。"""
        panel = self.recent_curve_splitter.widget(1)
        if panel is None:
            return
        if checked:
            panel.setVisible(True)
            self.recent_curve_splitter.setSizes(self._recent_panel_sizes)
            self.toggle_recent_btn.setText("Recent ◀")
        else:
            self._recent_panel_sizes = self.recent_curve_splitter.sizes()
            panel.setVisible(False)
            self.toggle_recent_btn.setText("Recent ▶")



    def _create_layout(self):
        root_layout = QVBoxLayout(self)
        root_layout.setContentsMargins(8, 6, 8, 8)
        root_layout.setSpacing(8)

        self.page = QFrame()
        self.page.setObjectName("page")
        page_layout = QVBoxLayout(self.page)
        page_layout.setContentsMargins(0, 0, 0, 0)
        page_layout.setSpacing(10)

        header_layout = QVBoxLayout()
        header_layout.setSpacing(2)


        title_label = QLabel("GPADC Automated Test")
        title_label.setObjectName("title_label")
        title_label.setStyleSheet("border: none")

        subtitle_label = QLabel("Evaluate GPADC performance including Linearity, ENOB, and Temperature Drift.")
        subtitle_label.setObjectName("subtitle_label")
        subtitle_label.setStyleSheet("border: none")

        header_layout.addWidget(title_label)
        header_layout.addWidget(subtitle_label)
        page_layout.addLayout(header_layout)

        body_layout = QHBoxLayout()
        body_layout.setSpacing(8)

        # 左侧滚动区
        left_wrapper = QVBoxLayout()
        left_wrapper.setContentsMargins(0, 0, 0, 0)
        left_wrapper.setSpacing(8)

        self.left_scroll = QScrollArea()
        self.left_scroll.setWidgetResizable(True)
        self.left_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.left_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.left_scroll.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Expanding)
        self.left_scroll.setMinimumWidth(310)
        self.left_scroll.setMaximumWidth(310)

        left_content = QFrame()
        left_content.setObjectName("left_scroll_content")
        left_content.setMinimumWidth(300)
        left_content.setMaximumWidth(300)
        left_content.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Preferred)

        left_col = QVBoxLayout(left_content)
        left_col.setContentsMargins(0, 0, 2, 0)
        left_col.setSpacing(12)

        # Test Item (下拉菜单)
        test_item_panel = QFrame()
        test_item_panel.setObjectName("panel")
        test_item_panel.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        test_item_layout = QVBoxLayout(test_item_panel)
        test_item_layout.setContentsMargins(12, 12, 12, 12)
        test_item_layout.setSpacing(8)

        test_item_title = QLabel("Test Item")
        test_item_title.setObjectName("section_title")
        test_item_title.setStyleSheet("border: none")
        test_item_layout.addWidget(test_item_title)

        self.test_item_combo = DarkComboBox(bg="#0a1733", border="#24365e")
        self.test_item_combo.addItem(self.TEST_1000CNT, self.TEST_1000CNT)
        self.test_item_combo.addItem(self.TEST_FORCE_VOLTAGE, self.TEST_FORCE_VOLTAGE)
        self.test_item_combo.addItem(self.TEST_HIGH_LOW_TEMP, self.TEST_HIGH_LOW_TEMP)
        self.test_item_combo.addItem(self.TEST_TEMP_CONSISTENCY, self.TEST_TEMP_CONSISTENCY)
        test_item_layout.addWidget(self.test_item_combo)
        left_col.addWidget(test_item_panel)

        # Instruments (动态显示)
        self.instruments_panel = QFrame()
        self.instruments_panel.setObjectName("panel")
        self.instruments_panel.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        instruments_layout = QVBoxLayout(self.instruments_panel)
        instruments_layout.setContentsMargins(12, 12, 12, 12)
        instruments_layout.setSpacing(10)

        instruments_title = QLabel("Instruments")
        instruments_title.setObjectName("section_title")
        instruments_title.setStyleSheet("border: none")
        instruments_layout.addWidget(instruments_title)

        self.n6705c_card = QFrame()
        self.n6705c_card.setObjectName("config_inner_panel")
        n6705c_card_layout = QVBoxLayout(self.n6705c_card)
        n6705c_card_layout.setContentsMargins(10, 10, 10, 10)
        n6705c_card_layout.setSpacing(6)
        n6705c_title_row = QHBoxLayout()
        n6705c_title_row.setSpacing(6)
        n6705c_title = QLabel("N6705C")
        n6705c_title.setStyleSheet("color: #c8d8ff; font-size: 11px; font-weight: 600; border: none;")
        n6705c_title_row.addWidget(n6705c_title)
        n6705c_title_row.addStretch()
        n6705c_card_layout.addLayout(n6705c_title_row)
        self.build_n6705c_connection_widgets(n6705c_card_layout, title_row=n6705c_title_row)

        self.n6705c_status = self.system_status_label
        self.n6705c_combo = self.visa_resource_combo
        self.n6705c_search_btn = self.search_btn
        self.n6705c_connect_btn = self.connect_btn
        self.n6705c_disconnect_btn = self.connect_btn

        instruments_layout.addWidget(self.n6705c_card)

        self.chamber_card = QFrame()
        self.chamber_card.setObjectName("config_inner_panel")
        chamber_card_layout = QVBoxLayout(self.chamber_card)
        chamber_card_layout.setContentsMargins(10, 10, 10, 10)
        chamber_card_layout.setSpacing(6)
        chamber_title_row = QHBoxLayout()
        chamber_title_row.setSpacing(6)
        chamber_title = QLabel("Chamber")
        chamber_title.setStyleSheet("color: #c8d8ff; font-size: 11px; font-weight: 600; border: none;")
        chamber_title_row.addWidget(chamber_title)
        chamber_title_row.addStretch()
        chamber_card_layout.addLayout(chamber_title_row)
        self.build_chamber_connection_widgets(chamber_card_layout)
        chamber_card_layout.removeWidget(self.chamber_status_label)
        chamber_title_row.addWidget(self.chamber_status_label)
        instruments_layout.addWidget(self.chamber_card)
        left_col.addWidget(self.instruments_panel)

        # Data Acquisition
        data_panel = QFrame()
        data_panel.setObjectName("panel")
        data_panel.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        data_layout = QVBoxLayout(data_panel)
        data_layout.setContentsMargins(12, 12, 12, 12)
        data_layout.setSpacing(10)

        data_title = QLabel("Data Acquisition")
        data_title.setObjectName("section_title")
        data_layout.addWidget(data_title)

        radio_row = QHBoxLayout()
        radio_row.setSpacing(12)

        self.iic_radio = QRadioButton("I2C")
        self.uart_radio = QRadioButton("UART Log")
        self.iic_radio.setChecked(True)

        self.data_acquisition_group = QButtonGroup()
        self.data_acquisition_group.addButton(self.iic_radio)
        self.data_acquisition_group.addButton(self.uart_radio)

        radio_row.addWidget(self.iic_radio)
        radio_row.addWidget(self.uart_radio)
        radio_row.addStretch()
        data_layout.addLayout(radio_row)

        self.data_stack = QStackedWidget()
        self.data_stack.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        self.iic_group = QFrame()
        self.iic_group.setObjectName("config_inner_panel")
        iic_layout = QGridLayout(self.iic_group)
        iic_layout.setContentsMargins(10, 10, 10, 10)
        iic_layout.setHorizontalSpacing(6)
        iic_layout.setVerticalSpacing(6)

        iic_layout.addWidget(QLabel("Device Address (Hex)"), 0, 0)
        self.iic_device_address = QLineEdit("0x17")
        iic_layout.addWidget(self.iic_device_address, 1, 0)

        iic_layout.addWidget(QLabel("Raw Data Register (Hex)"), 2, 0)
        self.iic_data_address = QLineEdit("0x57")
        iic_layout.addWidget(self.iic_data_address, 3, 0)

        iic_layout.addWidget(QLabel("IIC Width"), 4, 0)
        self.iic_width_combo = DarkComboBox(bg="#0a1733", border="#24365e")
        self.iic_width_combo.addItem("8-bit", 8)
        self.iic_width_combo.addItem("10-bit", 10)
        self.iic_width_combo.addItem("32-bit", 32)
        self.iic_width_combo.setCurrentIndex(1)
        iic_layout.addWidget(self.iic_width_combo, 5, 0)

        self.uart_group = QFrame()
        self.uart_group.setObjectName("config_inner_panel")
        uart_layout = QVBoxLayout(self.uart_group)
        uart_layout.setContentsMargins(10, 10, 10, 10)
        uart_layout.setSpacing(8)

        self.build_serial_connection_widgets(uart_layout)
        self.bind_serial_signals()

        keyword_label = QLabel("Search Keyword")
        keyword_label.setStyleSheet("border: none;")
        uart_layout.addWidget(keyword_label)
        self.uart_keyword = QLineEdit("raw/volt")
        uart_layout.addWidget(self.uart_keyword)

        self.data_stack.addWidget(self.iic_group)
        self.data_stack.addWidget(self.uart_group)
        data_layout.addWidget(self.data_stack)
        left_col.addWidget(data_panel)

        # Test Parameters
        params_panel = QFrame()
        params_panel.setObjectName("panel")
        params_panel.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        params_layout = QVBoxLayout(params_panel)
        params_layout.setContentsMargins(12, 12, 12, 12)
        params_layout.setSpacing(8)

        self.params_title = QLabel("Test Parameters")
        self.params_title.setObjectName("section_title")
        self.params_title.setStyleSheet("border: none;")
        params_layout.addWidget(self.params_title)

        self.params_mode_label = QLabel("VOLTAGE SWEEP")
        self.params_mode_label.setStyleSheet("color: #7e96bf; font-size: 11px; font-weight: 700; border: none;")
        params_layout.addWidget(self.params_mode_label)

        self.voltage_params_frame = QFrame()
        self.voltage_params_frame.setStyleSheet("QFrame { background: transparent; border: none; }")
        voltage_layout = QGridLayout(self.voltage_params_frame)
        voltage_layout.setContentsMargins(0, 0, 0, 0)
        voltage_layout.setHorizontalSpacing(6)
        voltage_layout.setVerticalSpacing(6)

        voltage_layout.addWidget(QLabel("Start (V)"), 0, 0)
        voltage_layout.addWidget(QLabel("End (V)"), 0, 1)
        voltage_layout.addWidget(QLabel("Step (V)"), 0, 2)

        self.voltage_min = QDoubleSpinBox()
        self.voltage_min.setRange(0.0, 5.0)
        self.voltage_min.setValue(4.0)
        self.voltage_min.setSingleStep(0.01)
        self.voltage_min.setDecimals(3)

        self.voltage_max = QDoubleSpinBox()
        self.voltage_max.setRange(0.0, 5.0)
        self.voltage_max.setValue(4.2)
        self.voltage_max.setSingleStep(0.01)
        self.voltage_max.setDecimals(3)

        self.voltage_step = QDoubleSpinBox()
        self.voltage_step.setRange(0.001, 1.0)
        self.voltage_step.setValue(0.05)
        self.voltage_step.setSingleStep(0.010)
        self.voltage_step.setDecimals(3)

        voltage_layout.addWidget(self.voltage_min, 1, 0)
        voltage_layout.addWidget(self.voltage_max, 1, 1)
        voltage_layout.addWidget(self.voltage_step, 1, 2)

        self.temp_params_frame = QFrame()
        self.temp_params_frame.setStyleSheet("QFrame { background: transparent; border: none; }")
        temp_layout = QGridLayout(self.temp_params_frame)
        temp_layout.setContentsMargins(0, 0, 0, 0)
        temp_layout.setHorizontalSpacing(6)
        temp_layout.setVerticalSpacing(6)

        temp_layout.addWidget(QLabel("Start (°C)"), 0, 0)
        temp_layout.addWidget(QLabel("End (°C)"), 0, 1)
        temp_layout.addWidget(QLabel("Step (°C)"), 0, 2)

        self.temp_min = QDoubleSpinBox()
        self.temp_min.setRange(-40.0, 125.0)
        self.temp_min.setValue(-40.0)
        self.temp_min.setSingleStep(1.0)
        self.temp_min.setDecimals(1)

        self.temp_max = QDoubleSpinBox()
        self.temp_max.setRange(-40.0, 125.0)
        self.temp_max.setValue(125.0)
        self.temp_max.setSingleStep(1.0)
        self.temp_max.setDecimals(1)

        self.temp_step = QDoubleSpinBox()
        self.temp_step.setRange(0.1, 50.0)
        self.temp_step.setValue(5)
        self.temp_step.setSingleStep(0.1)
        self.temp_step.setDecimals(2)

        temp_layout.addWidget(self.temp_min, 1, 0)
        temp_layout.addWidget(self.temp_max, 1, 1)
        temp_layout.addWidget(self.temp_step, 1, 2)

        temp_layout.addWidget(QLabel("Soak Time (s)"), 2, 0)
        self.soak_time = QSpinBox()
        self.soak_time.setRange(0, 3600)
        self.soak_time.setValue(180)
        self.soak_time.setSingleStep(30)
        temp_layout.addWidget(self.soak_time, 3, 0)

        params_layout.addWidget(self.voltage_params_frame)
        params_layout.addWidget(self.temp_params_frame)

        # 校准参数：默认留空 = 自动取 1/4、3/4 扫描点；两格都填写时作为手动校准两点
        self.calib_params_frame = QFrame()
        self.calib_params_frame.setStyleSheet("QFrame { background: transparent; border: none; }")
        calib_layout = QGridLayout(self.calib_params_frame)
        calib_layout.setContentsMargins(0, 0, 0, 0)
        calib_layout.setHorizontalSpacing(6)
        calib_layout.setVerticalSpacing(6)

        self.calib_low_label = QLabel("Calib Low (V)")
        self.calib_high_label = QLabel("Calib High (V)")
        calib_layout.addWidget(self.calib_low_label, 0, 0)
        calib_layout.addWidget(self.calib_high_label, 0, 1)

        self.calib_low = QLineEdit()
        self.calib_low.setPlaceholderText("Auto")
        self.calib_high = QLineEdit()
        self.calib_high.setPlaceholderText("Auto")
        calib_layout.addWidget(self.calib_low, 1, 0)
        calib_layout.addWidget(self.calib_high, 1, 1)

        params_layout.addWidget(self.calib_params_frame)

        self.sample_count_label = QLabel("Sample Count")
        self.sample_count_label.setObjectName("muted_label")
        params_layout.addWidget(self.sample_count_label)

        self.sample_count = QSpinBox()
        self.sample_count.setRange(1, 100000)
        self.sample_count.setValue(1000)
        self.sample_count.setSingleStep(100)
        params_layout.addWidget(self.sample_count)

        self.voltage_channel_label = QLabel("Voltage Channel")
        self.voltage_channel_label.setObjectName("muted_label")
        params_layout.addWidget(self.voltage_channel_label)

        self.voltage_channel = DarkComboBox(bg="#0a1733", border="#24365e")
        for ch in range(1, 5):
            self.voltage_channel.addItem(f"Channel {ch}", ch)
        self.voltage_channel.setCurrentIndex(3)
        params_layout.addWidget(self.voltage_channel)

        self.temp_hint_label = QLabel("Connect chamber to enable temperature testing.")
        self.temp_hint_label.setStyleSheet("color: #ff5a7a; font-size: 11px;")
        self.temp_hint_label.setWordWrap(True)
        params_layout.addWidget(self.temp_hint_label)

        left_col.addWidget(params_panel)

        left_col.addStretch()

        self.left_scroll.setWidget(left_content)
        left_wrapper.addWidget(self.left_scroll, 1)

        self.start_test_btn = QPushButton("▶ START TEST")
        self.start_test_btn.setObjectName("primaryStartBtn")
        # 宽度与上方面板可见宽对齐（左列内容宽 - 列右边距），防止长文本撑开左列
        self.start_test_btn.setMaximumWidth(
            left_content.maximumWidth() - left_col.contentsMargins().right()
        )
        left_wrapper.addWidget(self.start_test_btn)

        self.stop_test_btn = QPushButton("■")
        self.stop_test_btn.setObjectName("stopBtn")
        self.stop_test_btn.setEnabled(False)
        self.stop_test_btn.hide()

        # 右侧
        right_col = QVBoxLayout()
        right_col.setSpacing(12)

        metrics_layout = QHBoxLayout()
        metrics_layout.setSpacing(8)

        inl_card, self.inl_value = self._create_metric_card("INL", "---", "metric_value_green")
        dnl_card, self.dnl_value = self._create_metric_card("DNL", "---", "metric_value_green")
        enob_card, self.enob_value = self._create_metric_card("ENOB", "---", "metric_value_blue")
        offset_card, self.offset_error_value = self._create_metric_card("OFFSET ERR", "---", "metric_value_yellow")
        gain_card, self.gain_error_value = self._create_metric_card("GAIN ERR", "---", "metric_value_yellow")
        
        # 添加1000CNT TEST的指标卡片
        avg_card, self.avg_value = self._create_metric_card("AVG", "---", "metric_value_green")
        min_card, self.min_value = self._create_metric_card("MIN", "---", "metric_value_blue")
        max_card, self.max_value = self._create_metric_card("MAX", "---", "metric_value_yellow")
        std_card, self.std_value = self._create_metric_card("STD (code)", "---", "metric_value_blue")
        pp_card, self.pp_value = self._create_metric_card("P-P NOISE (code)", "---", "metric_value_yellow")
        samples_card, self.samples_value = self._create_metric_card("SAMPLES", "---", "metric_value_green")

        # 按测试项分组显隐：1000CNT 展示详细统计卡组，其余测试项展示线性度卡组
        self._cnt_cards = [avg_card, min_card, max_card, std_card, pp_card, samples_card]
        self._fv_cards = [inl_card, dnl_card, enob_card, offset_card, gain_card]

        self.linearity_value = QLabel("---")
        self.linearity_value.hide()

        metrics_layout.addWidget(inl_card)
        metrics_layout.addWidget(dnl_card)
        metrics_layout.addWidget(enob_card)
        metrics_layout.addWidget(offset_card)
        metrics_layout.addWidget(gain_card)
        metrics_layout.addWidget(avg_card)
        metrics_layout.addWidget(min_card)
        metrics_layout.addWidget(max_card)
        metrics_layout.addWidget(std_card)
        metrics_layout.addWidget(pp_card)
        metrics_layout.addWidget(samples_card)

        right_col.addLayout(metrics_layout)

        chart_panel = QFrame()
        chart_panel.setObjectName("chart_panel")
        chart_panel.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.chart_panel = chart_panel

        chart_layout = QVBoxLayout(chart_panel)
        chart_layout.setContentsMargins(14, 14, 14, 14)
        chart_layout.setSpacing(10)

        chart_top = QHBoxLayout()
        chart_title = QLabel("ADC Transfer Curve")
        chart_title.setObjectName("section_title")
        chart_title.setStyleSheet("border: none;")

        self.export_result_btn = QPushButton("Export Result")
        self.export_result_btn.setObjectName("tool_btn")

        # Recent 管理栏折叠/展开开关（面板在 Curve 右侧）
        self.toggle_recent_btn = QPushButton("Recent ◀")
        self.toggle_recent_btn.setObjectName("tool_btn")
        self.toggle_recent_btn.setCheckable(True)
        self.toggle_recent_btn.setChecked(True)
        self.toggle_recent_btn.setToolTip("显示/隐藏最近测试管理栏")

        chart_top.addWidget(chart_title)
        chart_top.addStretch()
        chart_top.addWidget(self.toggle_recent_btn)
        chart_top.addWidget(self.export_result_btn)
        chart_layout.addLayout(chart_top)

        self.chart_placeholder = QFrame()
        self.chart_placeholder.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.chart_placeholder.setMinimumHeight(200)
        self.chart_placeholder.setStyleSheet("""
            QFrame {
                background-color: #0a1735;
                border: none;
                border-radius: 8px;
            }
        """)
        # 图表区右键：弹 Curve View 显示项菜单（空图状态下也可用）
        self.chart_placeholder.setContextMenuPolicy(Qt.CustomContextMenu)
        self.chart_placeholder.customContextMenuRequested.connect(self._show_curve_view_menu)
        self._build_default_chart_placeholder()

        chart_layout.addWidget(self.chart_placeholder, 1)

        right_splitter, self.execution_logs = ExecutionLogsFrame.wrap_with(
            chart_panel, title="TEST LOG", show_progress=True, stretch=(3, 2)
        )
        self.log_text = self.execution_logs.log_edit
        self.progress_bar = self.execution_logs.progress_bar
        self.progress_text_label = self.execution_logs.progress_text_label
        self.clear_log_btn = self.execution_logs.clear_log_btn

        # Curve 右侧可折叠的最近测试管理栏：与图表区水平分栏
        self.recent_curve_splitter = QSplitter(Qt.Horizontal)
        self.recent_curve_splitter.setObjectName("recent_curve_splitter")
        self.recent_curve_splitter.addWidget(right_splitter)
        self.recent_curve_splitter.addWidget(self._create_recent_tests_panel())
        self.recent_curve_splitter.setStretchFactor(0, 1)
        self.recent_curve_splitter.setStretchFactor(1, 0)
        self.recent_curve_splitter.setSizes([1200, 260])

        right_col.addWidget(self.recent_curve_splitter, 1)

        self.export_params_btn = QPushButton("Export Parameters")
        self.export_params_btn.hide()

        self.save_config_btn = QPushButton("Save Config")
        self.load_config_btn = QPushButton("Load Config")
        self.save_config_btn.hide()
        self.load_config_btn.hide()

        body_layout.addLayout(left_wrapper, 0)
        body_layout.addLayout(right_col, 1)

        page_layout.addLayout(body_layout, 1)
        root_layout.addWidget(self.page, 1)

    def _init_ui_elements(self):
        self.current_test_item = self.TEST_1000CNT

        self.iic_radio.toggled.connect(self._update_data_acquisition_ui)
        self.uart_radio.toggled.connect(self._update_data_acquisition_ui)

        self.test_item_combo.currentIndexChanged.connect(self._on_test_item_changed)

        self.bind_n6705c_signals()
        self.bind_chamber_signals()
        self.serial_data_received.connect(self._on_uart_rx_data)

        self.start_test_btn.clicked.connect(self._on_start_or_stop)
        self.stop_test_btn.clicked.connect(self._stop_test)
        self.export_result_btn.clicked.connect(self.export_result)

        self.compare_recent_btn.clicked.connect(self._on_compare_recent_tests)
        self.load_recent_btn.clicked.connect(self._on_load_recent_test)
        self.clear_recent_btn.clicked.connect(self._on_clear_recent_tests)
        self.recent_test_list.itemDoubleClicked.connect(self._on_recent_item_double_clicked)
        self.recent_test_list.itemSelectionChanged.connect(self._on_recent_selection_changed)
        self.recent_test_list.customContextMenuRequested.connect(self._show_recent_item_menu)

        self.toggle_recent_btn.toggled.connect(self._on_toggle_recent_panel)
        for btn in (self.show_mean_btn, self.show_band_btn, self.show_error_btn):
            btn.toggled.connect(self._on_curve_view_changed)

        self._update_data_acquisition_ui()
        self._set_test_item(self.TEST_1000CNT)

    def _on_test_item_changed(self, index):
        test_item = self.test_item_combo.currentData()
        if test_item:
            self._set_test_item(test_item)

    def _set_status_label(self, label, text, status_type="err"):
        if status_type == "ok":
            label.setStyleSheet("color: #00d39a; font-weight: 600; border: none;")
        elif status_type == "warn":
            label.setStyleSheet("color: #ffb84d; font-weight: 600; border: none;")
        elif status_type == "warn":
            label.setStyleSheet("color: #ffb84d; font-weight: 600; border: none;")
        else:
            label.setStyleSheet("color: #ff5a7a; font-weight: 600; border: none;")
        label.setText(text)

    def _update_data_acquisition_ui(self):
        if self.iic_radio.isChecked():
            self.data_stack.setCurrentWidget(self.iic_group)
        else:
            self.data_stack.setCurrentWidget(self.uart_group)

    def _on_uart_rx_data(self, data):
        self._uart_rx_queue.put(bytes(data))

    def _drain_uart_rx_queue(self):
        while True:
            try:
                self._uart_rx_queue.get_nowait()
            except queue.Empty:
                break

    def _next_uart_log_line(self, partial, deadline, stop_check=None):
        while b"\n" not in partial:
            # 停止请求必须穿透内层等待：主线程一旦阻塞，串口数据泵（主线程槽）即断供，
            # 此循环会自旋到 deadline（默认 120s）且不查 stop_check，导致停止无响应、UI 卡死。
            if stop_check and stop_check():
                return partial, None
            remain = deadline - time.monotonic()
            if remain <= 0:
                return partial, None
            try:
                chunk = self._uart_rx_queue.get(timeout=min(0.1, remain))
            except queue.Empty:
                continue
            partial += chunk
        raw_line, partial = partial.split(b"\n", 1)
        return partial, raw_line.decode("utf-8", errors="replace").strip()

    def gpadc_uart_read_by_cnts(
        self,
        get_reg_cnt=1000,
        keyword="",
        timeout_s=120.0,
        return_raw=False,
        stop_check=None,
        progress_callback=None,
    ):
        if DEBUG_MOCK:
            rng = random.Random()
            raw_data = [max(0, int(rng.gauss(2844, 2.0))) for _ in range(get_reg_cnt)]
            if progress_callback:
                progress_callback(100)
            return compute_reg_stats(raw_data, return_raw=return_raw)

        raw_data = []
        partial = b""
        deadline = time.monotonic() + timeout_s
        while len(raw_data) < get_reg_cnt:
            if stop_check and stop_check():
                break
            if time.monotonic() >= deadline:
                self._test_worker.log.emit(
                    f"[WARN] UART 采集超时（{timeout_s:.0f}s），仅获取 {len(raw_data)}/{get_reg_cnt} 个样本，"
                    "请检查串口连接与 Search Keyword 是否匹配日志"
                )
                break
            partial, line = self._next_uart_log_line(partial, deadline, stop_check=stop_check)
            if line is None:
                continue
            value = parse_uart_gpadc_raw(line, keyword)
            if value is None:
                continue
            raw_data.append(value)
            if progress_callback:
                progress_callback(int(len(raw_data) * 100 / get_reg_cnt))

        if not raw_data:
            raise RuntimeError("未从 UART 日志提取到任何 GPADC 样本，请确认 DUT 日志输出与 Search Keyword 匹配")
        return compute_reg_stats(raw_data, return_raw=return_raw)

    def _gpadc_read_by_cnts(
        self,
        device_addr=0x17,
        reg_addr=0x56,
        iic_weight=10,
        get_reg_cnt=1000,
        return_raw=False,
        stop_check=None,
        progress_callback=None,
    ):
        if self._acq_mode_snapshot == 'UART':
            return self.gpadc_uart_read_by_cnts(
                get_reg_cnt=get_reg_cnt,
                keyword=self._uart_keyword_snapshot,
                return_raw=return_raw,
                stop_check=stop_check,
                progress_callback=progress_callback,
            )
        return self.gpadc_reg_read_by_cnts(
            device_addr=device_addr,
            reg_addr=reg_addr,
            iic_weight=iic_weight,
            get_reg_cnt=get_reg_cnt,
            return_raw=return_raw,
            stop_check=stop_check,
            progress_callback=progress_callback,
        )

    def _set_test_item(self, test_item):
        self.current_test_item = test_item

        required = self.INSTRUMENT_MAP.get(test_item, [])
        has_instruments = len(required) > 0
        self.instruments_panel.setVisible(has_instruments)
        self.n6705c_card.setVisible("n6705c" in required)
        self.chamber_card.setVisible("chamber" in required)

        # 1000CNT 用不到 ADC Transfer Curve，隐藏图表面板，改显详细统计卡组
        is_cnt_test = test_item == self.TEST_1000CNT
        self.chart_panel.setVisible(not is_cnt_test)
        for card in self._cnt_cards:
            card.setVisible(is_cnt_test)
        for card in self._fv_cards:
            card.setVisible(not is_cnt_test)

        if test_item == self.TEST_1000CNT:
            self.params_mode_label.setText("1000 COUNT TEST")
            self.voltage_params_frame.hide()
            self.temp_params_frame.hide()
            self.temp_hint_label.hide()
            self.voltage_channel_label.hide()
            self.voltage_channel.hide()
            self.start_test_btn.setText("▶ START TEST")
            self._start_btn_text = "▶ START TEST"
        elif test_item == self.TEST_FORCE_VOLTAGE:
            self.params_mode_label.setText("VOLTAGE SWEEP")
            self.voltage_params_frame.show()
            self.temp_params_frame.hide()
            self.temp_hint_label.hide()
            self.voltage_channel_label.show()
            self.voltage_channel.show()
            self.start_test_btn.setText("▶ START TEST")
            self._start_btn_text = "▶ START TEST"
        elif test_item == self.TEST_HIGH_LOW_TEMP:
            self.params_mode_label.setText("TEMPERATURE SWEEP")
            self.voltage_params_frame.hide()
            self.temp_params_frame.show()
            self.temp_hint_label.show()
            self.voltage_channel_label.show()
            self.voltage_channel.show()
            self.start_test_btn.setText("▶ START TEST")
            self._start_btn_text = "▶ START TEST"
        else:
            self.params_mode_label.setText("VOLTAGE + TEMPERATURE")
            self.voltage_params_frame.show()
            self.temp_params_frame.show()
            self.temp_hint_label.show()
            self.voltage_channel_label.show()
            self.voltage_channel.show()
            self.start_test_btn.setText("▶ START TEST")
            self._start_btn_text = "▶ START TEST"

        # 校准点仅在线性度类测试项生效；温度扫描时 x 轴为温度，单位随动
        self.calib_params_frame.setVisible(not is_cnt_test)
        calib_unit = "°C" if test_item == self.TEST_HIGH_LOW_TEMP else "V"
        self.calib_low_label.setText(f"Calib Low ({calib_unit})")
        self.calib_high_label.setText(f"Calib High ({calib_unit})")

    def _set_btn_connected(self, btn):
        update_connect_button_state(btn, connected=True)
        btn.setEnabled(True)

    def _set_btn_disconnected(self, btn):
        update_connect_button_state(btn, connected=False)
        btn.setEnabled(True)

    def _on_start_or_stop(self):
        if self.is_test_running:
            self._stop_test()
        else:
            self._start_test()

    def _start_test(self):
        if self.is_test_running:
            return

        self._acq_mode_snapshot = 'UART' if self.uart_radio.isChecked() else 'IIC'
        self._uart_keyword_snapshot = self.uart_keyword.text().strip()
        self._calib_points_snapshot = self._parse_calib_points()
        if self._calib_points_snapshot is not None:
            v_lo, v_hi = self._calib_points_snapshot
            self._append_log(f"[INFO] 使用手动校准点: low={v_lo}, high={v_hi}")
        if self._acq_mode_snapshot == 'UART':
            if not DEBUG_MOCK and not self._serial_connected:
                self._append_log("[ERROR] DUT 串口未连接，无法通过 UART Log 采集")
                self.set_system_status("错误: DUT串口未连接", is_error=True)
                return
            self._drain_uart_rx_queue()

        # 新测试开始：清除上一次的曲线与结果指标，避免新旧数据混显
        self._reset_result_display()

        self.is_test_running = True
        self.start_test_btn.setEnabled(True)
        self.stop_test_btn.setEnabled(True)
        self._test_stop_requested = False
        self._update_test_button_state(True)
        self._set_ui_enabled(False)
        self.set_progress(0)
        self._append_log(f"[INFO] Starting GPADC test... mode={self.current_test_item}")

        test_item = self.current_test_item
        iic_device_addr = int(self.iic_device_address.text(), 16)
        iic_reg_addr = int(self.iic_data_address.text(), 16)
        iic_width = self.iic_width_combo.currentData()
        sample_cnt = self.sample_count.value()

        if test_item == self.TEST_1000CNT:
            fn = self._run_1000cnt_test
            kwargs = dict(
                device_addr=iic_device_addr,
                reg_addr=iic_reg_addr,
                iic_weight=iic_width,
                sample_cnt=sample_cnt,
            )
        elif test_item == self.TEST_FORCE_VOLTAGE:
            fn = self._run_force_voltage_test
            kwargs = dict(
                device_addr=iic_device_addr,
                reg_addr=iic_reg_addr,
                iic_weight=iic_width,
                voltage_min=self.voltage_min.value(),
                voltage_max=self.voltage_max.value(),
                voltage_step=self.voltage_step.value(),
                voltage_channel=self.voltage_channel.currentData(),
                sample_cnt=sample_cnt,
            )
        elif test_item == self.TEST_HIGH_LOW_TEMP:
            fn = self._run_high_low_temp_test
            kwargs = dict(
                device_addr=iic_device_addr,
                reg_addr=iic_reg_addr,
                iic_weight=iic_width,
                temp_min=self.temp_min.value(),
                temp_max=self.temp_max.value(),
                temp_step=self.temp_step.value(),
                voltage_channel=self.voltage_channel.currentData(),
                soak_time=self.soak_time.value(),
                sample_cnt=sample_cnt,
            )
        elif test_item == self.TEST_TEMP_CONSISTENCY:
            fn = self._run_temp_consistency_test
            kwargs = dict(
                device_addr=iic_device_addr,
                reg_addr=iic_reg_addr,
                iic_weight=iic_width,
                temp_min=self.temp_min.value(),
                temp_max=self.temp_max.value(),
                temp_step=self.temp_step.value(),
                voltage_min=self.voltage_min.value(),
                voltage_max=self.voltage_max.value(),
                voltage_step=self.voltage_step.value(),
                voltage_channel=self.voltage_channel.currentData(),
                soak_time=self.soak_time.value(),
                sample_cnt=sample_cnt,
            )
        else:
            self._stop_test()
            return

        worker = _TestWorker(fn, kwargs)
        thread = QThread()
        worker.moveToThread(thread)

        thread.started.connect(worker.run)
        worker.finished.connect(self._on_test_done)
        worker.error.connect(self._on_test_error)
        worker.log.connect(self._append_log)
        worker.progress.connect(self.set_progress)
        worker.finished.connect(thread.quit)
        worker.error.connect(thread.quit)
        thread.finished.connect(worker.deleteLater)
        thread.finished.connect(self._on_test_thread_finished)

        self._test_worker = worker
        self.test_thread = thread
        thread.start()

    def _run_1000cnt_test(self, device_addr, reg_addr, sample_cnt=1000, iic_weight=10, stop_check=None):
        if self._acq_mode_snapshot == 'UART':
            self._test_worker.log.emit(f"[INFO] Starting 1000CNT TEST via UART Log, keyword='{self._uart_keyword_snapshot}', count={sample_cnt}")
        else:
            self._test_worker.log.emit(f"[INFO] Starting 1000CNT TEST with I2C address: 0x{device_addr:x} Register: 0x{reg_addr:x}, count={sample_cnt}")
        _, _, _, raw_data = self._gpadc_read_by_cnts(
            device_addr=device_addr,
            reg_addr=reg_addr,
            iic_weight=iic_weight,
            get_reg_cnt=sample_cnt,
            return_raw=True,
            stop_check=stop_check,
            progress_callback=lambda v: self._test_worker.progress.emit(v),
        )
        return ('1000cnt', compute_detailed_stats(raw_data))

    def _run_force_voltage_test(self, device_addr, reg_addr, voltage_min, voltage_max,
                                voltage_step, voltage_channel, sample_cnt=1000, iic_weight=10,
                                stop_check=None):
        result = self.gpadc_force_voltage_test(
            n6705c=self.n6705c,
            device_addr=device_addr,
            reg_addr=reg_addr,
            iic_weight=iic_weight,
            voltage_min=voltage_min,
            voltage_max=voltage_max,
            voltage_step=voltage_step,
            voltage_channel=voltage_channel,
            sample_cnt=sample_cnt,
            stop_check=stop_check,
            progress_callback=lambda v: self._test_worker.progress.emit(v),
        )
        return ('force_voltage', result)

    def _run_high_low_temp_test(self, device_addr, reg_addr, temp_min, temp_max,
                                temp_step, voltage_channel, sample_cnt=1000,
                                iic_weight=10, stop_check=None, soak_time=180):
        self._test_worker.log.emit("[INFO] RUN TEST_HIGH_LOW_TEMP TEST")
        result = self.gpadc_high_low_temp_test(
            device_addr=device_addr,
            reg_addr=reg_addr,
            iic_weight=iic_weight,
            temp_min=temp_min,
            temp_max=temp_max,
            temp_step=temp_step,
            voltage_channel=voltage_channel,
            soak_time=soak_time,
            sample_cnt=sample_cnt,
            stop_check=stop_check,
            progress_callback=lambda v: self._test_worker.progress.emit(v),
        )
        return ('high_low_temp', result)

    def _run_temp_consistency_test(self, device_addr, reg_addr, temp_min, temp_max,
                                   temp_step, voltage_min, voltage_max, voltage_step,
                                   voltage_channel, sample_cnt=1000, iic_weight=10,
                                   stop_check=None, soak_time=180):
        self._test_worker.log.emit("[INFO] RUN TEST_TEMP_CONSISTENCY TEST")
        result = self.gpadc_temp_consistency_test(
            device_addr=device_addr,
            reg_addr=reg_addr,
            iic_weight=iic_weight,
            temp_min=temp_min,
            temp_max=temp_max,
            temp_step=temp_step,
            voltage_min=voltage_min,
            voltage_max=voltage_max,
            voltage_step=voltage_step,
            voltage_channel=voltage_channel,
            soak_time=soak_time,
            sample_cnt=sample_cnt,
            stop_check=stop_check,
            progress_callback=lambda v: self._test_worker.progress.emit(v),
        )
        return ('temp_consistency', result)

    def _on_test_done(self, payload):
        if payload is None:
            return
        kind, result = payload

        if kind == '1000cnt':
            self.update_test_result(result)
            self._append_log(
                f"[RESULT] 1000CNT TEST: AVG={result.get('avg', 0):.3f}, MIN={result.get('min', 0):.3f}, "
                f"MAX={result.get('max', 0):.3f}, STD={result.get('std', 0):.3f}, "
                f"P-P={result.get('pp', 0):.0f} code, N={result.get('count', 0)}"
            )

        elif kind == 'force_voltage':
            if result is not None:
                # 新测试完成：默认显示单次全部波形（Voltage / Min-Max / Error）
                self._set_curve_view_all(True)
                result_after_calibration = self._calibration_data(result)
                voltage_data, mean_cali, adc_min_cali, adc_max_cali = result_after_calibration
                params = self._calculate_gpadc_parameters(result)
                self._plot_voltage_adc_curve(voltage_data, mean_cali, adc_min_cali, adc_max_cali)
                self.update_test_result(params)
                self._export_data = {
                    'params': params,
                    'raw': result,
                    'calibration': {
                        'voltage': voltage_data,
                        'mean_cali': mean_cali,
                        'min_cali': adc_min_cali,
                        'max_cali': adc_max_cali,
                    }
                }

        elif kind == 'high_low_temp':
            if result is not None:
                # 新测试完成：默认显示单次全部波形（Voltage / Min-Max / Error）
                self._set_curve_view_all(True)
                temp_data, mean_cali, adc_min_cali, adc_max_cali = self._calibration_data(result)
                params = self._calculate_gpadc_parameters(result)
                self._plot_voltage_adc_curve(temp_data, mean_cali, adc_min_cali, adc_max_cali, is_temp_mode=True)
                self.update_test_result(params)
                self._export_data = {
                    'params': params,
                    'raw': result,
                    'calibration': {
                        'voltage': temp_data,
                        'mean_cali': mean_cali,
                        'min_cali': adc_min_cali,
                        'max_cali': adc_max_cali,
                    }
                }
                self.set_system_status("GPADC温度测试完成")

        elif kind == 'temp_consistency':
            if result is not None:
                # 新测试完成：默认显示单次全部波形（Mean / Min-Max）
                self._set_curve_view_all(True)
                self._plot_temp_consistency_curves(result)
                self._export_data = {'raw': result}
                self._append_log("[RESULT] Temp Consistency Test completed")
                self.set_system_status("GPADC温度一致性测试完成")

        # 记入最近测试列表，供后续对比/载入
        self._record_recent_test(kind, result)
        # 曲线类测试完成后，当前图上显示的就是这条记录（右键切换 Curve View 时据此重绘）
        if self._recent_test_records and self._recent_test_records[-1]['kind'] in self._CURVE_KINDS:
            self._loaded_record = self._recent_test_records[-1]

    def _on_test_error(self, err):
        self._append_log(f"[ERROR] Test error: {err}")

    def _on_test_thread_finished(self):
        self._test_worker = None
        if self.test_thread is not None:
            self.test_thread.deleteLater()
            self.test_thread = None
        self.is_test_running = False
        self.start_test_btn.setEnabled(True)
        self.stop_test_btn.setEnabled(False)
        self._update_test_button_state(False)
        self._set_ui_enabled(True)
        # 通知 AI 异步动作层：测试结束，触发 pending 任务回灌续跑（§4 / S3-2）。
        stopped = bool(getattr(self, "_test_stop_requested", False))
        summary = "测试被中止" if stopped else "测试完成"
        self.sequence_execution_finished.emit(not stopped, summary)

    def _update_test_button_state(self, running):
        update_start_btn_state(self.start_test_btn, running,
                               start_text=self._start_btn_text,
                               stop_text="■ STOP")

    def _stop_test(self):
        if self._test_worker is not None:
            self._test_worker.request_stop()
        self._test_stop_requested = True
        if self.test_thread is not None and self.test_thread.isRunning():
            self.test_thread.quit()
        # 禁在 UI 线程 wait()：worker 收尾耗时不可控（UART 采集等待、仪器超时等），
        # 无限期阻塞会卡死窗口；线程收尾由 thread.finished → _on_test_thread_finished 异步复位。
        self._append_log("[INFO] Stopping GPADC test...")

    def export_result(self):
        import openpyxl
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                     Border, Side)
        from openpyxl.utils import get_column_letter
        from pathlib import Path
        import datetime

        if self._export_data is None:
            self._append_log("[WARN] No test result to export.")
            return

        params   = self._export_data['params']
        raw      = self._export_data['raw']
        calib    = self._export_data['calibration']

        results_dir = Path(__file__).parent.parent.parent.parent / "Results"
        results_dir.mkdir(exist_ok=True)

        ts       = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = results_dir / f"GPADC_Result_{ts}.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "GPADC Result"

        hdr_fill   = PatternFill("solid", fgColor="0A1735")
        hdr_font   = Font(bold=True, color="00D39A", size=11)
        val_font   = Font(color="000000", size=10)
        sub_fill   = PatternFill("solid", fgColor="0D1F40")
        sub_font   = Font(bold=True, color="7E96BF", size=10)
        thin_side  = Side(style="thin", color="1B2847")
        thin_border = Border(left=thin_side, right=thin_side,
                             top=thin_side, bottom=thin_side)

        def _hdr(ws, row, col, text):
            c = ws.cell(row=row, column=col, value=text)
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border
            return c

        def _val(ws, row, col, value):
            c = ws.cell(row=row, column=col, value=value)
            c.font = val_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border
            if isinstance(value, float):
                c.number_format = "0.000000"
            return c

        def _sub(ws, row, col, text):
            c = ws.cell(row=row, column=col, value=text)
            c.fill = sub_fill
            c.font = sub_font
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = thin_border
            return c

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 18

        row = 1
        _hdr(ws, row, 1, "GPADC ADC Parameters")
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=4)
        row += 1

        param_items = [
            ("Linearity (R²)",   params.get("linearity",    0.0)),
            ("INL (LSB)",        params.get("inl",          0.0)),
            ("DNL (LSB)",        params.get("dnl",          0.0)),
            ("ENOB (bits)",      params.get("enob",         0.0)),
            ("Gain Error (%)",   params.get("gain_error",   0.0)),
            ("Offset Error (LSB)", params.get("offset_error", 0.0)),
        ]
        _sub(ws, row, 1, "Parameter")
        _sub(ws, row, 2, "Value")
        _sub(ws, row, 3, "Parameter")
        _sub(ws, row, 4, "Value")
        row += 1
        for i in range(0, len(param_items), 2):
            _val(ws, row, 1, param_items[i][0])
            _val(ws, row, 2, param_items[i][1])
            if i + 1 < len(param_items):
                _val(ws, row, 3, param_items[i + 1][0])
                _val(ws, row, 4, param_items[i + 1][1])
            row += 1

        row += 1

        chart_data_start_row = row
        _hdr(ws, row, 1, "Chart Data")
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=4)
        row += 1
        _sub(ws, row, 1, "Input Voltage (V)")
        _sub(ws, row, 2, "Mean Cali (V)")
        _sub(ws, row, 3, "Min Cali (V)")
        _sub(ws, row, 4, "Max Cali (V)")
        row += 1
        chart_data_body_start = row
        v_list   = calib['voltage']
        mc_list  = calib['mean_cali']
        mn_list  = calib['min_cali']
        mx_list  = calib['max_cali']
        for i in range(len(v_list)):
            _val(ws, row, 1, float(v_list[i]))
            _val(ws, row, 2, float(mc_list[i]))
            _val(ws, row, 3, float(mn_list[i]))
            _val(ws, row, 4, float(mx_list[i]))
            row += 1
        chart_data_end = row - 1

        chart_anchor_col = get_column_letter(6)
        if self._chart_image_bytes is not None:
            try:
                from openpyxl.drawing.image import Image as XLImage
                self._chart_image_bytes.seek(0)
                img = XLImage(self._chart_image_bytes)
                img.width  = 600
                img.height = 350
                ws.add_image(img, f"{chart_anchor_col}{chart_data_start_row}")
            except Exception as ex:
                self._append_log(f"[WARN] Embed chart image failed: {ex}")
                self._embed_native_chart(ws, chart_data_start_row, chart_data_body_start,
                                         chart_data_end, chart_anchor_col)
        else:
            self._embed_native_chart(ws, chart_data_start_row, chart_data_body_start,
                                     chart_data_end, chart_anchor_col)

        row += 2

        _hdr(ws, row, 1, "Raw Data")
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=4)
        row += 1
        _sub(ws, row, 1, "Input Voltage (V)")
        _sub(ws, row, 2, "ADC Mean (LSB)")
        _sub(ws, row, 3, "ADC Min (LSB)")
        _sub(ws, row, 4, "ADC Max (LSB)")
        row += 1
        rv_list  = raw['voltage']
        rm_list  = raw['mean']
        rn_list  = raw['min']
        rx_list  = raw['max']
        for i in range(len(rv_list)):
            _val(ws, row, 1, float(rv_list[i]))
            _val(ws, row, 2, float(rm_list[i]))
            _val(ws, row, 3, float(rn_list[i]))
            _val(ws, row, 4, float(rx_list[i]))
            row += 1

        row += 2

        _hdr(ws, row, 1, "Calibrated Data")
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=4)
        row += 1
        _sub(ws, row, 1, "Input Voltage (V)")
        _sub(ws, row, 2, "Mean Cali (V)")
        _sub(ws, row, 3, "Min Cali (V)")
        _sub(ws, row, 4, "Max Cali (V)")
        row += 1
        for i in range(len(v_list)):
            _val(ws, row, 1, float(v_list[i]))
            _val(ws, row, 2, float(mc_list[i]))
            _val(ws, row, 3, float(mn_list[i]))
            _val(ws, row, 4, float(mx_list[i]))
            row += 1

        wb.save(str(filename))
        self._append_log(f"[INFO] Result exported to: {filename}")

    def _embed_native_chart(self, ws, chart_data_start_row, chart_data_body_start,
                            chart_data_end, chart_anchor_col):
        from openpyxl.chart import LineChart, Reference
        chart = LineChart()
        chart.title        = "ADC Transfer Curve (Calibrated)"
        chart.style        = 10
        chart.y_axis.title = "Calibrated Voltage (V)"
        chart.x_axis.title = "Input Voltage (V)"
        chart.width        = 22
        chart.height       = 14
        ref_x    = Reference(ws, min_col=1, min_row=chart_data_body_start, max_row=chart_data_end)
        ref_mean = Reference(ws, min_col=2, min_row=chart_data_body_start - 1, max_row=chart_data_end)
        ref_min  = Reference(ws, min_col=3, min_row=chart_data_body_start - 1, max_row=chart_data_end)
        ref_max  = Reference(ws, min_col=4, min_row=chart_data_body_start - 1, max_row=chart_data_end)
        chart.add_data(ref_mean, titles_from_data=True)
        chart.add_data(ref_min,  titles_from_data=True)
        chart.add_data(ref_max,  titles_from_data=True)
        chart.set_categories(ref_x)
        chart.series[0].graphicalProperties.line.solidFill = "00D39A"
        chart.series[0].graphicalProperties.line.width     = 20000
        chart.series[1].graphicalProperties.line.solidFill = "F0A040"
        chart.series[2].graphicalProperties.line.solidFill = "F0A040"
        ws.add_chart(chart, f"{chart_anchor_col}{chart_data_start_row}")

    def _set_ui_enabled(self, enabled):
        widgets = [
            self.n6705c_combo, self.n6705c_search_btn, self.n6705c_connect_btn,
            self.chamber_port_combo, self.chamber_search_btn, self.chamber_connect_btn,
            self.serial_combo, self.serial_search_btn, self.serial_connect_btn, self.uart_keyword,
            self.iic_radio, self.uart_radio,
            self.iic_device_address, self.iic_data_address, self.iic_width_combo,
            self.test_item_combo,
            self.voltage_channel,
            self.voltage_min, self.voltage_max, self.voltage_step,
            self.temp_min, self.temp_max, self.temp_step,
            self.soak_time, self.sample_count,
            self.calib_low, self.calib_high,
            self.recent_test_list, self.compare_recent_btn,
            self.load_recent_btn, self.clear_recent_btn,
        ]
        for widget in widgets:
            widget.setEnabled(enabled)

    def get_test_config(self):
        acquisition_mode = 'IIC' if self.iic_radio.isChecked() else 'UART'
        dut_port = self.get_selected_serial_port() or ""

        return {
            'n6705c_connected': self.is_connected,
            'chamber_connected': self.is_chamber_connected,
            'test_item': self.current_test_item,
            'data_acquisition_mode': acquisition_mode,
            'iic_device_address': self.iic_device_address.text(),
            'iic_data_address': self.iic_data_address.text(),
            'iic_width': self.iic_width_combo.currentData(),
            'dut_port': dut_port,
            'uart_keyword': self.uart_keyword.text(),
            'voltage_channel': self.voltage_channel.currentData(),
            'voltage_min': self.voltage_min.value(),
            'voltage_max': self.voltage_max.value(),
            'voltage_step': self.voltage_step.value(),
            'temp_min': self.temp_min.value(),
            'temp_max': self.temp_max.value(),
            'temp_step': self.temp_step.value(),
            'soak_time': self.soak_time.value(),
            'sample_count': self.sample_count.value(),
            'calib_low': self.calib_low.text(),
            'calib_high': self.calib_high.text()
        }

    def update_test_result(self, result):
        if self.current_test_item == self.TEST_1000CNT:
            self.avg_value.setText(f"{result['avg']:.3f}" if 'avg' in result else "---")
            self.min_value.setText(f"{result['min']:.3f}" if 'min' in result else "---")
            self.max_value.setText(f"{result['max']:.3f}" if 'max' in result else "---")
            self.std_value.setText(f"{result['std']:.3f}" if 'std' in result else "---")
            self.pp_value.setText(f"{result['pp']:.0f}" if 'pp' in result else "---")
            self.samples_value.setText(str(result['count']) if 'count' in result else "---")
            self.linearity_value.setText("---")
            self.enob_value.setText("---")
            self.dnl_value.setText("---")
            self.inl_value.setText("---")
            self.gain_error_value.setText("---")
            self.offset_error_value.setText("---")
        else:
            self.avg_value.setText("---")
            self.min_value.setText("---")
            self.max_value.setText("---")
            self.std_value.setText("---")
            self.pp_value.setText("---")
            self.samples_value.setText("---")
            self.linearity_value.setText(f"{result['linearity']:.3f}" if 'linearity' in result else "---")
            self.enob_value.setText(f"{result['enob']:.3f}" if 'enob' in result else "---")
            self.dnl_value.setText(f"{result['dnl']:.3f}" if 'dnl' in result else "---")
            self.inl_value.setText(f"{result['inl']:.3f}" if 'inl' in result else "---")
            self.gain_error_value.setText(f"{result['gain_error']:.3f}" if 'gain_error' in result else "---")
            self.offset_error_value.setText(f"{result['offset_error']:.3f}" if 'offset_error' in result else "---")

    def clear_results(self):
        self.linearity_value.setText("---")
        self.enob_value.setText("---")
        self.dnl_value.setText("---")
        self.inl_value.setText("---")
        self.gain_error_value.setText("---")
        self.offset_error_value.setText("---")
        # 清除1000CNT TEST的结果
        self.avg_value.setText("---")
        self.min_value.setText("---")
        self.max_value.setText("---")
        self.std_value.setText("---")
        self.pp_value.setText("---")
        self.samples_value.setText("---")

    # ------------------------------------------------------------------
    # 最近测试管理：记录 / 对比 / 载入 / 清空
    # ------------------------------------------------------------------
    def _record_recent_test(self, kind, result):
        """测试完成后记入最近测试列表（result 为 None 时不记录）。"""
        if result is None:
            return
        self._recent_test_seq += 1
        record = {
            'id': self._recent_test_seq,
            'time': datetime.datetime.now(),
            'test_item': self.current_test_item,
            'kind': kind,
            'label': '',
            'params': None,
            'raw': None,
            'calibration': None,
            'summary': "",
        }
        if kind == '1000cnt':
            record['params'] = dict(result)
            record['summary'] = (
                f"AVG={result.get('avg', 0):.3f} STD={result.get('std', 0):.3f} "
                f"P-P={result.get('pp', 0):.0f} N={result.get('count', 0)}"
            )
        elif kind in ('force_voltage', 'high_low_temp'):
            export_data = self._export_data or {}
            record['params'] = export_data.get('params')
            record['raw'] = export_data.get('raw')
            record['calibration'] = export_data.get('calibration')
            params = record['params'] or {}
            record['summary'] = (
                f"INL={params.get('inl', 0.0):.3f} DNL={params.get('dnl', 0.0):.3f} "
                f"ENOB={params.get('enob', 0.0):.3f} R²={params.get('linearity', 0.0):.4f}"
            )
        else:  # temp_consistency
            record['raw'] = result
            temps = result.get('temp') or []
            if temps:
                record['summary'] = f"T {temps[0]:.1f}~{temps[-1]:.1f}°C · {len(temps)} rows"
            else:
                record['summary'] = "no data"

        self._recent_test_records.append(record)
        while len(self._recent_test_records) > self.RECENT_TEST_LIMIT:
            self._recent_test_records.pop(0)
        self._refresh_recent_test_list()
        self._append_log(f"[INFO] 已记录最近测试 #{record['id']}：{record['summary']}")

    def _record_color(self, record_id: int) -> str:
        """按记录 id 稳定分配专属曲线颜色（列表装饰、图例、曲线共用）。"""
        return self._COMPARE_PALETTE[record_id % len(self._COMPARE_PALETTE)]

    def _record_display_name(self, record) -> str:
        """记录显示名：优先用户重命名的 label，否则测试项名。"""
        return record.get('label') or record['test_item']

    def _refresh_recent_test_list(self):
        # 记忆勾选与选中状态（按 record id），重建列表后恢复
        checked_ids = set()
        for i in range(self.recent_test_list.count()):
            item = self.recent_test_list.item(i)
            if item.checkState() == Qt.Checked:
                rec = item.data(Qt.UserRole)
                if rec is not None:
                    checked_ids.add(rec['id'])
        current_id = self._highlight_record_id
        self.recent_test_list.blockSignals(True)
        self.recent_test_list.clear()
        for record in self._recent_test_records:
            ts = record['time'].strftime("%m-%d %H:%M:%S")
            item = QListWidgetItem(
                f"● #{record['id']} [{ts}] {self._record_display_name(record)} · {record['summary']}"
            )
            color = self._record_color(record['id'])
            item.setForeground(QBrush(QColor(color)))
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Checked if record['id'] in checked_ids else Qt.Unchecked)
            item.setData(Qt.UserRole, record)
            self.recent_test_list.addItem(item)
            if record['id'] == current_id:
                self.recent_test_list.setCurrentItem(item)
        self.recent_test_list.blockSignals(False)

    def _on_recent_selection_changed(self):
        """列表选中变化：同步高亮目标并重绘对比图（选中记录加粗、其余淡化）。"""
        items = self.recent_test_list.selectedItems()
        record = items[0].data(Qt.UserRole) if items else None
        self._highlight_record_id = record['id'] if record is not None else None
        checked = self._get_checked_records()
        curve_records = [r for r in checked if r['kind'] in self._CURVE_KINDS]
        if len(curve_records) >= 2:
            self._plot_comparison_curves(curve_records)

    def _show_recent_item_menu(self, pos):
        """Recent 列表右键菜单：重命名 / 载入 / 勾选对比 / 删除 / 清空。"""
        item = self.recent_test_list.itemAt(pos)
        menu = QMenu(self)

        if item is not None:
            record = item.data(Qt.UserRole)
            act_rename = QAction("Rename…", menu)
            act_rename.triggered.connect(lambda: self._rename_recent_record(record))
            menu.addAction(act_rename)

            act_load = QAction("Load Curve", menu)
            act_load.triggered.connect(lambda: self._load_recent_record(record))
            menu.addAction(act_load)

            act_check = QAction("Uncheck" if item.checkState() == Qt.Checked else "Check for Compare", menu)
            act_check.triggered.connect(
                lambda: item.setCheckState(
                    Qt.Unchecked if item.checkState() == Qt.Checked else Qt.Checked)
            )
            menu.addAction(act_check)

            menu.addSeparator()
            act_remove = QAction("Remove", menu)
            act_remove.triggered.connect(lambda: self._delete_recent_record(record))
            menu.addAction(act_remove)
            menu.addSeparator()

        act_clear = QAction("Clear All", menu)
        act_clear.triggered.connect(self._on_clear_recent_tests)
        menu.addAction(act_clear)

        menu.exec(self.recent_test_list.mapToGlobal(pos))

    def _rename_recent_record(self, record):
        """重命名记录显示名（列表、对比图图例、单次图图例同步）。"""
        text, ok = QInputDialog.getText(
            self, "Rename", "Display name:",
            text=self._record_display_name(record),
        )
        if not ok:
            return
        record['label'] = text.strip()
        self._refresh_recent_test_list()
        self._append_log(f"[INFO] 记录 #{record['id']} 已重命名为「{record['label'] or record['test_item']}」")
        # 该记录在当前对比图中时重绘更新图例
        checked = self._get_checked_records()
        curve_records = [r for r in checked if r['kind'] in self._CURVE_KINDS]
        if len(curve_records) >= 2:
            self._plot_comparison_curves(curve_records)
        elif self._loaded_record is record:
            self._load_recent_record(record)

    def _delete_recent_record(self, record):
        """删除单条最近测试记录。"""
        try:
            self._recent_test_records.remove(record)
        except ValueError:
            return
        if self._loaded_record is record:
            self._loaded_record = None
        if self._highlight_record_id == record['id']:
            self._highlight_record_id = None
        self._refresh_recent_test_list()
        self._append_log(f"[INFO] 已删除记录 #{record['id']}")

    def _get_checked_records(self):
        records = []
        for i in range(self.recent_test_list.count()):
            item = self.recent_test_list.item(i)
            if item.checkState() == Qt.Checked:
                record = item.data(Qt.UserRole)
                if record is not None:
                    records.append(record)
        return records

    def _on_compare_recent_tests(self):
        records = self._get_checked_records()
        if len(records) < 2:
            self._append_log("[WARN] 对比至少需要勾选 2 条记录")
            return
        stat_records = [r for r in records if r['kind'] == '1000cnt']
        curve_records = [r for r in records if r['kind'] in self._CURVE_KINDS]
        if stat_records:
            self._compare_stats_to_log(stat_records)
        if not curve_records:
            self._append_log("[INFO] 勾选记录无曲线数据，统计对比已输出到日志")
            return
        if not self.chart_panel.isVisible():
            self._append_log(
                "[WARN] 当前测试项不显示曲线面板，请切换到曲线类测试项（如 Force Voltage Test）后重试"
            )
            return
        self._plot_comparison_curves(curve_records)

    def _compare_stats_to_log(self, records):
        self._append_log("===== 1000CNT 记录对比 =====")
        self._append_log(
            f"{'#':<4}{'Time':<14}{'AVG':>10}{'MIN':>10}{'MAX':>10}{'STD':>10}{'P-P':>8}{'N':>8}"
        )
        for r in records:
            p = r.get('params') or {}
            ts = r['time'].strftime("%m-%d %H:%M")
            self._append_log(
                f"#{r['id']:<3}{ts:<14}"
                f"{p.get('avg', 0):>10.3f}{p.get('min', 0):>10.3f}{p.get('max', 0):>10.3f}"
                f"{p.get('std', 0):>10.3f}{p.get('pp', 0):>8.0f}{p.get('count', 0):>8d}"
            )

    def _on_load_recent_test(self):
        records = self._get_checked_records()
        if not records:
            self._append_log("[WARN] 请先勾选一条要载入的记录")
            return
        if len(records) > 1:
            self._append_log("[WARN] 载入仅支持单条记录，已取勾选的第一条")
        self._load_recent_record(records[0])

    def _on_recent_item_double_clicked(self, item):
        record = item.data(Qt.UserRole)
        if record is not None:
            self._load_recent_record(record)

    def _load_recent_record(self, record):
        """载入历史记录：恢复其曲线、指标卡与可导出数据。"""
        kind = record['kind']
        if kind == '1000cnt':
            params = record.get('params') or {}
            self._append_log(
                f"[INFO] 已载入记录 #{record['id']}：AVG={params.get('avg', 0):.3f}, "
                f"STD={params.get('std', 0):.3f}, P-P={params.get('pp', 0):.0f}, N={params.get('count', 0)}"
            )
            if self.current_test_item == self.TEST_1000CNT:
                self.update_test_result(params)
            else:
                self._append_log("[INFO] 切换到 1000CNT TEST 测试项可在指标卡查看该记录统计")
            return
        if not self.chart_panel.isVisible():
            self._append_log(
                "[WARN] 当前测试项不显示曲线面板，请切换到曲线类测试项（如 Force Voltage Test）后重试"
            )
            return
        self._export_data = None
        self._chart_image_bytes = None
        self._loaded_record = record
        if kind in ('force_voltage', 'high_low_temp'):
            calib = record.get('calibration') or {}
            self._plot_voltage_adc_curve(
                calib.get('voltage'), calib.get('mean_cali'),
                calib.get('min_cali'), calib.get('max_cali'),
                is_temp_mode=(kind == 'high_low_temp'),
            )
            params = record.get('params')
            if params:
                self.update_test_result(params)
            self._export_data = {
                'params': record.get('params'),
                'raw': record.get('raw'),
                'calibration': record.get('calibration'),
            }
        else:  # temp_consistency
            raw = record.get('raw') or {}
            self._plot_temp_consistency_curves(raw)
            self._export_data = {'raw': raw}
        self._append_log(f"[INFO] 已载入记录 #{record['id']} 的曲线与指标（{record['test_item']}）")

    def _on_clear_recent_tests(self):
        if not self._recent_test_records:
            return
        self._recent_test_records.clear()
        self.recent_test_list.clear()
        self._highlight_record_id = None
        self._loaded_record = None
        self._append_log("[INFO] 已清空最近测试记录")

    def _plot_comparison_curves(self, records):
        """把勾选的多条记录曲线画到同一张对比图上（按 Curve View 选项过滤）。"""
        try:
            import pyqtgraph as pg
            import numpy as np

            self._clear_chart_placeholder()
            layout = self.chart_placeholder.layout()
            if layout is None:
                layout = QVBoxLayout(self.chart_placeholder)
            layout.setContentsMargins(14, 14, 14, 10)
            layout.setSpacing(8)

            show_mean = self._is_curve_view_enabled('mean')
            show_band = self._is_curve_view_enabled('band')
            show_error = self._is_curve_view_enabled('error')
            if not (show_mean or show_band or show_error):
                self._append_log("[WARN] Curve View 未勾选任何曲线类型，请至少勾选 Mean / Min-Max / Error 之一")
                return

            first_kind = records[0]['kind']
            if first_kind == 'high_low_temp':
                x_title, y_title = "Input Temperature (°C)", "Calibrated Temperature (°C)"
            elif first_kind == 'temp_consistency':
                x_title, y_title = "Input Voltage (V)", "ADC Code"
            else:
                x_title, y_title = "Input Voltage (V)", "Calibrated Voltage (V)"
            kinds = {r['kind'] for r in records}
            if len(kinds) > 1:
                self._append_log("[WARN] 所选记录测试类型不一致，坐标轴单位按首条记录显示，请谨慎解读")

            # 选中高亮：目标记录加粗，其余淡化（未选中任何记录时全部正常显示）
            highlight_id = self._highlight_record_id
            dim_others = highlight_id is not None and any(r['id'] == highlight_id for r in records)

            legend_row = QHBoxLayout()
            legend_row.addStretch()
            for r in records:
                color = self._record_color(r['id'])
                is_hl = dim_others and r['id'] == highlight_id
                lbl = QLabel(
                    f"● #{r['id']} {self._record_display_name(r)} {r['time'].strftime('%m-%d %H:%M')}"
                )
                style = f"color: {color}; font-size: 11px;"
                if is_hl:
                    style += " font-weight: bold;"
                lbl.setStyleSheet(style)
                legend_row.addWidget(lbl)
                legend_row.addSpacing(12)
            legend_row.addStretch()
            layout.addLayout(legend_row)

            pw = pg.PlotWidget()
            pw.setBackground("#0a1735")
            pw.showGrid(x=True, y=True, alpha=0.15)
            pw.setLabel("left", y_title, color="#a0b4d8")
            pw.setLabel("bottom", x_title, color="#a0b4d8")
            for axis_name in ("left", "bottom"):
                axis = pw.getAxis(axis_name)
                axis.setTextPen(pg.mkPen("#a0b4d8"))
                axis.setPen(pg.mkPen("#3a4f7a"))
            self._attach_curve_context_menu(pw)

            plotted = 0
            for r in records:
                plotted += self._plot_comparison_record(
                    pw, r, self._record_color(r['id']),
                    show_mean=show_mean, show_band=show_band, show_error=show_error,
                    dimmed=dim_others and r['id'] != highlight_id,
                )

            layout.addWidget(pw, 1)

            x_label = QLabel(x_title)
            x_label.setAlignment(Qt.AlignCenter)
            x_label.setObjectName("muted_label")
            layout.addWidget(x_label)

            self._append_log(f"[INFO] 对比图已绘制：{len(records)} 条记录共 {plotted} 条曲线")
        except Exception as e:
            self._append_log(f"[ERROR] 绘制对比曲线失败: {e}")
            logger.error("绘制对比曲线失败: %s", e, exc_info=True)

    def _plot_comparison_record(self, pw, record, color, show_mean=True, show_band=False, show_error=False, dimmed=False):
        """把单条记录的曲线画到对比图上，返回绘制的曲线条数。

        dimmed=True 时曲线/符号/包络带半透明（用于选中高亮时淡化其它记录）。
        """
        import pyqtgraph as pg
        import numpy as np

        line_color = QColor(color)
        sym_color = QColor(color)
        if dimmed:
            line_color.setAlpha(90)
            sym_color.setAlpha(90)

        kind = record['kind']
        if kind in ('force_voltage', 'high_low_temp'):
            calib = record.get('calibration') or {}
            raw = record.get('raw') or {}
            x_data = calib.get('voltage')
            y_mean = calib.get('mean_cali')
            y_min = calib.get('min_cali')
            y_max = calib.get('max_cali')
            x_raw = raw.get('voltage')
            y_raw_mean = raw.get('mean')
            y_raw_min = raw.get('min')
            y_raw_max = raw.get('max')
            if not x_data or not y_mean:
                x_data, y_mean, y_min, y_max = x_raw, y_raw_mean, y_raw_min, y_raw_max
            if not x_data or not y_mean:
                self._append_log(f"[WARN] 记录 #{record['id']} 缺少曲线数据，已跳过")
                return 0
            x = np.array(x_data, dtype=float)
            y = np.array(y_mean, dtype=float)
            n = min(len(x), len(y))
            x, y = x[:n], y[:n]
            count = 0
            if show_band and y_min and y_max:
                y_min_arr = np.array(y_min, dtype=float)[:n]
                y_max_arr = np.array(y_max, dtype=float)[:n]
                band_alpha = 15 if dimmed else 35
                pw.addItem(pg.FillBetweenItem(
                    pg.PlotDataItem(x, y_max_arr),
                    pg.PlotDataItem(x, y_min_arr),
                    brush=pg.mkBrush(240, 160, 64, band_alpha),
                ))
                pw.plot(x, y_max_arr, pen=pg.mkPen(line_color, width=1, style=pg.QtCore.Qt.DashLine))
                pw.plot(x, y_min_arr, pen=pg.mkPen(line_color, width=1, style=pg.QtCore.Qt.DashLine))
                count += 1
            if show_mean:
                pw.plot(x, y, pen=pg.mkPen(line_color, width=2),
                        symbol="o", symbolSize=4, symbolBrush=sym_color, symbolPen=None)
                count += 1
            if show_error:
                diff = y - x
                pw.plot(x, diff, pen=pg.mkPen(line_color, width=2, style=pg.QtCore.Qt.DashLine),
                        symbol="t", symbolSize=5, symbolBrush=sym_color, symbolPen=None)
                count += 1
            return count

        # temp_consistency：每个温度一条线（同一记录统一用其专属色）
        raw = record.get('raw') or {}
        voltage_pts = raw.get('voltage')
        mean_matrix = raw.get('mean')
        if not voltage_pts or not mean_matrix:
            self._append_log(f"[WARN] 记录 #{record['id']} 缺少曲线数据，已跳过")
            return 0
        x = np.array(voltage_pts, dtype=float)
        count = 0
        for row in mean_matrix:
            y = np.array(row, dtype=float)
            n = min(len(x), len(y))
            if n == 0:
                continue
            pw.plot(x[:n], y[:n], pen=pg.mkPen(line_color, width=1),
                    symbol="o", symbolSize=3, symbolBrush=sym_color, symbolPen=None)
            count += 1
        return count

    def cleanup_threads(self):
        try:
            from shiboken6 import isValid
        except ImportError:
            isValid = lambda obj: obj is not None

        for attr in ('_n6705c_search_thread', '_chamber_search_thread', '_serial_search_thread'):
            thread = getattr(self, attr, None)
            if thread is not None and isValid(thread) and thread.isRunning():
                thread.quit()
                thread.wait(3000)
            setattr(self, attr, None)

        if self._test_worker is not None:
            self._test_worker.request_stop()
        if self.test_thread is not None and isValid(self.test_thread) and self.test_thread.isRunning():
            self.test_thread.quit()
            self.test_thread.wait(3000)
        self.test_thread = None

    def _append_log(self, text):
        self.execution_logs.append_log(text)

    def append_log(self, message):
        self.execution_logs.append_log(message)

    def set_progress(self, value: int):
        self.execution_logs.set_progress(value)

    def set_system_status(self, status, is_error=False):
        """线程安全重写：非 GUI 线程调用时经信号队列化回主线程再操作 QWidget。

        Mixin 原实现直接 setText/setObjectName/unpolish/polish，从 worker 线程
        调用会破坏 Qt 线程亲和性，实测触发 SIGABRT 使整个进程崩溃
        （Force Voltage 响应保护路径即因此崩过）；主线程调用行为不变。
        """
        if QThread.currentThread() is not self._gui_thread:
            self.system_status_requested.emit(status, bool(is_error))
            return
        super().set_system_status(status, is_error)

    def _apply_system_status(self, status, is_error=False):
        """system_status_requested 的主线程槽：直连 Mixin 原实现（不重入路由判断）。"""
        super().set_system_status(status, is_error)

    def update_instrument_info(self, instrument_info):
        if self.is_connected:
            self.set_system_status("● Connected")

    def get_test_mode(self):
        return "GPADC Test"

    def set_test_mode(self, mode):
        pass

    def get_test_id(self):
        return "GPADC_TEST_001"

    def set_test_id(self, test_id):
        pass

    
    def gpadc_reg_read_by_cnts(
        self,
        device_addr=0x17,
        reg_addr=0x56,
        iic_weight=10,
        get_reg_cnt=1000,
        return_raw=False,
        stop_check=None,
        progress_callback=None,
    ):
        if DEBUG_MOCK:
            if not hasattr(self, "_mock_i2c"):
                self._mock_i2c = MockI2C()
            deviceI2C = self._mock_i2c
        else:
            if not hasattr(self, "deviceI2C"):
                self.deviceI2C = I2CInterface()
                self.set_system_status("I2C接口初始化成功")
            deviceI2C = self.deviceI2C

        raw_data = []

        for i in range(get_reg_cnt):
            if stop_check and stop_check():
                break
            temp = deviceI2C.read(device_addr, reg_addr, iic_weight)
            raw_data.append(temp)
            if progress_callback:
                progress_callback(int((i + 1) * 100 / get_reg_cnt))

        return compute_reg_stats(raw_data, return_raw=return_raw)

    def _check_sweep_response(self, voltage_data, adc_mean):
        """扫描响应保护：电压上升时前 N 点 Raw 均值应随之明显变化。

        返回 None 表示响应正常；返回字符串为告警原因（调用方据此中止测试）：
        - 极差 < 阈值：Raw 几乎不变（读数卡死 / 电压未施加 / 通道错误）
        - 净变化 <= 0：电压上升但 Raw 无净增量（方向异常 / DUT 无响应）
        """
        n = self._SWEEP_GUARD_POINTS
        if len(adc_mean) < n or len(voltage_data) < n:
            # 扫描点数不足（极小范围配置），不做判定
            return None
        head_v = voltage_data[:n]
        head_m = adc_mean[:n]
        spread = max(head_m) - min(head_m)
        net = head_m[-1] - head_m[0]
        detail = (
            f"前{n}点 V={[f'{v:.3f}' for v in head_v]}, "
            f"RawMean={[f'{m:.1f}' for m in head_m]}, "
            f"极差={spread:.1f} LSB, 净变化={net:+.1f} LSB"
        )
        if head_v[-1] > head_v[0] and net <= 0:
            return f"电压上升但 Raw 均值无净增量（疑似通道/地址/接线错误）。{detail}"
        if spread < self._SWEEP_GUARD_MIN_SPREAD_LSB:
            return f"Raw 均值几乎不随电压变化（疑似 DUT 无响应或电压未施加）。{detail}"
        return None

    def gpadc_force_voltage_test(
        self,
        n6705c=None,
        device_addr=0x17,
        reg_addr=0x56,
        iic_weight=10,
        voltage_min=0.1,
        voltage_max=1.8,
        voltage_step=0.05,
        voltage_channel=1,
        sample_cnt=1000,
        stop_check=None,
        progress_callback=None,
    ):
        self._test_worker.log.emit(f"[INFO] Running FORCE VOLTAGE TEST with I2C address: 0x{device_addr:x}, Register: 0x{reg_addr:x}")

        if DEBUG_MOCK:
            if not hasattr(self, "_mock_i2c"):
                self._mock_i2c = MockI2C()
            vol_source = MockN6705C()
            vol_source._mock_i2c = self._mock_i2c
        else:
            vol_source = n6705c if n6705c is not None else self.n6705c
            if vol_source is None or not self.is_connected:
                self._test_worker.log.emit("[ERROR] N6705C not connected")
                self.set_system_status("错误: N6705C未连接", is_error=True)
                return None

        settle_time = 0.0 if DEBUG_MOCK else 0.5
        step_time   = 0.0 if DEBUG_MOCK else 0.2

        voltage_data = []
        adc_mean = []
        adc_min = []
        adc_max = []

        vol_source.set_voltage(voltage_channel, voltage_min)
        time.sleep(settle_time)

        current_voltage = voltage_min
        total_points = max(1, int(round((voltage_max - voltage_min) / voltage_step)) + 1)
        point_idx = 0

        while current_voltage <= voltage_max + voltage_step * 0.001:
            if stop_check and stop_check():
                self._test_worker.log.emit("[INFO] Force voltage test stopped by user.")
                break
            vol_source.set_voltage(voltage_channel, current_voltage)
            time.sleep(step_time)

            avg, max_val, min_val = self._gpadc_read_by_cnts(
                device_addr,
                reg_addr,
                iic_weight,
                get_reg_cnt=sample_cnt,
                return_raw=False,
                stop_check=stop_check,
            )

            voltage_data.append(current_voltage)
            adc_mean.append(avg)
            adc_min.append(min_val)
            adc_max.append(max_val)

            # 响应保护：前 N 点 Raw 均值接近或方向异常 → 判定 DUT 无响应，中止测试
            if len(adc_mean) == self._SWEEP_GUARD_POINTS:
                guard_reason = self._check_sweep_response(voltage_data, adc_mean)
                if guard_reason is not None:
                    self._test_worker.log.emit(f"[WARN] GPADC response guard: {guard_reason}")
                    self._test_worker.log.emit("[WARN] Force voltage test 已中止，请检查 DUT 供电/接线/通道与采集配置")
                    self.set_system_status("警告: GPADC 无响应，测试已提前中止")
                    return None

            current_voltage = round(current_voltage + voltage_step, 6)
            point_idx += 1
            if progress_callback:
                progress_callback(int(point_idx * 100 / total_points))

        self._test_worker.log.emit("===== FORCE VOLTAGE TEST 结果 =====")

        result = {
            "voltage": voltage_data,
            "mean": adc_mean,
            "min": adc_min,
            "max": adc_max,
        }
        return result

    def _parse_calib_points(self):
        """解析用户填写的两个校准点；均留空返回 None 表示自动选取。"""
        low_text = self.calib_low.text().strip()
        high_text = self.calib_high.text().strip()
        if not low_text and not high_text:
            return None
        try:
            v_low = float(low_text)
            v_high = float(high_text)
        except ValueError:
            self._append_log("[WARN] 校准点格式无效（需填写两个数值），回退为自动选取")
            return None
        if not (math.isfinite(v_low) and math.isfinite(v_high)):
            self._append_log("[WARN] 校准点需为有限数值，回退为自动选取")
            return None
        if v_low == v_high:
            self._append_log("[WARN] 两个校准点相同，回退为自动选取")
            return None
        return (min(v_low, v_high), max(v_low, v_high))

    def _calibration_data(self, result):
        adc_raw_data = result["voltage"]
        adc_mean     = result["mean"]
        adc_min      = result["min"]
        adc_max      = result["max"]

        k, b, mean_cali, adc_min_cali, adc_max_cali, v_low, m_low, v_high, m_high = \
            compute_calibration(adc_raw_data, adc_mean, adc_min, adc_max,
                                calib_points=self._calib_points_snapshot)

        if k == 0.0:
            self._append_log(
                "[WARN] 标定斜率 k=0（ADC 读数无变化或采样点电压相同），跳过标定，使用原始数据"
            )
        self._append_log(f"[INFO] Calibration: k={k:.6f} (LSB/V), b={b:.6f} (LSB)")

        self._append_log(
            'Voltage,RawMean,mean_cali,Δmean(mV),RawMin,RawMax,min_cali,max_cali,Δmin(mV),Δmax(mV)'
        )
        n = len(adc_raw_data)
        for i in range(n):
            v_in = adc_raw_data[i]
            delta_mean_mv = (mean_cali[i] - v_in) * 1000
            delta_min_mv = (adc_min_cali[i] - v_in) * 1000
            delta_max_mv = (adc_max_cali[i] - v_in) * 1000
            self._append_log(
                f'{v_in:.4f},{adc_mean[i]:.3f},{mean_cali[i]:.6f},{delta_mean_mv:+.3f}'
                f',{adc_min[i]},{adc_max[i]}'
                f',{adc_min_cali[i]:.6f},{adc_max_cali[i]:.6f}'
                f',{delta_min_mv:+.3f},{delta_max_mv:+.3f}'
            )

        self._append_log(f"[INFO] calibration points: v_low={v_low:.4f}, m_low={m_low:.3f}, v_high={v_high:.4f}, m_high={m_high:.3f}")
        self._append_log(f"[INFO] Calibration: k={k:.6f} (LSB/V), b={b:.6f} (LSB)")

        return adc_raw_data, mean_cali, adc_min_cali, adc_max_cali

    def gpadc_high_low_temp_test(
        self,
        device_addr=0x17,
        reg_addr=0x56,
        iic_weight=10,
        temp_min=0,
        temp_max=100,
        temp_step=1,
        voltage_channel=100,
        soak_time=180,
        sample_cnt=1000,
        stop_check=None,
        progress_callback=None,
    ):
        try:
            if DEBUG_MOCK:
                if not hasattr(self, "_mock_i2c"):
                    self._mock_i2c = MockI2C()
                chamber = MockChamber()
                deviceI2C = self._mock_i2c
            else:
                if not hasattr(self, 'chamber') or not self.is_chamber_connected:
                    self._test_worker.log.emit("[ERROR] Chamber not connected")
                    self.set_system_status("错误: 温箱未连接", is_error=True)
                    return None
                if self._acq_mode_snapshot == 'UART':
                    deviceI2C = None
                else:
                    if not hasattr(self, "deviceI2C"):
                        self.deviceI2C = I2CInterface()
                        self.set_system_status("I2C接口初始化成功")
                    deviceI2C = self.deviceI2C
                chamber = self.chamber

            if deviceI2C is not None:
                test_data = deviceI2C.read(device_addr, reg_addr, iic_weight)
                self._test_worker.log.emit(f"[INFO] Test data: {test_data:x}")

            temp_data = []
            adc_mean = []
            adc_min = []
            adc_max = []
            adc_raw_all = []

            current_temp = temp_min
            total_points = max(1, int(round((temp_max - temp_min) / temp_step)) + 1)
            point_idx = 0

            while current_temp <= temp_max + 0.001:
                if stop_check and stop_check():
                    self._test_worker.log.emit("[INFO] High/Low temp test stopped by user.")
                    break

                chamber.set_temperature(current_temp)
                self.set_system_status(f"设置温箱温度到 {current_temp:.1f}°C")

                if DEBUG_MOCK:
                    self._test_worker.log.emit(f"[DEBUG] Temp set to {current_temp:.1f}°C (instant)")
                else:
                    if point_idx == 0:
                        try:
                            chamber.start()
                            self._test_worker.log.emit("[INFO] Chamber started (constant-temp run command sent)")
                        except Exception as e:
                            self._test_worker.log.emit(f"[WARN] Chamber start command failed: {e}")

                    stabilizer = TemperatureStabilizer(
                        chamber,
                        log_fn=self._test_worker.log.emit,
                        stop_check=stop_check,
                    )
                    result = stabilizer.wait_for_stable(current_temp)

                    if result.reason == "stopped":
                        self._test_worker.log.emit("[INFO] High/Low temp test stopped by user.")
                        break

                    actual_str = "N/A" if result.actual is None else f"{result.actual:.2f}"
                    self._test_worker.log.emit(
                        f"[INFO] Temperature {result.reason}: target={current_temp:.1f}, "
                        f"actual={actual_str}, waited {result.waited_s:.0f}s, polls={result.poll_count}"
                    )

                    self.set_system_status(f"DUT温度均衡中: {current_temp:.1f}°C")
                    for _ in range(int(soak_time)):
                        if stop_check and stop_check():
                            break
                        time.sleep(1)

                    if stop_check and stop_check():
                        self._test_worker.log.emit("[INFO] High/Low temp test stopped by user.")
                        break

                if DEBUG_MOCK:
                    self._mock_i2c.set_mock_voltage(current_temp / 100.0)

                avg, max_val, min_val, raw_data = self._gpadc_read_by_cnts(
                    device_addr,
                    reg_addr,
                    iic_weight,
                    get_reg_cnt=sample_cnt,
                    return_raw=True,
                    stop_check=stop_check,
                )

                temp_data.append(current_temp)
                adc_mean.append(avg)
                adc_min.append(min_val)
                adc_max.append(max_val)
                adc_raw_all.append(raw_data)

                self._test_worker.log.emit(f"[INFO] T={current_temp:.1f}°C, avg={avg:.3f}, min={min_val}, max={max_val}")

                current_temp += temp_step
                point_idx += 1
                if progress_callback:
                    progress_callback(int(point_idx * 100 / total_points))
                time.sleep(1)
            chamber.set_temperature(25.0)
            self._test_worker.log.emit("===== HIGH/LOW TEMP TEST 结果 =====")
            self._test_worker.log.emit("Temp, RawMean, RawMin, RawMax")
            for i in range(len(temp_data)):
                self._test_worker.log.emit(f"{temp_data[i]:.2f}, {adc_mean[i]:.3f}, {adc_min[i]}, {adc_max[i]}")

            return {
                "voltage": temp_data,
                "mean": adc_mean,
                "min": adc_min,
                "max": adc_max,
            }

        except Exception as e:
            self._test_worker.log.emit(f"[ERROR] {e}")
            logger.error("测试执行错误: %s", e, exc_info=True)
            self.set_system_status(f"错误: {e}", is_error=True)
            return None
    
    def gpadc_temp_consistency_test(
        self,
        device_addr=0x17,
        reg_addr=0x56,
        iic_weight=10,
        temp_min=0,
        temp_max=100,
        temp_step=1,
        voltage_min=0.1,
        voltage_max=1.8,
        voltage_step=0.05,
        voltage_channel=1,
        soak_time=180,
        sample_cnt=1000,
        stop_check=None,
        progress_callback=None,
    ):
        self._test_worker.log.emit(f"[INFO] Running TEMP CONSISTENCY TEST with I2C address: 0x{device_addr:x}, Register: 0x{reg_addr:x}")

        if DEBUG_MOCK:
            if not hasattr(self, "_mock_i2c"):
                self._mock_i2c = MockI2C()
                chamber = MockChamber()
            vol_source = MockN6705C()
            vol_source._mock_i2c = self._mock_i2c
        else:
            if not hasattr(self, 'chamber') or not self.is_chamber_connected:
                self._test_worker.log.emit("[ERROR] Chamber not connected")
                self.set_system_status("错误: 温箱未连接", is_error=True)
                return None
            if self.n6705c is None or not self.is_connected:
                self._test_worker.log.emit("[ERROR] N6705C not connected")
                self.set_system_status("错误: N6705C未连接", is_error=True)
                return None
            if self._acq_mode_snapshot != 'UART':
                if not hasattr(self, "deviceI2C"):
                    self.deviceI2C = I2CInterface()
                    self.set_system_status("I2C接口初始化成功")
            chamber = self.chamber
            vol_source = self.n6705c

        settle_time = 0.0 if DEBUG_MOCK else 0.5
        step_time   = 0.0 if DEBUG_MOCK else 0.2

        voltage_points = []
        v = voltage_min
        while v <= voltage_max + voltage_step * 0.001:
            voltage_points.append(round(v, 6))
            v = round(v + voltage_step, 6)

        temp_list   = []
        mean_matrix = []
        min_matrix  = []
        max_matrix  = []

        total_temp_points = max(1, int(round((temp_max - temp_min) / temp_step)) + 1)
        total_voltage_points = len(voltage_points)
        total_steps = total_temp_points * total_voltage_points
        completed_steps = 0

        current_temp = temp_min
        while current_temp <= temp_max + 0.001:
            if stop_check and stop_check():
                self._test_worker.log.emit("[INFO] Temp consistency test stopped by user.")
                break

            chamber.set_temperature(current_temp)
            self.set_system_status(f"设置温箱温度到 {current_temp:.1f}°C")

            if DEBUG_MOCK:
                self._test_worker.log.emit(f"[DEBUG] Temp set to {current_temp:.1f}°C (instant)")
            else:
                history = []
                stable_count = 0
                while True:
                    if stop_check and stop_check():
                        break
                    actual_temp = chamber.get_current_temp()
                    history.append(actual_temp)
                    if len(history) > 10:
                        history.pop(0)
                    if len(history) >= 5:
                        if max(history) - min(history) < 0.2:
                            stable_count += 1
                        else:
                            stable_count = 0
                        if stable_count >= 3:
                            break
                    self._test_worker.log.emit(f"[INFO] Temp stabilizing: target={current_temp:.1f}, actual={actual_temp:.2f}")
                    # 30s 轮询间隔切成 1s 片，保证停止请求可及时打断（日志节奏不变）
                    for _ in range(30):
                        if stop_check and stop_check():
                            break
                        time.sleep(1)

                if stop_check and stop_check():
                    self._test_worker.log.emit("[INFO] Temp consistency test stopped by user.")
                    break

                self.set_system_status(f"DUT温度均衡中: {current_temp:.1f}°C")
                for _ in range(int(soak_time)):
                    if stop_check and stop_check():
                        break
                    time.sleep(1)

                if stop_check and stop_check():
                    self._test_worker.log.emit("[INFO] Temp consistency test stopped by user.")
                    break

            mean_row = []
            min_row  = []
            max_row  = []

            vol_source.set_voltage(voltage_channel, voltage_min)
            time.sleep(settle_time)

            for vpt in voltage_points:
                if stop_check and stop_check():
                    break
                vol_source.set_voltage(voltage_channel, vpt)
                time.sleep(step_time)

                avg, max_val, min_val = self._gpadc_read_by_cnts(
                    device_addr,
                    reg_addr,
                    iic_weight,
                    get_reg_cnt=sample_cnt,
                    return_raw=False,
                    stop_check=stop_check,
                )
                mean_row.append(avg)
                min_row.append(min_val)
                max_row.append(max_val)
                completed_steps += 1
                if progress_callback:
                    progress_callback(int(completed_steps * 100 / total_steps))

            if stop_check and stop_check():
                break

            temp_list.append(current_temp)
            mean_matrix.append(mean_row)
            min_matrix.append(min_row)
            max_matrix.append(max_row)

            self._test_worker.log.emit(f"[INFO] T={current_temp:.1f}°C  voltage sweep done ({len(mean_row)} points)")
            current_temp = round(current_temp + temp_step, 6)

        vol_header = "  ".join(f"{v:.3f}" for v in voltage_points)

        self._test_worker.log.emit("===== TEMP CONSISTENCY TEST 结果 (Mean) =====")
        self._test_worker.log.emit("Temp\\Voltage  " + vol_header)
        for i, t in enumerate(temp_list):
            if i < len(mean_matrix):
                row = "  ".join(f"{mean_matrix[i][j]:.1f}" for j in range(len(mean_matrix[i])))
                self._test_worker.log.emit(f"T={t:.1f}  {row}")

        self._test_worker.log.emit("===== TEMP CONSISTENCY TEST 结果 (Min) =====")
        self._test_worker.log.emit("Temp\\Voltage  " + vol_header)
        for i, t in enumerate(temp_list):
            if i < len(min_matrix):
                row = "  ".join(f"{min_matrix[i][j]:.1f}" for j in range(len(min_matrix[i])))
                self._test_worker.log.emit(f"T={t:.1f}  {row}")

        self._test_worker.log.emit("===== TEMP CONSISTENCY TEST 结果 (Max) =====")
        self._test_worker.log.emit("Temp\\Voltage  " + vol_header)
        for i, t in enumerate(temp_list):
            if i < len(max_matrix):
                row = "  ".join(f"{max_matrix[i][j]:.1f}" for j in range(len(max_matrix[i])))
                self._test_worker.log.emit(f"T={t:.1f}  {row}")

        mean_cali_matrix = []
        min_cali_matrix = []
        max_cali_matrix = []
        for i, t in enumerate(temp_list):
            if i < len(mean_matrix):
                row_result = {
                    "voltage": voltage_points,
                    "mean": mean_matrix[i],
                    "min": min_matrix[i],
                    "max": max_matrix[i],
                }
                _, mean_cali, min_cali, max_cali = self._calibration_data(row_result)
                mean_cali_matrix.append(mean_cali)
                min_cali_matrix.append(min_cali)
                max_cali_matrix.append(max_cali)

        self._test_worker.log.emit("===== TEMP CONSISTENCY TEST 校准结果 (Mean Cali) =====")
        self._test_worker.log.emit("Temp\\Voltage  " + vol_header)
        for i, t in enumerate(temp_list):
            if i < len(mean_cali_matrix):
                row = "  ".join(f"{mean_cali_matrix[i][j]:.6f}" for j in range(len(mean_cali_matrix[i])))
                self._test_worker.log.emit(f"T={t:.1f}  {row}")

        self._test_worker.log.emit("===== TEMP CONSISTENCY TEST 校准结果 (Min Cali) =====")
        self._test_worker.log.emit("Temp\\Voltage  " + vol_header)
        for i, t in enumerate(temp_list):
            if i < len(min_cali_matrix):
                row = "  ".join(f"{min_cali_matrix[i][j]:.6f}" for j in range(len(min_cali_matrix[i])))
                self._test_worker.log.emit(f"T={t:.1f}  {row}")

        self._test_worker.log.emit("===== TEMP CONSISTENCY TEST 校准结果 (Max Cali) =====")
        self._test_worker.log.emit("Temp\\Voltage  " + vol_header)
        for i, t in enumerate(temp_list):
            if i < len(max_cali_matrix):
                row = "  ".join(f"{max_cali_matrix[i][j]:.6f}" for j in range(len(max_cali_matrix[i])))
                self._test_worker.log.emit(f"T={t:.1f}  {row}")

        return {
            "temp":    temp_list,
            "voltage": voltage_points,
            "mean":    mean_matrix,
            "min":     min_matrix,
            "max":     max_matrix,
            "mean_cali": mean_cali_matrix,
            "min_cali":  min_cali_matrix,
            "max_cali":  max_cali_matrix,
        }

    def _calculate_gpadc_parameters(self, force_voltage_result):
        """
        基于 gpadc_force_voltage_test 的返回值计算 GPADC 性能参数。

        force_voltage_result 格式:
            {
                "voltage": [v0, v1, ...],   # 输入电压 (V)
                "mean":    [m0, m1, ...],   # 每个电压点的 ADC 均值
                "min":     [n0, n1, ...],   # 每个电压点的 ADC 最小值
                "max":     [x0, x1, ...],   # 每个电压点的 ADC 最大值
            }

        返回 params dict，包含:
            linearity, dnl, inl, enob, gain_error, offset_error,
            avg, min, max
        """
        import numpy as np

        voltage = np.array(force_voltage_result["voltage"], dtype=float)
        adc_mean = np.array(force_voltage_result["mean"], dtype=float)
        adc_min = np.array(force_voltage_result["min"], dtype=float)
        adc_max = np.array(force_voltage_result["max"], dtype=float)

        n = len(voltage)
        params = {}

        params['avg'] = float(np.mean(adc_mean))
        params['max'] = float(np.max(adc_max))
        params['min'] = float(np.min(adc_min))

        if n < 2:
            params.update({'linearity': 0.0, 'dnl': 0.0, 'inl': 0.0,
                           'enob': 0.0, 'gain_error': 0.0, 'offset_error': 0.0})
            return params

        # ── 线性拟合 voltage → adc_mean ──────────────────────────────
        coeffs = np.polyfit(voltage, adc_mean, 1)
        slope = coeffs[0]       # LSB/V
        intercept = coeffs[1]
        adc_ideal = np.polyval(coeffs, voltage)

        # Linearity: R²（越接近1越好）
        ss_res = np.sum((adc_mean - adc_ideal) ** 2)
        ss_tot = np.sum((adc_mean - np.mean(adc_mean)) ** 2)
        params['linearity'] = float(1.0 - ss_res / ss_tot) if ss_tot > 0 else 1.0

        # ── INL (Integral Non-Linearity) ──────────────────────────────
        # 每个采样点与理想线性值的偏差，取峰峰值的一半（LSB）
        inl_values = adc_mean - adc_ideal
        params['inl'] = float(np.max(np.abs(inl_values)))

        # ── DNL (Differential Non-Linearity) ─────────────────────────
        # 相邻步进的实际ADC增量 vs 理想增量之差（LSB）
        ideal_step = slope * (voltage[1] - voltage[0])   # 理想单步增量（LSB）
        actual_steps = np.diff(adc_mean)
        if ideal_step != 0:
            dnl_values = (actual_steps - ideal_step) / abs(ideal_step)
            params['dnl'] = float(np.max(np.abs(dnl_values)))
        else:
            params['dnl'] = 0.0

        # ── ENOB (Effective Number of Bits) ──────────────────────────
        # 用每个采样点的噪声范围（max-min）估算 RMS 噪声，再计算 SINAD→ENOB
        noise_pp = adc_max - adc_min          # peak-to-peak 噪声
        noise_rms = np.mean(noise_pp) / (2 * np.sqrt(3))   # 均匀分布近似
        adc_full_range = params['max'] - params['min']
        if noise_rms > 0 and adc_full_range > 0:
            snr_db = 20 * np.log10(adc_full_range / (noise_rms * 2 * np.sqrt(3)))
            params['enob'] = float(max(0.0, (snr_db - 1.76) / 6.02))
        else:
            params['enob'] = 0.0

        # ── Gain Error ────────────────────────────────────────────────
        # 理想斜率 = (ADC满量程范围) / (电压满量程范围)
        volt_range = float(voltage[-1] - voltage[0])
        adc_range = float(adc_mean[-1] - adc_mean[0])
        if volt_range > 0:
            ideal_slope = adc_range / volt_range
            params['gain_error'] = float(((slope - ideal_slope) / ideal_slope) * 100) if ideal_slope != 0 else 0.0
        else:
            params['gain_error'] = 0.0

        # ── Offset Error ──────────────────────────────────────────────
        # 拟合线在最小电压处的预测值与实际测量值的偏差（LSB）
        params['offset_error'] = float(adc_ideal[0] - adc_mean[0])
        self._append_log(f"[RESULT] INL={params['inl']:.3f}, DNL={params['dnl']:.3f}, ENOB={params['enob']:.3f}, Gain Error: {params['gain_error']:.3f}%, Offset Error: {params['offset_error']:.3f} LSB")
        return params
    
    def _build_default_chart_placeholder(self):
        """构建/重置图表占位符为初始空坐标系（新测试开始时复用）。"""
        self._clear_chart_placeholder()
        layout = self.chart_placeholder.layout()
        if layout is None:
            layout = QVBoxLayout(self.chart_placeholder)
        layout.setContentsMargins(22, 22, 22, 18)
        layout.setSpacing(12)

        legend_row = QHBoxLayout()
        legend_row.addStretch()

        actual_legend = QLabel("↔ Actual Code")
        actual_legend.setStyleSheet("color: #00d39a; font-size: 12px;")
        ideal_legend = QLabel("↔ Ideal Code")
        ideal_legend.setStyleSheet("color: #7e96bf; font-size: 12px;")

        legend_row.addWidget(actual_legend)
        legend_row.addWidget(ideal_legend)
        legend_row.addStretch()
        layout.addLayout(legend_row)

        plot_area = QFrame()
        plot_area.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        plot_area.setStyleSheet("""
            QFrame {
                background-color: transparent;
                border-left: 1px solid #6f7fa5;
                border-bottom: 1px solid #6f7fa5;
                border-top: 1px dashed rgba(126,150,191,0.18);
                border-right: none;
                border-radius: 0px;
            }
        """)
        layout.addWidget(plot_area, 1)

        x_label = QLabel("Input Voltage (V)")
        x_label.setAlignment(Qt.AlignCenter)
        x_label.setObjectName("muted_label")
        layout.addWidget(x_label)

    def _reset_result_display(self):
        """新测试开始时清除上一次的曲线与结果，避免新旧数据混显。"""
        self.clear_results()
        self._export_data = None
        self._chart_image_bytes = None
        self._loaded_record = None
        self._build_default_chart_placeholder()

    def _clear_chart_placeholder(self):
        existing = self.chart_placeholder.layout()
        if existing is not None:
            while existing.count():
                item = existing.takeAt(0)
                widget = item.widget()
                if widget is not None:
                    widget.deleteLater()
                else:
                    sub = item.layout()
                    if sub is not None:
                        while sub.count():
                            sub_item = sub.takeAt(0)
                            sub_widget = sub_item.widget()
                            if sub_widget is not None:
                                sub_widget.deleteLater()

    def _plot_voltage_adc_curve(self, voltage_data, mean_cali, adc_min_cali=None, adc_max_cali=None, is_temp_mode=False):
        try:
            import pyqtgraph as pg
            import numpy as np

            # Curve View 过滤（载入 Recent 记录时生效；刚完成测试时三项默认全开）
            show_mean = self._is_curve_view_enabled('mean')
            show_band = self._is_curve_view_enabled('band')
            show_error = self._is_curve_view_enabled('error')
            if not (show_mean or show_band or show_error):
                show_mean = True

            self._clear_chart_placeholder()
            layout = self.chart_placeholder.layout()
            if layout is None:
                layout = QVBoxLayout(self.chart_placeholder)
            layout.setContentsMargins(14, 14, 14, 10)
            layout.setSpacing(8)

            # 当前记录专属色（有载入记录时用其色，否则默认绿）
            curve_color = "#00d39a"
            record_name = ""
            if self._loaded_record is not None:
                curve_color = self._record_color(self._loaded_record['id'])
                record_name = f" #{self._loaded_record['id']} {self._record_display_name(self._loaded_record)} · "

            legend_row = QHBoxLayout()
            legend_row.addStretch()
            if is_temp_mode:
                cali_legend = QLabel(f"● {record_name}Calibrated Temperature (°C)")
                x_axis_title = "Input Temperature (°C)"
                y_axis_title = "Calibrated Temperature (°C)"
            else:
                cali_legend = QLabel(f"● {record_name}Calibrated Voltage (V)")
                x_axis_title = "Input Voltage (V)"
                y_axis_title = "Calibrated Voltage (V)"
            cali_legend.setStyleSheet(f"color: {curve_color}; font-size: 12px;")
            ideal_legend = QLabel("● Ideal (y = x)")
            ideal_legend.setStyleSheet("color: #7e96bf; font-size: 12px;")
            band_legend = QLabel("▨ Max/Min Error Band")
            band_legend.setStyleSheet("color: #f0a040; font-size: 12px;")
            legend_row.addWidget(cali_legend)
            legend_row.addSpacing(16)
            legend_row.addWidget(ideal_legend)
            legend_row.addSpacing(16)
            legend_row.addWidget(band_legend)
            if not is_temp_mode:
                # 误差单位按量级自适应：<1V 显示 mV，否则 V（图例与轴标题共用）
                err_arr = np.array(mean_cali, dtype=float) - np.array(voltage_data, dtype=float)
                max_abs_err = float(np.max(np.abs(err_arr))) if err_arr.size else 0.0
                err_unit = "mV" if max_abs_err < 1.0 else "V"
                err_legend = QLabel(f"● Error (Actual - Ideal) ({err_unit})")
                err_legend.setStyleSheet("color: #e05c5c; font-size: 12px;")
                legend_row.addSpacing(16)
                legend_row.addWidget(err_legend)
            legend_row.addStretch()
            layout.addLayout(legend_row)

            pw = pg.PlotWidget()
            pw.setBackground("#0a1735")
            pw.showGrid(x=True, y=True, alpha=0.15)
            pw.setLabel("left",   y_axis_title, color="#a0b4d8")
            pw.setLabel("bottom", x_axis_title, color="#a0b4d8")
            self._attach_curve_context_menu(pw)

            for axis_name in ("left", "bottom"):
                axis = pw.getAxis(axis_name)
                axis.setTextPen(pg.mkPen("#a0b4d8"))
                axis.setPen(pg.mkPen("#3a4f7a"))

            x = np.array(voltage_data, dtype=float)
            y = np.array(mean_cali,    dtype=float)

            if show_band and adc_min_cali is not None and adc_max_cali is not None:
                y_min = np.array(adc_min_cali, dtype=float)
                y_max = np.array(adc_max_cali, dtype=float)

                fill = pg.PlotDataItem(
                    np.concatenate([x, x[::-1]]),
                    np.concatenate([y_max, y_min[::-1]]),
                    pen=pg.mkPen(color="#f0a040", width=1))
                fill_under = pg.FillBetweenItem(
                    pg.PlotDataItem(x, y_max),
                    pg.PlotDataItem(x, y_min),
                    brush=pg.mkBrush(240, 160, 64, 50)
                )
                pw.addItem(fill_under)

                pw.plot(x, y_max, pen=pg.mkPen(color="#f0a040", width=1,
                        style=pg.QtCore.Qt.DashLine))
                pw.plot(x, y_min, pen=pg.mkPen(color="#f0a040", width=1,
                        style=pg.QtCore.Qt.DashLine))

            if show_mean:
                pw.plot(x, x, pen=pg.mkPen(color="#7e96bf", width=1,
                        style=pg.QtCore.Qt.DashLine))

                pw.plot(x, y, pen=pg.mkPen(color=curve_color, width=2),
                        symbol="o", symbolSize=5,
                        symbolBrush=curve_color, symbolPen=None)

            if show_error and not is_temp_mode:
                # 右 Y 轴：Actual - Ideal 差值曲线（校准电压 - 输入电压）
                # 单位与图例共用 err_unit（mV 时数据放大 1000 倍）
                diff = y - x
                if err_unit == "mV":
                    diff = diff * 1000.0

                pw.showAxis('right')
                right_axis = pw.getAxis('right')
                right_axis.setLabel(f"Error ({err_unit})", color="#e05c5c")
                right_axis.setTextPen(pg.mkPen("#e05c5c"))
                right_axis.setPen(pg.mkPen("#3a4f7a"))

                right_view = pg.ViewBox()
                right_view.setMouseEnabled(x=False, y=False)
                pw.plotItem.scene().addItem(right_view)
                right_axis.linkToView(right_view)
                right_view.setXLink(pw.plotItem)
                pw.plotItem.vb.sigResized.connect(
                    lambda: right_view.setGeometry(pw.plotItem.vb.sceneBoundingRect()))
                right_view.setGeometry(pw.plotItem.vb.sceneBoundingRect())

                right_view.addItem(pg.PlotDataItem(
                    x, diff,
                    pen=pg.mkPen(color="#e05c5c", width=2),
                    symbol="o", symbolSize=4,
                    symbolBrush="#e05c5c", symbolPen=None))

            layout.addWidget(pw, 1)

            x_label = QLabel(x_axis_title)
            x_label.setAlignment(Qt.AlignCenter)
            x_label.setObjectName("muted_label")
            layout.addWidget(x_label)

            try:
                import io
                from pyqtgraph.exporters import ImageExporter
                from PySide6.QtCore import QBuffer, QIODevice
                from PySide6.QtGui import QImage
                exporter = ImageExporter(pw.plotItem)
                exporter.parameters()['width'] = 1200
                exporter.parameters()['height'] = 700
                result = exporter.export(toBytes=True)
                if isinstance(result, QImage):
                    qbuf = QBuffer()
                    qbuf.open(QIODevice.WriteOnly)
                    result.save(qbuf, "PNG")
                    raw = bytes(qbuf.data())
                    qbuf.close()
                else:
                    raw = bytes(result)
                buf = io.BytesIO(raw)
                self._chart_image_bytes = buf
            except Exception as ex:
                self._append_log(f"[WARN] Chart snapshot failed: {ex}")
                self._chart_image_bytes = None

        except Exception as e:
            self._append_log(f"[ERROR] Error plotting voltage-ADC curve: {e}")
            logger.error("Error plotting voltage-ADC curve: %s", e, exc_info=True)

    def _plot_temp_consistency_curves(self, result):
        try:
            import pyqtgraph as pg
            import numpy as np

            temp_list    = result["temp"]
            voltage_pts  = result["voltage"]
            mean_matrix  = result["mean"]
            min_matrix   = result["min"]
            max_matrix   = result["max"]

            show_band = self._is_curve_view_enabled('band')

            self._clear_chart_placeholder()
            layout = self.chart_placeholder.layout()
            if layout is None:
                layout = QVBoxLayout(self.chart_placeholder)
            layout.setContentsMargins(14, 14, 14, 10)
            layout.setSpacing(8)

            palette = [
                "#00d39a", "#f0a040", "#7e96bf", "#e05c5c",
                "#a78bfa", "#34d399", "#fb923c", "#60a5fa",
            ]

            legend_row = QHBoxLayout()
            legend_row.addStretch()
            for i, t in enumerate(temp_list):
                color = palette[i % len(palette)]
                lbl = QLabel(f"● {t:.1f}°C")
                lbl.setStyleSheet(f"color: {color}; font-size: 11px;")
                legend_row.addWidget(lbl)
                legend_row.addSpacing(8)
            legend_row.addStretch()
            layout.addLayout(legend_row)

            pw = pg.PlotWidget()
            pw.setBackground("#0a1735")
            pw.showGrid(x=True, y=True, alpha=0.15)
            pw.setLabel("left",   "ADC Code",         color="#a0b4d8")
            pw.setLabel("bottom", "Input Voltage (V)", color="#a0b4d8")
            self._attach_curve_context_menu(pw)

            for axis_name in ("left", "bottom"):
                axis = pw.getAxis(axis_name)
                axis.setTextPen(pg.mkPen("#a0b4d8"))
                axis.setPen(pg.mkPen("#3a4f7a"))

            x = np.array(voltage_pts, dtype=float)

            for i, t in enumerate(temp_list):
                if i >= len(mean_matrix):
                    break
                color = palette[i % len(palette)]
                mean_row = np.array(mean_matrix[i], dtype=float)
                min_row  = np.array(min_matrix[i],  dtype=float)
                max_row  = np.array(max_matrix[i],  dtype=float)

                n = min(len(x), len(mean_row))

                if show_band:
                    fill = pg.FillBetweenItem(
                        pg.PlotDataItem(x[:n], max_row[:n]),
                        pg.PlotDataItem(x[:n], min_row[:n]),
                        brush=pg.mkBrush(
                            int(color[1:3], 16),
                            int(color[3:5], 16),
                            int(color[5:7], 16),
                            35,
                        )
                    )
                    pw.addItem(fill)

                    pw.plot(x[:n], max_row[:n],
                            pen=pg.mkPen(color=color, width=1,
                                         style=pg.QtCore.Qt.DashLine))
                    pw.plot(x[:n], min_row[:n],
                            pen=pg.mkPen(color=color, width=1,
                                         style=pg.QtCore.Qt.DashLine))
                pw.plot(x[:n], mean_row[:n],
                        pen=pg.mkPen(color=color, width=2),
                        symbol="o", symbolSize=4,
                        symbolBrush=color, symbolPen=None,
                        name=f"{t:.1f}°C")

            layout.addWidget(pw, 1)

            x_label = QLabel("Input Voltage (V)")
            x_label.setAlignment(Qt.AlignCenter)
            x_label.setObjectName("muted_label")
            layout.addWidget(x_label)

            try:
                import io
                from pyqtgraph.exporters import ImageExporter
                from PySide6.QtCore import QBuffer, QIODevice
                from PySide6.QtGui import QImage
                exporter = ImageExporter(pw.plotItem)
                exporter.parameters()['width'] = 1200
                exporter.parameters()['height'] = 700
                snap = exporter.export(toBytes=True)
                if isinstance(snap, QImage):
                    qbuf = QBuffer()
                    qbuf.open(QIODevice.WriteOnly)
                    snap.save(qbuf, "PNG")
                    raw = bytes(qbuf.data())
                    qbuf.close()
                else:
                    raw = bytes(snap)
                self._chart_image_bytes = io.BytesIO(raw)
            except Exception as ex:
                self._append_log(f"[WARN] Chart snapshot failed: {ex}")
                self._chart_image_bytes = None

        except Exception as e:
            self._append_log(f"[ERROR] Error plotting temp consistency curves: {e}")
            logger.error("Error plotting temp consistency curves: %s", e, exc_info=True)

    def _plot_temperature_adc_curve(self, temp_data, adc_data):
        """绘制温度-ADC曲线到UI"""
        try:
            self._clear_chart_placeholder()
            chart_placeholder_layout = self.chart_placeholder.layout()
            if chart_placeholder_layout is None:
                chart_placeholder_layout = QVBoxLayout(self.chart_placeholder)
            chart_placeholder_layout.setContentsMargins(22, 22, 22, 18)
            chart_placeholder_layout.setSpacing(12)
            
            # 添加图例
            legend_row = QHBoxLayout()
            legend_row.addStretch()
            
            temp_legend = QLabel("↔ Temperature (°C)")
            temp_legend.setStyleSheet("color: #00d39a; font-size: 12px;")
            adc_legend = QLabel("↔ ADC Value")
            adc_legend.setStyleSheet("color: #59a8ff; font-size: 12px;")
            
            legend_row.addWidget(temp_legend)
            legend_row.addWidget(adc_legend)
            legend_row.addStretch()
            chart_placeholder_layout.addLayout(legend_row)
            
            # 创建绘图区域
            plot_area = QFrame()
            plot_area.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            plot_area.setStyleSheet("""
                QFrame {
                    background-color: transparent;
                    border-left: 1px solid #6f7fa5;
                    border-bottom: 1px solid #6f7fa5;
                    border-top: 1px dashed rgba(126,150,191,0.18);
                    border-right: none;
                    border-radius: 0px;
                }
            """)
            
            # 简单的曲线绘制（使用QLabel模拟，实际项目中应该使用专业绘图库）
            import numpy as np
            
            # 创建一个简单的文本显示，显示曲线数据
            curve_data = """
Temperature (°C) | ADC Value
----------------|----------
"""
            for t, adc in zip(temp_data, adc_data):
                curve_data += f"{t:>15.1f} | {adc:>10.3f}\n"
            
            # 添加统计信息
            curve_data += "\n" + "="*40 + "\n"
            curve_data += f"Min Temp: {min(temp_data):.1f}°C\n"
            curve_data += f"Max Temp: {max(temp_data):.1f}°C\n"
            curve_data += f"Min ADC:  {min(adc_data):.3f}\n"
            curve_data += f"Max ADC:  {max(adc_data):.3f}\n"
            curve_data += f"Avg ADC:  {np.mean(adc_data):.3f}\n"
            
            data_label = QLabel(curve_data)
            data_label.setStyleSheet(f"color: #d8e3ff; font-family: {FONT_MONO}; font-size: 10px;")
            data_label.setWordWrap(True)
            
            # 创建滚动区域以显示大量数据
            scroll_area = QScrollArea()
            scroll_area.setWidget(data_label)
            scroll_area.setWidgetResizable(True)
            scroll_area.setStyleSheet("background: transparent; border: none;" + SCROLL_AREA_STYLE)
            
            # 添加到布局
            chart_placeholder_layout.addWidget(scroll_area, 1)
            
            # 添加轴标签
            x_label = QLabel("Temperature (°C)")
            x_label.setAlignment(Qt.AlignCenter)
            x_label.setObjectName("muted_label")
            chart_placeholder_layout.addWidget(x_label)
            
            # 在实际项目中，应该使用matplotlib或PyQtGraph等专业绘图库
            # 这里只是一个简单的模拟实现
            
        except Exception as e:
            self._append_log(f"[ERROR] Error plotting temperature-ADC curve: {e}")
            logger.error("Error plotting temperature-ADC curve: %s", e, exc_info=True)
            
            # 如果绘图失败，显示错误信息
            error_label = QLabel(f"绘图失败: {str(e)}")
            error_label.setStyleSheet("color: #ff5a7a; font-size: 14px;")
            error_label.setAlignment(Qt.AlignCenter)

            self._clear_chart_placeholder()
            chart_placeholder_layout = self.chart_placeholder.layout()
            if chart_placeholder_layout is None:
                chart_placeholder_layout = QVBoxLayout(self.chart_placeholder)
            chart_placeholder_layout.addWidget(error_label, 1, Qt.AlignCenter)

    # ------------------------------------------------------------------
    # AIControllablePage 契约实现（AIAssist_PageScopedControlPlan.md §2 / Phase 5）
    #
    # GPADC 测试接入 AI 受控契约，薄封装既有方法：
    #   - ai_get_config 复用 get_test_config()
    #   - ai_apply_config 经 apply_config_to_controls() 单一写入口回填控件
    #   - ai_start_test/ai_stop_test 复用 _start_test/_stop_test
    # 枢纽（MainWindow.resolve_active_ai_page）经 Tab 子页下钻拿到本实例，
    # 鸭子调用契约方法，无需 core / handler 改动。
    # ------------------------------------------------------------------
    def ai_capabilities(self) -> set[str]:
        return {
            CAP_GET_CONFIG,
            CAP_APPLY_CONFIG,
            CAP_START_TEST,
            CAP_STOP_TEST,
            CAP_GET_RESULT,
        }

    def ai_get_config(self) -> dict[str, Any] | None:
        try:
            return self.get_test_config()
        except Exception:  # noqa: BLE001 - 快照失败降级为 None
            logger.error("AI 读取 GPADC 测试配置失败", exc_info=True)
            return None

    def ai_apply_config(self, payload: Any) -> tuple[bool, str]:
        """落地配置草案到控件（写操作，经确认+审计后由枢纽调用）。

        运行中拒绝改配置（§6.3），避免与正在执行的测试冲突。
        """
        if self.is_test_running:
            return False, "测试运行中，无法修改配置，请先停止测试。"
        return self.apply_config_to_controls(payload if isinstance(payload, dict) else {})

    def ai_start_test(self) -> tuple[bool, str]:
        if self.is_test_running:
            return False, "测试已在运行中。"
        # 按当前测试项的仪器需求前置校验（INSTRUMENT_MAP）
        required = self.INSTRUMENT_MAP.get(self.current_test_item, [])
        if "n6705c" in required and (not self.is_connected or self.n6705c is None):
            return False, f"当前测试项 {self.current_test_item} 需要 N6705C，请先连接再启动。"
        if "chamber" in required and (not self.is_chamber_connected or self.chamber is None):
            return False, f"当前测试项 {self.current_test_item} 需要温箱，请先连接温箱再启动。"
        # I2C 模式下校验地址可解析
        if self.iic_radio.isChecked():
            try:
                int(self.iic_device_address.text(), 16)
                int(self.iic_data_address.text(), 16)
            except ValueError:
                return False, "I2C 设备/寄存器地址格式无效（应为 16 进制）。"
        self.append_log(f"[AI] 请求启动 GPADC 测试：{self.current_test_item}。")
        try:
            self._start_test()
        except Exception:  # noqa: BLE001 - 启动异常转可读结果
            logger.error("AI 启动 GPADC 测试失败", exc_info=True)
            return False, "启动测试异常，请查看日志。"
        if self.is_test_running:
            return True, f"已请求启动 GPADC 测试：{self.current_test_item}。"
        return False, "启动未成功，请查看执行日志。"

    def ai_stop_test(self) -> tuple[bool, str]:
        if not self.is_test_running:
            return False, "当前未在运行测试。"
        self.append_log("[AI] 请求停止测试。")
        try:
            self._stop_test()
        except Exception:  # noqa: BLE001 - 停止异常转可读结果
            logger.error("AI 停止 GPADC 测试失败", exc_info=True)
            return False, "停止测试异常，请查看日志。"
        return True, "已发送停止请求。"

    def ai_get_result_summary(self) -> dict[str, Any] | None:
        summary: dict[str, Any] = {
            "available": True,
            "running": self.is_test_running,
            "test_item": self.current_test_item,
        }
        if not self._export_data:
            return summary
        params = self._export_data.get("params") if isinstance(self._export_data, dict) else None
        if params:
            summary["params"] = params
        return summary

    # ------------------------------------------------------------------
    # UI 回填单一写入口（AIAssist_PageScopedControlPlan.md §4.2）
    #
    # apply_config_to_controls(cfg) 是回填测试配置控件的唯一入口，
    # AI 回填与未来轮询/手动刷新共用，杜绝两套逻辑漂移。键名与
    # get_test_config() 输出对齐。
    # ------------------------------------------------------------------
    def apply_config_to_controls(self, cfg: dict) -> tuple[bool, str]:
        if not isinstance(cfg, dict):
            return False, "配置草案格式无效（期望 dict）。"

        # 线程边界（§4.2-2）：AI 决策在 QThread，回填须经主线程执行；
        # dispatcher 经 QTimer.singleShot(0) 已切回主线程，此处加防御性守卫，
        # 杜绝 worker 线程直接 setValue 违反「UI 禁阻塞 / 跨线程改控件」铁律。
        if threading.current_thread() is not threading.main_thread():
            logger.error(
                "apply_config_to_controls 在非主线程被调用，拒绝回填以防违反线程边界"
            )
            return False, "配置回填未在主线程执行，已拒绝。"

        applied: list[str] = []
        touched: list = []

        def _set_spin(spin, key):
            val = cfg.get(key)
            if val is None:
                return
            try:
                spin.setValue(float(val))
            except (TypeError, ValueError):
                return
            applied.append(key)
            touched.append(spin)

        def _set_int_spin(spin, key):
            val = cfg.get(key)
            if val is None:
                return
            try:
                spin.setValue(int(float(val)))
            except (TypeError, ValueError):
                return
            applied.append(key)
            touched.append(spin)

        def _set_text(edit, key):
            val = cfg.get(key)
            if val is None:
                return
            edit.setText(str(val))
            applied.append(key)
            touched.append(edit)

        def _set_combo_data(combo, key):
            val = cfg.get(key)
            if val is None:
                return
            idx = combo.findData(val)
            if idx >= 0:
                combo.setCurrentIndex(idx)
                applied.append(key)
                touched.append(combo)

        # 测试项（按 data 匹配，触发 _on_test_item_changed → _set_test_item）
        test_item = cfg.get("test_item")
        if test_item is not None:
            idx = self.test_item_combo.findData(test_item)
            if idx >= 0:
                self.test_item_combo.setCurrentIndex(idx)
                applied.append("test_item")
                touched.append(self.test_item_combo)

        # 数据采集模式
        mode = cfg.get("data_acquisition_mode")
        if mode is not None:
            is_iic = str(mode).upper().startswith("IIC")
            self.iic_radio.setChecked(is_iic)
            self.uart_radio.setChecked(not is_iic)
            self._update_data_acquisition_ui()
            applied.append("data_acquisition_mode")
            touched.extend([self.iic_radio, self.uart_radio])

        # I2C 地址
        _set_text(self.iic_device_address, "iic_device_address")
        _set_text(self.iic_data_address, "iic_data_address")
        _set_combo_data(self.iic_width_combo, "iic_width")

        # UART 配置
        _set_text(self.uart_keyword, "uart_keyword")

        # 电压扫描参数
        _set_combo_data(self.voltage_channel, "voltage_channel")
        _set_spin(self.voltage_min, "voltage_min")
        _set_spin(self.voltage_max, "voltage_max")
        _set_spin(self.voltage_step, "voltage_step")

        # 温度扫描参数
        _set_spin(self.temp_min, "temp_min")
        _set_spin(self.temp_max, "temp_max")
        _set_spin(self.temp_step, "temp_step")
        _set_int_spin(self.soak_time, "soak_time")

        # 采样次数
        _set_int_spin(self.sample_count, "sample_count")

        # 校准点（留空 = 自动选取）
        _set_text(self.calib_low, "calib_low")
        _set_text(self.calib_high, "calib_high")

        if not applied:
            return False, "配置草案未包含任何可识别的配置项。"
        # §4.2-3 可视化反馈：被 AI 修改的控件临时高亮（Phase 3）。
        self._highlight_widgets(touched)
        self.append_log(f"[AI] 已应用配置：{', '.join(applied)}")
        return True, f"已应用配置项：{', '.join(applied)}。"

    def _highlight_widgets(self, widgets: list) -> None:
        """被 AI 修改的控件临时高亮边框（§4.2-3 / Phase 3）。"""
        if not widgets:
            return
        for widget in widgets:
            if widget is None:
                continue
            widget.setStyleSheet(_AI_HIGHLIGHT_QSS)
            widget.setProperty("aiHighlighted", True)
            widget.style().unpolish(widget)
            widget.style().polish(widget)

            def _clear(_w=widget):
                try:
                    _w.setStyleSheet("")
                    _w.setProperty("aiHighlighted", False)
                    _w.style().unpolish(_w)
                    _w.style().polish(_w)
                except RuntimeError:  # noqa: BLE001 - widget 可能已销毁
                    pass
            QTimer.singleShot(_AI_HIGHLIGHT_MS, _clear)


if __name__ == "__main__":
    import sys
    import logging
    from PySide6.QtCore import qInstallMessageHandler, QtMsgType
    from log_config import setup_logging
    from ui.standalone import resize_and_center_window
    setup_logging()

    app = QApplication(sys.argv)
    app.setStyle("Fusion")

    def custom_message_handler(msg_type, context, message):
        if msg_type == QtMsgType.QtWarningMsg and "QPainter::end" in message:
            return
        logging.getLogger(__name__).debug("%s:%s - %s", context.file, context.line, message)
    qInstallMessageHandler(custom_message_handler)

    gpadc_test_ui = GPADCTestUI()
    gpadc_test_ui.setWindowTitle("GPADC Test System")
    resize_and_center_window(gpadc_test_ui)
    gpadc_test_ui.show()

    sys.exit(app.exec())



