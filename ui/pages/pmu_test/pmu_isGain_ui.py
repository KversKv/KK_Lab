#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PMU Is_gain测试UI组件
暗色卡片式重构版本（PySide6）
"""

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGridLayout, QPushButton, QComboBox,
    QLabel, QLineEdit, QSpinBox, QDoubleSpinBox, QFrame, QTextEdit,
    QProgressBar, QTableWidget, QTableWidgetItem, QHeaderView, QCheckBox,
    QFileDialog, QDialog, QRadioButton, QButtonGroup, QSizePolicy, QScrollArea,
)
from PySide6.QtCore import Qt, Signal, QThread, QTimer
from PySide6.QtGui import QFont, QColor
import os
import threading
from typing import Any
from ui.resource_path import get_resource_base
import base64
import csv
from datetime import datetime

from instruments.power.keysight.n6705c import N6705C
from ui.widgets.button import SpinningSearchButton, update_connect_button_state
from ui.modules.n6705c_module_frame import N6705CConnectionMixin
from ui.modules.oscilloscope_module_frame import OscilloscopeConnectionMixin
from ui.widgets.dark_combobox import DarkComboBox
from ui.styles import SCROLLBAR_STYLE, START_BTN_STYLE, update_start_btn_state
from ui.modules.execution_logs_module_frame import ExecutionLogsFrame
from ui.theme import Colors, FontSizes, Radius, Spacing, FONT_MONO
from ui.styles import get_page_base_qss
from core.pmu_test.isgain import IsGainTestWorker
from core.ai.page_contract import (
    CAP_APPLY_CONFIG,
    CAP_GET_CONFIG,
    CAP_GET_RESULT,
    CAP_START_TEST,
    CAP_STOP_TEST,
)
from log_config import get_logger

_logger = get_logger(__name__)

# AI 回填可视化（AIAssist_PageScopedControlPlan.md §4.2 / Phase 3）：
# 被 AI 修改的控件临时高亮边框色 + 持续时长。
_AI_HIGHLIGHT_QSS = "border: 1px solid #15d1a3;"
_AI_HIGHLIGHT_MS = 1500


class CardFrame(QFrame):
    def __init__(self, title="", parent=None):
        super().__init__(parent)
        self.setObjectName("cardFrame")
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.main_layout = QVBoxLayout(self)
        self.main_layout.setContentsMargins(16, 12, 16, 16)
        self.main_layout.setSpacing(12)

        if title:
            self.title_row = QHBoxLayout()
            self.title_row.setSpacing(8)

            self.accent_bar = QFrame()
            self.accent_bar.setFixedSize(3, 14)
            self.accent_bar.setStyleSheet("background-color: #5b5cf6; border-radius: 1px; border: none;")
            self.title_row.addWidget(self.accent_bar)

            self.title_label = QLabel(title)
            self.title_label.setObjectName("cardTitle")
            self.title_row.addWidget(self.title_label)
            self.title_row.addStretch()
            self.main_layout.addLayout(self.title_row)

            self.title_separator = QFrame()
            self.title_separator.setFixedHeight(1)
            self.title_separator.setStyleSheet("background-color: #152240; border: none;")
            self.main_layout.addWidget(self.title_separator)
        else:
            self.title_label = None
            self.title_row = None


class FixedPopupComboBox(DarkComboBox):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setSizeAdjustPolicy(
            DarkComboBox.SizeAdjustPolicy.AdjustToMinimumContentsLengthWithIcon
        )
        self.setMinimumContentsLength(10)
        self.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Fixed)

    def showPopup(self):
        super().showPopup()
        view = self.view()
        if view and view.window():
            popup = view.window()
            global_pos = self.mapToGlobal(self.rect().bottomLeft())
            popup.move(global_pos.x(), global_pos.y())


class PMUIsGainUI(N6705CConnectionMixin, OscilloscopeConnectionMixin, QWidget):
    """PMU Is_gain测试UI组件"""

    connection_status_changed = Signal(bool)
    # 测试结束 → AI 异步动作回灌续跑（与 Orchestrator 同契约，§4 / S3-2）。
    # MainWindow._ai_on_sequence_finished_resume 监听本信号，回灌 pending 任务。
    sequence_execution_finished = Signal(bool, str)

    def __init__(self, n6705c_top=None, mso64b_top=None, instrument_manager=None):
        super().__init__()

        self._instrument_manager = instrument_manager
        self.init_n6705c_connection(n6705c_top, instrument_manager=instrument_manager)
        self.init_oscilloscope_connection(mso64b_top, instrument_manager=instrument_manager)

        self.is_test_running = False
        self.test_thread = None
        self._test_worker = None
        self._test_result_data = []

        self._setup_style()
        self._create_layout()
        self._init_ui_elements()
        self._bind_signals()
        self.sync_n6705c_from_top()
        self.sync_oscilloscope_from_top()

    @staticmethod
    def _get_checkmark_path(accent_color):
        safe_name = accent_color.replace("#", "").replace(" ", "")
        icons_dir = os.path.join(
            get_resource_base(),
            "resources", "icons"
        )
        return {
            "checked": os.path.join(icons_dir, f"checked_{safe_name}.svg").replace("\\", "/"),
            "unchecked": os.path.join(icons_dir, f"unchecked_{safe_name}.svg").replace("\\", "/"),
        }

    def _setup_style(self):
        font = QFont("Segoe UI", 9)
        self.setFont(font)

        _cb_icons = self._get_checkmark_path("4f46e5")
        page_extra = (f"""
            QFrame#cardFrame:hover {{
                border: 1px solid #1f3460;
            }}

            QFrame#resultContainer {{
                background-color: {Colors.bg_deep};
                border: 1px solid {Colors.border_secondary};
                border-radius: {Radius.card}px;
            }}

            QSpinBox::up-button, QSpinBox::down-button,
            QDoubleSpinBox::up-button, QDoubleSpinBox::down-button {{
                width: 0px; height: 0px; border: none;
            }}

            QPushButton#smallActionBtn {{
                min-height: 28px;
                padding: 4px 10px;
                border-radius: {Radius.widget}px;
                background-color: #13254b;
                color: {Colors.text_secondary};
            }}

            QPushButton#smallActionBtn:hover {{
                background-color: {Colors.submenu_item_hover_bg};
                border: 1px solid #3c5fa1;
            }}
""" + START_BTN_STYLE + f"""
            QPushButton#exportBtn {{
                min-height: 28px;
                padding: 4px 14px;
                border-radius: {Radius.widget}px;
                background-color: #16284f;
                color: {Colors.text_secondary};
                border: 1px solid #233d6a;
            }}

            QPushButton#exportBtn:hover {{
                background-color: #1d3462;
                border: 1px solid #3c5fa1;
            }}

            QPushButton#exportBtn:pressed {{
                background-color: #122244;
            }}

            QTextEdit#logEdit {{
                background-color: #050d22;
                border: 1px solid #1f315d;
                border-radius: {Radius.widget}px;
                color: #7cecc8;
                font-family: {FONT_MONO};
                font-size: {FontSizes.caption};
            }}

            QTableWidget {{
                background-color: #040c20;
                border: 1px solid #17294f;
                border-radius: {Radius.widget}px;
                gridline-color: #111f3e;
                color: {Colors.text_secondary};
                alternate-background-color: #060f28;
            }}

            QHeaderView::section {{
                background-color: #0a1735;
                color: {Colors.text_muted};
                border: none;
                border-bottom: 2px solid #1c3362;
                border-right: 1px solid #111f3e;
                padding: 8px 8px;
                font-weight: 700;
                font-size: {FontSizes.tiny};
                text-transform: uppercase;
            }}

            QHeaderView::section:last {{
                border-right: none;
            }}

            QTableWidget::item {{
                padding: 6px 8px;
                border-bottom: 1px solid #0e1d3c;
            }}

            QTableWidget::item:selected {{
                background-color: #162a56;
            }}

            QProgressBar {{
                background-color: #152749;
                border: none;
                border-radius: 4px;
                text-align: center;
                color: #b7c8ea;
                min-height: 8px;
                max-height: 8px;
            }}

            QProgressBar::chunk {{
                background-color: {Colors.accent_primary};
                border-radius: 4px;
            }}

            QCheckBox::indicator {{
                width: 16px;
                height: 16px;
                image: url("__UNCHECKED__");
            }}

            QCheckBox::indicator:checked {{
                image: url("__CHECKED__");
            }}
        """).replace("__UNCHECKED__", _cb_icons['unchecked']).replace("__CHECKED__", _cb_icons['checked'])
        self.setStyleSheet(get_page_base_qss() + page_extra)

    def _create_layout(self):
        root_layout = QVBoxLayout(self)
        root_layout.setContentsMargins(16, 14, 16, 14)
        root_layout.setSpacing(12)

        title_layout = QVBoxLayout()
        title_layout.setSpacing(2)

        self.page_title = QLabel("Load Capability and Output Ripple Test")
        self.page_title.setObjectName("pageTitle")

        self.page_subtitle = QLabel("Configure and execute load capability and output ripple test sequences.")
        self.page_subtitle.setObjectName("pageSubtitle")

        title_layout.addWidget(self.page_title)
        title_layout.addWidget(self.page_subtitle)
        root_layout.addLayout(title_layout)

        content_layout = QHBoxLayout()
        content_layout.setSpacing(12)
        root_layout.addLayout(content_layout, 1)

        left_wrapper = QVBoxLayout()
        left_wrapper.setContentsMargins(0, 0, 0, 0)
        left_wrapper.setSpacing(8)

        self.left_scroll = QScrollArea()
        self.left_scroll.setWidgetResizable(True)
        self.left_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.left_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.left_scroll.setFixedWidth(320)
        self.left_scroll.setObjectName("leftScrollArea")
        self.left_scroll.setStyleSheet("""
            QScrollArea#leftScrollArea {
                background-color: #08132d;
                border: 1px solid #142240;
                border-radius: 16px;
            }
        """ + SCROLLBAR_STYLE)

        self.left_panel = QWidget()
        self.left_panel.setObjectName("leftPanelInner")

        left_layout = QVBoxLayout(self.left_panel)
        left_layout.setContentsMargins(8, 8, 8, 8)
        left_layout.setSpacing(8)

        self.connection_card = CardFrame("Instruments")
        self._build_connection_card()
        left_layout.addWidget(self.connection_card)

        self.channel_card = CardFrame("Channels")
        self._build_channel_card()
        left_layout.addWidget(self.channel_card)

        self.is_gain_card = CardFrame("Test Config")
        self._build_is_gain_card()
        left_layout.addWidget(self.is_gain_card)

        left_layout.addStretch()

        self.left_scroll.setWidget(self.left_panel)
        left_wrapper.addWidget(self.left_scroll, 1)

        self.start_test_btn = QPushButton("▷  Start Sequence")
        self.start_test_btn.setObjectName("primaryStartBtn")
        left_wrapper.addWidget(self.start_test_btn)

        self.stop_test_btn = QPushButton("Abort Test")
        self.stop_test_btn.hide()

        content_layout.addLayout(left_wrapper)

        right_layout = QVBoxLayout()
        right_layout.setSpacing(12)
        content_layout.addLayout(right_layout, 1)

        self.result_frame = QFrame()
        self.result_frame.setObjectName("resultContainer")
        result_layout = QVBoxLayout(self.result_frame)
        result_layout.setContentsMargins(14, 12, 14, 12)
        result_layout.setSpacing(8)

        result_header = QHBoxLayout()
        result_header.setSpacing(8)

        result_accent = QFrame()
        result_accent.setFixedSize(3, 14)
        result_accent.setStyleSheet("background-color: #5b5cf6; border-radius: 1px; border: none;")
        result_header.addWidget(result_accent)

        self.result_title = QLabel("Test Results")
        self.result_title.setObjectName("sectionTitle")
        result_header.addWidget(self.result_title)

        self.result_count_label = QLabel("0 rows")
        self.result_count_label.setObjectName("fieldLabel")
        self.result_count_label.setStyleSheet("color: #5c7a9e; font-size: 10px; background: transparent; border: none;")
        result_header.addWidget(self.result_count_label)

        result_header.addStretch()

        self.export_result_btn = QPushButton("Export")
        self.export_result_btn.setObjectName("exportBtn")
        result_header.addWidget(self.export_result_btn)

        result_layout.addLayout(result_header)

        result_separator = QFrame()
        result_separator.setFixedHeight(1)
        result_separator.setStyleSheet("background-color: #152240; border: none;")
        result_layout.addWidget(result_separator)

        self.result_table = QTableWidget(0, 7)
        self.result_table.setHorizontalHeaderLabels([
            "STEP", "LOAD (A)", "VOLTAGE (V)", "RIPPLE (mV)",
            "V_DROP (mV)", "SCREENSHOT", "REMARK"
        ])
        header = self.result_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        header.setSectionResizeMode(0, QHeaderView.Fixed)
        header.setSectionResizeMode(5, QHeaderView.Fixed)
        self.result_table.setColumnWidth(0, 60)
        self.result_table.setColumnWidth(5, 80)
        header.setDefaultAlignment(Qt.AlignCenter)
        header.setMinimumSectionSize(50)
        self.result_table.verticalHeader().setVisible(False)
        self.result_table.verticalHeader().setDefaultSectionSize(32)
        self.result_table.setAlternatingRowColors(True)
        self.result_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.result_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.result_table.setShowGrid(False)
        result_layout.addWidget(self.result_table)

        right_splitter, self.execution_logs = ExecutionLogsFrame.wrap_with(
            self.result_frame, show_progress=True, stretch=(5, 1)
        )
        self.log_edit = self.execution_logs.log_edit
        self.progress_bar = self.execution_logs.progress_bar
        self.progress_text_label = self.execution_logs.progress_text_label
        self.clear_log_btn = self.execution_logs.clear_log_btn
        right_layout.addWidget(right_splitter, 1)

    def _build_connection_card(self):
        layout = self.connection_card.main_layout

        self.build_n6705c_connection_widgets(
            layout,
            title_row=self.connection_card.title_row,
        )

        self.build_oscilloscope_connection_widgets(layout)

        conn_separator = QFrame()
        conn_separator.setFixedHeight(1)
        conn_separator.setStyleSheet("background-color: #152240; border: none;")
        layout.addWidget(conn_separator)

        test_selection_label = QLabel("Test Selection")
        test_selection_label.setObjectName("fieldLabel")
        layout.addWidget(test_selection_label)

        self.test_selection_combo = FixedPopupComboBox()
        self.test_selection_combo.addItems(["单次 Is_gain 测试", "遍历 Is_gain 测试"])
        layout.addWidget(self.test_selection_combo)

    def _build_channel_card(self):
        layout = self.channel_card.main_layout
        grid = QGridLayout()
        grid.setHorizontalSpacing(8)
        grid.setVerticalSpacing(6)

        lbl_ripple = QLabel("Ripple Channel (Oscilloscope)")
        lbl_ripple.setObjectName("fieldLabel")
        self.ripple_channel_combo = FixedPopupComboBox()
        self.ripple_channel_combo.addItems(["CH 1", "CH 2", "CH 3", "CH 4"])

        lbl_load = QLabel("Load Channel (N6705C)")
        lbl_load.setObjectName("fieldLabel")
        self.load_channel_combo = FixedPopupComboBox()
        self.load_channel_combo.addItems(["CH 1", "CH 2", "CH 3", "CH 4"])
        self.load_channel_combo.setCurrentIndex(1)

        grid.addWidget(lbl_ripple, 0, 0)
        grid.addWidget(self.ripple_channel_combo, 0, 1)

        grid.addWidget(lbl_load, 1, 0)
        grid.addWidget(self.load_channel_combo, 1, 1)

        layout.addLayout(grid)

    def _build_is_gain_card(self):
        layout = self.is_gain_card.main_layout

        top_row = QHBoxLayout()
        top_row.setSpacing(8)

        mode_label = QLabel("Method")
        mode_label.setObjectName("fieldLabel")
        self.is_gain_method_combo = FixedPopupComboBox()
        self.is_gain_method_combo.addItems(["CV测试法", "CC测试法"])

        top_row.addWidget(mode_label)
        top_row.addWidget(self.is_gain_method_combo, 1)
        layout.addLayout(top_row)

        grid = QGridLayout()
        grid.setHorizontalSpacing(8)
        grid.setVerticalSpacing(6)

        self.lbl_dev = QLabel("Device Addr")
        self.lbl_dev.setObjectName("fieldLabel")
        self.is_gain_device_addr_edit = QLineEdit("0x60")

        self.lbl_reg = QLabel("Reg Addr")
        self.lbl_reg.setObjectName("fieldLabel")
        self.is_gain_reg_addr_edit = QLineEdit("0x01")

        self.lbl_msb = QLabel("MSB")
        self.lbl_msb.setObjectName("fieldLabel")
        self.is_gain_msb_spin = QSpinBox()
        self.is_gain_msb_spin.setRange(0, 255)
        self.is_gain_msb_spin.setValue(7)

        self.lbl_lsb = QLabel("LSB")
        self.lbl_lsb.setObjectName("fieldLabel")
        self.is_gain_lsb_spin = QSpinBox()
        self.is_gain_lsb_spin.setRange(0, 255)
        self.is_gain_lsb_spin.setValue(4)

        grid.addWidget(self.lbl_dev, 0, 0)
        grid.addWidget(self.lbl_reg, 0, 1)
        grid.addWidget(self.is_gain_device_addr_edit, 1, 0)
        grid.addWidget(self.is_gain_reg_addr_edit, 1, 1)

        grid.addWidget(self.lbl_msb, 2, 0)
        grid.addWidget(self.lbl_lsb, 2, 1)
        grid.addWidget(self.is_gain_msb_spin, 3, 0)
        grid.addWidget(self.is_gain_lsb_spin, 3, 1)

        self._traverse_only_widgets = [
            self.lbl_dev, self.lbl_reg,
            self.is_gain_device_addr_edit, self.is_gain_reg_addr_edit,
            self.lbl_msb, self.lbl_lsb,
            self.is_gain_msb_spin, self.is_gain_lsb_spin,
        ]

        current_separator = QFrame()
        current_separator.setFixedHeight(1)
        current_separator.setStyleSheet("background-color: #152240; border: none;")
        grid.addWidget(current_separator, 4, 0, 1, 2)

        lbl_start_current = QLabel("Start Current (A)")
        lbl_start_current.setObjectName("fieldLabel")
        self.is_gain_start_current_spin = QDoubleSpinBox()
        self.is_gain_start_current_spin.setDecimals(3)
        self.is_gain_start_current_spin.setRange(-9999.0, 9999.0)
        self.is_gain_start_current_spin.setSingleStep(0.001)
        self.is_gain_start_current_spin.setValue(0.000)

        lbl_end_current = QLabel("End Current (A)")
        lbl_end_current.setObjectName("fieldLabel")
        self.is_gain_end_current_spin = QDoubleSpinBox()
        self.is_gain_end_current_spin.setDecimals(3)
        self.is_gain_end_current_spin.setRange(-9999.0, 9999.0)
        self.is_gain_end_current_spin.setSingleStep(0.001)
        self.is_gain_end_current_spin.setValue(0.250)

        lbl_step_current = QLabel("Step Current (A)")
        lbl_step_current.setObjectName("fieldLabel")
        self.is_gain_step_current_spin = QDoubleSpinBox()
        self.is_gain_step_current_spin.setDecimals(3)
        self.is_gain_step_current_spin.setRange(-9999.0, 9999.0)
        self.is_gain_step_current_spin.setSingleStep(0.001)
        self.is_gain_step_current_spin.setValue(0.01)

        grid.addWidget(lbl_start_current, 5, 0)
        grid.addWidget(lbl_end_current, 5, 1)
        grid.addWidget(self.is_gain_start_current_spin, 6, 0)
        grid.addWidget(self.is_gain_end_current_spin, 6, 1)

        grid.addWidget(lbl_step_current, 7, 0)
        grid.addWidget(self.is_gain_step_current_spin, 8, 0)

        self.save_screenshot_cb = QCheckBox("Save Screenshot")
        self.save_screenshot_cb.setChecked(True)
        grid.addWidget(self.save_screenshot_cb, 8, 1)

        layout.addLayout(grid)

    def _init_ui_elements(self):
        self._update_n6705c_connect_button_state(False)
        update_connect_button_state(self.scope_connect_btn, False)
        self._update_test_button_state(False)
        self._on_test_selection_changed()
        self.append_log("[SYSTEM] Ready. Waiting for instrument connection.")
        self.set_progress(0)
        self.stop_test_btn.setEnabled(False)

    def _bind_signals(self):
        self.bind_n6705c_signals()
        self.bind_oscilloscope_signals()
        self.test_selection_combo.currentIndexChanged.connect(self._on_test_selection_changed)
        self.start_test_btn.clicked.connect(self._on_start_or_abort_clicked)
        self.stop_test_btn.clicked.connect(self._abort_test_from_external)
        self.export_result_btn.clicked.connect(self._on_export)
        self.clear_log_btn.clicked.connect(self._on_clear_log)

    def _abort_test_from_external(self):
        if self.is_test_running:
            if self._test_worker is not None:
                self._test_worker.stop()
            self._test_stop_requested = True
            self.append_log("[TEST] Abort requested by external stop button.")

    def _on_test_selection_changed(self):
        is_traverse = self.test_selection_combo.currentText() == "遍历 Is_gain 测试"
        for w in self._traverse_only_widgets:
            w.setVisible(is_traverse)

    def _update_test_button_state(self, running: bool):
        self.is_test_running = running
        update_start_btn_state(self.start_test_btn, running,
                               start_text="▷  Start Sequence",
                               stop_text="□  Abort Test")

    def append_log(self, message):
        self.execution_logs.append_log(message)

    def _on_clear_log(self):
        self.execution_logs.clear_log()

    def set_progress(self, value: int):
        self.execution_logs.set_progress(value)

    def get_test_config(self):
        return {
            "ripple_channel": self.ripple_channel_combo.currentText(),
            "load_channel": self.load_channel_combo.currentText(),
            "is_gain_method": self.is_gain_method_combo.currentText(),
            "is_gain_device_addr": self.is_gain_device_addr_edit.text().strip(),
            "is_gain_reg_addr": self.is_gain_reg_addr_edit.text().strip(),
            "is_gain_msb": self.is_gain_msb_spin.value(),
            "is_gain_lsb": self.is_gain_lsb_spin.value(),
            "is_gain_start_current": self.is_gain_start_current_spin.value(),
            "is_gain_end_current": self.is_gain_end_current_spin.value(),
            "is_gain_step_current": self.is_gain_step_current_spin.value(),
            "save_screenshot": self.save_screenshot_cb.isChecked(),
        }

    def set_test_running(self, running):
        self._update_test_button_state(running)
        self.stop_test_btn.setEnabled(running)

        widgets = [
            self.ripple_channel_combo,
            self.load_channel_combo,
            self.is_gain_method_combo,
            self.is_gain_device_addr_edit,
            self.is_gain_reg_addr_edit,
            self.is_gain_msb_spin,
            self.is_gain_lsb_spin,
            self.is_gain_start_current_spin,
            self.is_gain_end_current_spin,
            self.is_gain_step_current_spin,
            self.save_screenshot_cb,

            self.visa_resource_combo,
            self.scope_type_combo,
            self.scope_resource_combo,
            self.search_btn,
            self.scope_search_btn,
            self.connect_btn,
            self.scope_connect_btn,
            self.test_selection_combo
        ]

        for widget in widgets:
            widget.setEnabled(not running)

        if running:
            self.append_log("[TEST] Starting Is_gain Test Sequence...")
        else:
            self.append_log("[TEST] Test stopped or completed.")

    def clear_results(self):
        self.result_table.setRowCount(0)
        self._test_result_data = []
        self.set_progress(0)
        self.result_count_label.setText("0 rows")
        self.append_log("[SYSTEM] Results cleared.")

    def add_result_row(self, step, load_current, voltage, ripple, v_drop=None, has_screenshot=False, remark=""):
        row = self.result_table.rowCount()
        self.result_table.insertRow(row)

        voltage_str = f"{voltage:.6f}" if voltage is not None else "N/A"
        ripple_mv = ripple * 1000 if ripple is not None else None
        ripple_str = f"{ripple_mv:.3f}" if ripple_mv is not None else "N/A"
        v_drop_mv = v_drop * 1000 if v_drop is not None else None
        v_drop_str = f"{v_drop_mv:.3f}" if v_drop_mv is not None else "N/A"
        screenshot_str = "✓" if has_screenshot else "—"

        items = [
            QTableWidgetItem(str(step)),
            QTableWidgetItem(f"{load_current:.3f}"),
            QTableWidgetItem(voltage_str),
            QTableWidgetItem(ripple_str),
            QTableWidgetItem(v_drop_str),
            QTableWidgetItem(screenshot_str),
            QTableWidgetItem(remark),
        ]

        colors = ["#8ea8d4", "#6b9fff", "#00d6a2", "#f0a030", "#ff6b6b", "#4cc9f0", "#9b8ec8"]
        for col, item in enumerate(items):
            item.setTextAlignment(Qt.AlignCenter)
            item.setForeground(QColor(colors[col]))
            self.result_table.setItem(row, col, item)

        total_rows = self.result_table.rowCount()
        self.result_count_label.setText(f"{total_rows} row{'s' if total_rows != 1 else ''}")

    def _on_test_result_row(self, row_data):
        self._test_result_data.append(row_data)

        step = row_data.get("step", "")
        load_current = row_data.get("load_current", 0)
        voltage = row_data.get("voltage")
        ripple = row_data.get("ripple")
        v_drop = row_data.get("v_drop")
        reg_value = row_data.get("reg_value")

        voltage_str = f"{voltage:.6f}" if voltage is not None else "N/A"
        ripple_mv = ripple * 1000 if ripple is not None else None
        ripple_str = f"{ripple_mv:.3f}" if ripple_mv is not None else "N/A"
        v_drop_mv = v_drop * 1000 if v_drop is not None else None
        v_drop_str = f"{v_drop_mv:.3f}" if v_drop_mv is not None else "N/A"

        if reg_value is not None:
            self.append_log(
                f"[DATA] Step={step}\tReg={reg_value}\tLoad={load_current:.3f}A\t"
                f"V={voltage_str}V\tRipple={ripple_str}mV\tVdrop={v_drop_str}mV"
            )
        else:
            self.append_log(
                f"[DATA] Step={step}\tLoad={load_current:.3f}A\t"
                f"V={voltage_str}V\tRipple={ripple_str}mV\tVdrop={v_drop_str}mV"
            )

        self.add_result_row(
            step,
            load_current,
            voltage,
            ripple,
            v_drop,
            row_data.get("screenshot_b64") is not None,
            row_data.get("remark", ""),
        )

    def _on_test_summary(self, analysis):
        v0 = analysis.get("v0")
        max_load = analysis.get("max_load_current")
        max_rp = analysis.get("max_ripple")
        max_rp_cur = analysis.get("max_ripple_current")

        self.append_log("=" * 50)
        self.append_log("[SUMMARY] Test Analysis Results:")
        if v0 is not None:
            self.append_log(f"  0-load baseline voltage: {v0:.6f} V")
        if max_load is not None:
            self.append_log(f"  Max load capacity (Vdrop <= 30mV): {max_load:.3f} A")
        else:
            self.append_log("  Max load capacity: N/A (voltage drop > 30mV at all loads)")
        if max_rp is not None:
            self.append_log(f"  Max ripple: {max_rp * 1000:.3f} mV @ {max_rp_cur:.3f} A")
        self.append_log("=" * 50)

    def _on_export(self):
        if not self._test_result_data:
            self.set_page_status("No data to export", is_error=True)
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Export Test Results")
        dialog.setFixedWidth(460)
        dialog.setStyleSheet("""
            QDialog {
                background-color: #0a1628;
                color: #c8daf5;
            }
            QLabel {
                color: #8eb0e3;
                font-size: 12px;
                background: transparent;
                border: none;
            }
            QLineEdit {
                background-color: #0c1a35;
                border: 1px solid #1a2d52;
                border-radius: 6px;
                color: #eaf2ff;
                padding: 6px 10px;
                font-size: 12px;
            }
            QLineEdit:focus {
                border-color: #4f6db8;
            }
            QRadioButton {
                color: #dbe7ff;
                font-size: 12px;
                spacing: 6px;
                background: transparent;
            }
            QRadioButton::indicator {
                width: 14px;
                height: 14px;
                border-radius: 7px;
                border: 2px solid #2a4272;
                background-color: #0c1a35;
            }
            QRadioButton::indicator:checked {
                background-color: #5b5cf6;
                border-color: #7b7cff;
            }
            QRadioButton::indicator:hover {
                border-color: #5b5cf6;
            }
            QPushButton {
                background-color: #162d55;
                border: 1px solid #1e3460;
                border-radius: 6px;
                color: #c8daf5;
                padding: 6px 18px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #1e3460;
                border-color: #3a6fd4;
            }
            QPushButton:pressed {
                background-color: #112445;
            }
            QFrame#exportSeparator {
                background-color: #152240;
                max-height: 1px;
                border: none;
            }
        """)

        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(16)

        title = QLabel("Export Test Results")
        title.setStyleSheet("font-size: 14px; font-weight: bold; color: #f4f7ff;")
        layout.addWidget(title)

        sep1 = QFrame()
        sep1.setObjectName("exportSeparator")
        sep1.setFixedHeight(1)
        layout.addWidget(sep1)

        info_grid = QGridLayout()
        info_grid.setHorizontalSpacing(12)
        info_grid.setVerticalSpacing(10)

        lbl_dut = QLabel("DUT Name")
        self._export_dut_edit = QLineEdit()
        self._export_dut_edit.setPlaceholderText("e.g. PMU_A1")
        info_grid.addWidget(lbl_dut, 0, 0)
        info_grid.addWidget(self._export_dut_edit, 0, 1)

        lbl_module = QLabel("Module")
        self._export_module_edit = QLineEdit()
        self._export_module_edit.setPlaceholderText("e.g. LDO1")
        info_grid.addWidget(lbl_module, 1, 0)
        info_grid.addWidget(self._export_module_edit, 1, 1)

        lbl_cond = QLabel("Condition")
        self._export_condition_edit = QLineEdit()
        self._export_condition_edit.setPlaceholderText("e.g. Room Temp 25°C")
        info_grid.addWidget(lbl_cond, 2, 0)
        info_grid.addWidget(self._export_condition_edit, 2, 1)

        lbl_operator = QLabel("Operator")
        self._export_operator_edit = QLineEdit()
        self._export_operator_edit.setPlaceholderText("e.g. Zhang San")
        info_grid.addWidget(lbl_operator, 3, 0)
        info_grid.addWidget(self._export_operator_edit, 3, 1)

        lbl_remark = QLabel("Remark")
        self._export_remark_edit = QLineEdit()
        self._export_remark_edit.setPlaceholderText("(optional)")
        info_grid.addWidget(lbl_remark, 4, 0)
        info_grid.addWidget(self._export_remark_edit, 4, 1)

        layout.addLayout(info_grid)

        sep2 = QFrame()
        sep2.setObjectName("exportSeparator")
        sep2.setFixedHeight(1)
        layout.addWidget(sep2)

        fmt_label = QLabel("Export Format")
        fmt_label.setStyleSheet("font-size: 12px; font-weight: 600; color: #c0d0f0;")
        layout.addWidget(fmt_label)

        fmt_group = QButtonGroup(dialog)
        fmt_layout = QHBoxLayout()
        fmt_layout.setSpacing(20)

        self._rb_xlsx = QRadioButton("Excel (.xlsx)")
        self._rb_pdf = QRadioButton("PDF (.pdf)")
        self._rb_csv = QRadioButton("CSV (.csv)")
        self._rb_xlsx.setChecked(True)

        fmt_group.addButton(self._rb_xlsx)
        fmt_group.addButton(self._rb_pdf)
        fmt_group.addButton(self._rb_csv)

        fmt_layout.addWidget(self._rb_xlsx)
        fmt_layout.addWidget(self._rb_pdf)
        fmt_layout.addWidget(self._rb_csv)
        fmt_layout.addStretch()
        layout.addLayout(fmt_layout)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(dialog.reject)
        export_btn = QPushButton("Export")
        export_btn.setStyleSheet("""
            QPushButton {
                background-color: qlineargradient(
                    x1:0, y1:0, x2:1, y2:0,
                    stop:0 #5b5cf6,
                    stop:1 #6a38ff
                );
                border: 1px solid #5b5cf6;
                border-radius: 6px;
                color: #ffffff;
                padding: 7px 24px;
                font-size: 12px;
                font-weight: 600;
            }
            QPushButton:hover {
                background-color: qlineargradient(
                    x1:0, y1:0, x2:1, y2:0,
                    stop:0 #6b6cff,
                    stop:1 #7d4cff
                );
                border-color: #7b7cff;
            }
            QPushButton:pressed {
                background-color: #4a3dd4;
            }
        """)
        export_btn.clicked.connect(dialog.accept)
        btn_layout.addWidget(cancel_btn)
        btn_layout.addWidget(export_btn)
        layout.addLayout(btn_layout)

        if dialog.exec() != QDialog.Accepted:
            return

        export_info = {
            "dut_name": self._export_dut_edit.text().strip(),
            "module": self._export_module_edit.text().strip(),
            "condition": self._export_condition_edit.text().strip(),
            "operator": self._export_operator_edit.text().strip(),
            "remark": self._export_remark_edit.text().strip(),
            "test_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

        dut = export_info["dut_name"] or "DUT"
        module = export_info["module"] or "Module"
        base_name = f"{dut}_{module}_is_gainAndRipple_Test"

        if self._rb_xlsx.isChecked():
            fmt_filter = "Excel Files (*.xlsx)"
            default_name = f"{base_name}.xlsx"
        elif self._rb_pdf.isChecked():
            fmt_filter = "PDF Files (*.pdf)"
            default_name = f"{base_name}.pdf"
        else:
            fmt_filter = "CSV Files (*.csv)"
            default_name = f"{base_name}.csv"

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export Results", default_name, fmt_filter
        )
        if not file_path:
            return

        try:
            if file_path.endswith(".xlsx"):
                self._export_xlsx(file_path, export_info)
            elif file_path.endswith(".pdf"):
                self._export_pdf(file_path, export_info)
            else:
                self._export_csv(file_path, export_info)
            self.append_log(f"[EXPORT] Results saved to: {file_path}")
            self.set_page_status(f"Exported to {os.path.basename(file_path)}")
        except Exception as e:
            self.append_log(f"[ERROR] Export failed: {e}")
            self.set_page_status("Export failed", is_error=True)

    def _export_csv(self, file_path, export_info=None):
        with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            if export_info:
                writer.writerow(["DUT Name", export_info.get("dut_name", "")])
                writer.writerow(["Module", export_info.get("module", "")])
                writer.writerow(["Condition", export_info.get("condition", "")])
                writer.writerow(["Operator", export_info.get("operator", "")])
                writer.writerow(["Remark", export_info.get("remark", "")])
                writer.writerow(["Test Time", export_info.get("test_time", "")])
                writer.writerow([])
            writer.writerow(["Step", "Load Current (A)", "Voltage (V)", "Ripple (mV)", "V_Drop (mV)", "Remark"])
            for r in self._test_result_data:
                rp = r.get("ripple")
                vd = r.get("v_drop")
                writer.writerow([
                    r.get("step", ""),
                    f"{r.get('load_current', 0):.3f}",
                    f"{r['voltage']:.6f}" if r.get("voltage") is not None else "N/A",
                    f"{rp * 1000:.3f}" if rp is not None else "N/A",
                    f"{vd * 1000:.3f}" if vd is not None else "N/A",
                    r.get("remark", ""),
                ])

    def _export_xlsx(self, file_path, export_info=None):
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XlImage
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font as XlFont, PatternFill, Alignment, Border, Side
        import io

        xlsx_dir = os.path.dirname(os.path.abspath(file_path))
        xlsx_name = os.path.splitext(os.path.basename(file_path))[0]
        screenshots_dir = os.path.join(xlsx_dir, f"{xlsx_name}_screenshots")

        has_screenshots = any(r.get("screenshot_b64") for r in self._test_result_data)
        if has_screenshots:
            os.makedirs(screenshots_dir, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Is_gain Results"

        header_font = XlFont(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F4070", end_color="2F4070", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        info_key_font = XlFont(bold=True, color="2F4070", size=11)
        info_val_font = XlFont(color="333333", size=11)
        thin_border = Border(
            left=Side(style="thin", color="C0C0C0"),
            right=Side(style="thin", color="C0C0C0"),
            top=Side(style="thin", color="C0C0C0"),
            bottom=Side(style="thin", color="C0C0C0"),
        )

        current_row = 1

        if export_info:
            title_cell = ws.cell(row=current_row, column=1, value="Is_gain Test Report")
            title_cell.font = XlFont(bold=True, color="1a1a2e", size=16)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
            current_row += 2

            info_items = [
                ("DUT Name", export_info.get("dut_name", "")),
                ("Module", export_info.get("module", "")),
                ("Condition", export_info.get("condition", "")),
                ("Operator", export_info.get("operator", "")),
                ("Test Time", export_info.get("test_time", "")),
            ]
            if export_info.get("remark"):
                info_items.append(("Remark", export_info["remark"]))

            for key, val in info_items:
                ws.cell(row=current_row, column=1, value=key).font = info_key_font
                ws.cell(row=current_row, column=2, value=val).font = info_val_font
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
                current_row += 1

            current_row += 1

        headers = ["Step", "Load Current (A)", "Voltage (V)", "Ripple (mV)", "V_Drop (mV)", "Screenshot", "Remark"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        img_col_width = 45
        ws.column_dimensions[get_column_letter(6)].width = img_col_width
        img_row_height = 180
        data_start_row = current_row + 1

        saved_screenshots = []

        for idx, r in enumerate(self._test_result_data):
            row_idx = data_start_row + idx

            ws.cell(row=row_idx, column=1, value=r.get("step", "")).border = thin_border
            ws.cell(row=row_idx, column=2, value=round(r.get("load_current", 0), 3)).border = thin_border
            v = r.get("voltage")
            ws.cell(row=row_idx, column=3, value=round(v, 6) if v is not None else "N/A").border = thin_border
            rp = r.get("ripple")
            ws.cell(row=row_idx, column=4, value=round(rp * 1000, 3) if rp is not None else "N/A").border = thin_border
            vd = r.get("v_drop")
            ws.cell(row=row_idx, column=5, value=round(vd * 1000, 3) if vd is not None else "N/A").border = thin_border
            ws.cell(row=row_idx, column=7, value=r.get("remark", "")).border = thin_border

            for c in range(1, 8):
                ws.cell(row=row_idx, column=c).alignment = Alignment(horizontal="center", vertical="center")

            sc_b64 = r.get("screenshot_b64")
            if sc_b64:
                try:
                    img_data = base64.b64decode(sc_b64)
                    png_sig = b'\x89PNG\r\n\x1a\n'
                    bmp_sig = b'BM'
                    if img_data[:8] != png_sig and img_data[:2] != bmp_sig:
                        ws.cell(row=row_idx, column=6, value="(invalid image data)")
                        continue

                    step_val = r.get("step", idx)
                    reg_val = r.get("reg_value")
                    if reg_val is not None:
                        png_name = f"step{step_val}_reg{reg_val}.png"
                    else:
                        png_name = f"step{step_val}.png"
                    png_path = os.path.join(screenshots_dir, png_name)
                    with open(png_path, "wb") as f:
                        f.write(img_data)
                    saved_screenshots.append(png_path)

                    cell = ws.cell(row=row_idx, column=6, value=png_name)
                    cell.hyperlink = png_path
                    cell.font = XlFont(color="4472C4", underline="single")
                    cell.border = thin_border

                    img_stream = io.BytesIO(img_data)
                    img = XlImage(img_stream)

                    col_px = img_col_width * 7
                    row_px = img_row_height * 1.33
                    scale = min(col_px / img.width, row_px / img.height)
                    img.width = int(img.width * scale)
                    img.height = int(img.height * scale)

                    ws.add_image(img, f"F{row_idx}")
                    ws.row_dimensions[row_idx].height = img_row_height
                except Exception as e:
                    ws.cell(row=row_idx, column=6, value=f"(image error: {e})")
            else:
                cell = ws.cell(row=row_idx, column=6, value="")
                cell.border = thin_border

        wb.save(file_path)

        if saved_screenshots:
            self.append_log(f"[EXPORT] {len(saved_screenshots)} screenshots saved to: {screenshots_dir}")

    def _export_pdf(self, file_path, export_info=None):
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import io

        font_name = "Helvetica"
        font_name_bold = "Helvetica-Bold"
        try:
            import platform
            if platform.system() == "Windows":
                win_font = os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Fonts", "msyh.ttc")
                if os.path.exists(win_font):
                    pdfmetrics.registerFont(TTFont("MSYH", win_font, subfontIndex=0))
                    pdfmetrics.registerFont(TTFont("MSYH-Bold", win_font, subfontIndex=1))
                    font_name = "MSYH"
                    font_name_bold = "MSYH-Bold"
        except Exception:
            pass

        doc = SimpleDocTemplate(
            file_path,
            pagesize=landscape(A4),
            leftMargin=15 * mm, rightMargin=15 * mm,
            topMargin=15 * mm, bottomMargin=15 * mm,
        )

        elements = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            "ReportTitle", parent=styles["Title"],
            fontName=font_name_bold, fontSize=18, textColor=colors.HexColor("#1a1a2e"),
            spaceAfter=6 * mm,
        )
        elements.append(Paragraph("Is_gain Test Report", title_style))

        if export_info:
            info_style = ParagraphStyle(
                "InfoStyle", parent=styles["Normal"],
                fontName=font_name, fontSize=10, textColor=colors.HexColor("#333333"),
                leading=16,
            )
            info_bold = ParagraphStyle(
                "InfoBold", parent=info_style,
                fontName=font_name_bold,
            )

            info_data = [
                [Paragraph("DUT Name", info_bold), Paragraph(export_info.get("dut_name", ""), info_style),
                 Paragraph("Module", info_bold), Paragraph(export_info.get("module", ""), info_style)],
                [Paragraph("Condition", info_bold), Paragraph(export_info.get("condition", ""), info_style),
                 Paragraph("Operator", info_bold), Paragraph(export_info.get("operator", ""), info_style)],
                [Paragraph("Test Time", info_bold), Paragraph(export_info.get("test_time", ""), info_style),
                 Paragraph("", info_style), Paragraph("", info_style)],
            ]
            if export_info.get("remark"):
                info_data.append([
                    Paragraph("Remark", info_bold), Paragraph(export_info["remark"], info_style),
                    Paragraph("", info_style), Paragraph("", info_style),
                ])

            info_table = Table(info_data, colWidths=[25 * mm, 60 * mm, 25 * mm, 60 * mm])
            info_table.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ("LINEBELOW", (0, -1), (-1, -1), 0.5, colors.HexColor("#cccccc")),
            ]))
            elements.append(info_table)
            elements.append(Spacer(1, 6 * mm))

        cell_style = ParagraphStyle(
            "CellStyle", parent=styles["Normal"],
            fontName=font_name, fontSize=8, textColor=colors.HexColor("#222222"),
            alignment=1, leading=11,
        )
        header_para_style = ParagraphStyle(
            "HeaderPara", parent=cell_style,
            fontName=font_name_bold, fontSize=9, textColor=colors.white,
        )

        headers = ["Step", "Load Current (A)", "Voltage (V)", "Ripple (mV)", "V_Drop (mV)", "Remark"]
        header_row = [Paragraph(h, header_para_style) for h in headers]
        table_data = [header_row]

        for r in self._test_result_data:
            rp = r.get("ripple")
            vd = r.get("v_drop")
            row = [
                Paragraph(str(r.get("step", "")), cell_style),
                Paragraph(f"{r.get('load_current', 0):.3f}", cell_style),
                Paragraph(f"{r['voltage']:.6f}" if r.get("voltage") is not None else "N/A", cell_style),
                Paragraph(f"{rp * 1000:.3f}" if rp is not None else "N/A", cell_style),
                Paragraph(f"{vd * 1000:.3f}" if vd is not None else "N/A", cell_style),
                Paragraph(r.get("remark", ""), cell_style),
            ]
            table_data.append(row)

        col_widths = [18 * mm, 35 * mm, 35 * mm, 30 * mm, 30 * mm, 40 * mm]
        table = Table(table_data, colWidths=col_widths, repeatRows=1)

        tbl_style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F4070")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#C0C0C0")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FA")]),
        ])
        table.setStyle(tbl_style)
        elements.append(table)

        doc.build(elements)

    def _on_test_error(self, err_msg):
        self.append_log(f"[ERROR] {err_msg}")
        self.set_page_status(f"Test error: {err_msg}", is_error=True)

    def _on_test_finished(self):
        self.set_test_running(False)
        self.set_page_status("Test completed")
        # 通知 AI 异步动作层：测试结束，触发 pending 任务回灌续跑（§4 / S3-2）。
        stopped = bool(getattr(self, "_test_stop_requested", False))
        summary = "测试被中止" if stopped else "测试完成"
        self.sequence_execution_finished.emit(not stopped, summary)

    def _cleanup_test_thread(self):
        if self.test_thread is not None:
            self.test_thread.wait(5000)
            self.test_thread.deleteLater()
            self.test_thread = None
        if self._test_worker is not None:
            self._test_worker.deleteLater()
            self._test_worker = None

    def set_page_status(self, status, is_error=False):
        self.page_subtitle.setText(status)
        if is_error:
            self.page_subtitle.setObjectName("statusErr")
        elif "Connecting" in status or "Searching" in status or "Running" in status:
            self.page_subtitle.setObjectName("statusWarn")
        else:
            self.page_subtitle.setObjectName("pageSubtitle")

        self.page_subtitle.style().unpolish(self.page_subtitle)
        self.page_subtitle.style().polish(self.page_subtitle)
        self.page_subtitle.update()

    def update_instrument_info(self, instrument_info):
        if self.is_connected:
            self.set_system_status("● Connected")

    def _on_start_or_abort_clicked(self):
        if self.is_test_running:
            if self._test_worker is not None:
                self._test_worker.stop()
            self._test_stop_requested = True
            self.append_log("[TEST] Abort requested by user.")
            return

        if not self.is_connected or self.n6705c is None:
            self.set_page_status("Please connect N6705C first", is_error=True)
            self.append_log("[ERROR] N6705C not connected.")
            return

        if not self.scope_connected or self.Osc_ins is None:
            self.set_page_status("Please connect Oscilloscope first", is_error=True)
            self.append_log("[ERROR] Oscilloscope not connected.")
            return

        config = self.get_test_config()

        if abs(config["is_gain_step_current"]) < 1e-9:
            self.set_page_status("Step Current must be > 0", is_error=True)
            self.append_log("[ERROR] Step Current must be greater than 0.")
            return

        selected_test = self.test_selection_combo.currentText()

        if selected_test == "单次 Is_gain 测试":
            self._start_is_gain_test(config)
        elif selected_test == "遍历 Is_gain 测试":
            self._start_traverse_is_gain_test(config)

    def _start_is_gain_test(self, config):
        self._launch_test_thread(config, IsGainTestWorker.MODE_SINGLE, "Running single Is_gain test...")

    def _start_traverse_is_gain_test(self, config):
        self._launch_test_thread(config, IsGainTestWorker.MODE_TRAVERSE, "Running traverse Is_gain test...")

    def _launch_test_thread(self, config, test_mode, status_msg):
        self.clear_results()
        self._test_stop_requested = False
        self.set_test_running(True)
        self.set_page_status(status_msg)

        self._test_worker = IsGainTestWorker(self.n6705c, self.Osc_ins, config, test_mode=test_mode)
        self.test_thread = QThread()
        self._test_worker.moveToThread(self.test_thread)

        self.test_thread.started.connect(self._test_worker.run)
        self._test_worker.log.connect(self.append_log)
        self._test_worker.progress.connect(self.set_progress)
        self._test_worker.result_row.connect(self._on_test_result_row)
        self._test_worker.summary.connect(self._on_test_summary)
        self._test_worker.error.connect(self._on_test_error)
        self._test_worker.finished.connect(self._on_test_finished)
        self._test_worker.finished.connect(self.test_thread.quit)
        self.test_thread.finished.connect(self._cleanup_test_thread)

        self.test_thread.start()

    # ------------------------------------------------------------------
    # AIControllablePage 契约实现（AIAssist_PageScopedControlPlan.md §2 / Phase 5）
    #
    # PMU Is_gain 测试接入 AI 受控契约，薄封装既有方法：
    #   - ai_get_config 复用 get_test_config()
    #   - ai_apply_config 经 apply_config_to_controls() 单一写入口回填控件
    #   - ai_start_test/ai_stop_test 复用 _on_start_or_abort_clicked/_abort_test_from_external
    # 枢纽（MainWindow.resolve_active_ai_page）经 Tab 子页下钻拿到本实例，
    # 鸭子调用契约方法，无需 core / handler 改动。
    # ------------------------------------------------------------------
    def ai_capabilities(self) -> set[str]:
        return {
            CAP_GET_CONFIG,
            CAP_APPLY_CONFIG,
            CAP_START_TEST,
            CAP_STOP_TEST,
            CAP_GET_RESULT,
        }

    def ai_get_config(self) -> dict[str, Any] | None:
        try:
            return self.get_test_config()
        except Exception:  # noqa: BLE001 - 快照失败降级为 None
            _logger.error("AI 读取 Is_gain 测试配置失败", exc_info=True)
            return None

    def ai_apply_config(self, payload: Any) -> tuple[bool, str]:
        """落地配置草案到控件（写操作，经确认+审计后由枢纽调用）。

        运行中拒绝改配置（§6.3），避免与正在执行的扫描冲突。
        """
        if self.is_test_running:
            return False, "测试运行中，无法修改配置，请先停止测试。"
        return self.apply_config_to_controls(payload if isinstance(payload, dict) else {})

    def ai_start_test(self) -> tuple[bool, str]:
        if not self.is_connected or self.n6705c is None:
            return False, "未连接 N6705C 仪器，请先连接再启动测试。"
        if not self.scope_connected or self.Osc_ins is None:
            return False, "未连接示波器，请先连接示波器再启动 Is_gain 测试。"
        if self.is_test_running:
            return False, "测试已在运行中。"
        cfg = self.get_test_config()
        if abs(cfg.get("is_gain_step_current", 0)) < 1e-9:
            return False, "Step Current 必须大于 0。"
        selected = self.test_selection_combo.currentText()
        self.append_log(
            f"[AI] 请求启动 {selected}："
            f"电流 {cfg.get('is_gain_start_current')}~{cfg.get('is_gain_end_current')}A。"
        )
        try:
            self._on_start_or_abort_clicked()
        except Exception:  # noqa: BLE001 - 启动异常转可读结果
            _logger.error("AI 启动 Is_gain 测试失败", exc_info=True)
            return False, "启动测试异常，请查看日志。"
        if self.is_test_running:
            return True, f"已请求启动 {selected}。"
        return False, "启动未成功，请查看执行日志。"

    def ai_stop_test(self) -> tuple[bool, str]:
        if not self.is_test_running:
            return False, "当前未在运行测试。"
        self.append_log("[AI] 请求停止测试。")
        try:
            self._abort_test_from_external()
        except Exception:  # noqa: BLE001 - 停止异常转可读结果
            _logger.error("AI 停止 Is_gain 测试失败", exc_info=True)
            return False, "停止测试异常，请查看日志。"
        return True, "已发送停止请求。"

    def ai_get_result_summary(self) -> dict[str, Any] | None:
        summary: dict[str, Any] = {
            "available": True,
            "running": self.is_test_running,
            "rows": len(self._test_result_data),
        }
        if not self._test_result_data:
            return summary
        # 汇总首末行负载电流与电压（禁止臆造，直接回读结构化行）
        first = self._test_result_data[0]
        last = self._test_result_data[-1]
        summary["first_load_current"] = first.get("load_current")
        summary["last_load_current"] = last.get("load_current")
        return summary

    # ------------------------------------------------------------------
    # UI 回填单一写入口（AIAssist_PageScopedControlPlan.md §4.2）
    #
    # apply_config_to_controls(cfg) 是回填测试配置控件的唯一入口，
    # AI 回填与未来轮询/手动刷新共用，杜绝两套逻辑漂移。键名与
    # get_test_config() 输出对齐。
    # ------------------------------------------------------------------
    def apply_config_to_controls(self, cfg: dict) -> tuple[bool, str]:
        if not isinstance(cfg, dict):
            return False, "配置草案格式无效（期望 dict）。"

        # 线程边界（§4.2-2）：AI 决策在 QThread，回填须经主线程执行；
        # dispatcher 经 QTimer.singleShot(0) 已切回主线程，此处加防御性守卫，
        # 杜绝 worker 线程直接 setValue 违反「UI 禁阻塞 / 跨线程改控件」铁律。
        if threading.current_thread() is not threading.main_thread():
            _logger.error(
                "apply_config_to_controls 在非主线程被调用，拒绝回填以防违反线程边界"
            )
            return False, "配置回填未在主线程执行，已拒绝。"

        applied: list[str] = []
        touched: list = []

        def _set_combo(combo, key):
            val = cfg.get(key)
            if val is None:
                return
            idx = combo.findText(str(val))
            if idx >= 0:
                combo.setCurrentIndex(idx)
                applied.append(key)
                touched.append(combo)

        def _set_spin(spin, key):
            val = cfg.get(key)
            if val is None:
                return
            try:
                spin.setValue(float(val))
            except (TypeError, ValueError):
                return
            applied.append(key)
            touched.append(spin)

        def _set_int_spin(spin, key):
            val = cfg.get(key)
            if val is None:
                return
            try:
                spin.setValue(int(float(val)))
            except (TypeError, ValueError):
                return
            applied.append(key)
            touched.append(spin)

        def _set_text(edit, key):
            val = cfg.get(key)
            if val is None:
                return
            edit.setText(str(val))
            applied.append(key)
            touched.append(edit)

        def _set_check(check, key):
            val = cfg.get(key)
            if val is None:
                return
            check.setChecked(bool(val))
            applied.append(key)
            touched.append(check)

        _set_combo(self.ripple_channel_combo, "ripple_channel")
        _set_combo(self.load_channel_combo, "load_channel")
        _set_combo(self.is_gain_method_combo, "is_gain_method")
        _set_text(self.is_gain_device_addr_edit, "is_gain_device_addr")
        _set_text(self.is_gain_reg_addr_edit, "is_gain_reg_addr")
        _set_int_spin(self.is_gain_msb_spin, "is_gain_msb")
        _set_int_spin(self.is_gain_lsb_spin, "is_gain_lsb")
        _set_spin(self.is_gain_start_current_spin, "is_gain_start_current")
        _set_spin(self.is_gain_end_current_spin, "is_gain_end_current")
        _set_spin(self.is_gain_step_current_spin, "is_gain_step_current")
        _set_check(self.save_screenshot_cb, "save_screenshot")

        if not applied:
            return False, "配置草案未包含任何可识别的配置项。"
        # §4.2-3 可视化反馈：被 AI 修改的控件临时高亮（Phase 3）。
        self._highlight_widgets(touched)
        self.append_log(f"[AI] 已应用配置：{', '.join(applied)}")
        return True, f"已应用配置项：{', '.join(applied)}。"

    def _highlight_widgets(self, widgets: list) -> None:
        """被 AI 修改的控件临时高亮边框（§4.2-3 / Phase 3）。"""
        if not widgets:
            return
        for widget in widgets:
            if widget is None:
                continue
            widget.setStyleSheet(_AI_HIGHLIGHT_QSS)
            widget.setProperty("aiHighlighted", True)
            widget.style().unpolish(widget)
            widget.style().polish(widget)

            def _clear(_w=widget):
                try:
                    _w.setStyleSheet("")
                    _w.setProperty("aiHighlighted", False)
                    _w.style().unpolish(_w)
                    _w.style().polish(_w)
                except RuntimeError:  # noqa: BLE001 - widget 可能已销毁
                    pass
            QTimer.singleShot(_AI_HIGHLIGHT_MS, _clear)


def main():
    from ui.standalone import run_standalone_widget

    return run_standalone_widget(
        lambda: PMUIsGainUI(),
        "PMU Is_gain",
    )


if __name__ == "__main__":
    raise SystemExit(main())
