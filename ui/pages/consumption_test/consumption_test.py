#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Consumption Test UI组件
用于对DUT进行固件下载和功耗测试
"""

import sys
import os
import threading
from ui.resource_path import get_resource_base
import re
import json

try:
    import yaml
except ImportError:
    yaml = None

sys.path.append(get_resource_base())

from ui.modules.n6705c_module_frame import N6705CConnectionMixin
from ui.widgets.button import SpinningSearchButton, update_connect_button_state
from ui.modules.serialCom_module.serialCom_module_frame import SerialComMixin, MODE_INLINE
from ui.modules.execution_logs_module_frame import ExecutionLogsFrame
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
    QPushButton, QLabel, QLineEdit, QPlainTextEdit,
    QFrame, QApplication, QFileDialog,
    QCheckBox, QSizePolicy, QMessageBox, QScrollArea,
    QTableWidget, QTableWidgetItem, QHeaderView,
)
from PySide6.QtCore import (
    Qt, QTimer, Signal, QThread, QObject, QSize,
    QRectF, QRect, QPropertyAnimation, QEasingCurve, Property
)
from PySide6.QtGui import (
    QFont, QIcon, QPixmap, QPainter, QColor, QPen,
    QFontMetrics
)
from PySide6.QtSvg import QSvgRenderer

from lib.download_tools.download_script import download_bin, DownloadMode, DownloadState, DownloadResult, detect_chip_from_bin
from chips.bes_chip_configs.bes_chip_configs import SUPPORTED_CHIPS, get_chip_config
from ui.widgets.dark_combobox import DarkComboBox
from ui.widgets.progress_button import ProgressButton
from log_config import get_logger
from ui.theme import Colors, FontSizes, Radius, Spacing, FONT_FAMILY, FONT_MONO
from ui.styles import get_page_base_qss, SCROLLBAR_STYLE
from core.ai.ui_action_registry import UIActionSpec
from core.ai.page_contract import (
    CAP_GET_CONFIG,
    CAP_APPLY_CONFIG,
    CAP_START_TEST,
    CAP_STOP_TEST,
    CAP_GET_RESULT,
)

from core.consumption_test.workers import (
    CURRENT_UNIT,
    _UNIT_CONFIG,
    _format_current_unified,
    _ChipCheckWorker,
    _ConsumptionTestForceHighWorker,
    _ConsumptionTestForceWorker,
)
from ui.pages.consumption_test.widgets import (
    DownloadModeToggle, ControlMethodToggle, PolarityToggle,
)
from ui.pages.consumption_test.view_config import ConsumptionTestViewConfigMixin
from ui.pages.consumption_test.view_results import ConsumptionTestViewResultsMixin
from ui.pages.consumption_test.view_panels import ConsumptionTestViewPanelsMixin
from core.consumption_test.consumption_controller import ConsumptionController

logger = get_logger(__name__)


class _SearchMcuPortWorker(QObject):
    finished = Signal(list)
    error = Signal(str)

    def __init__(self, mcu_type="yd_rp2040"):
        super().__init__()
        self._mcu_type = mcu_type

    def run(self):
        try:
            if self._mcu_type == "ch9114f":
                from instruments.MCU_IO.ch9114f import (
                    list_ch9114f_ports, CH9114_USB_VID, CH9114_USB_PID,
                )
                import serial.tools.list_ports
                ports = list_ch9114f_ports() or []
                # 构建 device -> description 映射，补全下拉显示的完整端口名
                desc_map = {}
                for info in serial.tools.list_ports.comports():
                    if info.vid == CH9114_USB_VID and info.pid == CH9114_USB_PID:
                        desc_map[info.device] = info.description or info.device
                labels = [f"{p} - {desc_map.get(p, p)}" for p in ports]
                self.finished.emit(labels)
            else:
                import serial.tools.list_ports
                ports = serial.tools.list_ports.comports()
                self.finished.emit([f"{p.device} - {p.description}" for p in ports])
        except Exception as e:
            logger.error("MCU port scan failed: %s", e, exc_info=True)
            self.error.emit(str(e))


class _ConnectMcuWorker(QObject):
    finished = Signal(object, str)
    error = Signal(str)

    def __init__(self, port, mcu_type="yd_rp2040", baudrate=921600):
        super().__init__()
        self._port = port
        self._mcu_type = mcu_type
        self._baudrate = baudrate

    def run(self):
        try:
            from instruments.factory import create_mcu_io
            if self._mcu_type == "ch9114f":
                inst = create_mcu_io("ch9114f", port=self._port)
            else:
                inst = create_mcu_io("yd_rp2040", port=self._port, baudrate=self._baudrate)
            ok = inst.connect()
            if not ok:
                self.error.emit(f"Failed to connect {self._port}")
                return
            self.finished.emit(inst, inst.identify())
        except Exception as e:
            logger.error("MCU IO connection failed: %s", e, exc_info=True)
            self.error.emit(str(e))


_ICONS_DIR = os.path.join(
    get_resource_base(),
    "resources", "icons"
)

_PAGE_SVGS_DIR = os.path.join(
    get_resource_base(),
    "resources", "pages", "consumption_test_SVGs"
)

_MAIN_CHIP_CONFIGS_DIR = os.path.join(
    get_resource_base(),
    "chips", "bes_chip_configs", "main_chip_configs"
)

from ui.utils.icon_utils import tinted_svg_icon as _tinted_svg_icon


class ConsumptionTestUI(QWidget, ConsumptionTestViewConfigMixin, ConsumptionTestViewResultsMixin, ConsumptionTestViewPanelsMixin, N6705CConnectionMixin, SerialComMixin):
    connection_status_changed = Signal(bool)
    serial_connection_changed = Signal(bool)
    serial_data_received = Signal(bytes)

    CHANNEL_COLORS_LIST = [
        {"accent": "#d4a514", "bg": "#1a1708", "border": "#3d2e08"},
        {"accent": "#18b67a", "bg": "#081a14", "border": "#0a3d28"},
        {"accent": "#2f6fed", "bg": "#081028", "border": "#0c2a5e"},
        {"accent": "#d14b72", "bg": "#1a080e", "border": "#3d0c22"},
        {"accent": "#a855f7", "bg": "#150a20", "border": "#3a1a5e"},
        {"accent": "#06b6d4", "bg": "#081a1e", "border": "#0a3d4a"},
        {"accent": "#f97316", "bg": "#1a1008", "border": "#3d2808"},
        {"accent": "#ec4899", "bg": "#1a0812", "border": "#3d0c28"},
    ]

    NAME_OPTIONS = ["Vbat", "Vcore", "VcoreM", "VcoreL", "VANA", "VHPPA", "Vusb"]

    SINGLE_DEVICE_CHANNEL_CONFIGS = [
        {"name": "Vbat", "channel": "A-CH1", "enabled": True},
        {"name": "Vcore", "channel": "A-CH2", "enabled": True},
        {"name": "VANA", "channel": "A-CH3", "enabled": True},
        {"name": "VHPPA", "channel": "A-CH4", "enabled": True},
    ]

    DUAL_DEVICE_CHANNEL_CONFIGS = [
        {"name": "Vbat", "channel": "A-CH1", "enabled": True},
        {"name": "Vcore", "channel": "A-CH2", "enabled": True},
        {"name": "VANA", "channel": "A-CH3", "enabled": True},
        {"name": "VHPPA", "channel": "A-CH4", "enabled": True},
        {"name": "CH5", "channel": "B-CH1", "enabled": False},
        {"name": "CH6", "channel": "B-CH2", "enabled": False},
        {"name": "CH7", "channel": "B-CH3", "enabled": False},
        {"name": "CH8", "channel": "B-CH4", "enabled": False},
    ]

    def __init__(self, n6705c_top=None, instrument_manager=None, ui_action_registry=None):
        super().__init__()

        self._n6705c_top = n6705c_top
        self._instrument_manager = instrument_manager
        self._ui_action_registry = ui_action_registry
        self.n6705c_a = None
        self.n6705c_b = None
        self.is_connected_a = False
        self.is_connected_b = False
        self.mcu_io = None
        self.is_mcu_connected = False
        self._mcu_search_thread = None
        self._mcu_search_worker = None
        self._mcu_connect_thread = None
        self._mcu_connect_worker = None
        self._default_mcu_type = "ch9114f"

        self.init_n6705c_connection(n6705c_top, instrument_manager=instrument_manager)
        self.init_serial_connection(mode=MODE_INLINE, prefix="DUT Serial")

        self.firmware_path = ""
        self.firmware_paths = []
        self.config_content = ""
        self.selected_chip_config = None
        self.is_testing = False

        self._test_thread = None
        self._test_worker = None
        self._download_thread = None
        self._download_worker = None
        self._chip_check_thread = None
        self._chip_check_worker = None
        self._auto_test_thread = None
        self._auto_test_worker = None
        self._bin_results_data = []
        self._current_total_bins = 0

        self._channel_configs = []
        self._channel_config_widgets = []
        self._syncing = False

        self._pending_channel_selections = None
        self._pending_aux_selections = None
        self._suppress_preset_channels = False

        self.poweron_channel_combo = None
        self.reset_channel_combo = None
        self._saved_control_channels = {
            "N6705C": {"poweron": "B-CH1", "reset": "B-CH2"},
            "MCU": {"poweron": "GPIO0", "reset": "GPIO1"},
        }
        self._current_control_method = "MCU"

        self._controller = ConsumptionController(parent=self)
        self._controller.log_message.connect(self.append_log)
        self._controller.download_started.connect(self._on_ctrl_download_started)
        self._controller.download_state_changed.connect(self._on_ctrl_download_state_changed)
        self._controller.download_finished.connect(self._on_ctrl_download_finished)
        self._controller.download_error.connect(self._on_ctrl_download_error)
        self._controller.download_cleaned.connect(self._on_ctrl_download_cleaned)

        self._controller.auto_test_channel_result.connect(self._on_force_high_channel_result)
        self._controller.auto_test_summary.connect(self._on_test_summary)
        self._controller.auto_test_progress.connect(lambda v: self.auto_test_btn.setProgress(v))
        self._controller.auto_test_error.connect(self._on_auto_test_error)
        self._controller.auto_test_finished.connect(self._on_auto_test_finished)
        self._controller.auto_test_cleaned.connect(self._on_auto_test_thread_cleaned)

        self._setup_style()
        self._create_layout()
        self._sync_n6705c_dual_from_top()

        # §5b：登记本页无专用接口的按钮为具名 UI 动作（白名单制，handler 复用原槽）
        self._register_ai_ui_actions()

    def _register_ai_ui_actions(self):
        """§5b.5：登记本页按钮为 AI 可触发的具名 UI 动作（白名单制）。

        handler 直接复用按钮原 clicked.connect 的槽（_on_chip_check /
        _on_auto_test 等），行为与人点按钮完全一致；enabled_when 校验最小前置条件，
        不满足时 list_ui_actions 不返回、ui_invoke 明示不可用（不盲点）。
        """
        registry = self._ui_action_registry
        if registry is None:
            return

        def _has_n6705c() -> bool:
            return self.is_connected_a or self.is_connected_b

        def _has_mcu() -> bool:
            return bool(self.is_mcu_connected and self.mcu_io is not None)

        def _has_firmware() -> bool:
            return bool(getattr(self, "firmware_paths", []) or self.firmware_path)

        def _has_serial_port() -> bool:
            return bool(self.get_selected_serial_port())

        def _has_enabled_channel() -> bool:
            return any(cfg.get("enabled") for cfg in self._channel_configs)

        def _has_chip() -> bool:
            return self.selected_chip_config is not None

        def _is_testing() -> bool:
            return bool(self.is_testing)

        def _wrap(label, fn, pre_msg=None):
            def _run() -> tuple[bool, str]:
                try:
                    if pre_msg is not None:
                        self.append_log(f"[AI] 触发 {label}")
                    fn()
                    return True, f"{label} 已触发。"
                except Exception as exc:  # noqa: BLE001
                    logger.error("%s 执行失败", label, exc_info=True)
                    return False, f"{label} 执行失败：{exc}"
            return _run

        registry.register_many([
            UIActionSpec(
                id="consumption_test.chip_check",
                label="Chip Check (I2C)",
                page_key="consumption_test",
                handler=_wrap("Chip Check", self._on_chip_check, pre_msg=True),
                risk="medium",
                confirm=True,
                enabled_when=lambda: not _is_testing(),
                description=(
                    "经 I2C 探测 DUT 芯片型号并自动匹配 chip_combo。"
                    "需 MCU/I2C 接口可用（CH9114F/YD-RP2040 已连接或 I2C 接口已初始化）。"
                ),
            ),
            UIActionSpec(
                id="consumption_test.chip_save",
                label="Save Chip Rails",
                page_key="consumption_test",
                handler=_wrap("Save Chip Rails", self._on_chip_save, pre_msg=True),
                risk="low",
                confirm=True,
                enabled_when=_has_chip,
                description=(
                    "把当前 5 个电源轨 (Vcore/VcoreM/VcoreL/VANA/VHPPA) 的 YAML 配置"
                    "写入 chips/bes_chip_configs/main_chip_configs/<chip>.yaml。"
                    "需先在 Chip 下拉中选择芯片。"
                ),
            ),
            UIActionSpec(
                id="consumption_test.start_test",
                label="Start Consumption Test",
                page_key="consumption_test",
                handler=_wrap("Start Consumption Test", self._on_start_test, pre_msg=True),
                risk="high",
                confirm=True,
                enabled_when=lambda: (
                    _has_n6705c() and _has_enabled_channel() and not _is_testing()
                ),
                description=(
                    "启动 Force-High 功耗测试：对启用的非 Vbat 通道拉高电压，"
                    "在 test_time 时间内采样电流并统计 Vbat_remain。"
                    "需至少一台 N6705C 已连接、Vbat 通道已配置、且至少一个通道启用。"
                ),
            ),
            UIActionSpec(
                id="consumption_test.auto_test",
                label="Start Auto Test",
                page_key="consumption_test",
                handler=_wrap("Start Auto Test", self._on_auto_test, pre_msg=True),
                risk="high",
                confirm=True,
                enabled_when=lambda: (
                    _has_n6705c() and _has_firmware() and _has_serial_port()
                    and _has_enabled_channel() and not _is_testing()
                ),
                description=(
                    "启动 Auto Test 全流程：固件下载 → POWERON → RESET → I2C 配置 → 功耗测试。"
                    "需 N6705C 已连接、固件已选、DUT 串口已选、通道已配置；"
                    "MCU 控制方式下另需 MCU 已连接。"
                ),
            ),
            UIActionSpec(
                id="consumption_test.stop_test",
                label="Stop Running Test",
                page_key="consumption_test",
                handler=_wrap("Stop Running Test", self._stop_running_test, pre_msg=True),
                risk="medium",
                confirm=True,
                enabled_when=_is_testing,
                description=(
                    "停止当前正在运行的功耗测试或 Auto Test。"
                    "包括 _test_worker（force-high/force-auto）与 _auto_test_worker。"
                ),
            ),
        ])

    def _stop_running_test(self):
        """AI 触发 stop_test 的统一入口：自动判别 force-high / auto_test 两类 worker。"""
        if getattr(self, "_auto_test_worker", None) is not None or \
                getattr(self, "_controller", None) and self._controller.is_auto_test_running():
            self._stop_auto_test()
            return
        if self._test_worker is not None:
            self._stop_test()
            return
        # 兜底：直接复位状态
        self._stop_test()

    def _setup_style(self):
        self.setFont(QFont("Segoe UI", 9))
        self.setObjectName("ConsumptionTestRoot")
        _cb_icons = self._get_checkmark_path("5d45ff")
        page_extra = f"""
        QWidget#ConsumptionTestRoot {{
            background-color: {Colors.bg_secondary};
        }}

        QFrame#logContainer {{
            background-color: {Colors.bg_deep};
            border: 1px solid {Colors.border_secondary};
            border-radius: {Radius.container}px;
        }}

        QCheckBox::indicator {{
            width: 16px;
            height: 16px;
            image: url("__UNCHECKED__");
        }}

        QCheckBox::indicator:checked {{
            image: url("__CHECKED__");
        }}
        """.replace("__UNCHECKED__", _cb_icons['unchecked']).replace("__CHECKED__", _cb_icons['checked'])
        self.setStyleSheet(get_page_base_qss() + page_extra)

    def _on_device_search(self, label):
        top = self._n6705c_top
        if top:
            is_conn = getattr(top, f"is_connected_{label.lower()}", False)
            if is_conn:
                return

        from debug_config import DEBUG_MOCK
        w = self._n6705c_conn_widgets[label]
        if DEBUG_MOCK:
            w["combo"].clear()
            w["combo"].addItem(f"DEBUG::MOCK::N6705C::{label}")
            w["status"].setText("● Mock Ready")
            w["status"].setStyleSheet("color: #ff9800; font-weight: bold; background: transparent; border: none;")
            self.append_log(f"[DEBUG] Mock device {label} loaded, skip real VISA scan.")
            return

        w["status"].setText("● Searching")
        w["status"].setStyleSheet("color: #ff9800; font-weight: bold; background: transparent; border: none;")
        w["search_btn"].setEnabled(False)
        w["connect_btn"].setEnabled(False)
        self.append_log(f"[SYSTEM] Scanning VISA resources for N6705C-{label}...")

        from ui.modules.n6705c_module_frame import _SearchN6705CWorker
        worker = _SearchN6705CWorker()
        thread = QThread()
        worker.moveToThread(thread)
        thread.started.connect(worker.run)
        worker.finished.connect(lambda devs, lbl=label: self._on_device_search_done(lbl, devs))
        worker.error.connect(lambda err, lbl=label: self._on_device_search_error(lbl, err))
        worker.finished.connect(thread.quit)
        worker.error.connect(thread.quit)
        thread.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)

        setattr(self, f"_search_thread_{label}", thread)
        setattr(self, f"_search_worker_{label}", worker)
        thread.start()

    def _on_device_search_done(self, label, devices):
        w = self._n6705c_conn_widgets[label]
        w["combo"].setEnabled(True)
        w["combo"].clear()
        if devices:
            for dev in devices:
                w["combo"].addItem(dev)
            w["status"].setText(f"● Found {len(devices)}")
            w["status"].setStyleSheet("color: #00a859; font-weight: bold; background: transparent; border: none;")
            self.append_log(f"[SYSTEM] Found {len(devices)} N6705C device(s) for slot {label}.")
        else:
            w["combo"].addItem("No N6705C device found")
            w["combo"].setEnabled(False)
            w["status"].setText("● Not Found")
            w["status"].setStyleSheet("color: #e53935; font-weight: bold; background: transparent; border: none;")
        w["search_btn"].setEnabled(True)
        w["connect_btn"].setEnabled(True)

    def _on_device_search_error(self, label, err):
        w = self._n6705c_conn_widgets[label]
        w["status"].setText("● Failed")
        w["status"].setStyleSheet("color: #e53935; font-weight: bold; background: transparent; border: none;")
        self.append_log(f"[ERROR] Search failed for N6705C-{label}: {err}")
        w["search_btn"].setEnabled(True)
        w["connect_btn"].setEnabled(True)

    def _on_device_connect_or_disconnect(self, label):
        attr = label.lower()
        is_conn = getattr(self, f"is_connected_{attr}", False)
        if is_conn:
            self._disconnect_device(label)
        else:
            self._connect_device(label)

    def _connect_device(self, label):
        attr = label.lower()
        w = self._n6705c_conn_widgets[label]
        from debug_config import DEBUG_MOCK
        from ui.modules.n6705c_module_frame import _update_n6705c_btn_state
        prev_count = self._connected_device_count()

        if DEBUG_MOCK:
            from instruments.mock.mock_instruments import MockN6705C
            inst = MockN6705C()
            setattr(self, f"n6705c_{attr}", inst)
            setattr(self, f"is_connected_{attr}", True)
            _update_n6705c_btn_state(w["connect_btn"], connected=True)
            w["search_btn"].setEnabled(False)
            w["status"].setText("● Connected")
            w["status"].setStyleSheet("color: #00a859; font-weight: bold; background: transparent; border: none;")
            self.append_log(f"[DEBUG] Mock N6705C-{label} connected.")
            visa = w["combo"].currentText()
            self._syncing = True
            try:
                if self._n6705c_top:
                    getattr(self._n6705c_top, f"connect_{attr}")(visa, inst, f"MOCK-{label}")
            finally:
                self._syncing = False
            new_count = self._connected_device_count()
            self._apply_preset_channels(prev_count, new_count)
            self._update_available_channels()
            return

        w["status"].setText("● Connecting")
        w["status"].setStyleSheet("color: #ff9800; font-weight: bold; background: transparent; border: none;")
        w["connect_btn"].setEnabled(False)
        self.append_log(f"[SYSTEM] Connecting N6705C-{label}...")

        if self._instrument_manager:
            from core.instruments import InstrumentSpec
            visa = w["combo"].currentText()
            self._instrument_manager.connect_async(InstrumentSpec(
                instrument_type="n6705c",
                role="power_analyzer",
                connection_kind="visa",
                slot=label,
                resource=visa,
            ))
            return

        try:
            from instruments.power.keysight.n6705c import N6705C
            visa = w["combo"].currentText()
            inst = N6705C(visa)
            idn = inst.instr.query("*IDN?")
            if "N6705C" in idn:
                setattr(self, f"n6705c_{attr}", inst)
                setattr(self, f"is_connected_{attr}", True)
                _update_n6705c_btn_state(w["connect_btn"], connected=True)
                w["search_btn"].setEnabled(False)
                w["status"].setText("● Connected")
                w["status"].setStyleSheet("color: #00a859; font-weight: bold; background: transparent; border: none;")
                self.append_log(f"[SYSTEM] N6705C-{label} connected. IDN: {idn.strip()}")
                self._syncing = True
                try:
                    if self._n6705c_top:
                        serial = ""
                        try:
                            serial = idn.strip().split(",")[2].strip()
                        except Exception:
                            pass
                        getattr(self._n6705c_top, f"connect_{attr}")(visa, inst, serial)
                finally:
                    self._syncing = False
                new_count = self._connected_device_count()
                self._apply_preset_channels(prev_count, new_count)
                self._update_available_channels()
            else:
                w["status"].setText("● Mismatch")
                w["status"].setStyleSheet("color: #e53935; font-weight: bold; background: transparent; border: none;")
                self.append_log(f"[ERROR] Connected device on {label} is not N6705C.")
        except Exception as e:
            w["status"].setText("● Failed")
            w["status"].setStyleSheet("color: #e53935; font-weight: bold; background: transparent; border: none;")
            self.append_log(f"[ERROR] Connection failed for N6705C-{label}: {e}")
        finally:
            w["connect_btn"].setEnabled(True)

    def _disconnect_device(self, label):
        attr = label.lower()
        w = self._n6705c_conn_widgets[label]
        from ui.modules.n6705c_module_frame import _update_n6705c_btn_state
        prev_count = self._connected_device_count()

        try:
            self._syncing = True
            try:
                if self._n6705c_top:
                    getattr(self._n6705c_top, f"disconnect_{attr}")()
                else:
                    inst = getattr(self, f"n6705c_{attr}", None)
                    if inst:
                        if hasattr(inst, 'disconnect'):
                            inst.disconnect()
                        else:
                            if hasattr(inst, 'instr') and inst.instr:
                                inst.instr.close()
                            if hasattr(inst, 'rm') and inst.rm:
                                inst.rm.close()
            finally:
                self._syncing = False
            setattr(self, f"n6705c_{attr}", None)
            setattr(self, f"is_connected_{attr}", False)
            _update_n6705c_btn_state(w["connect_btn"], connected=False)
            w["search_btn"].setEnabled(True)
            w["combo"].setEnabled(True)
            w["status"].setText("● Disconnected")
            w["status"].setStyleSheet("color: #8ea6cf; font-weight: bold; background: transparent; border: none;")
            self.append_log(f"[SYSTEM] N6705C-{label} disconnected.")
            new_count = self._connected_device_count()
            self._apply_preset_channels(prev_count, new_count)
            self._update_available_channels()
        except Exception as e:
            w["status"].setText("● Failed")
            w["status"].setStyleSheet("color: #e53935; font-weight: bold; background: transparent; border: none;")
            self.append_log(f"[ERROR] Disconnect failed for N6705C-{label}: {e}")

    def _sync_n6705c_dual_from_top(self):
        top = self._n6705c_top
        if not top:
            self._update_test_panel_state()
            return
        from ui.modules.n6705c_module_frame import _update_n6705c_btn_state
        prev_count = self._connected_device_count()
        for label, attr in [("A", "a"), ("B", "b")]:
            n6705c = getattr(top, f"n6705c_{attr}", None)
            is_conn = getattr(top, f"is_connected_{attr}", False)
            visa_res = getattr(top, f"visa_resource_{attr}", "")
            if label not in self._n6705c_conn_widgets:
                continue
            w = self._n6705c_conn_widgets[label]
            if is_conn and n6705c:
                setattr(self, f"n6705c_{attr}", n6705c)
                setattr(self, f"is_connected_{attr}", True)
                _update_n6705c_btn_state(w["connect_btn"], connected=True)
                w["search_btn"].setEnabled(False)
                if visa_res:
                    w["combo"].clear()
                    w["combo"].addItem(visa_res)
                w["status"].setText("● Connected")
                w["status"].setStyleSheet("color: #00a859; font-weight: bold; background: transparent; border: none;")
            else:
                setattr(self, f"n6705c_{attr}", None)
                setattr(self, f"is_connected_{attr}", False)
                _update_n6705c_btn_state(w["connect_btn"], connected=False)
                w["search_btn"].setEnabled(True)
                w["combo"].setEnabled(True)
                w["status"].setText("● Disconnected")
                w["status"].setStyleSheet("color: #8ea6cf; font-weight: bold; background: transparent; border: none;")
        self.n6705c = self.n6705c_a
        self.is_connected = self.is_connected_a
        new_count = self._connected_device_count()
        self._apply_preset_channels(prev_count, new_count)
        self._update_available_channels()

    def sync_n6705c_from_top(self):
        if self._syncing:
            return
        self._sync_n6705c_dual_from_top()

    def set_system_status(self, status, is_error=False):
        pass

    @staticmethod
    def _label_from_n6705c_session_id(session_id):
        # session_id 格式: "n6705c:A" / "n6705c:B"
        if not session_id or ":" not in session_id:
            return None
        prefix, slot = session_id.split(":", 1)
        if prefix != "n6705c" or slot not in ("A", "B"):
            return None
        return slot

    def _on_mixin_manager_connected(self, session_id: str):
        # ConsumptionTestUI 使用 per-label 控件字典(self._n6705c_conn_widgets),
        # 不使用基类的 self.connect_btn / self.search_btn, 故此处覆盖基类实现。
        label = self._label_from_n6705c_session_id(session_id)
        if label is None or label not in self._n6705c_conn_widgets:
            return
        if not self._n6705c_instrument_manager:
            return
        session = self._n6705c_instrument_manager.get_session(session_id)
        if not session or not session.connected:
            return

        from ui.modules.n6705c_module_frame import _update_n6705c_btn_state
        attr = label.lower()
        prev_count = self._connected_device_count()

        setattr(self, f"n6705c_{attr}", session.instance)
        setattr(self, f"is_connected_{attr}", True)
        w = self._n6705c_conn_widgets[label]
        _update_n6705c_btn_state(w["connect_btn"], connected=True)
        w["connect_btn"].setEnabled(True)
        w["search_btn"].setEnabled(False)
        w["status"].setText("● Connected")
        w["status"].setStyleSheet("color: #00a859; font-weight: bold; background: transparent; border: none;")
        self.append_log(f"[SYSTEM] N6705C-{label} connected via manager.")

        # N6705CTop 已通过自身 _on_session_connected 订阅同一 manager 的
        # session_connected 信号完成状态同步; 此处不可再调用 connect_a/connect_b,
        # 否则会触发 attach_external -> 重复 emit session_connected -> 无限递归。
        new_count = self._connected_device_count()
        self._apply_preset_channels(prev_count, new_count)
        self._update_available_channels()
        self.connection_status_changed.emit(True)

    def _on_mixin_manager_connect_failed(self, session_id: str, error: str):
        label = self._label_from_n6705c_session_id(session_id)
        if label is None or label not in self._n6705c_conn_widgets:
            return
        w = self._n6705c_conn_widgets[label]
        w["connect_btn"].setEnabled(True)
        w["status"].setText("● Failed")
        w["status"].setStyleSheet("color: #e53935; font-weight: bold; background: transparent; border: none;")
        self.append_log(f"[ERROR] Connection failed for N6705C-{label}: {error}")

    def _on_mixin_manager_disconnected(self, session_id: str):
        label = self._label_from_n6705c_session_id(session_id)
        if label is None or label not in self._n6705c_conn_widgets:
            return

        from ui.modules.n6705c_module_frame import _update_n6705c_btn_state
        attr = label.lower()
        prev_count = self._connected_device_count()

        setattr(self, f"n6705c_{attr}", None)
        setattr(self, f"is_connected_{attr}", False)
        w = self._n6705c_conn_widgets[label]
        _update_n6705c_btn_state(w["connect_btn"], connected=False)
        w["connect_btn"].setEnabled(True)
        w["search_btn"].setEnabled(True)
        w["combo"].setEnabled(True)
        w["status"].setText("● Disconnected")
        w["status"].setStyleSheet("color: #8ea6cf; font-weight: bold; background: transparent; border: none;")
        self.append_log(f"[SYSTEM] N6705C-{label} disconnected.")
        new_count = self._connected_device_count()
        self._apply_preset_channels(prev_count, new_count)
        self._update_available_channels()
        self.connection_status_changed.emit(False)

    def _get_available_channel_options(self):
        options = []
        for label in ["A", "B"]:
            for ch in range(1, 5):
                options.append(f"{label}-CH{ch}")
        return options

    def _connected_device_count(self):
        count = 0
        if self.is_connected_a:
            count += 1
        if self.is_connected_b:
            count += 1
        return count

    def _update_available_channels(self):
        options = self._get_available_channel_options()
        pending_ch = self._pending_channel_selections or {}
        for i, wdata in enumerate(self._channel_config_widgets):
            combo = wdata["channel_combo"]
            # 目标值:优先使用待恢复的导入值,否则使用 _channel_configs 里记录的通道
            desired = None
            if i in pending_ch:
                desired = pending_ch[i]
            elif i < len(self._channel_configs):
                desired = self._channel_configs[i].get("channel", "")
            if not desired:
                desired = combo.currentText()
            combo.blockSignals(True)
            combo.clear()
            for opt in options:
                combo.addItem(opt)
            matched = False
            for j in range(combo.count()):
                if combo.itemText(j) == desired:
                    combo.setCurrentIndex(j)
                    matched = True
                    break
            combo.blockSignals(False)
            # 若期望值在选项里找到了,同步回 _channel_configs,并清理 pending
            if matched and i < len(self._channel_configs):
                self._channel_configs[i]["channel"] = desired
            if i in pending_ch and matched:
                pending_ch.pop(i, None)
        if self._pending_channel_selections is not None and not pending_ch:
            self._pending_channel_selections = None

        pending_aux = self._pending_aux_selections or {}
        aux_options = self._get_control_channel_options()
        for key, extra_combo in (("poweron", self.poweron_channel_combo),
                                 ("reset", self.reset_channel_combo)):
            if extra_combo is None:
                continue
            desired = pending_aux.get(key) if pending_aux else None
            if not desired:
                desired = extra_combo.currentText()
            extra_combo.blockSignals(True)
            extra_combo.clear()
            for opt in aux_options:
                extra_combo.addItem(opt)
            matched = False
            for j in range(extra_combo.count()):
                if extra_combo.itemText(j) == desired:
                    extra_combo.setCurrentIndex(j)
                    matched = True
                    break
            extra_combo.blockSignals(False)
            if matched and pending_aux:
                pending_aux.pop(key, None)
        if self._pending_aux_selections is not None and not pending_aux:
            self._pending_aux_selections = None

        self._refresh_result_cards()
        self._update_test_panel_state()

    def _apply_preset_channels(self, prev_count, new_count):
        if prev_count == new_count:
            return
        # 导入配置流程中不允许 preset 覆盖导入后的通道配置
        if getattr(self, "_suppress_preset_channels", False):
            return

        if new_count == 0:
            self._clear_all_channel_configs()
        elif new_count == 1:
            self._clear_all_channel_configs()
            for cfg in self.SINGLE_DEVICE_CHANNEL_CONFIGS:
                self._add_channel_config_card(cfg["name"], cfg["channel"], cfg["enabled"])
        elif new_count >= 2 and prev_count < 2:
            self._clear_all_channel_configs()
            for cfg in self.DUAL_DEVICE_CHANNEL_CONFIGS:
                self._add_channel_config_card(cfg["name"], cfg["channel"], cfg["enabled"])

    def _clear_all_channel_configs(self):
        for wdata in reversed(self._channel_config_widgets):
            wdata["card"].hide()
            wdata["card"].deleteLater()
        self._channel_configs.clear()
        self._channel_config_widgets.clear()
        while self.result_cards_layout.count():
            item = self.result_cards_layout.takeAt(0)
            w = item.widget()
            if w:
                w.hide()
                w.deleteLater()
        self.channel_cards = {}

    def _update_test_panel_state(self):
        has_device = self._connected_device_count() > 0
        if hasattr(self, '_disabled_overlay'):
            if has_device:
                self._disabled_overlay.hide()
            else:
                self._disabled_overlay.show()
                self._disabled_overlay.raise_()

    def _get_checkmark_path(self, accent_color):
        safe_name = accent_color.replace("#", "").replace(" ", "")
        icons_dir = os.path.join(
            get_resource_base(),
            "resources", "icons"
        )
        return {
            "checked": os.path.join(icons_dir, f"checked_{safe_name}.svg").replace("\\", "/"),
            "unchecked": os.path.join(icons_dir, f"unchecked_{safe_name}.svg").replace("\\", "/"),
        }

    def _browse_firmware(self):
        file_paths, _ = QFileDialog.getOpenFileNames(
            self, "Select Firmware File(s)", "",
            "Firmware Files (*.bin *.hex);;All Files (*)"
        )
        if file_paths:
            self.firmware_paths = file_paths
            self.firmware_path = file_paths[0]
            names = [os.path.basename(p) for p in file_paths]
            self.firmware_file_input.setText("; ".join(names))
            for fp in file_paths:
                self.append_log(f"[SYSTEM] Firmware file selected: {os.path.basename(fp)}")

    def _download_to_dut(self):
        if not self.firmware_path:
            logger.warning("No firmware file selected")
            self.append_log("[WARNING] No firmware file selected.")
            return

        if self._controller.is_download_running():
            logger.warning("Download already in progress")
            self.append_log("[WARNING] Download already in progress.")
            return

        port_text = self.get_selected_serial_port()
        if not port_text:
            logger.warning("No serial port selected")
            self.append_log("[WARNING] No serial port selected.")
            return

        m = re.search(r'(\d+)', port_text)
        com_port = m.group(1) if m else port_text

        mode_str = self.download_mode_toggle.value().lower()
        mode = DownloadMode.FLASH if mode_str == "flash" else DownloadMode.RAMRUN

        self._controller.start_download(com_port, self.firmware_path, mode)

    def _on_ctrl_download_started(self, file_size):
        self.download_btn.setFileSize(file_size)
        self.download_btn.setStateWaiting()

    def _on_ctrl_download_state_changed(self, state_value):
        if state_value in (DownloadState.WAITING_SYNC.value, DownloadState.SYNCING.value):
            if self.download_btn.state() != ProgressButton.STATE_WAITING:
                self.download_btn.setStateWaiting()
        elif state_value == DownloadState.PROGRAMMING.value:
            self.download_btn.setStateProgramming()

    def _on_ctrl_download_finished(self, result):
        if result.success:
            self.download_btn.setStateComplete()
        else:
            self.download_btn.setStateFailed()

    def _on_ctrl_download_error(self, err_msg):
        self.download_btn.setStateFailed()

    def _on_ctrl_download_cleaned(self):
        pass

    def _stop_download(self):
        self._controller.stop_download()
        self.download_btn.setStateFailed()
        self.append_log("[DOWNLOAD] Download stopped by user.")
        logger.info("Download stopped by user")

    def _on_chip_selected(self, index):
        if index <= 0:
            self.selected_chip_config = None
            self._clear_rail_config_edits()
            return
        chip_name = self.chip_combo.currentText()
        cfg = get_chip_config(chip_name)
        self.selected_chip_config = cfg
        if cfg:
            logger.info("Chip selected: %s", chip_name)
            self.append_log(f"[SYSTEM] Chip selected: {chip_name}")
        else:
            logger.warning("No config found for chip: %s", chip_name)
            self.append_log(f"[WARNING] No config found for chip: {chip_name}")
        self._load_rail_configs_from_chip(chip_name)

    def _clear_rail_config_edits(self):
        """清空所有电源 YAML 文本框。"""
        edits = getattr(self, "_rail_config_edits", {}) or {}
        for rail, edit in edits.items():
            edit.clear()

    def _load_rail_configs_from_chip(self, chip_name):
        """根据所选芯片, 从 main_chip_configs/<chip>.yaml 加载 5 个电源轨配置。

        YAML 文件顶层 key 与 _RAIL_NAMES(Vcore/VcoreM/VcoreL/VANA/VHPPA)
        大小写不敏感匹配; 每个 key 的值应为命令字符串列表, 会用换行拼接后
        填入对应的电源 YAML 文本框。
        """
        edits = getattr(self, "_rail_config_edits", {}) or {}
        if not edits:
            return
        # 先清空
        for edit in edits.values():
            edit.clear()

        if not chip_name:
            return

        yaml_path = os.path.join(_MAIN_CHIP_CONFIGS_DIR, f"{chip_name}.yaml")
        if not os.path.isfile(yaml_path):
            self.append_log(f"[CONFIG] No saved YAML for {chip_name} (file not found: {chip_name}.yaml).")
            return

        if yaml is None:
            self.append_log("[WARNING] PyYAML not installed; cannot parse chip rail configs.")
            return

        try:
            with open(yaml_path, "r", encoding="utf-8") as f:
                parsed = yaml.safe_load(f.read()) or {}
        except Exception as e:
            logger.warning("Failed to parse chip YAML %s: %s", yaml_path, e)
            self.append_log(f"[WARNING] Failed to parse chip YAML for {chip_name}: {e}")
            return

        if not isinstance(parsed, dict):
            self.append_log(f"[CONFIG] {chip_name}.yaml has no top-level dict, rail edits left empty.")
            return

        # 构造大小写不敏感的 key → rail 映射
        lower_to_rail = {rail.lower(): rail for rail in self._RAIL_NAMES}
        loaded_rails = []
        for key, val in parsed.items():
            rail = lower_to_rail.get(str(key).strip().lower())
            if rail is None:
                continue
            if isinstance(val, list):
                text = "\n".join(str(line) for line in val)
            elif isinstance(val, str):
                text = val
            else:
                text = str(val)
            edits[rail].setPlainText(text)
            loaded_rails.append(rail)

        total = len(self._RAIL_NAMES)
        loaded_count = len(loaded_rails)
        if loaded_count > 0:
            self.append_log(
                f"[CONFIG] Auto-loaded {loaded_count}/{total} rail configs from {chip_name}.yaml: "
                f"{', '.join(loaded_rails)}"
            )
        else:
            # 没有匹配到任何 rail key, 列出 yaml 中的实际 key 供用户判断
            available_keys = list(parsed.keys())
            keys_str = ", ".join(str(k) for k in available_keys) if available_keys else "(empty)"
            self.append_log(
                f"[CONFIG] {chip_name}.yaml exists but has no rail-name keys "
                f"(expected: {', '.join(self._RAIL_NAMES)}; found: {keys_str}). "
                f"Rail edits left empty."
            )

    def _get_rail_config_text(self, rail_name):
        """获取指定电源轨的 YAML 配置文本。"""
        edits = getattr(self, "_rail_config_edits", {}) or {}
        edit = edits.get(rail_name)
        if edit is None:
            return ""
        return edit.toPlainText().strip()

    def _get_combined_rail_config_text(self):
        """合并所有非空电源轨配置(用换行分隔),用于 Exec 立即执行。"""
        parts = []
        for rail in self._RAIL_NAMES:
            text = self._get_rail_config_text(rail)
            if text:
                parts.append(text)
        return "\n".join(parts)

    def _build_config_text_for_standard_mode(self, enabled_configs):
        """标准电压模式: 根据启用的通道 Name 匹配电源配置, 合并成 config_text。

        仅包含与启用通道 Name 匹配的电源轨配置(Vcore 通道 → Vcore 配置)。
        匹配规则: 通道 Name 与电源轨名称大小写不敏感完全匹配。
        """
        matched_rails = set()
        for _i, cfg in enabled_configs:
            name = cfg.get("name", "").strip().lower()
            for rail in self._RAIL_NAMES:
                if rail.lower() == name:
                    matched_rails.add(rail)
                    break
        parts = []
        for rail in self._RAIL_NAMES:
            if rail in matched_rails:
                text = self._get_rail_config_text(rail)
                if text:
                    parts.append(text)
        return "\n".join(parts)

    def _on_chip_check(self):
        if self._chip_check_thread is not None and self._chip_check_thread.isRunning():
            self.append_log("[WARNING] Chip check already in progress.")
            return

        self.chip_check_btn.setEnabled(False)
        self.append_log("[SYSTEM] Starting chip check via I2C...")

        worker = _ChipCheckWorker()
        thread = QThread()
        worker.moveToThread(thread)

        thread.started.connect(worker.run)
        worker.finished.connect(self._on_chip_check_finished)
        worker.error.connect(self._on_chip_check_error)
        worker.finished.connect(thread.quit)
        worker.error.connect(thread.quit)
        thread.finished.connect(self._on_chip_check_thread_cleaned)
        thread.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)

        self._chip_check_thread = thread
        self._chip_check_worker = worker
        thread.start()

    def _on_chip_check_finished(self, chip_info):
        self.chip_check_btn.setEnabled(True)

        self.append_log(
            f"[CHIP_CHECK] chip={chip_info.get('chip_name') or 'N/A'}"
            f"  main_die={chip_info.get('main_die') or 'N/A'}({chip_info.get('main_die_version') or '?'}, addr={chip_info.get('main_die_i2c_addr') or 'N/A'}, {chip_info.get('main_die_i2c_width') or 'N/A'}bit)"
            f"  main_die_pmu={chip_info.get('main_die_pmu') or 'N/A'}(addr={chip_info.get('main_die_pmu_i2c_addr') or 'N/A'}, {chip_info.get('main_die_pmu_i2c_width') or 'N/A'}bit)"
            f"  has_pmu={chip_info.get('has_pmu', False)}"
            f"  pmu={chip_info.get('pmu') or 'N/A'}({chip_info.get('pmu_version') or '?'}, addr={chip_info.get('pmu_i2c_addr') or 'N/A'}, {chip_info.get('pmu_i2c_width') or 'N/A'}bit)"
        )

        warning = chip_info.get("warning")
        if warning:
            self.append_log(f"[CHIP_CHECK] ⚠ {warning}")

        detected_name = chip_info.get("chip_name")
        if not detected_name:
            self.append_log("[WARNING] Chip check: no chip detected.")
            return

        exact_idx = self.chip_combo.findText(detected_name, Qt.MatchExactly)
        if exact_idx >= 0:
            self.chip_combo.setCurrentIndex(exact_idx)
            self.append_log(f"[CHIP_CHECK] Chip matched: {detected_name}")
            return

        prefix_match = detected_name.split("_")[0] if "_" in detected_name else detected_name
        for i in range(1, self.chip_combo.count()):
            item = self.chip_combo.itemText(i)
            if item == prefix_match or item.startswith(detected_name):
                self.chip_combo.setCurrentIndex(i)
                self.append_log(f"[CHIP_CHECK] Chip matched (prefix): {item}")
                return

        self.append_log(f"[WARNING] No matching chip found in list for: {detected_name}")

    def _on_chip_check_error(self, err_msg):
        self.chip_check_btn.setEnabled(True)
        logger.error("Chip check error: %s", err_msg)
        self.append_log(f"[ERROR] Chip check failed: {err_msg}")

    def _on_chip_check_thread_cleaned(self):
        self._chip_check_worker = None
        self._chip_check_thread = None

    def _on_chip_save(self):
        """把当前 5 个 YAML 文本框内容一次性写入 <chip>.yaml 文件。

        会保留文件中已有的非 rail-name 顶层 key(例如 voltage_low/voltage_default
        等模板配置),仅按 _RAIL_NAMES 逐轨覆盖/新增对应的 rail key。
        空 YAML 文本框也会写入空列表,等效于清空对应 rail key。
        """
        if self.chip_combo.currentIndex() <= 0:
            self.append_log("[WARNING] No chip selected. Please select a chip before saving.")
            return
        chip_name = self.chip_combo.currentText()

        if yaml is None:
            self.append_log("[WARNING] PyYAML not installed; cannot save rail configs.")
            return

        yaml_path = os.path.join(_MAIN_CHIP_CONFIGS_DIR, f"{chip_name}.yaml")

        # 读取已有内容(保留非 rail key)
        try:
            if os.path.isfile(yaml_path):
                with open(yaml_path, "r", encoding="utf-8") as f:
                    existing_text = f.read()
                parsed = yaml.safe_load(existing_text) or {}
                if not isinstance(parsed, dict):
                    parsed = {}
            else:
                parsed = {}
        except Exception as e:
            logger.warning("Failed to read existing YAML %s, will overwrite: %s", yaml_path, e)
            parsed = {}

        # 逐轨写入(空文本写入空列表)
        saved_rails = []
        for rail in self._RAIL_NAMES:
            text = self._get_rail_config_text(rail)
            config_lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
            parsed[rail] = config_lines
            if config_lines:
                saved_rails.append(f"{rail}({len(config_lines)})")

        try:
            os.makedirs(os.path.dirname(yaml_path), exist_ok=True)
            with open(yaml_path, "w", encoding="utf-8") as f:
                yaml.safe_dump(
                    parsed, f,
                    default_flow_style=False,
                    allow_unicode=True,
                    sort_keys=False,
                )
            saved_summary = ", ".join(saved_rails) if saved_rails else "(all empty)"
            logger.info("All rail configs saved to %s", yaml_path)
            self.append_log(
                f"[CONFIG] Saved 5 rail configs to {chip_name}.yaml: {saved_summary}"
            )
        except Exception as e:
            logger.error("Failed to write chip YAML %s: %s", yaml_path, e)
            self.append_log(f"[ERROR] Failed to save chip config: {e}")

    def _execute_rail_configuration(self, rail_name):
        """执行单个电源轨的 I2C 配置(跳过其它 rail)。

        仅根据指定 rail 的 YAML 文本框内容解析命令并下发 I2C,
        不会触碰其它 rail 的配置,也不会修改 N6705C 通道设置。
        """
        if rail_name not in self._RAIL_NAMES:
            logger.warning("Unknown rail name: %s", rail_name)
            return

        chip_name = self.chip_combo.currentText()
        if self.chip_combo.currentIndex() <= 0 or self.selected_chip_config is None:
            logger.warning("No chip selected for %s execution", rail_name)
            self.append_log(f"[WARNING] No chip selected. Please select a chip first for {rail_name}.")
            return

        refreshed = get_chip_config(chip_name, force_reload=True)
        if refreshed:
            self.selected_chip_config = refreshed

        config_text = self._get_rail_config_text(rail_name)
        if not config_text:
            self.append_log(f"[WARNING] {rail_name} config is empty, nothing to execute.")
            return

        config_commands = self._parse_config_commands(config_text)
        if not config_commands:
            self.append_log(f"[WARNING] {rail_name} config has no valid commands.")
            return

        self.append_log(
            f"[EXECUTE] Starting {rail_name} configuration for chip: {chip_name} "
            f"({len(config_commands)} commands)"
        )

        try:
            from lib.i2c.i2c_interface_x64 import I2CInterface
            i2c = I2CInterface()
            if not i2c.initialize():
                self.append_log("[ERROR] I2C interface initialization failed.")
                return
            self.append_log("[EXECUTE] I2C interface initialized successfully.")
        except Exception as e:
            logger.error("I2C initialization error: %s", e)
            self.append_log(f"[ERROR] I2C initialization error: {e}")
            return

        try:
            chip_info = i2c.bes_chip_check()
            self.append_log(f"[EXECUTE] Chip detected: {chip_info.get('chip_name', 'N/A')}")
        except Exception as e:
            logger.error("bes_chip_check failed: %s", e)
            self.append_log(f"[ERROR] Chip check failed: {e}")
            return

        self._compare_chip_info(chip_info, self.selected_chip_config)
        self._run_config_commands(i2c, chip_info, config_commands)
        self.append_log(f"[EXECUTE] {rail_name} configuration execution completed.")

    def _update_chip_rail_yaml(self, chip_name, rail_name, config_text):
        """将单轨配置写入 <chip>.yaml 文件中对应 rail 的 key 下。

        若文件不存在则创建; 若存在则只更新该 rail key, 其它 rail key 保留。
        """
        if rail_name not in self._RAIL_NAMES:
            return
        if yaml is None:
            self.append_log("[WARNING] PyYAML not installed; cannot save rail config.")
            return

        # 解析为命令行列表(过滤空行)
        config_lines = []
        for raw_line in config_text.strip().splitlines():
            line = raw_line.strip()
            if line:
                config_lines.append(line)

        yaml_path = os.path.join(_MAIN_CHIP_CONFIGS_DIR, f"{chip_name}.yaml")

        # 读取已有内容(若文件存在)
        try:
            if os.path.isfile(yaml_path):
                with open(yaml_path, "r", encoding="utf-8") as f:
                    existing_text = f.read()
                parsed = yaml.safe_load(existing_text) or {}
                if not isinstance(parsed, dict):
                    parsed = {}
            else:
                parsed = {}
        except Exception as e:
            logger.warning("Failed to read existing YAML %s, will overwrite: %s", yaml_path, e)
            parsed = {}

        # 仅更新该 rail 的 key
        parsed[rail_name] = config_lines

        try:
            os.makedirs(os.path.dirname(yaml_path), exist_ok=True)
            with open(yaml_path, "w", encoding="utf-8") as f:
                yaml.safe_dump(
                    parsed, f,
                    default_flow_style=False,
                    allow_unicode=True,
                    sort_keys=False,
                )
            logger.info("Rail %s config saved to %s", rail_name, yaml_path)
            self.append_log(
                f"[SYSTEM] {rail_name} config saved to {chip_name}.yaml ({len(config_lines)} lines)"
            )
        except Exception as e:
            logger.error("Failed to write rail YAML %s: %s", yaml_path, e)
            self.append_log(f"[ERROR] Failed to save {rail_name} config: {e}")

    def _compare_chip_info(self, detected, config):
        compare_keys = [
            "chip_name", "main_die", "main_die_version",
            "main_die_i2c_width", "main_die_i2c_addr",
            "main_die_pmu", "main_die_pmu_version",
            "main_die_pmu_i2c_width", "main_die_pmu_i2c_addr",
            "has_pmu", "pmu", "pmu_version", "pmu_i2c_width", "pmu_i2c_addr",
        ]
        for key in compare_keys:
            det_val = detected.get(key)
            cfg_val = config.get(key)
            if cfg_val in (None, "", {}):
                continue
            if self._chip_values_equal(det_val, cfg_val):
                continue
            logger.warning(
                "Chip info mismatch [%s]: detected=%s, config=%s",
                key, det_val, cfg_val
            )
            self.append_log(
                f"[WARNING] Chip info mismatch [{key}]: detected={det_val}, config={cfg_val}"
            )

    @staticmethod
    def _chip_values_equal(a, b):
        if a == b:
            return True
        if a is None or b is None:
            return False
        str_a = str(a).strip().lower()
        str_b = str(b).strip().lower()
        if str_a == str_b:
            return True
        try:
            return int(str_a, 0) == int(str_b, 0)
        except (ValueError, TypeError):
            pass
        if isinstance(a, bool) or isinstance(b, bool):
            truthy = {"true", "1", "yes"}
            falsy = {"false", "0", "no", ""}
            a_bool = str_a in truthy
            b_bool = str_b in truthy
            a_is_bool = str_a in truthy or str_a in falsy
            b_is_bool = str_b in truthy or str_b in falsy
            if a_is_bool and b_is_bool:
                return a_bool == b_bool
        return False

    @staticmethod
    def _parse_config_commands(text):
        commands = []
        lines = text.strip().splitlines()
        for raw_line in lines:
            line = raw_line.strip()
            if line.startswith("-"):
                line = line[1:].strip()
            if line.startswith("'") or line.startswith('"'):
                line = line[1:]
            if line.endswith("'") or line.endswith('"'):
                line = line[:-1]
            line = line.strip()

            comment_idx = line.find("//")
            if comment_idx >= 0:
                line = line[:comment_idx].strip()

            if not line:
                continue

            upper = line.upper()
            if not any(kw in upper for kw in ("WRITE_BITS", "WRITE", "READ")):
                continue

            target = "NO_PREFIX"
            if ":" in line:
                prefix, rest = line.split(":", 1)
                prefix_upper = prefix.strip().upper()
                rest_upper = rest.strip().upper()
                has_command = any(kw in rest_upper for kw in ("WRITE_BITS", "WRITE", "READ"))
                if has_command:
                    if prefix_upper == "DUT":
                        target = "DUT"
                        line = rest.strip()
                    elif prefix_upper == "PMU":
                        target = "MAIN_DIE_PMU"
                        line = rest.strip()
                    elif prefix_upper == "MAIN_DIE_PMU":
                        target = "MAIN_DIE_PMU"
                        line = rest.strip()
                    elif prefix_upper.endswith("_PMU"):
                        target = "EXT_PMU"
                        line = rest.strip()
                    elif prefix_upper.endswith("_DUT") or prefix_upper.endswith("_MAIN"):
                        target = "DUT"
                        line = rest.strip()

            parts = line.split()
            if len(parts) < 2:
                continue

            op = parts[0].upper()
            if op == "WRITE_BITS" and len(parts) >= 5:
                reg_addr = int(parts[1], 0)
                msb = int(parts[2], 0)
                lsb = int(parts[3], 0)
                value = int(parts[4], 0)
                commands.append({
                    "op": "WRITE_BITS",
                    "target": target,
                    "reg_addr": reg_addr,
                    "msb": msb,
                    "lsb": lsb,
                    "value": value,
                })
            elif op == "WRITE" and len(parts) >= 3:
                reg_addr = int(parts[1], 0)
                value = int(parts[2], 0)
                commands.append({
                    "op": "WRITE",
                    "target": target,
                    "reg_addr": reg_addr,
                    "value": value,
                })
            elif op == "READ" and len(parts) >= 2:
                reg_addr = int(parts[1], 0)
                commands.append({
                    "op": "READ",
                    "target": target,
                    "reg_addr": reg_addr,
                })

        return commands

    @staticmethod
    def _to_int_addr(addr):
        if addr is None:
            return None
        if isinstance(addr, int):
            return addr
        if isinstance(addr, str):
            return int(addr, 0)
        return None

    def _resolve_device(self, chip_info, target):
        if target == "DUT":
            addr = self._to_int_addr(chip_info.get("main_die_i2c_addr"))
            width = chip_info.get("main_die_i2c_width")
            return addr, width
        if target == "EXT_PMU":
            addr = self._to_int_addr(chip_info.get("pmu_i2c_addr"))
            width = chip_info.get("pmu_i2c_width")
            return addr, width
        if target == "MAIN_DIE_PMU":
            addr = self._to_int_addr(chip_info.get("main_die_pmu_i2c_addr"))
            width = chip_info.get("main_die_pmu_i2c_width")
            return addr, width
        if chip_info.get("has_pmu") and chip_info.get("pmu_i2c_addr"):
            addr = self._to_int_addr(chip_info.get("pmu_i2c_addr"))
            width = chip_info.get("pmu_i2c_width")
        else:
            addr = self._to_int_addr(chip_info.get("main_die_pmu_i2c_addr"))
            width = chip_info.get("main_die_pmu_i2c_width")
        return addr, width

    def _run_config_commands(self, i2c, chip_info, commands):
        for idx, cmd in enumerate(commands):
            op = cmd["op"]
            target = cmd.get("target", "NO_PREFIX")
            reg_addr = cmd["reg_addr"]

            device_addr, width = self._resolve_device(chip_info, target)
            if device_addr is None or width is None:
                self.append_log(
                    f"[ERROR] Cannot resolve device address for target={target}, skipping command #{idx+1}"
                )
                continue

            try:
                if op == "WRITE_BITS":
                    msb = cmd["msb"]
                    lsb = cmd["lsb"]
                    value = cmd["value"]
                    current_val = i2c.read(device_addr, reg_addr, width)
                    bit_mask = ((1 << (msb - lsb + 1)) - 1) << lsb
                    new_val = (current_val & ~bit_mask) | ((value << lsb) & bit_mask)
                    i2c.write(device_addr, reg_addr, new_val, width)
                    self.append_log(
                        f"[EXECUTE] #{idx+1} WRITE_BITS dev=0x{device_addr:02X} "
                        f"reg=0x{reg_addr:08X} [{msb}:{lsb}]=0x{value:X} "
                        f"(0x{current_val:X} -> 0x{new_val:X})"
                    )

                elif op == "WRITE":
                    value = cmd["value"]
                    i2c.write(device_addr, reg_addr, value, width)
                    self.append_log(
                        f"[EXECUTE] #{idx+1} WRITE dev=0x{device_addr:02X} "
                        f"reg=0x{reg_addr:08X} data=0x{value:X}"
                    )

                elif op == "READ":
                    read_val = i2c.read(device_addr, reg_addr, width)
                    self.append_log(
                        f"[EXECUTE] #{idx+1} READ dev=0x{device_addr:02X} "
                        f"reg=0x{reg_addr:08X} => 0x{read_val:X}"
                    )

            except Exception as e:
                logger.error("Command #%d failed: %s", idx + 1, e)
                self.append_log(f"[ERROR] Command #{idx+1} failed: {e}")

    def _update_chip_config_file(self, chip_name, config_text):
        try:
            config_lines = []
            for raw_line in config_text.strip().splitlines():
                line = raw_line.strip()
                if line:
                    config_lines.append(line)

            chips_dir = os.path.join(
                get_resource_base(),
                "chips", "bes_chip_configs"
            )
            if chip_name.startswith("pmu_"):
                config_file = os.path.join(chips_dir, "pmu_chips", f"{chip_name}.py")
            else:
                config_file = os.path.join(chips_dir, "main_chips", f"{chip_name}.py")

            if not os.path.exists(config_file):
                logger.warning("Chip config file not found: %s", config_file)
                self.append_log(f"[WARNING] Chip config file not found: {config_file}")
                return

            with open(config_file, "r", encoding="utf-8") as f:
                content = f.read()

            import ast
            tree = ast.parse(content)
            chip_config_dict = None
            for node in ast.walk(tree):
                if isinstance(node, ast.Assign):
                    for t in node.targets:
                        if isinstance(t, ast.Name) and t.id == "CHIP_CONFIG":
                            chip_config_dict = ast.literal_eval(content[node.value.col_offset:].split("\n}")[0] + "\n}")
                            break

            if chip_config_dict is None:
                chip_config_dict = {}

            chip_config_dict["power_distribution"] = {"user_config": config_lines}

            lines = ["CHIP_CONFIG = {\n"]
            for key, val in chip_config_dict.items():
                lines.append(f"    {key!r}: {val!r},\n")
            lines.append("}\n")

            with open(config_file, "w", encoding="utf-8") as f:
                f.writelines(lines)

            logger.info("Chip config updated: %s", config_file)
            self.append_log(f"[SYSTEM] Chip config updated: {chip_name}")

            refreshed = get_chip_config(chip_name, force_reload=True)
            if refreshed:
                self.selected_chip_config = refreshed
        except Exception as e:
            logger.error("Failed to update chip config: %s", e)
            self.append_log(f"[ERROR] Failed to update chip config: {e}")

    def _parse_channel_key(self, channel_key):
        m = re.match(r'^([AB])-CH(\d+)$', channel_key)
        if m:
            return m.group(1), int(m.group(2))
        m = re.match(r'^GPIO(\d+)$', channel_key, re.IGNORECASE)
        if m:
            return "MCU", int(m.group(1))
        return None, None

    def _build_channel_force_configs(self):
        """构建每通道的 force 配置字典。

        返回:
            {(device_label, hw_ch): {
                "force_mode": "force"/"auto",
                "force_value": float,        # force 模式下的目标电压(V)
                "boost_mode": "constant"/"percent",
                "boost_value": float,        # auto 模式下的增压值
            }}
            Force 模式且 force_value 为空/非法时跳过该通道(走 Auto 默认值)。
            Auto 模式下 boost_value 为空/非法时使用默认 0.02。
        """
        configs = {}
        for cfg in self._channel_configs:
            if not cfg["enabled"]:
                continue
            device_label, hw_ch = self._parse_channel_key(cfg["channel"])
            if device_label is None or hw_ch is None:
                continue

            force_mode = cfg.get("force_mode", "auto")
            entry = {
                "force_mode": force_mode,
                "force_value": None,
                "boost_mode": cfg.get("boost_mode", "constant"),
                "boost_value": 0.02,
            }

            if force_mode == "force":
                val_text = cfg.get("force_value", "").strip()
                if not val_text:
                    continue
                try:
                    entry["force_value"] = float(val_text)
                except ValueError:
                    continue
            else:
                val_text = cfg.get("boost_value", "").strip()
                if val_text:
                    try:
                        entry["boost_value"] = float(val_text)
                    except ValueError:
                        pass

            configs[(device_label, hw_ch)] = entry
        return configs

    def _current_mcu_type(self):
        if hasattr(self, "mcu_type_combo") and self.mcu_type_combo is not None:
            data = self.mcu_type_combo.currentData()
            if data in ("yd_rp2040", "ch9114f"):
                return data
        return getattr(self, "_default_mcu_type", "ch9114f")

    def _get_mcu_gpio_options(self):
        if self._current_mcu_type() == "ch9114f":
            return [f"GPIO{i}" for i in (0, 1, 6, 7, 2, 8, 14, 20)]
        return [f"GPIO{i}" for i in range(0, 30)]

    def _get_control_channel_options(self, method=None):
        method = method or (
            self.control_method_toggle.value()
            if getattr(self, "control_method_toggle", None) else "N6705C"
        )
        if method == "MCU":
            return self._get_mcu_gpio_options()
        return self._get_available_channel_options()

    @staticmethod
    def _set_combo_options(combo, options, desired=None):
        if combo is None:
            return False
        if desired is None:
            desired = combo.currentText()
        combo.blockSignals(True)
        combo.clear()
        for opt in options:
            combo.addItem(opt)
        matched = False
        if desired:
            for i in range(combo.count()):
                if combo.itemText(i) == desired:
                    combo.setCurrentIndex(i)
                    matched = True
                    break
        if not matched and combo.count() > 0:
            combo.setCurrentIndex(0)
        combo.blockSignals(False)
        return matched

    def _selected_mcu_port(self):
        text = self.mcu_port_combo.currentText() if getattr(self, "mcu_port_combo", None) else ""
        if not text or text in (
            "No serial ports found",
            "No CH9114F ports found",
            "Select MCU COM...",
            "Select CH9114F COM...",
        ):
            return None
        return text.split()[0]

    def _on_mcu_search(self):
        from debug_config import DEBUG_MOCK
        mcu_type = self._current_mcu_type()
        if DEBUG_MOCK:
            self.mcu_port_combo.clear()
            if mcu_type == "ch9114f":
                self.mcu_port_combo.addItem("[MOCK] COM99 - Mock CH9114F")
            else:
                self.mcu_port_combo.addItem("[MOCK] COM98 - Mock YD RP2040")
            self.mcu_status_label.setText("● Mock Ready")
            self.mcu_status_label.setStyleSheet(
                "color: #ff9800; font-size: 10px; font-weight: bold; background: transparent; border: none;"
            )
            self.mcu_connect_btn.setEnabled(True)
            self.append_log(f"[DEBUG] Mock {mcu_type} port loaded.")
            return

        if self._mcu_search_thread is not None and self._mcu_search_thread.isRunning():
            return
        self.mcu_status_label.setText("● Searching")
        self.mcu_status_label.setStyleSheet(
            "color: #ff9800; font-size: 10px; font-weight: bold; background: transparent; border: none;"
        )
        self.mcu_search_btn.setEnabled(False)
        self.mcu_connect_btn.setEnabled(False)
        if mcu_type == "ch9114f":
            self.append_log("[MCU] Scanning for CH9114F ports...")
        else:
            self.append_log("[MCU] Scanning serial ports for YD RP2040...")

        worker = _SearchMcuPortWorker(mcu_type)
        thread = QThread()
        worker.moveToThread(thread)
        thread.started.connect(worker.run)
        worker.finished.connect(self._on_mcu_search_done)
        worker.error.connect(self._on_mcu_search_error)
        worker.finished.connect(thread.quit)
        worker.error.connect(thread.quit)
        thread.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)
        thread.finished.connect(lambda: self._on_mcu_search_thread_cleanup())
        self._mcu_search_worker = worker
        self._mcu_search_thread = thread
        thread.start()

    def _on_mcu_search_done(self, ports):
        mcu_type = self._current_mcu_type()
        self.mcu_port_combo.clear()
        self.mcu_port_combo.setEnabled(True)
        if ports:
            for port in ports:
                self.mcu_port_combo.addItem(port)
            self.mcu_status_label.setText(f"● Found {len(ports)}")
            self.mcu_status_label.setStyleSheet(
                "color: #00a859; font-size: 10px; font-weight: bold; background: transparent; border: none;"
            )
            label = "CH9114F port(s)" if mcu_type == "ch9114f" else "serial port(s)"
            self.append_log(f"[MCU] Found {len(ports)} {label}.")
        else:
            if mcu_type == "ch9114f":
                self.mcu_port_combo.addItem("No CH9114F ports found")
            else:
                self.mcu_port_combo.addItem("No serial ports found")
            self.mcu_port_combo.setEnabled(False)
            self.mcu_status_label.setText("● Not Found")
            self.mcu_status_label.setStyleSheet(
                "color: #e53935; font-size: 10px; font-weight: bold; background: transparent; border: none;"
            )
        self.mcu_search_btn.setEnabled(True)
        self.mcu_connect_btn.setEnabled(bool(ports))

    def _on_mcu_search_error(self, err):
        self.mcu_status_label.setText("● Search Failed")
        self.mcu_status_label.setStyleSheet(
            "color: #e53935; font-size: 10px; font-weight: bold; background: transparent; border: none;"
        )
        self.append_log(f"[MCU] Search failed: {err}")
        self.mcu_search_btn.setEnabled(True)
        self.mcu_connect_btn.setEnabled(True)

    def _on_mcu_search_thread_cleanup(self):
        self._mcu_search_thread = None
        self._mcu_search_worker = None

    def _on_mcu_connect_or_disconnect(self):
        if self.is_mcu_connected:
            self._disconnect_mcu()
        else:
            self._connect_mcu()

    def _connect_mcu(self):
        port = self._selected_mcu_port()
        if not port:
            self.append_log("[MCU] No valid MCU port selected.")
            return
        mcu_type = self._current_mcu_type()
        if self._mcu_connect_thread is not None and self._mcu_connect_thread.isRunning():
            return
        self.mcu_status_label.setText("● Connecting")
        self.mcu_status_label.setStyleSheet(
            "color: #ff9800; font-size: 10px; font-weight: bold; background: transparent; border: none;"
        )
        self.mcu_search_btn.setEnabled(False)
        self.mcu_connect_btn.setEnabled(False)
        type_label = "CH9114F" if mcu_type == "ch9114f" else "YD RP2040"
        self.append_log(f"[MCU] Connecting {type_label} on {port}...")

        worker = _ConnectMcuWorker(port, mcu_type=mcu_type)
        thread = QThread()
        worker.moveToThread(thread)
        thread.started.connect(worker.run)
        worker.finished.connect(self._on_mcu_connected)
        worker.error.connect(self._on_mcu_connect_error)
        worker.finished.connect(thread.quit)
        worker.error.connect(thread.quit)
        thread.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)
        thread.finished.connect(lambda: self._on_mcu_connect_thread_cleanup())
        self._mcu_connect_worker = worker
        self._mcu_connect_thread = thread
        thread.start()

    def _on_mcu_connected(self, inst, idn):
        self.mcu_io = inst
        self.is_mcu_connected = True
        self.mcu_status_label.setText("● Connected")
        self.mcu_status_label.setStyleSheet(
            "color: #00a859; font-size: 10px; font-weight: bold; background: transparent; border: none;"
        )
        self.mcu_search_btn.setEnabled(False)
        self.mcu_connect_btn.setEnabled(True)
        self.mcu_connect_btn.setText("Disconnect")
        self.append_log(f"[MCU] Connected: {idn}")
        self._suppress_preset_channels = False

    def _on_mcu_connect_error(self, err):
        self.mcu_status_label.setText("● Failed")
        self.mcu_status_label.setStyleSheet(
            "color: #e53935; font-size: 10px; font-weight: bold; background: transparent; border: none;"
        )
        self.mcu_search_btn.setEnabled(True)
        self.mcu_connect_btn.setEnabled(True)
        self.append_log(f"[MCU] Connection failed: {err}")
        self._suppress_preset_channels = False

    def _on_mcu_connect_thread_cleanup(self):
        self._mcu_connect_thread = None
        self._mcu_connect_worker = None

    def _disconnect_mcu(self):
        try:
            if self.mcu_io is not None:
                self.mcu_io.disconnect()
        except Exception as e:
            logger.error("MCU disconnect failed: %s", e, exc_info=True)
            self.append_log(f"[MCU] Disconnect failed: {e}")
        finally:
            self.mcu_io = None
            self.is_mcu_connected = False
            self.mcu_status_label.setText("● Disconnected")
            self.mcu_status_label.setStyleSheet(
                "color: #8ea6cf; font-size: 10px; font-weight: bold; background: transparent; border: none;"
            )
            self.mcu_search_btn.setEnabled(True)
            self.mcu_connect_btn.setEnabled(True)
            self.mcu_connect_btn.setText("Connect")
            self.append_log("[MCU] Disconnected.")

    def _on_start_test(self):
        self._start_test()

    def _start_test(self):
        if self.is_testing:
            return

        enabled_configs = [
            (i, cfg) for i, cfg in enumerate(self._channel_configs) if cfg["enabled"]
        ]
        if not enabled_configs:
            self.append_log("[ERROR] No channel enabled.")
            return

        vbat_idx = None
        vbat_cfg = None
        for i, cfg in enabled_configs:
            if cfg["name"].lower().startswith("vbat"):
                vbat_idx = i
                vbat_cfg = cfg
                break
        if vbat_cfg is None:
            vbat_idx, vbat_cfg = enabled_configs[0]

        vbat_device_label, vbat_hw_ch = self._parse_channel_key(vbat_cfg["channel"])
        if vbat_device_label is None or vbat_hw_ch is None:
            self.append_log(f"[ERROR] Invalid Vbat channel key: {vbat_cfg['channel']}")
            return

        vbat_attr = vbat_device_label.lower()
        vbat_inst = getattr(self, f"n6705c_{vbat_attr}", None)
        vbat_conn = getattr(self, f"is_connected_{vbat_attr}", False)
        if not vbat_conn or not vbat_inst:
            self.append_log(f"[ERROR] N6705C-{vbat_device_label} is not connected (required by Vbat).")
            return

        force_high_map = {}
        config_index_map = {vbat_device_label: {vbat_hw_ch: vbat_idx}}

        sub_configs = [(i, cfg) for i, cfg in enabled_configs if i != vbat_idx]
        for i, cfg in sub_configs:
            device_label, hw_ch = self._parse_channel_key(cfg["channel"])
            if device_label is None or hw_ch is None:
                self.append_log(f"[ERROR] Invalid channel key: {cfg['channel']}")
                return
            attr = device_label.lower()
            inst = getattr(self, f"n6705c_{attr}", None)
            is_conn = getattr(self, f"is_connected_{attr}", False)
            if not is_conn or not inst:
                self.append_log(f"[ERROR] N6705C-{device_label} is not connected (required by {cfg['name']}).")
                return
            if device_label not in force_high_map:
                force_high_map[device_label] = (inst, [])
            if hw_ch not in force_high_map[device_label][1]:
                force_high_map[device_label][1].append(hw_ch)
            config_index_map.setdefault(device_label, {})[hw_ch] = i

        try:
            test_time = float(self.test_time_input.text())
        except ValueError:
            self.append_log("[ERROR] Invalid test time.")
            return
        sample_period = 20.0 / 1_000_000

        self.is_testing = True
        self._config_index_map = config_index_map
        self._current_total_bins = 0
        self.bin_result_table.hide()
        self.start_test_btn.setStateWaiting()
        self.append_log(
            f"[TEST] Starting force-high consumption test: "
            f"Vbat={vbat_cfg['name']}({vbat_cfg['channel']}), "
            f"time={test_time}s, base_period={sample_period*1e6:.0f}us"
        )

        for idx in self.channel_cards:
            self.channel_cards[idx]["value_label"].setText("- - -")
        if self._vbat_remain_card is not None:
            self._vbat_remain_card["value_label"].setText("- - -")

        channel_names = {}
        channel_names[(vbat_device_label, vbat_hw_ch)] = vbat_cfg["name"]
        for i, cfg in sub_configs:
            device_label, hw_ch = self._parse_channel_key(cfg["channel"])
            if device_label is not None and hw_ch is not None:
                channel_names[(device_label, hw_ch)] = cfg["name"]

        worker = _ConsumptionTestForceHighWorker(
            vbat_device_label, vbat_inst, vbat_hw_ch,
            force_high_map, test_time, sample_period,
            channel_names=channel_names,
            channel_force_configs=self._build_channel_force_configs(),
        )
        thread = QThread()
        worker.moveToThread(thread)

        thread.started.connect(worker.run)
        worker.log_message.connect(self.append_log)
        worker.channel_result.connect(self._on_force_high_channel_result)
        worker.test_summary.connect(self._on_test_summary)
        worker.progress.connect(self.start_test_btn.setProgress)
        worker.error.connect(self._on_test_error)
        worker.finished.connect(self._on_test_finished)
        worker.finished.connect(thread.quit)
        thread.finished.connect(self._on_test_thread_cleaned)
        thread.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)

        self._test_thread = thread
        self._test_worker = worker
        self.start_test_btn.setStateProgramming()
        self.start_test_btn._progress_timer.stop()
        thread.start()

    def consumption_test_force(self):
        if self.is_testing:
            return

        enabled_configs = [
            (i, cfg) for i, cfg in enumerate(self._channel_configs) if cfg["enabled"]
        ]
        if not enabled_configs:
            self.append_log("[ERROR] No channel enabled.")
            return

        vbat_idx = None
        vbat_cfg = None
        for i, cfg in enabled_configs:
            if cfg["name"].lower().startswith("vbat"):
                vbat_idx = i
                vbat_cfg = cfg
                break
        if vbat_cfg is None:
            vbat_idx, vbat_cfg = enabled_configs[0]

        vbat_device_label, vbat_hw_ch = self._parse_channel_key(vbat_cfg["channel"])
        if vbat_device_label is None or vbat_hw_ch is None:
            self.append_log(f"[ERROR] Invalid Vbat channel key: {vbat_cfg['channel']}")
            return

        vbat_attr = vbat_device_label.lower()
        vbat_inst = getattr(self, f"n6705c_{vbat_attr}", None)
        vbat_conn = getattr(self, f"is_connected_{vbat_attr}", False)
        if not vbat_conn or not vbat_inst:
            self.append_log(f"[ERROR] N6705C-{vbat_device_label} is not connected (required by Vbat).")
            return

        force_map = {}
        config_index_map = {vbat_device_label: {vbat_hw_ch: vbat_idx}}

        sub_configs = [(i, cfg) for i, cfg in enabled_configs if i != vbat_idx]
        for i, cfg in sub_configs:
            device_label, hw_ch = self._parse_channel_key(cfg["channel"])
            if device_label is None or hw_ch is None:
                self.append_log(f"[ERROR] Invalid channel key: {cfg['channel']}")
                return
            attr = device_label.lower()
            inst = getattr(self, f"n6705c_{attr}", None)
            is_conn = getattr(self, f"is_connected_{attr}", False)
            if not is_conn or not inst:
                self.append_log(f"[ERROR] N6705C-{device_label} is not connected (required by {cfg['name']}).")
                return
            if device_label not in force_map:
                force_map[device_label] = (inst, [])
            if hw_ch not in force_map[device_label][1]:
                force_map[device_label][1].append(hw_ch)
            config_index_map.setdefault(device_label, {})[hw_ch] = i

        try:
            test_time = float(self.test_time_input.text())
        except ValueError:
            self.append_log("[ERROR] Invalid test time.")
            return
        sample_period = 20.0 / 1_000_000

        self.is_testing = True
        self._config_index_map = config_index_map
        self._current_total_bins = 0
        self.bin_result_table.hide()
        self.start_test_btn.setStateWaiting()
        self.append_log(
            f"[TEST] Starting force-auto consumption test: "
            f"Vbat={vbat_cfg['name']}({vbat_cfg['channel']}), "
            f"time={test_time}s, base_period={sample_period*1e6:.0f}us"
        )

        for idx in self.channel_cards:
            self.channel_cards[idx]["value_label"].setText("- - -")
        if self._vbat_remain_card is not None:
            self._vbat_remain_card["value_label"].setText("- - -")

        channel_names = {}
        channel_names[(vbat_device_label, vbat_hw_ch)] = vbat_cfg["name"]
        for i, cfg in sub_configs:
            device_label, hw_ch = self._parse_channel_key(cfg["channel"])
            if device_label is not None and hw_ch is not None:
                channel_names[(device_label, hw_ch)] = cfg["name"]

        worker = _ConsumptionTestForceWorker(
            vbat_device_label, vbat_inst, vbat_hw_ch,
            force_map, test_time, sample_period,
            channel_names=channel_names,
            channel_force_configs=self._build_channel_force_configs(),
        )
        thread = QThread()
        worker.moveToThread(thread)

        thread.started.connect(worker.run)
        worker.log_message.connect(self.append_log)
        worker.channel_result.connect(self._on_force_high_channel_result)
        worker.test_summary.connect(self._on_test_summary)
        worker.progress.connect(self.start_test_btn.setProgress)
        worker.error.connect(self._on_test_error)
        worker.finished.connect(self._on_test_finished)
        worker.finished.connect(thread.quit)
        thread.finished.connect(self._on_test_thread_cleaned)
        thread.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)

        self._test_thread = thread
        self._test_worker = worker
        self.start_test_btn.setStateProgramming()
        self.start_test_btn._progress_timer.stop()
        thread.start()

    def _on_force_high_channel_result(self, device_label, hw_channel, avg_current, phase):
        cfg_idx = self._config_index_map.get(device_label, {}).get(hw_channel)
        if cfg_idx is not None and cfg_idx in self.channel_cards:
            label = self.channel_cards[cfg_idx]["value_label"]
            label.setText(self._format_current(avg_current))

    def _on_test_summary(self, summary):
        vbat_remain = summary.get("vbat_remain")
        if vbat_remain is not None and self._vbat_remain_card is not None:
            self._vbat_remain_card["value_label"].setText(self._format_current(vbat_remain))

        if self._current_total_bins > 1:
            self._bin_results_data.append(summary)
            self._add_bin_result_row(summary)

    def _on_test_error(self, err_msg):
        self.append_log(f"[ERROR] {err_msg}")

    def _on_test_finished(self):
        self.is_testing = False
        self.start_test_btn.setStateComplete()
        self.append_log("[TEST] Test completed.")

    def _on_test_thread_cleaned(self):
        self._test_worker = None
        self._test_thread = None

    def _stop_test(self):
        if self._test_worker:
            self._test_worker.stop()
        self.is_testing = False
        self.start_test_btn.setStateFailed()
        self.append_log("[TEST] Test stopped.")

    def _on_auto_test(self):
        if self.is_testing:
            self.append_log("[WARNING] A test is already running.")
            return

        firmware_paths = getattr(self, 'firmware_paths', [])
        if not firmware_paths:
            if self.firmware_path:
                firmware_paths = [self.firmware_path]
            else:
                self.append_log("[ERROR] No firmware file selected.")
                return

        port_text = self.get_selected_serial_port()
        if not port_text:
            self.append_log("[ERROR] No serial port selected.")
            return
        m = re.search(r'(\d+)', port_text)
        com_port = m.group(1) if m else port_text

        mode_str = self.download_mode_toggle.value().lower()
        download_mode = DownloadMode.FLASH if mode_str == "flash" else DownloadMode.RAMRUN

        enabled_configs = [
            (i, cfg) for i, cfg in enumerate(self._channel_configs) if cfg["enabled"]
        ]
        if not enabled_configs:
            self.append_log("[ERROR] No channel enabled.")
            return

        vbat_idx = None
        vbat_cfg = None
        for i, cfg in enabled_configs:
            if cfg["name"].lower().startswith("vbat"):
                vbat_idx = i
                vbat_cfg = cfg
                break
        if vbat_cfg is None:
            vbat_idx, vbat_cfg = enabled_configs[0]

        vbat_device_label, vbat_hw_ch = self._parse_channel_key(vbat_cfg["channel"])
        if vbat_device_label is None or vbat_hw_ch is None:
            self.append_log(f"[ERROR] Invalid Vbat channel key: {vbat_cfg['channel']}")
            return

        vbat_attr = vbat_device_label.lower()
        vbat_inst = getattr(self, f"n6705c_{vbat_attr}", None)
        vbat_conn = getattr(self, f"is_connected_{vbat_attr}", False)
        if not vbat_conn or not vbat_inst:
            self.append_log(f"[ERROR] N6705C-{vbat_device_label} is not connected (required by Vbat).")
            return

        poweron_key = self.poweron_channel_combo.currentText() if self.poweron_channel_combo else ""
        reset_enabled = (
            self.reset_enable_cb.isChecked()
            if getattr(self, "reset_enable_cb", None) is not None else False
        )
        control_method = (
            self.control_method_toggle.value()
            if getattr(self, "control_method_toggle", None) else "N6705C"
        )
        reset_key = self.reset_channel_combo.currentText() if self.reset_channel_combo else ""
        if not poweron_key:
            self.append_log("[ERROR] PowerON channel not configured.")
            return
        if reset_enabled and not reset_key:
            self.append_log("[ERROR] RESET channel not configured.")
            return

        poweron_dl, poweron_hw = self._parse_channel_key(poweron_key)
        if poweron_dl is None:
            self.append_log("[ERROR] Invalid PowerON channel key.")
            return
        if reset_enabled:
            reset_dl, reset_hw = self._parse_channel_key(reset_key)
            if reset_dl is None:
                self.append_log("[ERROR] Invalid RESET channel key.")
                return
        else:
            reset_dl, reset_hw = None, None

        if control_method == "MCU":
            if poweron_dl != "MCU" or (reset_enabled and reset_dl != "MCU"):
                self.append_log("[ERROR] MCU control requires GPIO channels for PowerON/RESET.")
                return
            if not self.is_mcu_connected or self.mcu_io is None:
                self.append_log("[ERROR] MCU IO is not connected (required by PowerON/RESET).")
                return
            poweron_inst = self.mcu_io
            poweron_dl = "MCU"
            if reset_enabled:
                reset_inst = self.mcu_io
                reset_dl = "MCU"
            else:
                reset_inst = None
        else:
            poweron_attr = poweron_dl.lower()
            poweron_inst = getattr(self, f"n6705c_{poweron_attr}", None)
            poweron_conn = getattr(self, f"is_connected_{poweron_attr}", False)
            if not poweron_conn or not poweron_inst:
                self.append_log(f"[ERROR] N6705C-{poweron_dl} is not connected (required by PowerON).")
                return

            if reset_enabled:
                reset_attr = reset_dl.lower()
                reset_inst = getattr(self, f"n6705c_{reset_attr}", None)
                reset_conn = getattr(self, f"is_connected_{reset_attr}", False)
                if not reset_conn or not reset_inst:
                    self.append_log(f"[ERROR] N6705C-{reset_dl} is not connected (required by RESET).")
                    return
            else:
                reset_inst = None

        poweron_polarity = self.poweron_polarity_toggle.value()
        reset_polarity = (
            self.reset_polarity_toggle.value() if reset_enabled else "rising"
        )

        force_map = {}
        config_index_map = {vbat_device_label: {vbat_hw_ch: vbat_idx}}
        sub_configs = [(i, cfg) for i, cfg in enabled_configs if i != vbat_idx]
        for i, cfg in sub_configs:
            device_label, hw_ch = self._parse_channel_key(cfg["channel"])
            if device_label is None or hw_ch is None:
                self.append_log(f"[ERROR] Invalid channel key: {cfg['channel']}")
                return
            attr = device_label.lower()
            inst = getattr(self, f"n6705c_{attr}", None)
            is_conn = getattr(self, f"is_connected_{attr}", False)
            if not is_conn or not inst:
                self.append_log(f"[ERROR] N6705C-{device_label} is not connected (required by {cfg['name']}).")
                return
            if device_label not in force_map:
                force_map[device_label] = (inst, [])
            if hw_ch not in force_map[device_label][1]:
                force_map[device_label][1].append(hw_ch)
            config_index_map.setdefault(device_label, {})[hw_ch] = i

        try:
            test_time = float(self.test_time_input.text())
        except ValueError:
            self.append_log("[ERROR] Invalid test time.")
            return
        try:
            stabilization_delay_sec = float(self.stable_delay_input.text())
        except ValueError:
            self.append_log("[ERROR] Invalid stable delay.")
            return
        try:
            download_pgm_rate = int(self.download_baudrate_input.text())
        except ValueError:
            self.append_log("[ERROR] Invalid baudrate.")
            return
        sample_period = 20.0 / 1_000_000

        channel_names = {}
        channel_names[(vbat_device_label, vbat_hw_ch)] = vbat_cfg["name"]
        for i, cfg in sub_configs:
            dl, hw = self._parse_channel_key(cfg["channel"])
            if dl is not None and hw is not None:
                channel_names[(dl, hw)] = cfg["name"]

        # 根据测试模式决定 config_text 来源:
        #   standard(标准电压) → 按启用通道 Name 匹配电源轨配置合并
        #   high_voltage(外供高压) → 不需要 I2C 配置, config_text 留空
        test_mode = getattr(self, "_test_mode", "high_voltage")
        if test_mode == "standard":
            config_text = self._build_config_text_for_standard_mode(enabled_configs)
        else:
            config_text = ""

        chip_combo_text = self.chip_combo.currentText() if self.chip_combo.currentIndex() > 0 else None

        self.is_testing = True
        self._config_index_map = config_index_map
        self._current_total_bins = len(firmware_paths)
        self.auto_test_btn.setStateWaiting()

        for idx in self.channel_cards:
            self.channel_cards[idx]["value_label"].setText("- - -")
        if self._vbat_remain_card is not None:
            self._vbat_remain_card["value_label"].setText("- - -")

        if self._current_total_bins > 1:
            self._setup_bin_result_table()
        else:
            self.bin_result_table.hide()

        reset_desc = f"{reset_key}({reset_polarity})" if reset_enabled else "DISABLED"
        mode_desc = "Std V (配置模式)" if test_mode == "standard" else "High V (外供高压)"
        self.append_log(
            f"[AUTO_TEST] Starting auto test: {len(firmware_paths)} BIN(s), "
            f"Mode={mode_desc}, "
            f"Vbat={vbat_cfg['name']}({vbat_cfg['channel']}), "
            f"Control={control_method}, "
            f"PowerON={poweron_key}({poweron_polarity}), "
            f"RESET={reset_desc}, "
            f"StableDelay={stabilization_delay_sec:.1f}s, "
            f"Baudrate={download_pgm_rate}"
        )

        worker_kwargs = dict(
            com_port=com_port,
            firmware_paths=firmware_paths,
            download_mode=download_mode,
            poweron_device_label=poweron_dl,
            poweron_inst=poweron_inst,
            poweron_hw_ch=poweron_hw,
            poweron_polarity=poweron_polarity,
            reset_device_label=reset_dl,
            reset_inst=reset_inst,
            reset_hw_ch=reset_hw,
            reset_polarity=reset_polarity,
            vbat_device_label=vbat_device_label,
            vbat_inst=vbat_inst,
            vbat_hw_ch=vbat_hw_ch,
            force_map=force_map,
            test_time=test_time,
            sample_period=sample_period,
            channel_names=channel_names,
            chip_combo_text=chip_combo_text,
            selected_chip_config=self.selected_chip_config,
            config_text=config_text,
            parse_config_commands_fn=self._parse_config_commands,
            resolve_device_fn=self._resolve_device,
            channel_force_configs=self._build_channel_force_configs(),
            force_config_enabled=(test_mode == "standard"),
            test_mode=test_mode,
            control_method=control_method,
            stabilization_delay_sec=stabilization_delay_sec,
            download_pgm_rate=download_pgm_rate,
        )
        self._controller.start_auto_test(worker_kwargs)
        self.auto_test_btn.setStateProgramming()
        self.auto_test_btn._progress_timer.stop()

    def _on_auto_test_error(self, err_msg):
        pass

    def _on_auto_test_finished(self):
        self.is_testing = False
        self.auto_test_btn.setStateComplete()

    def _on_auto_test_thread_cleaned(self):
        self._auto_test_worker = None
        self._auto_test_thread = None

    def _stop_auto_test(self):
        self._controller.stop_auto_test()
        self.is_testing = False
        self.auto_test_btn.setStateFailed()
        self.append_log("[AUTO_TEST] Auto test stopped by user.")

    def _save_datalog(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save DataLog", "",
            "CSV Files (*.csv);;All Files (*)"
        )
        if file_path:
            logger.info("Saving datalog to: %s", file_path)
            self.append_log(f"[SYSTEM] DataLog saved to: {file_path}")

    def _export_bin_results_to_excel(self):
        """把 BIN 结果导出为 .xlsx。

        - 电流列(每个 enabled 通道 + Vbat_remain)统一换算为 µA,
          且单位只写在表头(形如 "Vbat (µA)"),数据单元格里只保留裸数值,
          方便后续在 Excel 里继续计算 / 画图。
        - 非电流列(BIN / Voltage)保持原样(Voltage 仍为 "3.800 | 3.800" 这类
          由 UI 组装好的字符串,与 QTableWidget 一致)。
        - 数据源优先使用 self._bin_results_data(即测试时原始 summary,含浮点数),
          因此不受 UI 单位自动换算格式(mA/µA/nA)的影响。
        """
        table = getattr(self, "bin_result_table", None)
        summaries = list(getattr(self, "_bin_results_data", []) or [])
        if table is None or table.columnCount() == 0 or not summaries:
            QMessageBox.information(
                self, "Nothing to Export",
                "There are no BIN results to export yet.\n"
                "Run an Auto Test first and try again."
            )
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.critical(
                self, "Export Failed",
                "Python package 'openpyxl' is required to export Excel.\n"
                "Please install it first:\n\n    pip install openpyxl"
            )
            return

        from datetime import datetime
        default_name = f"consumption_bin_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        default_path = os.path.join(os.getcwd(), default_name)
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export BIN Results",
            default_path,
            "Excel Files (*.xlsx);;All Files (*)"
        )
        if not file_path:
            return
        if not file_path.lower().endswith(".xlsx"):
            file_path += ".xlsx"

        try:
            # ---- 先根据 enabled 通道配置组装列定义 ----
            # column 元组: (title, kind)
            #   kind = "bin" | "voltage" | "current" | "vbat_remain"
            unit_suffix = " (\u00b5A)"  # µA
            columns = [("BIN", "bin"), ("Voltage", "voltage")]
            enabled_cfgs = [cfg for cfg in self._channel_configs if cfg["enabled"]]
            for cfg in enabled_cfgs:
                columns.append((f"{cfg['name']}{unit_suffix}", "current", cfg))
            has_sub = any(
                not cfg["name"].lower().startswith("vbat") for cfg in enabled_cfgs
            )
            if has_sub:
                columns.append((f"Vbat_remain{unit_suffix}", "vbat_remain"))

            def _to_ua(value):
                """A -> µA;None / 非数值返回空串。"""
                if value is None:
                    return ""
                try:
                    return round(float(value) * 1e6, 3)
                except (TypeError, ValueError):
                    return ""

            def _voltage_text(summary):
                channel_voltages = summary.get("channel_voltages", {}) or {}
                parts = []
                for cfg in enabled_cfgs:
                    device_label, hw_ch = self._parse_channel_key(cfg["channel"])
                    v = channel_voltages.get((device_label, hw_ch))
                    parts.append(f"{v:.4g}" if v is not None else "N/A")
                return " | ".join(parts) if parts else "- - -"

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "BIN Results"

            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill("solid", fgColor="1F3864")
            header_align = Alignment(horizontal="center", vertical="center")
            cell_align = Alignment(horizontal="center", vertical="center")
            thin_side = Side(style="thin", color="B4C7E7")
            thin_border = Border(
                left=thin_side, right=thin_side,
                top=thin_side, bottom=thin_side,
            )
            row_fill_even = PatternFill("solid", fgColor="F2F6FC")

            for col_idx, col_def in enumerate(columns, start=1):
                title = col_def[0]
                cell = ws.cell(row=1, column=col_idx, value=title)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            # ---- 数据行 ----
            for r, summary in enumerate(summaries):
                excel_row = r + 2
                bin_name = summary.get("bin_name", f"BIN-{r + 1}")
                channels = summary.get("channels", {}) or {}
                vbat_current = summary.get("vbat")
                vbat_remain = summary.get("vbat_remain")

                for col_idx, col_def in enumerate(columns, start=1):
                    kind = col_def[1]
                    if kind == "bin":
                        value = bin_name
                    elif kind == "voltage":
                        value = _voltage_text(summary)
                    elif kind == "current":
                        cfg = col_def[2]
                        if cfg["name"].lower().startswith("vbat"):
                            value = _to_ua(vbat_current)
                        else:
                            device_label, hw_ch = self._parse_channel_key(cfg["channel"])
                            value = _to_ua(channels.get((device_label, hw_ch)))
                    elif kind == "vbat_remain":
                        value = _to_ua(vbat_remain)
                    else:
                        value = ""

                    cell = ws.cell(row=excel_row, column=col_idx, value=value)
                    cell.alignment = cell_align
                    cell.border = thin_border
                    if isinstance(value, (int, float)):
                        cell.number_format = "0.000"
                    if r % 2 == 1:
                        cell.fill = row_fill_even

            # ---- 列宽自适应 ----
            for col_idx, col_def in enumerate(columns, start=1):
                title = col_def[0]
                max_len = len(str(title))
                for row_cells in ws.iter_rows(
                    min_row=2, max_row=len(summaries) + 1,
                    min_col=col_idx, max_col=col_idx,
                ):
                    for c in row_cells:
                        tl = len(str(c.value)) if c.value is not None else 0
                        if tl > max_len:
                            max_len = tl
                ws.column_dimensions[get_column_letter(col_idx)].width = min(
                    max(max_len + 4, 10), 40
                )
            ws.row_dimensions[1].height = 22
            ws.freeze_panes = "A2"

            wb.save(file_path)
            self.append_log(f"[EXPORT] BIN results exported to: {file_path}")
            QMessageBox.information(
                self, "Export Succeeded",
                f"BIN results have been exported to:\n{file_path}"
            )
        except Exception as e:
            logger.exception("Export BIN results failed")
            self.append_log(f"[ERROR] Export BIN results failed: {e}")
            QMessageBox.critical(
                self, "Export Failed",
                f"Failed to export BIN results:\n{e}"
            )

    @staticmethod
    def _format_current(current_A):
        abs_i = abs(current_A)
        if abs_i >= 1:
            return f"{current_A:.3f} A"
        elif abs_i >= 1e-3:
            return f"{current_A*1e3:.3f} mA"
        elif abs_i >= 1e-6:
            return f"{current_A*1e6:.3f} µA"
        elif abs_i >= 1e-9:
            return f"{current_A*1e9:.3f} nA"
        else:
            return f"{current_A:.3e} A"

    def update_channel_current(self, channel_idx, avg_current):
        if channel_idx in self.channel_cards:
            label = self.channel_cards[channel_idx]["value_label"]
            if avg_current is not None:
                label.setText(self._format_current(avg_current))
            else:
                label.setText("- - -")

    def get_selected_channels(self):
        return [
            cfg["channel"] for cfg in self._channel_configs if cfg["enabled"]
        ]

    def get_test_config(self):
        """读取当前页配置快照（同时供 AI ai_get_config 复用，键名与
        apply_config_to_controls 对齐）。

        AI 可修改的字段（与 apply_config_to_controls 对齐，禁臆造键名）：
          - chip_name                  : str        当前所选芯片名（chip_combo 文本）
          - test_mode                  : str        "high_voltage" / "standard"
          - test_time_s                : float      Test Time (s)
          - stable_delay_s             : float      Stable Delay (s)
          - download_baudrate          : int        固件下载波特率
          - download_mode              : str        "flash" / "ramrun"
          - control_method             : str        "N6705C" / "MCU"
          - mcu_type                   : str        "ch9114f" / "yd_rp2040"
          - poweron_channel            : str        PowerON 通道键（如 "B-CH1" / "GPIO0"）
          - poweron_polarity           : str        "rising" / "falling"
          - reset_enabled              : bool       是否启用 RESET
          - reset_channel              : str        RESET 通道键（reset_enabled=False 时为 ""）
          - reset_polarity             : str        "rising" / "falling"
          - firmware_paths             : list[str]  已选固件路径列表
          - channel_configs            : list[dict] 每通道 {name, channel, enabled,
                                                   force_mode, force_value, boost_mode, boost_value}
        只读字段（仅供 AI 观察状态，apply_config_to_controls 不接受）：
          - n6705c_a/b_connected       : bool
          - mcu_connected              : bool
          - serial_port                : str
          - selected_chip_config       : dict | None  当前芯片完整配置（含 power_distribution）
        """
        try:
            test_time = float(self.test_time_input.text())
        except (ValueError, AttributeError):
            test_time = None
        try:
            stable_delay = float(self.stable_delay_input.text())
        except (ValueError, AttributeError):
            stable_delay = None
        try:
            baudrate = int(self.download_baudrate_input.text())
        except (ValueError, AttributeError):
            baudrate = None

        download_mode = (
            self.download_mode_toggle.value().lower()
            if getattr(self, "download_mode_toggle", None) is not None else "ramrun"
        )
        control_method = (
            self.control_method_toggle.value()
            if getattr(self, "control_method_toggle", None) is not None else "N6705C"
        )
        mcu_type = self._current_mcu_type()

        poweron_channel = (
            self.poweron_channel_combo.currentText()
            if getattr(self, "poweron_channel_combo", None) is not None else ""
        )
        poweron_polarity = (
            self.poweron_polarity_toggle.value()
            if getattr(self, "poweron_polarity_toggle", None) is not None else "rising"
        )
        reset_enabled = bool(
            self.reset_enable_cb.isChecked()
            if getattr(self, "reset_enable_cb", None) is not None else False
        )
        reset_channel = (
            self.reset_channel_combo.currentText()
            if getattr(self, "reset_channel_combo", None) is not None and reset_enabled else ""
        )
        reset_polarity = (
            self.reset_polarity_toggle.value()
            if getattr(self, "reset_polarity_toggle", None) is not None and reset_enabled else "rising"
        )

        # 深拷贝 channel_configs，避免外部修改回写本页状态
        chan_snapshot = [
            {
                "name": cfg.get("name", ""),
                "channel": cfg.get("channel", ""),
                "enabled": bool(cfg.get("enabled", False)),
                "force_mode": cfg.get("force_mode", "auto"),
                "force_value": cfg.get("force_value", ""),
                "boost_mode": cfg.get("boost_mode", "constant"),
                "boost_value": cfg.get("boost_value", ""),
            }
            for cfg in self._channel_configs
        ]

        chip_name = ""
        if getattr(self, "chip_combo", None) is not None and self.chip_combo.currentIndex() > 0:
            chip_name = self.chip_combo.currentText()

        return {
            # 只读状态
            'n6705c_a_connected': self.is_connected_a,
            'n6705c_b_connected': self.is_connected_b,
            'mcu_connected': bool(self.is_mcu_connected),
            'serial_port': self.get_selected_serial_port(),
            'selected_chip_config': self.selected_chip_config,
            'is_testing': bool(self.is_testing),
            # AI 可写键
            'chip_name': chip_name,
            'test_mode': getattr(self, "_test_mode", "high_voltage"),
            'test_time_s': test_time,
            'stable_delay_s': stable_delay,
            'download_baudrate': baudrate,
            'download_mode': download_mode,
            'control_method': control_method,
            'mcu_type': mcu_type,
            'poweron_channel': poweron_channel,
            'poweron_polarity': poweron_polarity,
            'reset_enabled': reset_enabled,
            'reset_channel': reset_channel,
            'reset_polarity': reset_polarity,
            'firmware_paths': list(self.firmware_paths or ([self.firmware_path] if self.firmware_path else [])),
            'channel_configs': chan_snapshot,
            # 兼容旧键（保持向后兼容，等同 selected_chip_config）
            'firmware_path': self.firmware_path,
            'config_content': self.config_content,
            'selected_chip': self.selected_chip_config,
            'channel_configs_legacy': [c["channel"] for c in chan_snapshot if c["enabled"]],
        }

    # ------------------------------------------------------------------
    # AIControllablePage 契约实现（AIAssist_PageScopedControlPlan.md §2 / Phase 5）
    #
    # Consumption Test 接入 AI 受控契约，薄封装既有方法：
    #   - ai_get_config 复用 get_test_config()
    #   - ai_apply_config 经 apply_config_to_controls() 单一写入口回填控件
    #   - ai_start_test 复用 _on_auto_test()（最完整流程：下载 + POWERON + 测试）
    #       若未配置固件/串口，则回退到 _on_start_test()（仅功耗测试）
    #   - ai_stop_test 复用 _stop_running_test()（统一判别 worker 类型）
    # 枢纽（MainWindow.resolve_active_ai_page）经 Tab 子页下钻拿到本实例，
    # 鸭子调用契约方法，无需 core / handler 改动。
    # ------------------------------------------------------------------
    def ai_capabilities(self) -> set[str]:
        return {
            CAP_GET_CONFIG,
            CAP_APPLY_CONFIG,
            CAP_START_TEST,
            CAP_STOP_TEST,
            CAP_GET_RESULT,
        }

    def ai_get_config(self) -> dict | None:
        try:
            return self.get_test_config()
        except Exception:  # noqa: BLE001 - 快照失败降级为 None
            logger.error("AI 读取 Consumption Test 配置失败", exc_info=True)
            return None

    def ai_apply_config(self, payload) -> tuple[bool, str]:
        """落地配置草案到控件（写操作，经确认+审计后由枢纽调用）。

        运行中拒绝改配置（§6.3），避免与正在执行的测试/下载冲突。
        """
        if self.is_testing:
            return False, "测试运行中，无法修改配置，请先停止测试。"
        if self._controller is not None and (
            self._controller.is_download_running() or self._controller.is_auto_test_running()
        ):
            return False, "下载或 Auto Test 运行中，无法修改配置。"
        return self.apply_config_to_controls(payload if isinstance(payload, dict) else {})

    def ai_start_test(self) -> tuple[bool, str]:
        """启动本页测试：优先 Auto Test，前置条件不足时回退到 Force-High 测试。"""
        if not (self.is_connected_a or self.is_connected_b):
            return False, "未连接 N6705C 仪器，请先连接再启动测试。"
        if self.is_testing:
            return False, "测试已在运行中。"
        if not any(cfg.get("enabled") for cfg in self._channel_configs):
            return False, "未启用任何通道，请先在 Channel Config 中启用至少一个通道。"

        has_firmware = bool(getattr(self, "firmware_paths", []) or self.firmware_path)
        has_serial = bool(self.get_selected_serial_port())
        cfg = self.get_test_config()

        if has_firmware and has_serial:
            mode_desc = "Std V (配置模式)" if cfg.get("test_mode") == "standard" else "High V (外供高压)"
            self.append_log(
                f"[AI] 请求启动 Auto Test：模式={mode_desc}，"
                f"Vbat 通道={cfg.get('poweron_channel', '')}，"
                f"test_time={cfg.get('test_time_s')}s。"
            )
            try:
                self._on_auto_test()
            except Exception:  # noqa: BLE001 - 启动异常转可读结果
                logger.error("AI 启动 Auto Test 失败", exc_info=True)
                return False, "启动 Auto Test 异常，请查看日志。"
            if self.is_testing:
                return True, "已请求启动 Auto Test。"
            return False, "Auto Test 启动未成功，请查看执行日志。"

        # 无固件或无串口 → 回退到仅功耗测试
        self.append_log(
            f"[AI] 固件或串口未配置，回退到 Force-High 功耗测试："
            f"test_time={cfg.get('test_time_s')}s。"
        )
        try:
            self._on_start_test()
        except Exception:  # noqa: BLE001
            logger.error("AI 启动 Force-High 测试失败", exc_info=True)
            return False, "启动测试异常，请查看日志。"
        if self.is_testing:
            return True, "已请求启动 Force-High 功耗测试。"
        return False, "启动未成功，请查看执行日志。"

    def ai_stop_test(self) -> tuple[bool, str]:
        if not self.is_testing:
            return False, "当前未在运行测试。"
        self.append_log("[AI] 请求停止测试。")
        try:
            self._stop_running_test()
        except Exception:  # noqa: BLE001
            logger.error("AI 停止测试失败", exc_info=True)
            return False, "停止测试异常，请查看日志。"
        return True, "已发送停止请求。"

    def ai_get_result_summary(self) -> dict | None:
        """读最近一次测试结果摘要（禁止臆造，直接回读结构化数据）。"""
        cfg = self.get_test_config()
        summary: dict = {
            "available": True,
            "running": bool(self.is_testing),
            "test_mode": cfg.get("test_mode"),
            "enabled_channels": [
                {"name": c["name"], "channel": c["channel"]}
                for c in cfg.get("channel_configs", []) if c["enabled"]
            ],
            "bin_count": len(self._bin_results_data),
            "expected_total_bins": self._current_total_bins,
        }
        if not self._bin_results_data:
            return summary
        # 汇总首末 BIN 的 Vbat / Vbat_remain（单位 A，由 AI 端按需换算）
        first = self._bin_results_data[0]
        last = self._bin_results_data[-1]
        summary["first_bin"] = {
            "bin_name": first.get("bin_name"),
            "vbat_current_A": first.get("vbat"),
            "vbat_remain_A": first.get("vbat_remain"),
        }
        summary["last_bin"] = {
            "bin_name": last.get("bin_name"),
            "vbat_current_A": last.get("vbat"),
            "vbat_remain_A": last.get("vbat_remain"),
        }
        return summary

    # ------------------------------------------------------------------
    # UI 回填单一写入口（AIAssist_PageScopedControlPlan.md §4.2）
    #
    # apply_config_to_controls(cfg) 是回填测试配置控件的唯一入口，
    # AI 回填与未来轮询/手动刷新共用，杜绝两套逻辑漂移。键名与
    # get_test_config() 输出对齐。
    # ------------------------------------------------------------------
    def apply_config_to_controls(self, cfg: dict) -> tuple[bool, str]:
        if not isinstance(cfg, dict):
            return False, "配置草案格式无效（期望 dict）。"

        # 线程边界（§4.2-2）：AI 决策在 QThread，回填须经主线程执行；
        # dispatcher 经 QTimer.singleShot(0) 已切回主线程，此处加防御性守卫，
        # 杜绝 worker 线程直接 setValue 违反「UI 禁阻塞 / 跨线程改控件」铁律。
        if threading.current_thread() is not threading.main_thread():
            logger.error(
                "apply_config_to_controls 在非主线程被调用，拒绝回填以防违反线程边界"
            )
            return False, "配置回填未在主线程执行，已拒绝。"

        applied: list[str] = []

        def _set_line_edit(edit, key, cast_fn=None):
            val = cfg.get(key)
            if val is None or edit is None:
                return
            try:
                text = str(cast_fn(val)) if cast_fn else str(val)
            except (TypeError, ValueError):
                return
            if edit.text() != text:
                edit.blockSignals(True)
                edit.setText(text)
                edit.blockSignals(False)
            applied.append(key)

        def _set_combo_text(combo, key):
            val = cfg.get(key)
            if val is None or combo is None:
                return
            idx = combo.findText(str(val))
            if idx >= 0 and combo.currentIndex() != idx:
                combo.blockSignals(True)
                combo.setCurrentIndex(idx)
                combo.blockSignals(False)
                applied.append(key)

        def _set_combo_data(combo, key):
            val = cfg.get(key)
            if val is None or combo is None:
                return
            for i in range(combo.count()):
                if combo.itemData(i) == val:
                    if combo.currentIndex() != i:
                        combo.blockSignals(True)
                        combo.setCurrentIndex(i)
                        combo.blockSignals(False)
                        applied.append(key)
                    return

        def _set_toggle(toggle, key):
            val = cfg.get(key)
            if val is None or toggle is None:
                return
            try:
                cur = toggle.value()
            except Exception:  # noqa: BLE001
                return
            if cur != val:
                toggle.blockSignals(True)
                toggle.setValue(val)
                toggle.blockSignals(False)
                applied.append(key)

        # 1. chip_name → chip_combo（触发 _on_chip_selected 自动加载 rail YAML）
        _set_combo_text(getattr(self, "chip_combo", None), "chip_name")

        # 2. 测试模式 / 控制方式 / 下载模式 / MCU 类型
        _set_toggle(getattr(self, "test_mode_toggle", None), "test_mode")
        _set_toggle(getattr(self, "control_method_toggle", None), "control_method")
        _set_toggle(getattr(self, "download_mode_toggle", None), "download_mode")
        _set_combo_data(getattr(self, "mcu_type_combo", None), "mcu_type")

        # 3. 数值输入
        _set_line_edit(getattr(self, "test_time_input", None), "test_time_s", cast_fn=float)
        _set_line_edit(getattr(self, "stable_delay_input", None), "stable_delay_s", cast_fn=float)
        _set_line_edit(getattr(self, "download_baudrate_input", None), "download_baudrate", cast_fn=int)

        # 4. PowerON / RESET 通道与极性
        _set_combo_text(getattr(self, "poweron_channel_combo", None), "poweron_channel")
        _set_toggle(getattr(self, "poweron_polarity_toggle", None), "poweron_polarity")

        reset_enabled = cfg.get("reset_enabled")
        if reset_enabled is not None and getattr(self, "reset_enable_cb", None) is not None:
            if self.reset_enable_cb.isChecked() != bool(reset_enabled):
                self.reset_enable_cb.blockSignals(True)
                self.reset_enable_cb.setChecked(bool(reset_enabled))
                self.reset_enable_cb.blockSignals(False)
                # 手动触发联动（blockSignals 跳过了原槽）
                if hasattr(self, "_on_reset_enable_toggled"):
                    self._on_reset_enable_toggled(bool(reset_enabled))
                applied.append("reset_enabled")
        _set_combo_text(getattr(self, "reset_channel_combo", None), "reset_channel")
        _set_toggle(getattr(self, "reset_polarity_toggle", None), "reset_polarity")

        # 5. 通道配置（按 index 顺序回填，不增删通道，只改字段）
        chan_cfgs = cfg.get("channel_configs")
        if isinstance(chan_cfgs, list) and chan_cfgs:
            applied_chan: list[str] = []
            for idx, src in enumerate(chan_cfgs):
                if idx >= len(self._channel_configs) or idx >= len(self._channel_config_widgets):
                    break
                if not isinstance(src, dict):
                    continue
                dst = self._channel_configs[idx]
                wdata = self._channel_config_widgets[idx]

                name = src.get("name")
                if name is not None:
                    name_combo = wdata.get("name_input")
                    if name_combo is not None:
                        i = name_combo.findText(str(name))
                        if i >= 0 and name_combo.currentIndex() != i:
                            name_combo.blockSignals(True)
                            name_combo.setCurrentIndex(i)
                            name_combo.blockSignals(False)
                channel = src.get("channel")
                if channel is not None:
                    ch_combo = wdata.get("channel_combo")
                    if ch_combo is not None:
                        i = ch_combo.findText(str(channel))
                        if i >= 0 and ch_combo.currentIndex() != i:
                            ch_combo.blockSignals(True)
                            ch_combo.setCurrentIndex(i)
                            ch_combo.blockSignals(False)
                enabled = src.get("enabled")
                if enabled is not None:
                    enable_cb = wdata.get("enable_cb")
                    if enable_cb is not None and enable_cb.isChecked() != bool(enabled):
                        enable_cb.blockSignals(True)
                        enable_cb.setChecked(bool(enabled))
                        enable_cb.blockSignals(False)
                # 同步 dst dict（与 _on_config_*_changed 槽保持一致的字段集）
                if name is not None:
                    dst["name"] = str(name)
                if channel is not None:
                    dst["channel"] = str(channel)
                if enabled is not None:
                    dst["enabled"] = bool(enabled)
                if "force_mode" in src:
                    dst["force_mode"] = str(src["force_mode"])
                if "force_value" in src:
                    dst["force_value"] = str(src.get("force_value", ""))
                if "boost_mode" in src:
                    dst["boost_mode"] = str(src["boost_mode"])
                if "boost_value" in src:
                    dst["boost_value"] = str(src.get("boost_value", ""))
                # 回填 force/boost 输入控件文本
                fv_input = wdata.get("force_value_input")
                if fv_input is not None and "force_value" in src:
                    fv_input.blockSignals(True)
                    fv_input.setText(str(src.get("force_value", "")))
                    fv_input.blockSignals(False)
                bv_input = wdata.get("boost_value_input")
                if bv_input is not None and "boost_value" in src:
                    bv_input.blockSignals(True)
                    bv_input.setText(str(src.get("boost_value", "")))
                    bv_input.blockSignals(False)
                applied_chan.append(f"#{idx}")
            if applied_chan:
                applied.append(f"channel_configs[{','.join(applied_chan)}]")

        # 6. 固件路径列表（仅回填到 firmware_paths，不自动选文件）
        fw_paths = cfg.get("firmware_paths")
        if isinstance(fw_paths, list) and fw_paths:
            self.firmware_paths = [str(p) for p in fw_paths if p]
            if self.firmware_paths:
                self.firmware_path = self.firmware_paths[0]
                if hasattr(self, "firmware_file_input") and self.firmware_file_input is not None:
                    names = "; ".join(os.path.basename(p) for p in self.firmware_paths)
                    self.firmware_file_input.setText(names)
                applied.append("firmware_paths")

        if not applied:
            return True, "配置草案无可应用字段，控件保持原状。"
        return True, f"已应用配置草案，受影响字段：{', '.join(applied)}。"

    @property
    def logs_frame(self):
        """logs_frame 属性别名：暴露 execution_logs 给 AI 枢纽读取执行日志。

        枢纽 _get_ai_execution_logs 读 page.logs_frame._all_logs；
        本页 execution_logs 就是 ExecutionLogsFrame 实例。
        """
        return getattr(self, "execution_logs", None)

    def update_test_result(self, result):
        if isinstance(result, dict):
            for idx, cfg in enumerate(self._channel_configs):
                if not cfg["enabled"]:
                    continue
                key = cfg["channel"]
                if key in result:
                    self.update_channel_current(idx, result[key])

    def append_log(self, message):
        self.execution_logs.append_log(message)

    def _on_clear_log(self):
        self.execution_logs.clear_log()

    def clear_results(self):
        for idx in self.channel_cards:
            self.channel_cards[idx]["value_label"].setText("- - -")
        if self._vbat_remain_card is not None:
            self._vbat_remain_card["value_label"].setText("- - -")
        self._bin_results_data = []
        self._current_total_bins = 0
        self.bin_result_table.setRowCount(0)
        self.bin_result_table.hide()
        if hasattr(self, "_bin_result_header"):
            self._bin_result_header.hide()

    def get_test_mode(self):
        return "Consumption Test"

    def set_test_mode(self, mode):
        pass

    def get_test_id(self):
        return "CONSUMPTION_TEST_001"

    def set_test_id(self, test_id):
        pass


if __name__ == "__main__":
    from ui.standalone import resize_and_center_window

    app = QApplication(sys.argv)
    app.setStyle("Fusion")

    win = ConsumptionTestUI()
    win.setWindowTitle("Consumption Test")
    resize_and_center_window(win)
    win.show()

    sys.exit(app.exec())
