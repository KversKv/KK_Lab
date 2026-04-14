#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PMU Is_gain测试UI组件
暗色卡片式重构版本（PySide6）
"""

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGridLayout, QPushButton, QComboBox,
    QLabel, QLineEdit, QSpinBox, QDoubleSpinBox, QFrame, QTextEdit,
    QProgressBar, QTableWidget, QTableWidgetItem, QHeaderView, QCheckBox,
    QFileDialog, QDialog, QRadioButton, QButtonGroup, QSizePolicy, QScrollArea
)
from PySide6.QtCore import Qt, Signal, QThread, QObject
from PySide6.QtGui import QFont, QColor
import os
import time
import base64
import csv
from datetime import datetime
import pyvisa

from instruments.power.keysight.n6705c import N6705C
from ui.styles.button import SpinningSearchButton, update_connect_button_state
from debug_config import DEBUG_MOCK
from instruments.mock.mock_instruments import MockN6705C, MockMSO64B
from instruments.scopes.tektronix.mso64b import MSO64B
from instruments.scopes.keysight.dsox4034a import DSOX4034A
from ui.widgets.dark_combobox import DarkComboBox
from ui.styles import SCROLLBAR_STYLE


class _InstrumentWorker(QObject):
    log = Signal(str)
    finished = Signal(dict)

    def __init__(self, task, kwargs=None):
        super().__init__()
        self._task = task
        self._kwargs = kwargs or {}

    def run(self):
        try:
            result = self._task(**self._kwargs)
            self.finished.emit(result if isinstance(result, dict) else {})
        except Exception as e:
            self.finished.emit({"error": str(e)})


class _IsGainTestWorker(QObject):
    log = Signal(str)
    progress = Signal(int)
    result_row = Signal(dict)
    summary = Signal(dict)
    finished = Signal()
    error = Signal(str)

    MODE_SINGLE = "single"
    MODE_TRAVERSE = "traverse"

    _YSCALE_SEQUENCE = [
        0.001, 0.002, 0.005,
        0.010, 0.020, 0.050,
        0.100, 0.200, 0.500,
        1.000, 2.000, 5.000,
    ]

    _RECOVERY_SCALE = 1.0

    def __init__(self, n6705c, scope, config, test_mode="single"):
        super().__init__()
        self.n6705c = n6705c
        self.scope = scope
        self.config = config
        self.test_mode = test_mode
        self._stop_flag = False

    def stop(self):
        self._stop_flag = True

    def _interruptible_sleep(self, seconds):
        interval = 0.1
        elapsed = 0.0
        while elapsed < seconds and not self._stop_flag:
            time.sleep(min(interval, seconds - elapsed))
            elapsed += interval

    def _parse_channel(self, ch_text):
        return int(ch_text.replace("CH ", "").strip())

    def _prev_scale(self, current_scale):
        prev = None
        for s in self._YSCALE_SEQUENCE:
            if s >= current_scale * 0.99:
                return prev
            prev = s
        return prev

    def _measure_with_autoscale(self, ripple_ch, label=""):
        voltage = None
        ripple = None

        try:
            voltage = self.scope.get_channel_mean(ripple_ch)
        except Exception as e:
            self.log.emit(f"  [WARN] Voltage measurement failed: {e}")
            voltage = None

        if self._stop_flag:
            return voltage, ripple

        try:
            ripple = self.scope.get_channel_pk2pk(ripple_ch)
        except Exception as e:
            self.log.emit(f"  [WARN] Ripple measurement failed: {e}")
            ripple = None

        if voltage is not None and ripple is not None:
            self.log.emit(f"  Voltage (CH{ripple_ch} Mean) = {voltage:.6f} V")
            self.log.emit(f"  Ripple  (CH{ripple_ch} Pk2Pk) = {ripple:.6f} V")
            return voltage, ripple

        if self._stop_flag:
            return voltage, ripple

        self.log.emit(f"  [AUTO-SCALE] {label}Measurement failed, starting recovery...")

        try:
            orig_scale = self.scope.get_channel_scale(ripple_ch)
        except Exception:
            orig_scale = 0.020

        try:
            self.scope.set_channel_scale(ripple_ch, self._RECOVERY_SCALE)
            self.scope.set_channel_offset(ripple_ch, 0.0)
            self.log.emit(f"  [AUTO-SCALE] {label}Set scale={self._RECOVERY_SCALE} V/div, offset=0 for coarse measurement")
            self._interruptible_sleep(0.3)
            if self._stop_flag:
                return voltage, ripple

            self.scope.run()
            self._interruptible_sleep(0.8)
            if self._stop_flag:
                return voltage, ripple
            self.scope.stop()
            self._interruptible_sleep(0.2)
        except Exception as e:
            self.log.emit(f"  [WARN] Recovery setup failed: {e}")
            return voltage, ripple

        voltage_t = None
        ripple_t = None
        try:
            voltage_t = self.scope.get_channel_mean(ripple_ch)
        except Exception:
            pass
        try:
            ripple_t = self.scope.get_channel_pk2pk(ripple_ch)
        except Exception:
            pass

        if voltage_t is None:
            self.log.emit(f"  [WARN] {label}Cannot get voltage even at {self._RECOVERY_SCALE} V/div, aborting auto-scale")
            try:
                self.scope.set_channel_scale(ripple_ch, orig_scale)
                self.scope.set_channel_offset(ripple_ch, 0.0)
            except Exception:
                pass
            return voltage, ripple

        rip_str = f"{ripple_t:.6f}" if ripple_t is not None else "N/A"
        self.log.emit(f"  [AUTO-SCALE] {label}Coarse Voltage_T = {voltage_t:.6f} V, Ripple_T = {rip_str} V")

        voltage = voltage_t
        ripple = ripple_t

        scale = self._RECOVERY_SCALE
        while True:
            if self._stop_flag:
                break

            smaller = self._prev_scale(scale)
            if smaller is None or smaller < orig_scale * 0.99:
                break

            offset_val = voltage_t - 0.03
            try:
                self.scope.set_channel_scale(ripple_ch, smaller)
                self.scope.set_channel_offset(ripple_ch, offset_val)
                self.log.emit(f"  [AUTO-SCALE] {label}Trying scale={smaller} V/div, offset={offset_val:.4f} V")
                self._interruptible_sleep(0.2)
                if self._stop_flag:
                    break

                self.scope.run()
                self._interruptible_sleep(0.6)
                if self._stop_flag:
                    break
                self.scope.stop()
                self._interruptible_sleep(0.2)
            except Exception as e:
                self.log.emit(f"  [WARN] {label}Scale adjustment failed: {e}")
                break

            v_try = None
            r_try = None
            try:
                v_try = self.scope.get_channel_mean(ripple_ch)
            except Exception:
                pass
            try:
                r_try = self.scope.get_channel_pk2pk(ripple_ch)
            except Exception:
                pass

            if v_try is not None and r_try is not None:
                voltage = v_try
                ripple = r_try
                scale = smaller
            else:
                self.log.emit(f"  [AUTO-SCALE] {label}Measurement failed at scale={smaller} V/div, using previous result")
                try:
                    self.scope.set_channel_scale(ripple_ch, scale)
                    self.scope.set_channel_offset(ripple_ch, voltage_t - 0.03)
                except Exception:
                    pass
                break

        if voltage is not None:
            self.log.emit(f"  Voltage (CH{ripple_ch} Mean) = {voltage:.6f} V")
        if ripple is not None:
            self.log.emit(f"  Ripple  (CH{ripple_ch} Pk2Pk) = {ripple:.6f} V")

        return voltage, ripple

    def _run_single_sweep(self, load_ch, ripple_ch, current_steps, step_offset=0, total_override=None, reg_value=None):
        total = total_override if total_override else (len(current_steps) + 1)
        results = []
        save_screenshot = self.config.get("save_screenshot", True)

        self.log.emit("[STEP 0] Measuring 0-load baseline (channel OFF)...")
        self.n6705c.channel_off(load_ch)
        self._interruptible_sleep(1.0)
        if self._stop_flag:
            self.log.emit("[TEST] Test aborted by user.")
            return False, results, None

        try:
            self.scope.set_AutoRipple_test(ripple_ch)
        except Exception as e:
            self.log.emit(f"  [WARN] set_AutoRipple_test failed: {e}")

        self._interruptible_sleep(0.5)

        v0 = None
        r0 = None

        screenshot_b64 = None
        if not self._stop_flag:
            try:
                self._interruptible_sleep(0.1)
                if save_screenshot and not self._stop_flag:
                    self.scope.stop()
                    self._interruptible_sleep(0.5)

                if not self._stop_flag:
                    v0, r0 = self._measure_with_autoscale(ripple_ch, label="0-load ")

                if save_screenshot and not self._stop_flag:
                    png_data = self.scope.capture_screen_png()
                    screenshot_b64 = base64.b64encode(png_data).decode("ascii")
                    self.log.emit("  Screenshot captured (0-load)")
            except Exception as e:
                self.log.emit(f"  [WARN] Measurement/Screenshot failed: {e}")
            finally:
                if save_screenshot:
                    try:
                        self.scope.run()
                    except Exception:
                        pass

        zero_row = {
            "step": step_offset,
            "load_current": 0.0,
            "voltage": v0,
            "ripple": r0,
            "v_drop": 0.0,
            "screenshot_b64": screenshot_b64,
            "remark": "0-load baseline",
        }
        if reg_value is not None:
            zero_row["reg_value"] = reg_value
        self.result_row.emit(zero_row)
        results.append(zero_row)
        self.progress.emit(int((step_offset + 1) / total * 100))

        self.n6705c.channel_on(load_ch)
        self._interruptible_sleep(0.5)

        for i, current_a in enumerate(current_steps):
            if self._stop_flag:
                self.log.emit("[TEST] Test aborted by user.")
                return False, results, v0

            global_step = step_offset + i + 1
            prefix = f"Reg={reg_value}, " if reg_value is not None else ""
            self.log.emit(f"[STEP {global_step}/{total}] {prefix}Load = {current_a} A")
            self.n6705c.set_current(load_ch, -abs(current_a))
            self._interruptible_sleep(0.1)

            voltage = None
            ripple = None
            sc_b64 = None

            if not self._stop_flag:
                try:
                    self._interruptible_sleep(0.1)
                    if save_screenshot and not self._stop_flag:
                        self.scope.stop()
                        self._interruptible_sleep(0.1)

                    if not self._stop_flag:
                        voltage, ripple = self._measure_with_autoscale(ripple_ch)

                    if save_screenshot and not self._stop_flag:
                        png_data = self.scope.capture_screen_png()
                        sc_b64 = base64.b64encode(png_data).decode("ascii")
                        self.log.emit(f"  Screenshot captured (step {global_step})")
                except Exception as e:
                    self.log.emit(f"  [WARN] Measurement/Screenshot failed: {e}")
                finally:
                    if save_screenshot:
                        try:
                            self.scope.run()
                        except Exception:
                            pass

            v_drop = None
            if v0 is not None and voltage is not None:
                v_drop = v0 - voltage

            row = {
                "step": global_step,
                "load_current": current_a,
                "voltage": voltage,
                "ripple": ripple,
                "v_drop": v_drop,
                "screenshot_b64": sc_b64,
            }
            if reg_value is not None:
                row["reg_value"] = reg_value
            self.result_row.emit(row)
            results.append(row)
            self.progress.emit(int((global_step + 1) / total * 100))

        return True, results, v0

    def _analyze_results(self, results, v0):
        max_ripple = None
        max_ripple_current = None
        max_load_current = None

        if v0 is not None:
            for r in results:
                v = r.get("voltage")
                cur = r.get("load_current", 0)
                rp = r.get("ripple")

                if rp is not None:
                    if max_ripple is None or rp > max_ripple:
                        max_ripple = rp
                        max_ripple_current = cur

                if v is not None and cur > 0:
                    drop = v0 - v
                    if drop <= 0.030:
                        max_load_current = cur

        return {
            "v0": v0,
            "max_load_current": max_load_current,
            "max_ripple": max_ripple,
            "max_ripple_current": max_ripple_current,
        }

    def run(self):
        load_ch = None
        try:
            cfg = self.config
            load_ch = self._parse_channel(cfg["load_channel"])
            ripple_ch = self._parse_channel(cfg["ripple_channel"])

            start_i = abs(cfg["is_gain_start_current"])
            end_i = abs(cfg["is_gain_end_current"])
            step_i = abs(cfg["is_gain_step_current"])

            if start_i > end_i:
                start_i, end_i = end_i, start_i

            if step_i < 1e-9:
                self.error.emit("Step Current must be greater than 0.")
                return

            current_steps = []
            cur = start_i
            while cur <= end_i + 1e-9:
                current_steps.append(round(cur, 6))
                cur += step_i

            if not current_steps:
                self.error.emit("No current steps generated. Check Start/End/Step Current.")
                return

            self.n6705c.set_mode(load_ch, "CCLoad")
            self.n6705c.set_current(load_ch, 0)

            t_start = time.time()
            total_steps = 0

            if self.test_mode == self.MODE_SINGLE:
                total = len(current_steps) + 1
                total_steps = total
                self.log.emit(f"[TEST] Single Is_gain test: {total} steps (including 0-load)")
                self.log.emit(f"[TEST] Load CH{load_ch}, Ripple CH{ripple_ch}")
                ok, results, v0 = self._run_single_sweep(load_ch, ripple_ch, current_steps, total_override=total)

                if v0 is None and results:
                    v0 = results[0].get("voltage")
                analysis = self._analyze_results(results, v0)
                self.summary.emit(analysis)

            elif self.test_mode == self.MODE_TRAVERSE:
                msb = cfg["is_gain_msb"]
                lsb = cfg["is_gain_lsb"]
                reg_values = list(range(lsb, msb + 1))

                if not reg_values:
                    self.error.emit("No register values to traverse. Check MSB >= LSB.")
                    return

                steps_per_reg = len(current_steps) + 1
                total = len(reg_values) * steps_per_reg
                total_steps = total
                self.log.emit(f"[TEST] Traverse Is_gain: {len(reg_values)} reg values x {steps_per_reg} steps = {total} total")
                self.log.emit(f"[TEST] Register range: {lsb} ~ {msb}, Current: {start_i} A ~ {end_i} A step {step_i} A")

                all_results = []
                for reg_idx, reg_val in enumerate(reg_values):
                    if self._stop_flag:
                        self.log.emit("[TEST] Test aborted by user.")
                        break

                    self.log.emit(f"[REG] Setting register value = {reg_val}")
                    offset = reg_idx * steps_per_reg
                    ok, results, v0 = self._run_single_sweep(
                        load_ch, ripple_ch, current_steps,
                        step_offset=offset, total_override=total, reg_value=reg_val,
                    )
                    all_results.extend(results)
                    if not ok:
                        break

                if all_results:
                    v0 = all_results[0].get("voltage")
                    analysis = self._analyze_results(all_results, v0)
                    self.summary.emit(analysis)

            elapsed = time.time() - t_start
            minutes, seconds = divmod(elapsed, 60)
            if total_steps > 0:
                avg = elapsed / total_steps
                self.log.emit(f"[TIME] Total: {int(minutes)}m {seconds:.1f}s | Steps: {total_steps} | Avg: {avg:.2f}s/step")
            else:
                self.log.emit(f"[TIME] Total: {int(minutes)}m {seconds:.1f}s")

        except Exception as e:
            self.error.emit(str(e))
        finally:
            if load_ch is not None:
                try:
                    self.n6705c.set_current(load_ch, 0)
                    self.n6705c.channel_off(load_ch)
                    self.log.emit("[TEST] Load current reset to 0 A. Test completed.")
                except Exception as e:
                    self.log.emit(f"[WARN] Cleanup failed: {e}")
            self.finished.emit()


class CardFrame(QFrame):
    def __init__(self, title="", parent=None):
        super().__init__(parent)
        self.setObjectName("cardFrame")
        self.main_layout = QVBoxLayout(self)
        self.main_layout.setContentsMargins(14, 14, 14, 14)
        self.main_layout.setSpacing(12)

        if title:
            self.title_label = QLabel(title)
            self.title_label.setObjectName("cardTitle")
            self.main_layout.addWidget(self.title_label)


class FixedPopupComboBox(DarkComboBox):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setSizeAdjustPolicy(
            DarkComboBox.SizeAdjustPolicy.AdjustToMinimumContentsLengthWithIcon
        )
        self.setMinimumContentsLength(10)
        self.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Fixed)

    def showPopup(self):
        super().showPopup()
        view = self.view()
        if view and view.window():
            popup = view.window()
            global_pos = self.mapToGlobal(self.rect().bottomLeft())
            popup.move(global_pos.x(), global_pos.y())


class PMUIsGainUI(QWidget):
    """PMU Is_gain测试UI组件"""

    connection_status_changed = Signal(bool)

    def __init__(self, n6705c_top=None, mso64b_top=None):
        super().__init__()

        self._n6705c_top = n6705c_top
        self._mso64b_top = mso64b_top
        self.rm = None
        self.n6705c = None
        self.available_devices = []
        self.is_connected = False

        self.scope_connected = False
        self.scope_resource = None
        self.Osc_ins = None

        self.is_test_running = False
        self.test_thread = None
        self._test_worker = None
        self._test_result_data = []

        self._instr_thread = None
        self._instr_worker = None

        self._setup_style()
        self._create_layout()
        self._init_ui_elements()
        self._bind_signals()
        self._sync_from_top()

        if self._mso64b_top is not None:
            self._mso64b_top.connection_changed.connect(self._on_mso64b_top_changed)

    @staticmethod
    def _get_checkmark_path(accent_color):
        safe_name = accent_color.replace("#", "").replace(" ", "")
        icons_dir = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))),
            "resources", "icons"
        )
        return {
            "checked": os.path.join(icons_dir, f"checked_{safe_name}.svg").replace("\\", "/"),
            "unchecked": os.path.join(icons_dir, f"unchecked_{safe_name}.svg").replace("\\", "/"),
        }

    def _setup_style(self):
        font = QFont("Segoe UI", 9)
        self.setFont(font)

        _cb_icons = self._get_checkmark_path("4f46e5")
        full_style = ("""
            QWidget {
                background-color: #020817;
                color: #dbe7ff;
            }

            QLabel {
                background-color: transparent;
                color: #dbe7ff;
                border: none;
            }

            QLabel#pageTitle {
                font-size: 18px;
                font-weight: 700;
                color: #f8fbff;
                background-color: transparent;
            }

            QLabel#pageSubtitle {
                font-size: 12px;
                color: #7da2d6;
                background-color: transparent;
            }

            QLabel#cardTitle {
                font-size: 11px;
                font-weight: 700;
                color: #f4f7ff;
                letter-spacing: 0.5px;
                background-color: transparent;
            }

            QLabel#sectionTitle {
                font-size: 12px;
                font-weight: 700;
                color: #f4f7ff;
                background-color: transparent;
            }

            QLabel#fieldLabel {
                color: #8eb0e3;
                font-size: 11px;
                background-color: transparent;
            }

            QLabel#statusErr {
                color: #ff5e7a;
                font-weight: 600;
                background-color: transparent;
            }

            QLabel#statusWarn {
                color: #ffb84d;
                font-weight: 600;
                background-color: transparent;
            }

            QFrame#panelFrame {
                background-color: #08132d;
                border: 1px solid #16274d;
                border-radius: 18px;
            }

            QFrame#cardFrame {
                background-color: #071127;
                border: 1px solid #1a2b52;
                border-radius: 14px;
            }

            QFrame#resultContainer, QFrame#logContainer {
                background-color: #09142e;
                border: 1px solid #1a2d57;
                border-radius: 16px;
            }

            QLineEdit, QComboBox, QSpinBox, QDoubleSpinBox, QTextEdit, QTableWidget {
                background-color: #0a1733;
                color: #eaf2ff;
                border: 1px solid #27406f;
                border-radius: 8px;
                padding: 6px 10px;
                selection-background-color: #4f46e5;
            }
            QSpinBox::up-button, QSpinBox::down-button,
            QDoubleSpinBox::up-button, QDoubleSpinBox::down-button {
                width: 0px; height: 0px; border: none;
            }

            QLineEdit:focus, QComboBox:focus, QSpinBox:focus, QDoubleSpinBox:focus, QTextEdit:focus, QTableWidget:focus {
                border: 1px solid #4cc9f0;
            }

            QComboBox {
                padding-right: 24px;
            }

            QComboBox::drop-down {
                border: none;
                width: 22px;
                background: transparent;
                subcontrol-origin: padding;
                subcontrol-position: top right;
            }

            QComboBox QAbstractItemView {
                background-color: #0a1733;
                color: #eaf2ff;
                border: 1px solid #27406f;
                selection-background-color: #334a7d;
                outline: 0px;
            }

            QComboBox QAbstractItemView::item {
                background-color: #0a1733;
                color: #eaf2ff;
                padding: 4px 8px;
            }

            QComboBox QAbstractItemView::item:hover {
                background-color: #1a3260;
            }

            QComboBox QAbstractItemView::item:selected {
                background-color: #334a7d;
            }

            QComboBox QFrame {
                background-color: #0a1733;
                border: 1px solid #27406f;
            }

            QPushButton {
                min-height: 34px;
                border-radius: 9px;
                padding: 6px 14px;
                border: 1px solid #2a4272;
                background-color: #102042;
                color: #dfeaff;
                font-weight: 600;
            }

            QPushButton:hover {
                background-color: #162a56;
                border: 1px solid #3c5fa1;
            }

            QPushButton:pressed {
                background-color: #0d1a37;
            }

            QPushButton:disabled {
                background-color: #0b1430;
                color: #5c7096;
                border: 1px solid #1a2850;
            }

            QPushButton#smallActionBtn {
                min-height: 28px;
                padding: 4px 10px;
                border-radius: 8px;
                background-color: #13254b;
                color: #dce7ff;
            }

            QPushButton#primaryActionBtn {
                min-height: 36px;
                border-radius: 10px;
                font-size: 14px;
                font-weight: 800;
                color: white;
            }

            QPushButton#primaryActionBtn[running="false"] {
                border: 1px solid #645bff;
                background-color: qlineargradient(
                    x1:0, y1:0, x2:1, y2:0,
                    stop:0 #5b5cf6,
                    stop:1 #6a38ff
                );
            }

            QPushButton#primaryActionBtn[running="false"]:hover {
                background-color: qlineargradient(
                    x1:0, y1:0, x2:1, y2:0,
                    stop:0 #6b6cff,
                    stop:1 #7d4cff
                );
            }

            QPushButton#primaryActionBtn[running="true"] {
                background-color: #8d0f3e;
                border: 1px solid #df4a7a;
                color: #ffd9e6;
            }

            QPushButton#primaryActionBtn[running="true"]:hover {
                background-color: #a11247;
                border: 1px solid #f05a8c;
            }

            QPushButton#exportBtn {
                min-height: 28px;
                padding: 4px 12px;
                border-radius: 8px;
                background-color: #16284f;
                color: #dfe8ff;
            }

            QTextEdit#logEdit {
                background-color: #050d22;
                border: 1px solid #1f315d;
                border-radius: 8px;
                color: #7cecc8;
                font-family: Consolas, "Courier New", monospace;
                font-size: 11px;
            }

            QTableWidget {
                background-color: #030b1f;
                border: 1px solid #1f315d;
                border-radius: 8px;
                gridline-color: #15284f;
                color: #dbe7ff;
            }

            QHeaderView::section {
                background-color: #08142f;
                color: #c7d7f6;
                border: none;
                border-bottom: 1px solid #16305c;
                padding: 6px;
                font-weight: 700;
            }

            QTableWidget::item {
                padding: 6px;
                border-bottom: 1px solid #102448;
            }

            QProgressBar {
                background-color: #152749;
                border: none;
                border-radius: 4px;
                text-align: center;
                color: #b7c8ea;
                min-height: 8px;
                max-height: 8px;
            }

            QProgressBar::chunk {
                background-color: #5b5cf6;
                border-radius: 4px;
            }

            QCheckBox {
                color: #dbe7ff;
                background: transparent;
                spacing: 6px;
            }

            QCheckBox::indicator {
                width: 16px;
                height: 16px;
                image: url("__UNCHECKED__");
            }

            QCheckBox::indicator:checked {
                image: url("__CHECKED__");
            }
        """ + SCROLLBAR_STYLE).replace("__UNCHECKED__", _cb_icons['unchecked']).replace("__CHECKED__", _cb_icons['checked'])
        self.setStyleSheet(full_style)

    def _create_layout(self):
        root_layout = QVBoxLayout(self)
        root_layout.setContentsMargins(6, 6, 6, 6)
        root_layout.setSpacing(10)

        title_layout = QVBoxLayout()
        title_layout.setSpacing(2)

        self.page_title = QLabel("Load Capability and Output Ripple Test")
        self.page_title.setObjectName("pageTitle")

        self.page_subtitle = QLabel("Configure and execute load capability and output ripple test sequences.")
        self.page_subtitle.setObjectName("pageSubtitle")

        title_layout.addWidget(self.page_title)
        title_layout.addWidget(self.page_subtitle)
        root_layout.addLayout(title_layout)

        content_layout = QHBoxLayout()
        content_layout.setSpacing(10)
        root_layout.addLayout(content_layout, 1)

        left_wrapper = QVBoxLayout()
        left_wrapper.setContentsMargins(0, 0, 0, 0)
        left_wrapper.setSpacing(8)

        self.left_scroll = QScrollArea()
        self.left_scroll.setWidgetResizable(True)
        self.left_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.left_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.left_scroll.setFixedWidth(340)
        self.left_scroll.setObjectName("leftScrollArea")
        self.left_scroll.setStyleSheet("""
            QScrollArea#leftScrollArea {
                background-color: #08132d;
                border: 1px solid #16274d;
                border-radius: 18px;
            }
        """ + SCROLLBAR_STYLE)

        self.left_panel = QWidget()
        self.left_panel.setObjectName("leftPanelInner")

        left_layout = QVBoxLayout(self.left_panel)
        left_layout.setContentsMargins(10, 10, 10, 10)
        left_layout.setSpacing(10)

        self.connection_card = CardFrame("⚡ INSTRUMENTS CONNECTION")
        self._build_connection_card()
        left_layout.addWidget(self.connection_card)

        self.channel_card = CardFrame("☷ CHANNEL SELECTION")
        self._build_channel_card()
        left_layout.addWidget(self.channel_card)

        self.is_gain_card = CardFrame("◉ IS_GAIN CONFIG")
        self._build_is_gain_card()
        left_layout.addWidget(self.is_gain_card)

        left_layout.addStretch()

        self.left_scroll.setWidget(self.left_panel)
        left_wrapper.addWidget(self.left_scroll, 1)

        self.start_test_btn = QPushButton("▷  Start Sequence")
        self.start_test_btn.setObjectName("primaryActionBtn")
        self.start_test_btn.setProperty("running", "false")
        left_wrapper.addWidget(self.start_test_btn)

        self.stop_test_btn = QPushButton("Abort Test")
        self.stop_test_btn.hide()

        content_layout.addLayout(left_wrapper)

        right_layout = QVBoxLayout()
        right_layout.setSpacing(10)
        content_layout.addLayout(right_layout, 1)

        self.result_frame = QFrame()
        self.result_frame.setObjectName("resultContainer")
        result_layout = QVBoxLayout(self.result_frame)
        result_layout.setContentsMargins(12, 12, 12, 12)
        result_layout.setSpacing(10)

        result_header = QHBoxLayout()
        self.result_title = QLabel("▦ Test Results")
        self.result_title.setObjectName("sectionTitle")
        result_header.addWidget(self.result_title)
        result_header.addStretch()

        self.export_result_btn = QPushButton("⇩ Export")
        self.export_result_btn.setObjectName("exportBtn")
        result_header.addWidget(self.export_result_btn)

        result_layout.addLayout(result_header)

        self.result_table = QTableWidget(0, 7)
        self.result_table.setHorizontalHeaderLabels([
            "STEP", "LOAD CURRENT (A)", "VOLTAGE (V)", "RIPPLE (mV)",
            "V_DROP (mV)", "SCREENSHOT", "REMARK"
        ])
        self.result_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.result_table.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)
        self.result_table.verticalHeader().setVisible(False)
        self.result_table.setAlternatingRowColors(False)
        self.result_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.result_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.result_table.setShowGrid(False)
        result_layout.addWidget(self.result_table)

        right_layout.addWidget(self.result_frame, 5)

        self.log_frame = QFrame()
        self.log_frame.setObjectName("logContainer")
        log_layout = QVBoxLayout(self.log_frame)
        log_layout.setContentsMargins(12, 12, 12, 12)
        log_layout.setSpacing(10)

        log_header = QHBoxLayout()
        self.log_title = QLabel("›_ EXECUTION LOGS")
        self.log_title.setObjectName("sectionTitle")
        log_header.addWidget(self.log_title)
        log_header.addStretch()

        self.progress_text_label = QLabel("0% Complete")
        self.progress_text_label.setObjectName("fieldLabel")
        log_header.addWidget(self.progress_text_label)

        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.progress_bar.setTextVisible(False)
        self.progress_bar.setFixedWidth(120)
        log_header.addWidget(self.progress_bar)

        self.clear_log_btn = QPushButton("Clear")
        self.clear_log_btn.setObjectName("smallActionBtn")
        log_header.addWidget(self.clear_log_btn)

        log_layout.addLayout(log_header)

        self.log_edit = QTextEdit()
        self.log_edit.setObjectName("logEdit")
        self.log_edit.setReadOnly(True)
        self.log_edit.setMinimumHeight(110)
        log_layout.addWidget(self.log_edit)

        right_layout.addWidget(self.log_frame, 1)

    def _build_connection_card(self):
        layout = self.connection_card.main_layout

        n6705c_label = QLabel("N6705C DC Power Analyzer")
        n6705c_label.setObjectName("fieldLabel")
        layout.addWidget(n6705c_label)

        self.visa_resource_combo = FixedPopupComboBox()
        self.visa_resource_combo.addItem("TCPIP0::K-N6705C-06098.local::hislip0::INSTR")
        layout.addWidget(self.visa_resource_combo)

        n670_row = QHBoxLayout()
        n670_row.setSpacing(8)

        self.search_btn = SpinningSearchButton()

        self.connect_btn = QPushButton()
        update_connect_button_state(self.connect_btn, connected=False)

        n670_row.addWidget(self.search_btn)
        n670_row.addWidget(self.connect_btn)
        layout.addLayout(n670_row)

        scope_label = QLabel("Oscilloscope")
        scope_label.setObjectName("fieldLabel")
        layout.addWidget(scope_label)

        self.scope_type_combo = FixedPopupComboBox()
        self.scope_type_combo.addItems(["DSOX4034A", "MSO64B"])
        layout.addWidget(self.scope_type_combo)

        self.scope_resource_combo = FixedPopupComboBox()
        self.scope_resource_combo.setEditable(True)
        self.scope_resource_combo.addItem("USB0::0x0957::0x17A4::MY61500152::INSTR")
        layout.addWidget(self.scope_resource_combo)

        scope_row = QHBoxLayout()
        scope_row.setSpacing(8)

        self.scope_search_btn = SpinningSearchButton()

        self.scope_connect_btn = QPushButton()
        update_connect_button_state(self.scope_connect_btn, connected=False)

        scope_row.addWidget(self.scope_search_btn)
        scope_row.addWidget(self.scope_connect_btn)
        layout.addLayout(scope_row)

        test_selection_label = QLabel("Test Selection")
        test_selection_label.setObjectName("fieldLabel")
        layout.addWidget(test_selection_label)

        self.test_selection_combo = FixedPopupComboBox()
        self.test_selection_combo.addItems(["单次 Is_gain 测试", "遍历 Is_gain 测试"])
        layout.addWidget(self.test_selection_combo)

    def _build_channel_card(self):
        layout = self.channel_card.main_layout
        grid = QGridLayout()
        grid.setHorizontalSpacing(10)
        grid.setVerticalSpacing(10)

        lbl_ripple = QLabel("Ripple Channel (Oscilloscope)")
        lbl_ripple.setObjectName("fieldLabel")
        self.ripple_channel_combo = FixedPopupComboBox()
        self.ripple_channel_combo.addItems(["CH 1", "CH 2", "CH 3", "CH 4"])

        lbl_load = QLabel("Load Channel (N6705C)")
        lbl_load.setObjectName("fieldLabel")
        self.load_channel_combo = FixedPopupComboBox()
        self.load_channel_combo.addItems(["CH 1", "CH 2", "CH 3", "CH 4"])
        self.load_channel_combo.setCurrentIndex(1)

        grid.addWidget(lbl_ripple, 0, 0)
        grid.addWidget(self.ripple_channel_combo, 0, 1)

        grid.addWidget(lbl_load, 1, 0)
        grid.addWidget(self.load_channel_combo, 1, 1)

        layout.addLayout(grid)

    def _build_is_gain_card(self):
        layout = self.is_gain_card.main_layout

        top_row = QHBoxLayout()

        mode_label = QLabel("测试方法")
        mode_label.setObjectName("fieldLabel")
        self.is_gain_method_combo = FixedPopupComboBox()
        self.is_gain_method_combo.addItems(["CV测试法", "CC测试法"])

        top_row.addWidget(mode_label)
        top_row.addWidget(self.is_gain_method_combo, 1)
        layout.addLayout(top_row)

        grid = QGridLayout()
        grid.setHorizontalSpacing(10)
        grid.setVerticalSpacing(6)

        self.lbl_dev = QLabel("Device Addr")
        self.lbl_dev.setObjectName("fieldLabel")
        self.is_gain_device_addr_edit = QLineEdit("0x60")

        self.lbl_reg = QLabel("Reg Addr")
        self.lbl_reg.setObjectName("fieldLabel")
        self.is_gain_reg_addr_edit = QLineEdit("0x01")

        self.lbl_msb = QLabel("MSB")
        self.lbl_msb.setObjectName("fieldLabel")
        self.is_gain_msb_spin = QSpinBox()
        self.is_gain_msb_spin.setRange(0, 255)
        self.is_gain_msb_spin.setValue(7)

        self.lbl_lsb = QLabel("LSB")
        self.lbl_lsb.setObjectName("fieldLabel")
        self.is_gain_lsb_spin = QSpinBox()
        self.is_gain_lsb_spin.setRange(0, 255)
        self.is_gain_lsb_spin.setValue(4)

        grid.addWidget(self.lbl_dev, 0, 0)
        grid.addWidget(self.lbl_reg, 0, 1)
        grid.addWidget(self.is_gain_device_addr_edit, 1, 0)
        grid.addWidget(self.is_gain_reg_addr_edit, 1, 1)

        grid.addWidget(self.lbl_msb, 2, 0)
        grid.addWidget(self.lbl_lsb, 2, 1)
        grid.addWidget(self.is_gain_msb_spin, 3, 0)
        grid.addWidget(self.is_gain_lsb_spin, 3, 1)

        self._traverse_only_widgets = [
            self.lbl_dev, self.lbl_reg,
            self.is_gain_device_addr_edit, self.is_gain_reg_addr_edit,
            self.lbl_msb, self.lbl_lsb,
            self.is_gain_msb_spin, self.is_gain_lsb_spin,
        ]

        lbl_start_current = QLabel("Start Current (A)")
        lbl_start_current.setObjectName("fieldLabel")
        self.is_gain_start_current_spin = QDoubleSpinBox()
        self.is_gain_start_current_spin.setDecimals(3)
        self.is_gain_start_current_spin.setRange(-9999.0, 9999.0)
        self.is_gain_start_current_spin.setSingleStep(0.001)
        self.is_gain_start_current_spin.setValue(0.000)

        lbl_end_current = QLabel("End Current (A)")
        lbl_end_current.setObjectName("fieldLabel")
        self.is_gain_end_current_spin = QDoubleSpinBox()
        self.is_gain_end_current_spin.setDecimals(3)
        self.is_gain_end_current_spin.setRange(-9999.0, 9999.0)
        self.is_gain_end_current_spin.setSingleStep(0.001)
        self.is_gain_end_current_spin.setValue(0.250)

        lbl_step_current = QLabel("Step Current (A)")
        lbl_step_current.setObjectName("fieldLabel")
        self.is_gain_step_current_spin = QDoubleSpinBox()
        self.is_gain_step_current_spin.setDecimals(3)
        self.is_gain_step_current_spin.setRange(-9999.0, 9999.0)
        self.is_gain_step_current_spin.setSingleStep(0.001)
        self.is_gain_step_current_spin.setValue(0.01)

        grid.addWidget(lbl_start_current, 4, 0)
        grid.addWidget(lbl_end_current, 4, 1)
        grid.addWidget(self.is_gain_start_current_spin, 5, 0)
        grid.addWidget(self.is_gain_end_current_spin, 5, 1)

        grid.addWidget(lbl_step_current, 6, 0)
        grid.addWidget(self.is_gain_step_current_spin, 7, 0)

        self.save_screenshot_cb = QCheckBox("Save Screenshot")
        self.save_screenshot_cb.setChecked(True)
        grid.addWidget(self.save_screenshot_cb, 7, 1)

        layout.addLayout(grid)

    def _init_ui_elements(self):
        self._update_connect_button_state(self.connect_btn, False)
        self._update_connect_button_state(self.scope_connect_btn, False)
        self._update_test_button_state(False)
        self._on_test_selection_changed()
        self.append_log("[SYSTEM] Ready. Waiting for instrument connection.")
        self.set_progress(0)
        self.stop_test_btn.setEnabled(False)

    def _bind_signals(self):
        self.search_btn.clicked.connect(self._on_search)
        self.connect_btn.clicked.connect(self._on_connect_or_disconnect_n6705c)
        self.scope_search_btn.clicked.connect(self._on_scope_search)
        self.scope_connect_btn.clicked.connect(self._on_connect_or_disconnect_scope)
        self.test_selection_combo.currentIndexChanged.connect(self._on_test_selection_changed)
        self.start_test_btn.clicked.connect(self._on_start_or_abort_clicked)
        self.stop_test_btn.clicked.connect(self._abort_test_from_external)
        self.export_result_btn.clicked.connect(self._on_export)
        self.clear_log_btn.clicked.connect(self._on_clear_log)

    def _abort_test_from_external(self):
        if self.is_test_running:
            if self._test_worker is not None:
                self._test_worker.stop()
            self.append_log("[TEST] Abort requested by external stop button.")

    def _on_test_selection_changed(self):
        is_traverse = self.test_selection_combo.currentText() == "遍历 Is_gain 测试"
        for w in self._traverse_only_widgets:
            w.setVisible(is_traverse)

    def _update_connect_button_state(self, button: QPushButton, connected: bool):
        update_connect_button_state(button, connected)

    def _sync_from_top(self):
        if self._n6705c_top:
            if self._n6705c_top.is_connected_a and self._n6705c_top.n6705c_a:
                self.n6705c = self._n6705c_top.n6705c_a
                self.is_connected = True
                self._update_connect_button_state(self.connect_btn, True)
                self.search_btn.setEnabled(False)
                if self._n6705c_top.visa_resource_a:
                    self.visa_resource_combo.clear()
                    self.visa_resource_combo.addItem(self._n6705c_top.visa_resource_a)
            elif not self.is_connected:
                self._update_connect_button_state(self.connect_btn, False)

        if self._mso64b_top:
            if self._mso64b_top.is_connected and self._mso64b_top.mso64b:
                self.Osc_ins = self._mso64b_top.mso64b
                self.scope_resource = self._mso64b_top.visa_resource
                self.scope_connected = True
                self._update_connect_button_state(self.scope_connect_btn, True)
                self.scope_search_btn.setEnabled(False)
                scope_type = getattr(self._mso64b_top, 'scope_type', 'MSO64B') or 'MSO64B'
                idx = self.scope_type_combo.findText(scope_type)
                if idx >= 0:
                    self.scope_type_combo.setCurrentIndex(idx)
                self.scope_type_combo.setEnabled(False)
                if self._mso64b_top.visa_resource:
                    self.scope_resource_combo.clear()
                    self.scope_resource_combo.addItem(self._mso64b_top.visa_resource)
            elif not self.scope_connected:
                self._update_connect_button_state(self.scope_connect_btn, False)

    def _on_mso64b_top_changed(self):
        if self._mso64b_top is None:
            return
        if self.is_test_running:
            return
        if self._mso64b_top.is_connected and self._mso64b_top.mso64b:
            if self.Osc_ins is self._mso64b_top.mso64b and self.scope_connected:
                return
            self.Osc_ins = self._mso64b_top.mso64b
            self.scope_resource = self._mso64b_top.visa_resource
            self.scope_connected = True
            self._update_connect_button_state(self.scope_connect_btn, True)
            self.scope_search_btn.setEnabled(False)
            scope_type = getattr(self._mso64b_top, 'scope_type', 'MSO64B') or 'MSO64B'
            idx = self.scope_type_combo.findText(scope_type)
            if idx >= 0:
                self.scope_type_combo.setCurrentIndex(idx)
            self.scope_type_combo.setEnabled(False)
            if self._mso64b_top.visa_resource:
                self.scope_resource_combo.clear()
                self.scope_resource_combo.addItem(self._mso64b_top.visa_resource)
            self.append_log(f"[SYSTEM] {scope_type} synced from external connection.")
        else:
            if not self.scope_connected:
                return
            self.Osc_ins = None
            self.scope_resource = None
            self.scope_connected = False
            self._update_connect_button_state(self.scope_connect_btn, False)
            self.scope_type_combo.setEnabled(True)
            self.scope_search_btn.setEnabled(True)
            self.append_log("[SYSTEM] Oscilloscope disconnected externally.")

    def _update_test_button_state(self, running: bool):
        self.is_test_running = running
        self.start_test_btn.setProperty("running", "true" if running else "false")
        self.start_test_btn.setText("□  Abort Test" if running else "▷  Start Sequence")
        self.start_test_btn.style().unpolish(self.start_test_btn)
        self.start_test_btn.style().polish(self.start_test_btn)
        self.start_test_btn.update()

    def append_log(self, message):
        self.log_edit.append(message)

    def _on_clear_log(self):
        self.log_edit.clear()

    def set_progress(self, value: int):
        value = max(0, min(100, int(value)))
        self.progress_bar.setValue(value)
        self.progress_text_label.setText(f"{value}% Complete")

    def get_test_config(self):
        return {
            "ripple_channel": self.ripple_channel_combo.currentText(),
            "load_channel": self.load_channel_combo.currentText(),
            "is_gain_method": self.is_gain_method_combo.currentText(),
            "is_gain_device_addr": self.is_gain_device_addr_edit.text().strip(),
            "is_gain_reg_addr": self.is_gain_reg_addr_edit.text().strip(),
            "is_gain_msb": self.is_gain_msb_spin.value(),
            "is_gain_lsb": self.is_gain_lsb_spin.value(),
            "is_gain_start_current": self.is_gain_start_current_spin.value(),
            "is_gain_end_current": self.is_gain_end_current_spin.value(),
            "is_gain_step_current": self.is_gain_step_current_spin.value(),
            "save_screenshot": self.save_screenshot_cb.isChecked(),
        }

    def set_test_running(self, running):
        self._update_test_button_state(running)
        self.stop_test_btn.setEnabled(running)

        widgets = [
            self.ripple_channel_combo,
            self.load_channel_combo,
            self.is_gain_method_combo,
            self.is_gain_device_addr_edit,
            self.is_gain_reg_addr_edit,
            self.is_gain_msb_spin,
            self.is_gain_lsb_spin,
            self.is_gain_start_current_spin,
            self.is_gain_end_current_spin,
            self.is_gain_step_current_spin,
            self.save_screenshot_cb,

            self.visa_resource_combo,
            self.scope_type_combo,
            self.scope_resource_combo,
            self.search_btn,
            self.scope_search_btn,
            self.connect_btn,
            self.scope_connect_btn,
            self.test_selection_combo
        ]

        for widget in widgets:
            widget.setEnabled(not running)

        if running:
            self.append_log("[TEST] Starting Is_gain Test Sequence...")
        else:
            self.append_log("[TEST] Test stopped or completed.")

    def clear_results(self):
        self.result_table.setRowCount(0)
        self._test_result_data = []
        self.set_progress(0)
        self.append_log("[SYSTEM] Results cleared.")

    def add_result_row(self, step, load_current, voltage, ripple, v_drop=None, has_screenshot=False, remark=""):
        row = self.result_table.rowCount()
        self.result_table.insertRow(row)

        voltage_str = f"{voltage:.6f}" if voltage is not None else "N/A"
        ripple_mv = ripple * 1000 if ripple is not None else None
        ripple_str = f"{ripple_mv:.3f}" if ripple_mv is not None else "N/A"
        v_drop_mv = v_drop * 1000 if v_drop is not None else None
        v_drop_str = f"{v_drop_mv:.3f}" if v_drop_mv is not None else "N/A"
        screenshot_str = "✔" if has_screenshot else ""

        items = [
            QTableWidgetItem(str(step)),
            QTableWidgetItem(f"{load_current:.3f}"),
            QTableWidgetItem(voltage_str),
            QTableWidgetItem(ripple_str),
            QTableWidgetItem(v_drop_str),
            QTableWidgetItem(screenshot_str),
            QTableWidgetItem(remark),
        ]

        colors = ["#eaf2ff", "#5f8cff", "#00d6a2", "#f5a623", "#ff6b6b", "#38bdf8", "#c084fc"]
        for col, item in enumerate(items):
            item.setTextAlignment(Qt.AlignCenter)
            item.setForeground(QColor(colors[col]))
            self.result_table.setItem(row, col, item)

    def _on_test_result_row(self, row_data):
        self._test_result_data.append(row_data)

        step = row_data.get("step", "")
        load_current = row_data.get("load_current", 0)
        voltage = row_data.get("voltage")
        ripple = row_data.get("ripple")
        v_drop = row_data.get("v_drop")
        reg_value = row_data.get("reg_value")

        voltage_str = f"{voltage:.6f}" if voltage is not None else "N/A"
        ripple_mv = ripple * 1000 if ripple is not None else None
        ripple_str = f"{ripple_mv:.3f}" if ripple_mv is not None else "N/A"
        v_drop_mv = v_drop * 1000 if v_drop is not None else None
        v_drop_str = f"{v_drop_mv:.3f}" if v_drop_mv is not None else "N/A"

        if reg_value is not None:
            self.append_log(
                f"[DATA] Step={step}\tReg={reg_value}\tLoad={load_current:.3f}A\t"
                f"V={voltage_str}V\tRipple={ripple_str}mV\tVdrop={v_drop_str}mV"
            )
        else:
            self.append_log(
                f"[DATA] Step={step}\tLoad={load_current:.3f}A\t"
                f"V={voltage_str}V\tRipple={ripple_str}mV\tVdrop={v_drop_str}mV"
            )

        self.add_result_row(
            step,
            load_current,
            voltage,
            ripple,
            v_drop,
            row_data.get("screenshot_b64") is not None,
            row_data.get("remark", ""),
        )

    def _on_test_summary(self, analysis):
        v0 = analysis.get("v0")
        max_load = analysis.get("max_load_current")
        max_rp = analysis.get("max_ripple")
        max_rp_cur = analysis.get("max_ripple_current")

        self.append_log("=" * 50)
        self.append_log("[SUMMARY] Test Analysis Results:")
        if v0 is not None:
            self.append_log(f"  0-load baseline voltage: {v0:.6f} V")
        if max_load is not None:
            self.append_log(f"  Max load capacity (Vdrop <= 30mV): {max_load:.3f} A")
        else:
            self.append_log("  Max load capacity: N/A (voltage drop > 30mV at all loads)")
        if max_rp is not None:
            self.append_log(f"  Max ripple: {max_rp * 1000:.3f} mV @ {max_rp_cur:.3f} A")
        self.append_log("=" * 50)

    def _on_export(self):
        if not self._test_result_data:
            self.set_system_status("No data to export", is_error=True)
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Export Test Results")
        dialog.setFixedWidth(460)
        dialog.setStyleSheet("""
            QDialog {
                background-color: #0a1628;
                color: #c8daf5;
            }
            QLabel {
                color: #8eb0e3;
                font-size: 12px;
                background: transparent;
                border: none;
            }
            QLineEdit {
                background-color: #0c1a35;
                border: 1px solid #1e3460;
                border-radius: 6px;
                color: #eaf2ff;
                padding: 6px 10px;
                font-size: 12px;
            }
            QLineEdit:focus {
                border-color: #3a6fd4;
            }
            QRadioButton {
                color: #dbe7ff;
                font-size: 12px;
                spacing: 6px;
                background: transparent;
            }
            QRadioButton::indicator {
                width: 14px;
                height: 14px;
                border-radius: 7px;
                border: 2px solid #3a5a8c;
                background-color: #0c1a35;
            }
            QRadioButton::indicator:checked {
                background-color: #5b5cf6;
                border-color: #7b7cff;
            }
            QRadioButton::indicator:hover {
                border-color: #5b5cf6;
            }
            QPushButton {
                background-color: #162d55;
                border: 1px solid #1e3460;
                border-radius: 6px;
                color: #c8daf5;
                padding: 6px 18px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #1e3460;
                border-color: #3a6fd4;
            }
            QFrame#exportSeparator {
                background-color: #1e3460;
                max-height: 1px;
                border: none;
            }
        """)

        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(14)

        title = QLabel("⇩  Export Test Results")
        title.setStyleSheet("font-size: 15px; font-weight: bold; color: #f4f7ff;")
        layout.addWidget(title)

        sep1 = QFrame()
        sep1.setObjectName("exportSeparator")
        sep1.setFixedHeight(1)
        layout.addWidget(sep1)

        info_grid = QGridLayout()
        info_grid.setHorizontalSpacing(12)
        info_grid.setVerticalSpacing(10)

        lbl_dut = QLabel("DUT Name")
        self._export_dut_edit = QLineEdit()
        self._export_dut_edit.setPlaceholderText("e.g. PMU_A1")
        info_grid.addWidget(lbl_dut, 0, 0)
        info_grid.addWidget(self._export_dut_edit, 0, 1)

        lbl_module = QLabel("Module")
        self._export_module_edit = QLineEdit()
        self._export_module_edit.setPlaceholderText("e.g. LDO1")
        info_grid.addWidget(lbl_module, 1, 0)
        info_grid.addWidget(self._export_module_edit, 1, 1)

        lbl_cond = QLabel("Condition")
        self._export_condition_edit = QLineEdit()
        self._export_condition_edit.setPlaceholderText("e.g. Room Temp 25°C")
        info_grid.addWidget(lbl_cond, 2, 0)
        info_grid.addWidget(self._export_condition_edit, 2, 1)

        lbl_operator = QLabel("Operator")
        self._export_operator_edit = QLineEdit()
        self._export_operator_edit.setPlaceholderText("e.g. Zhang San")
        info_grid.addWidget(lbl_operator, 3, 0)
        info_grid.addWidget(self._export_operator_edit, 3, 1)

        lbl_remark = QLabel("Remark")
        self._export_remark_edit = QLineEdit()
        self._export_remark_edit.setPlaceholderText("(optional)")
        info_grid.addWidget(lbl_remark, 4, 0)
        info_grid.addWidget(self._export_remark_edit, 4, 1)

        layout.addLayout(info_grid)

        sep2 = QFrame()
        sep2.setObjectName("exportSeparator")
        sep2.setFixedHeight(1)
        layout.addWidget(sep2)

        fmt_label = QLabel("Export Format")
        fmt_label.setStyleSheet("font-size: 12px; font-weight: bold; color: #f4f7ff;")
        layout.addWidget(fmt_label)

        fmt_group = QButtonGroup(dialog)
        fmt_layout = QHBoxLayout()
        fmt_layout.setSpacing(20)

        self._rb_xlsx = QRadioButton("Excel (.xlsx)")
        self._rb_pdf = QRadioButton("PDF (.pdf)")
        self._rb_csv = QRadioButton("CSV (.csv)")
        self._rb_xlsx.setChecked(True)

        fmt_group.addButton(self._rb_xlsx)
        fmt_group.addButton(self._rb_pdf)
        fmt_group.addButton(self._rb_csv)

        fmt_layout.addWidget(self._rb_xlsx)
        fmt_layout.addWidget(self._rb_pdf)
        fmt_layout.addWidget(self._rb_csv)
        fmt_layout.addStretch()
        layout.addLayout(fmt_layout)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(dialog.reject)
        export_btn = QPushButton("⇩  Export")
        export_btn.setStyleSheet("""
            QPushButton {
                background-color: #4f46e5;
                border: 1px solid #645bff;
                border-radius: 6px;
                color: #ffffff;
                padding: 7px 24px;
                font-size: 12px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #5b5cf6;
                border-color: #7b7cff;
            }
        """)
        export_btn.clicked.connect(dialog.accept)
        btn_layout.addWidget(cancel_btn)
        btn_layout.addWidget(export_btn)
        layout.addLayout(btn_layout)

        if dialog.exec() != QDialog.Accepted:
            return

        export_info = {
            "dut_name": self._export_dut_edit.text().strip(),
            "module": self._export_module_edit.text().strip(),
            "condition": self._export_condition_edit.text().strip(),
            "operator": self._export_operator_edit.text().strip(),
            "remark": self._export_remark_edit.text().strip(),
            "test_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

        dut = export_info["dut_name"] or "DUT"
        module = export_info["module"] or "Module"
        base_name = f"{dut}_{module}_is_gainAndRipple_Test"

        if self._rb_xlsx.isChecked():
            fmt_filter = "Excel Files (*.xlsx)"
            default_name = f"{base_name}.xlsx"
        elif self._rb_pdf.isChecked():
            fmt_filter = "PDF Files (*.pdf)"
            default_name = f"{base_name}.pdf"
        else:
            fmt_filter = "CSV Files (*.csv)"
            default_name = f"{base_name}.csv"

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export Results", default_name, fmt_filter
        )
        if not file_path:
            return

        try:
            if file_path.endswith(".xlsx"):
                self._export_xlsx(file_path, export_info)
            elif file_path.endswith(".pdf"):
                self._export_pdf(file_path, export_info)
            else:
                self._export_csv(file_path, export_info)
            self.append_log(f"[EXPORT] Results saved to: {file_path}")
            self.set_system_status(f"Exported to {os.path.basename(file_path)}")
        except Exception as e:
            self.append_log(f"[ERROR] Export failed: {e}")
            self.set_system_status("Export failed", is_error=True)

    def _export_csv(self, file_path, export_info=None):
        with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            if export_info:
                writer.writerow(["DUT Name", export_info.get("dut_name", "")])
                writer.writerow(["Module", export_info.get("module", "")])
                writer.writerow(["Condition", export_info.get("condition", "")])
                writer.writerow(["Operator", export_info.get("operator", "")])
                writer.writerow(["Remark", export_info.get("remark", "")])
                writer.writerow(["Test Time", export_info.get("test_time", "")])
                writer.writerow([])
            writer.writerow(["Step", "Load Current (A)", "Voltage (V)", "Ripple (mV)", "V_Drop (mV)", "Remark"])
            for r in self._test_result_data:
                rp = r.get("ripple")
                vd = r.get("v_drop")
                writer.writerow([
                    r.get("step", ""),
                    f"{r.get('load_current', 0):.3f}",
                    f"{r['voltage']:.6f}" if r.get("voltage") is not None else "N/A",
                    f"{rp * 1000:.3f}" if rp is not None else "N/A",
                    f"{vd * 1000:.3f}" if vd is not None else "N/A",
                    r.get("remark", ""),
                ])

    def _export_xlsx(self, file_path, export_info=None):
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XlImage
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font as XlFont, PatternFill, Alignment, Border, Side
        import io

        xlsx_dir = os.path.dirname(os.path.abspath(file_path))
        xlsx_name = os.path.splitext(os.path.basename(file_path))[0]
        screenshots_dir = os.path.join(xlsx_dir, f"{xlsx_name}_screenshots")

        has_screenshots = any(r.get("screenshot_b64") for r in self._test_result_data)
        if has_screenshots:
            os.makedirs(screenshots_dir, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Is_gain Results"

        header_font = XlFont(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F4070", end_color="2F4070", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        info_key_font = XlFont(bold=True, color="2F4070", size=11)
        info_val_font = XlFont(color="333333", size=11)
        thin_border = Border(
            left=Side(style="thin", color="C0C0C0"),
            right=Side(style="thin", color="C0C0C0"),
            top=Side(style="thin", color="C0C0C0"),
            bottom=Side(style="thin", color="C0C0C0"),
        )

        current_row = 1

        if export_info:
            title_cell = ws.cell(row=current_row, column=1, value="Is_gain Test Report")
            title_cell.font = XlFont(bold=True, color="1a1a2e", size=16)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
            current_row += 2

            info_items = [
                ("DUT Name", export_info.get("dut_name", "")),
                ("Module", export_info.get("module", "")),
                ("Condition", export_info.get("condition", "")),
                ("Operator", export_info.get("operator", "")),
                ("Test Time", export_info.get("test_time", "")),
            ]
            if export_info.get("remark"):
                info_items.append(("Remark", export_info["remark"]))

            for key, val in info_items:
                ws.cell(row=current_row, column=1, value=key).font = info_key_font
                ws.cell(row=current_row, column=2, value=val).font = info_val_font
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
                current_row += 1

            current_row += 1

        headers = ["Step", "Load Current (A)", "Voltage (V)", "Ripple (mV)", "V_Drop (mV)", "Screenshot", "Remark"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        img_col_width = 45
        ws.column_dimensions[get_column_letter(6)].width = img_col_width
        img_row_height = 180
        data_start_row = current_row + 1

        saved_screenshots = []

        for idx, r in enumerate(self._test_result_data):
            row_idx = data_start_row + idx

            ws.cell(row=row_idx, column=1, value=r.get("step", "")).border = thin_border
            ws.cell(row=row_idx, column=2, value=round(r.get("load_current", 0), 3)).border = thin_border
            v = r.get("voltage")
            ws.cell(row=row_idx, column=3, value=round(v, 6) if v is not None else "N/A").border = thin_border
            rp = r.get("ripple")
            ws.cell(row=row_idx, column=4, value=round(rp * 1000, 3) if rp is not None else "N/A").border = thin_border
            vd = r.get("v_drop")
            ws.cell(row=row_idx, column=5, value=round(vd * 1000, 3) if vd is not None else "N/A").border = thin_border
            ws.cell(row=row_idx, column=7, value=r.get("remark", "")).border = thin_border

            for c in range(1, 8):
                ws.cell(row=row_idx, column=c).alignment = Alignment(horizontal="center", vertical="center")

            sc_b64 = r.get("screenshot_b64")
            if sc_b64:
                try:
                    img_data = base64.b64decode(sc_b64)
                    png_sig = b'\x89PNG\r\n\x1a\n'
                    bmp_sig = b'BM'
                    if img_data[:8] != png_sig and img_data[:2] != bmp_sig:
                        ws.cell(row=row_idx, column=6, value="(invalid image data)")
                        continue

                    step_val = r.get("step", idx)
                    reg_val = r.get("reg_value")
                    if reg_val is not None:
                        png_name = f"step{step_val}_reg{reg_val}.png"
                    else:
                        png_name = f"step{step_val}.png"
                    png_path = os.path.join(screenshots_dir, png_name)
                    with open(png_path, "wb") as f:
                        f.write(img_data)
                    saved_screenshots.append(png_path)

                    cell = ws.cell(row=row_idx, column=6, value=png_name)
                    cell.hyperlink = png_path
                    cell.font = XlFont(color="4472C4", underline="single")
                    cell.border = thin_border

                    img_stream = io.BytesIO(img_data)
                    img = XlImage(img_stream)

                    col_px = img_col_width * 7
                    row_px = img_row_height * 1.33
                    scale = min(col_px / img.width, row_px / img.height)
                    img.width = int(img.width * scale)
                    img.height = int(img.height * scale)

                    ws.add_image(img, f"F{row_idx}")
                    ws.row_dimensions[row_idx].height = img_row_height
                except Exception as e:
                    ws.cell(row=row_idx, column=6, value=f"(image error: {e})")
            else:
                cell = ws.cell(row=row_idx, column=6, value="")
                cell.border = thin_border

        wb.save(file_path)

        if saved_screenshots:
            self.append_log(f"[EXPORT] {len(saved_screenshots)} screenshots saved to: {screenshots_dir}")

    def _export_pdf(self, file_path, export_info=None):
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import io

        font_name = "Helvetica"
        font_name_bold = "Helvetica-Bold"
        try:
            import platform
            if platform.system() == "Windows":
                win_font = os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Fonts", "msyh.ttc")
                if os.path.exists(win_font):
                    pdfmetrics.registerFont(TTFont("MSYH", win_font, subfontIndex=0))
                    pdfmetrics.registerFont(TTFont("MSYH-Bold", win_font, subfontIndex=1))
                    font_name = "MSYH"
                    font_name_bold = "MSYH-Bold"
        except Exception:
            pass

        doc = SimpleDocTemplate(
            file_path,
            pagesize=landscape(A4),
            leftMargin=15 * mm, rightMargin=15 * mm,
            topMargin=15 * mm, bottomMargin=15 * mm,
        )

        elements = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            "ReportTitle", parent=styles["Title"],
            fontName=font_name_bold, fontSize=18, textColor=colors.HexColor("#1a1a2e"),
            spaceAfter=6 * mm,
        )
        elements.append(Paragraph("Is_gain Test Report", title_style))

        if export_info:
            info_style = ParagraphStyle(
                "InfoStyle", parent=styles["Normal"],
                fontName=font_name, fontSize=10, textColor=colors.HexColor("#333333"),
                leading=16,
            )
            info_bold = ParagraphStyle(
                "InfoBold", parent=info_style,
                fontName=font_name_bold,
            )

            info_data = [
                [Paragraph("DUT Name", info_bold), Paragraph(export_info.get("dut_name", ""), info_style),
                 Paragraph("Module", info_bold), Paragraph(export_info.get("module", ""), info_style)],
                [Paragraph("Condition", info_bold), Paragraph(export_info.get("condition", ""), info_style),
                 Paragraph("Operator", info_bold), Paragraph(export_info.get("operator", ""), info_style)],
                [Paragraph("Test Time", info_bold), Paragraph(export_info.get("test_time", ""), info_style),
                 Paragraph("", info_style), Paragraph("", info_style)],
            ]
            if export_info.get("remark"):
                info_data.append([
                    Paragraph("Remark", info_bold), Paragraph(export_info["remark"], info_style),
                    Paragraph("", info_style), Paragraph("", info_style),
                ])

            info_table = Table(info_data, colWidths=[25 * mm, 60 * mm, 25 * mm, 60 * mm])
            info_table.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ("LINEBELOW", (0, -1), (-1, -1), 0.5, colors.HexColor("#cccccc")),
            ]))
            elements.append(info_table)
            elements.append(Spacer(1, 6 * mm))

        cell_style = ParagraphStyle(
            "CellStyle", parent=styles["Normal"],
            fontName=font_name, fontSize=8, textColor=colors.HexColor("#222222"),
            alignment=1, leading=11,
        )
        header_para_style = ParagraphStyle(
            "HeaderPara", parent=cell_style,
            fontName=font_name_bold, fontSize=9, textColor=colors.white,
        )

        headers = ["Step", "Load Current (A)", "Voltage (V)", "Ripple (mV)", "V_Drop (mV)", "Remark"]
        header_row = [Paragraph(h, header_para_style) for h in headers]
        table_data = [header_row]

        for r in self._test_result_data:
            rp = r.get("ripple")
            vd = r.get("v_drop")
            row = [
                Paragraph(str(r.get("step", "")), cell_style),
                Paragraph(f"{r.get('load_current', 0):.3f}", cell_style),
                Paragraph(f"{r['voltage']:.6f}" if r.get("voltage") is not None else "N/A", cell_style),
                Paragraph(f"{rp * 1000:.3f}" if rp is not None else "N/A", cell_style),
                Paragraph(f"{vd * 1000:.3f}" if vd is not None else "N/A", cell_style),
                Paragraph(r.get("remark", ""), cell_style),
            ]
            table_data.append(row)

        col_widths = [18 * mm, 35 * mm, 35 * mm, 30 * mm, 30 * mm, 40 * mm]
        table = Table(table_data, colWidths=col_widths, repeatRows=1)

        tbl_style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F4070")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#C0C0C0")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FA")]),
        ])
        table.setStyle(tbl_style)
        elements.append(table)

        doc.build(elements)

    def _on_test_error(self, err_msg):
        self.append_log(f"[ERROR] {err_msg}")
        self.set_system_status(f"Test error: {err_msg}", is_error=True)

    def _on_test_finished(self):
        self.set_test_running(False)
        self.set_system_status("Test completed")

    def _cleanup_test_thread(self):
        if self.test_thread is not None:
            self.test_thread.wait(5000)
            self.test_thread.deleteLater()
            self.test_thread = None
        if self._test_worker is not None:
            self._test_worker.deleteLater()
            self._test_worker = None

    def set_system_status(self, status, is_error=False):
        self.page_subtitle.setText(status)
        if is_error:
            self.page_subtitle.setObjectName("statusErr")
        elif "Connecting" in status or "Searching" in status or "Running" in status:
            self.page_subtitle.setObjectName("statusWarn")
        else:
            self.page_subtitle.setObjectName("pageSubtitle")

        self.page_subtitle.style().unpolish(self.page_subtitle)
        self.page_subtitle.style().polish(self.page_subtitle)
        self.page_subtitle.update()

    def update_instrument_info(self, instrument_info):
        pass

    def _run_instrument_task(self, task_func, on_finished, kwargs=None):
        if self._instr_thread is not None and self._instr_thread.isRunning():
            self.append_log("[WARN] Another instrument operation is in progress.")
            return

        self._instr_worker = _InstrumentWorker(task_func, kwargs)
        self._instr_thread = QThread()
        self._instr_worker.moveToThread(self._instr_thread)

        self._instr_thread.started.connect(self._instr_worker.run)
        self._instr_worker.log.connect(self.append_log)
        self._instr_worker.finished.connect(on_finished)
        self._instr_worker.finished.connect(self._instr_thread.quit)
        self._instr_thread.finished.connect(self._cleanup_instr_thread)

        self._instr_thread.start()

    def _cleanup_instr_thread(self):
        if self._instr_thread is not None:
            self._instr_thread.wait(5000)
            self._instr_thread.deleteLater()
            self._instr_thread = None
        if self._instr_worker is not None:
            self._instr_worker.deleteLater()
            self._instr_worker = None

    def _on_search(self):
        if self._n6705c_top and self._n6705c_top.is_connected_a:
            return
        self.set_system_status("Searching VISA resources...")
        self.append_log("[SYSTEM] Scanning VISA resources...")
        self.search_btn.setEnabled(False)
        self._run_instrument_task(self._search_devices_task, self._on_search_finished)

    def _search_devices_task(self):
        if self.rm is None:
            try:
                self.rm = pyvisa.ResourceManager()
            except Exception:
                self.rm = pyvisa.ResourceManager('@ni')

        self.available_devices = list(self.rm.list_resources()) or []
        n6705c_devices = []

        for dev in self.available_devices:
            try:
                instr = self.rm.open_resource(dev, timeout=1000)
                idn = instr.query('*IDN?').strip()
                instr.close()
                if "N6705C" in idn:
                    n6705c_devices.append(dev)
            except Exception:
                pass

        return {"devices": n6705c_devices}

    def _on_search_finished(self, result):
        self.search_btn.setEnabled(True)
        if "error" in result:
            self.set_system_status("Search failed", is_error=True)
            self.append_log(f"[ERROR] Search failed: {result['error']}")
            return

        n6705c_devices = result.get("devices", [])
        self.visa_resource_combo.setEnabled(True)
        self.visa_resource_combo.clear()

        if n6705c_devices:
            for dev in n6705c_devices:
                self.visa_resource_combo.addItem(dev)
            self.append_log(f"[SYSTEM] Found {len(n6705c_devices)} compatible N6705C device(s).")
            self.set_system_status(f"Found {len(n6705c_devices)} device(s)")
        else:
            self.visa_resource_combo.addItem("No N6705C device found")
            self.visa_resource_combo.setEnabled(False)
            self.set_system_status("No N6705C device found", is_error=True)
            self.append_log("[SYSTEM] No compatible N6705C instrument found.")

    def _on_scope_search(self):
        if self._mso64b_top and self._mso64b_top.is_connected:
            return
        self.set_system_status("Searching scope resources...")
        self.append_log("[SYSTEM] Scanning for oscilloscope resources (LAN & USB)...")
        self.scope_search_btn.setEnabled(False)
        self._run_instrument_task(self._search_scope_task, self._on_scope_search_finished)

    def _search_scope_task(self):
        if self.rm is None:
            try:
                self.rm = pyvisa.ResourceManager()
            except Exception:
                self.rm = pyvisa.ResourceManager('@ni')

        all_resources = list(self.rm.list_resources()) or []
        scope_devices = []

        for dev in all_resources:
            try:
                instr = self.rm.open_resource(dev, timeout=2000)
                idn = instr.query('*IDN?').strip()
                instr.close()
                if any(kw in idn.upper() for kw in ["MSO", "DSO", "SCOPE", "OSCILLOSCOPE", "DSOX", "MSOX"]):
                    scope_devices.append(dev)
            except Exception:
                pass

        return {"devices": scope_devices}

    def _on_scope_search_finished(self, result):
        self.scope_search_btn.setEnabled(True)
        if "error" in result:
            self.set_system_status("Scope search failed", is_error=True)
            self.append_log(f"[ERROR] Scope search failed: {result['error']}")
            return

        scope_devices = result.get("devices", [])
        self.scope_resource_combo.clear()

        if scope_devices:
            for dev in scope_devices:
                self.scope_resource_combo.addItem(dev)
            self.append_log(f"[SYSTEM] Found {len(scope_devices)} oscilloscope(s).")
            self.set_system_status(f"Found {len(scope_devices)} scope(s)")
        else:
            self.scope_resource_combo.addItem("USB0::0x0957::0x17A4::MY61500152::INSTR")
            self.set_system_status("No oscilloscope found", is_error=True)
            self.append_log("[SYSTEM] No oscilloscope found. Default resource restored.")

    def _on_connect_or_disconnect_n6705c(self):
        if self.is_connected:
            self._on_disconnect_n6705c()
        else:
            self._on_connect_n6705c()

    def _on_connect_or_disconnect_scope(self):
        if self.scope_connected:
            self._on_disconnect_scope()
        else:
            self._on_connect_scope()

    def _on_connect_n6705c(self):
        self.set_system_status("Connecting N6705C...")
        self.append_log("[SYSTEM] Attempting N6705C connection...")
        self.connect_btn.setEnabled(False)

        if DEBUG_MOCK:
            self.n6705c = MockN6705C()
            self.is_connected = True
            self._update_connect_button_state(self.connect_btn, True)
            self.search_btn.setEnabled(False)
            self.append_log("[DEBUG] Mock N6705C connected.")
            self.set_system_status("N6705C connected (Mock)")
            device_address = self.visa_resource_combo.currentText()
            if self._n6705c_top:
                self._n6705c_top.connect_a(device_address, self.n6705c)
            self.connection_status_changed.emit(True)
            self.connect_btn.setEnabled(True)
            return

        device_address = self.visa_resource_combo.currentText()
        self._run_instrument_task(
            self._connect_n6705c_task,
            self._on_connect_n6705c_finished,
            kwargs={"device_address": device_address},
        )

    def _connect_n6705c_task(self, device_address):
        n6705c = N6705C(device_address)
        idn = n6705c.instr.query("*IDN?").strip()
        return {"n6705c": n6705c, "idn": idn}

    def _on_connect_n6705c_finished(self, result):
        self.connect_btn.setEnabled(True)
        if "error" in result:
            self.set_system_status("N6705C connection failed", is_error=True)
            self.append_log(f"[ERROR] N6705C connection failed: {result['error']}")
            return

        idn = result.get("idn", "")
        if "N6705C" in idn:
            self.n6705c = result["n6705c"]
            self.is_connected = True
            self._update_connect_button_state(self.connect_btn, True)
            self.search_btn.setEnabled(False)
            self.append_log("[SYSTEM] N6705C connected.")
            self.append_log(f"[IDN] {idn}")
            self.set_system_status("N6705C connected")

            if self._n6705c_top:
                device_address = self.visa_resource_combo.currentText()
                self._n6705c_top.connect_a(device_address, self.n6705c)

            self.connection_status_changed.emit(True)
        else:
            self.set_system_status("Device mismatch", is_error=True)
            self.append_log("[ERROR] Connected device is not N6705C.")

    def _on_disconnect_n6705c(self):
        self.set_system_status("Disconnecting N6705C...")
        self.append_log("[SYSTEM] Disconnecting N6705C...")
        self.connect_btn.setEnabled(False)
        if self._n6705c_top:
            self._n6705c_top.disconnect_a()
            self.n6705c = None
            self._on_disconnect_n6705c_finished({})
        else:
            n6705c_ref = self.n6705c
            self.n6705c = None
            self._run_instrument_task(
                self._disconnect_n6705c_task,
                self._on_disconnect_n6705c_finished,
                kwargs={"n6705c_ref": n6705c_ref},
            )

    def _disconnect_n6705c_task(self, n6705c_ref):
        if n6705c_ref is not None:
            if hasattr(n6705c_ref, 'instr') and n6705c_ref.instr:
                n6705c_ref.instr.close()
            if hasattr(n6705c_ref, 'rm') and n6705c_ref.rm:
                n6705c_ref.rm.close()
        return {}

    def _on_disconnect_n6705c_finished(self, result):
        self.connect_btn.setEnabled(True)
        if "error" in result:
            self.set_system_status("N6705C disconnect failed", is_error=True)
            self.append_log(f"[ERROR] N6705C disconnect failed: {result['error']}")
            return

        self.is_connected = False
        self._update_connect_button_state(self.connect_btn, False)
        self.search_btn.setEnabled(True)
        self.append_log("[SYSTEM] N6705C disconnected.")
        self.set_system_status("N6705C disconnected")
        self.connection_status_changed.emit(False)

    def _on_connect_scope(self):
        scope_type = self.scope_type_combo.currentText()
        resource = self.scope_resource_combo.currentText().strip()
        if not resource:
            self.set_system_status("Invalid scope resource", is_error=True)
            self.append_log("[ERROR] Invalid scope resource.")
            return

        if DEBUG_MOCK:
            self.Osc_ins = MockMSO64B()
            self.is_scope_connected = True
            self._update_connect_button_state(self.scope_connect_btn, True)
            self.scope_search_btn.setEnabled(False)
            self.append_log("[DEBUG] Mock scope connected.")
            self.set_system_status("Scope connected (Mock)")
            if self._mso64b_top:
                self._mso64b_top.connect_instrument(resource, self.Osc_ins, scope_type="MSO64B")
            self.scope_connection_changed.emit(True)
            return

        self.set_system_status(f"Connecting {scope_type}...")
        self.append_log(f"[SYSTEM] Attempting {scope_type} connection...")
        self.scope_connect_btn.setEnabled(False)
        self._run_instrument_task(
            self._connect_scope_task,
            self._on_connect_scope_finished,
            kwargs={"scope_type": scope_type, "resource": resource},
        )

    def _connect_scope_task(self, scope_type, resource):
        if scope_type == "MSO64B":
            osc = MSO64B(resource)
        elif scope_type == "DSOX4034A":
            osc = DSOX4034A(resource)
        else:
            return {"error": f"Unknown scope type: {scope_type}"}

        idn = osc.identify_instrument()
        return {"osc": osc, "idn": idn, "resource": resource, "scope_type": scope_type}

    def _on_connect_scope_finished(self, result):
        self.scope_connect_btn.setEnabled(True)
        if "error" in result:
            scope_type = result.get("scope_type", self.scope_type_combo.currentText())
            self.Osc_ins = None
            self.set_system_status(f"{scope_type} connection failed", is_error=True)
            self.append_log(f"[ERROR] {scope_type} connection failed: {result['error']}")
            return

        scope_type = result["scope_type"]
        self.Osc_ins = result["osc"]
        self.scope_resource = result["resource"]
        self.scope_connected = True
        self._update_connect_button_state(self.scope_connect_btn, True)
        self.scope_type_combo.setEnabled(False)
        self.scope_search_btn.setEnabled(False)
        self.append_log(f"[SYSTEM] {scope_type} connected.")
        self.append_log(f"[IDN] {result['idn']}")
        self.set_system_status(f"{scope_type} connected")

        if self._mso64b_top:
            self._mso64b_top.connect_instrument(result["resource"], self.Osc_ins, scope_type=scope_type)

    def _on_disconnect_scope(self):
        scope_type = self.scope_type_combo.currentText()
        self.set_system_status(f"Disconnecting {scope_type}...")
        self.append_log(f"[SYSTEM] Disconnecting {scope_type}...")
        self.scope_connect_btn.setEnabled(False)

        if self._mso64b_top and self._mso64b_top.is_connected:
            self._mso64b_top.disconnect()
            self.Osc_ins = None
            self._on_disconnect_scope_finished({"scope_type": scope_type})
        else:
            osc_ref = self.Osc_ins
            self.Osc_ins = None
            self._run_instrument_task(
                self._disconnect_scope_task,
                self._on_disconnect_scope_finished,
                kwargs={"osc_ref": osc_ref, "scope_type": scope_type},
            )

    def _disconnect_scope_task(self, osc_ref, scope_type):
        if osc_ref is not None:
            if hasattr(osc_ref, 'disconnect'):
                osc_ref.disconnect()
            elif hasattr(osc_ref, 'instrument') and osc_ref.instrument:
                osc_ref.instrument.close()
        return {"scope_type": scope_type}

    def _on_disconnect_scope_finished(self, result):
        self.scope_connect_btn.setEnabled(True)
        scope_type = result.get("scope_type", self.scope_type_combo.currentText())
        if "error" in result:
            self.set_system_status(f"{scope_type} disconnect failed", is_error=True)
            self.append_log(f"[ERROR] {scope_type} disconnect failed: {result['error']}")
            return

        self.scope_resource = None
        self.scope_connected = False
        self._update_connect_button_state(self.scope_connect_btn, False)
        self.scope_type_combo.setEnabled(True)
        self.scope_search_btn.setEnabled(True)
        self.append_log(f"[SYSTEM] {scope_type} disconnected.")
        self.set_system_status(f"{scope_type} disconnected")

    def _on_start_or_abort_clicked(self):
        if self.is_test_running:
            if self._test_worker is not None:
                self._test_worker.stop()
            self.append_log("[TEST] Abort requested by user.")
            return

        if not self.is_connected or self.n6705c is None:
            self.set_system_status("Please connect N6705C first", is_error=True)
            self.append_log("[ERROR] N6705C not connected.")
            return

        if not self.scope_connected or self.Osc_ins is None:
            self.set_system_status("Please connect Oscilloscope first", is_error=True)
            self.append_log("[ERROR] Oscilloscope not connected.")
            return

        config = self.get_test_config()

        if abs(config["is_gain_step_current"]) < 1e-9:
            self.set_system_status("Step Current must be > 0", is_error=True)
            self.append_log("[ERROR] Step Current must be greater than 0.")
            return

        selected_test = self.test_selection_combo.currentText()

        if selected_test == "单次 Is_gain 测试":
            self._start_is_gain_test(config)
        elif selected_test == "遍历 Is_gain 测试":
            self._start_traverse_is_gain_test(config)

    def _start_is_gain_test(self, config):
        self._launch_test_thread(config, _IsGainTestWorker.MODE_SINGLE, "Running single Is_gain test...")

    def _start_traverse_is_gain_test(self, config):
        self._launch_test_thread(config, _IsGainTestWorker.MODE_TRAVERSE, "Running traverse Is_gain test...")

    def _launch_test_thread(self, config, test_mode, status_msg):
        self.clear_results()
        self.set_test_running(True)
        self.set_system_status(status_msg)

        self._test_worker = _IsGainTestWorker(self.n6705c, self.Osc_ins, config, test_mode=test_mode)
        self.test_thread = QThread()
        self._test_worker.moveToThread(self.test_thread)

        self.test_thread.started.connect(self._test_worker.run)
        self._test_worker.log.connect(self.append_log)
        self._test_worker.progress.connect(self.set_progress)
        self._test_worker.result_row.connect(self._on_test_result_row)
        self._test_worker.summary.connect(self._on_test_summary)
        self._test_worker.error.connect(self._on_test_error)
        self._test_worker.finished.connect(self._on_test_finished)
        self._test_worker.finished.connect(self.test_thread.quit)
        self.test_thread.finished.connect(self._cleanup_test_thread)

        self.test_thread.start()

    def get_n6705c_instance(self):
        return self.n6705c

    def is_n6705c_connected(self):
        return self.is_connected

    def get_scope_instance(self):
        return self.Osc_ins

    def is_scope_connected(self):
        return self.scope_connected
